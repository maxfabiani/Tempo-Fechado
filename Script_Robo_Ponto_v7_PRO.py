
import csv
import json
import logging
import re
import shutil
import sys
from collections import defaultdict
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Optional

import pdfplumber
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill

try:
    import win32com.client as win32
except ImportError:
    win32 = None


# ============================================================
# PAINEL RÁPIDO DE CONTROLE - V6 PRO / Consolidado Oficial + Histórico
# ============================================================

MODO_EXECUCAO = "PRODUCAO"          # "PRODUCAO" | "TESTE"
ENVIAR_EMAIL = True                # True | False
EXIBIR_EMAIL_ANTES_DE_ENVIAR = True
MODO_DATA = "d-1-util"             # "ultima" | "penultima" | "d-1-util" | "manual"
DATA_MANUAL = ""                   # Ex.: "14/04/2026" quando MODO_DATA = "manual"
LOG_NIVEL = "INFO"                 # "INFO" | "DEBUG"
MOVER_ARQUIVOS_PROCESSADOS = True  # Em TESTE, o ideal é False

# Alertas / gestão
LIMITE_HE_ALERTA_MINUTOS = 120     # 120 = 02:00
TOP_N_COLABORADORES_HE = 15
TOP_N_DIAS_HE = 15


@dataclass(frozen=True)
class Config:
    pasta_raiz: Path = Path.home() / "ponto_pdfs"
    pasta_entrada: Path = pasta_raiz / "entrada"
    pasta_processados: Path = pasta_raiz / "processados"
    pasta_erro: Path = pasta_raiz / "erro"
    pasta_saida: Path = pasta_raiz / "saida"
    pasta_controle: Path = pasta_raiz / "controle"

    arquivo_log: Path = pasta_controle / "robo_ponto.log"
    arquivo_lock: Path = pasta_controle / "robo_ponto.lock"
    arquivo_resultado: Path = pasta_controle / "robo_ponto_resultado.json"
    arquivo_erros: Path = pasta_controle / "erros_processamento.txt"
    arquivo_historico_execucoes: Path = pasta_controle / "historico_execucoes.csv"

    caminho_saida_excel: Path = pasta_saida / "resultado_lote_consultas_validado.xlsx"

    enviar_email_automaticamente: bool = ENVIAR_EMAIL
    exibir_email_antes_de_enviar: bool = EXIBIR_EMAIL_ANTES_DE_ENVIAR
    destinatarios_email: tuple[str, ...] = ()
    cc_email: tuple[str, ...] = ()
    cco_email: tuple[str, ...] = ()
    assunto_email_base: str = "Arquivo consolidado"

    modo_data_arquivo: str = MODO_DATA
    data_manual_arquivo: str = DATA_MANUAL

    modo_execucao: str = MODO_EXECUCAO
    log_nivel: str = LOG_NIVEL
    mover_arquivos_processados: bool = MOVER_ARQUIVOS_PROCESSADOS

    limite_he_alerta_minutos: int = LIMITE_HE_ALERTA_MINUTOS
    top_n_colaboradores_he: int = TOP_N_COLABORADORES_HE
    top_n_dias_he: int = TOP_N_DIAS_HE

    feriados_fixos: tuple[str, ...] = (
        "01/01",
        "21/04",
        "01/05",
        "07/09",
        "12/10",
        "02/11",
        "15/11",
        "20/11",
        "25/12",
    )

    feriados_especificos: tuple[str, ...] = (
        # "09/07/2026",
        # "20/01/2026",
        # "12/02/2026",
    )


CONFIG = Config()

CONFIG_EMAIL_ROBO_JSON_V81727 = Path.home() / "ponto_pdfs" / "config" / "config_email_robo.json"

def _lista_emails_v81727(valor):
    if isinstance(valor, (list, tuple, set)):
        base = valor
    else:
        base = str(valor or "").replace(";", ",").split(",")
    return tuple(str(x).strip() for x in base if str(x).strip())

def carregar_config_email_robo_v81727(config_atual):
    try:
        if not CONFIG_EMAIL_ROBO_JSON_V81727.exists():
            return config_atual
        with open(CONFIG_EMAIL_ROBO_JSON_V81727, "r", encoding="utf-8") as f:
            dados = json.load(f) or {}
        overrides = {}
        if "enviar_email_automaticamente" in dados:
            overrides["enviar_email_automaticamente"] = bool(dados.get("enviar_email_automaticamente"))
        if "exibir_email_antes_de_enviar" in dados:
            overrides["exibir_email_antes_de_enviar"] = bool(dados.get("exibir_email_antes_de_enviar"))
        if "destinatarios_email" in dados:
            overrides["destinatarios_email"] = _lista_emails_v81727(dados.get("destinatarios_email"))
        if "cc_email" in dados:
            overrides["cc_email"] = _lista_emails_v81727(dados.get("cc_email"))
        if "cco_email" in dados:
            overrides["cco_email"] = _lista_emails_v81727(dados.get("cco_email"))
        if "assunto_email_base" in dados:
            overrides["assunto_email_base"] = str(dados.get("assunto_email_base") or "Arquivo consolidado").strip() or "Arquivo consolidado"
        if not overrides:
            return config_atual
        from dataclasses import replace
        return replace(config_atual, **overrides)
    except Exception as e:
        try:
            print(f"[TempoFechado] Falha ao carregar config_email_robo.json: {e}")
        except Exception:
            pass
        return config_atual

CONFIG = carregar_config_email_robo_v81727(CONFIG)



COLUNAS_BASE = [
    "Arquivo Origem", "Matrícula", "Nome", "cc", "Função", "Data", "Dia",
    "1a E.", "1a S.", "2a E.", "2a S.", "3a E.", "3a S.", "4a E.", "4a S.",
    "Abono", "H.E.", "Absent.", "Jornada", "Ad. Not.", "Observação",
    "Saldo Atual", "Saldo Calculado", "Saldo Informado PDF", "Data Emissão PDF", "Saldo Anterior PDF", "Debito PDF", "Credito PDF",
    "Turno",
]

COLUNAS_SAIDA = [
    "cc", "Nome", "Função", "Turno", "Data", "Dia",
    "1a E.", "1a S.", "2a E.", "2a S.", "3a E.", "3a S.", "4a E.", "4a S.",
    "Abono", "H.E.", "Absent.", "Jornada", "Ad. Not.", "Observação",
    "Saldo Atual", "Saldo Calculado", "Saldo Informado PDF", "Data Emissão PDF", "Saldo Anterior PDF", "Debito PDF", "Credito PDF",
]

COLUNAS_EXPORTACAO = COLUNAS_SAIDA + [
    "Status Validação", "Inconsistências", "Jornada Calculada", "Qtd Batidas",
]

COLUNAS_PDF = [
    "cc", "CC", "Nome", "Função", "Turno", "Data", "Dia",
    "1a E.", "1a S.", "2a E.", "2a S.", "3a E.", "3a S.", "4a E.", "4a S.",
    "Abono", "H.E.", "Absent.", "Jornada", "Ad. Not.", "Observação",
    "Saldo Atual", "Saldo Calculado", "Saldo Informado PDF", "Data Emissão PDF", "Saldo Anterior PDF", "Debito PDF", "Credito PDF",
]

COLUNAS_CALCULADAS = [
    "Status Validação", "Inconsistências", "Jornada Calculada", "Qtd Batidas",
]

MAPEAMENTO_TURNOS = {
    "003 - 12X36 - 06:00": "12x36M6",
    "005 - SEG A QUI - 07:00": "Comercial7",
    "011 - 12X36 - 07:00": "12x36M7",
    "001 - SEG A SEX - 08:00": "Comercial8",
    "120 - 12X36 - 08:00": "12x36M8",
    "041 - 12X36 - 19:00": "12x36N19",
    "012 - 12X36 - 19:00": "12x36N19.1",
    "034 - SEG A SEX - 21:00": "Noturno1",
    "047 - DOM A QUI - 21:00": "Noturno2",
}

RENOMEAR_COLUNAS_EXCEL = {"cc": "CC"}

LABEL_SPECS_TABELA = [
    ("Data", 1),
    ("Dia", 1),
    ("1a E.", 2),
    ("1a S.", 2),
    ("2a E.", 2),
    ("2a S.", 2),
    ("3a E.", 2),
    ("3a S.", 2),
    ("4a E.", 2),
    ("4a S.", 2),
    ("Abono", 1),
    ("H.E.", 1),
    ("Absent.", 1),
    ("Jornada", 1),
    ("Ad. Not.", 2),
    ("Observação", 1),
]


def garantir_estrutura() -> None:
    CONFIG.pasta_entrada.mkdir(parents=True, exist_ok=True)
    CONFIG.pasta_processados.mkdir(parents=True, exist_ok=True)
    CONFIG.pasta_erro.mkdir(parents=True, exist_ok=True)
    CONFIG.pasta_saida.mkdir(parents=True, exist_ok=True)
    CONFIG.pasta_controle.mkdir(parents=True, exist_ok=True)


def configurar_logger() -> logging.Logger:
    garantir_estrutura()

    logger = logging.getLogger("robo_ponto")
    nivel = getattr(logging, str(CONFIG.log_nivel).upper(), logging.INFO)
    logger.setLevel(nivel)
    logger.handlers.clear()

    formatter = logging.Formatter(
        "[%(asctime)s] [%(levelname)s] %(message)s",
        "%Y-%m-%d %H:%M:%S",
    )

    file_handler = RotatingFileHandler(
        CONFIG.arquivo_log,
        maxBytes=2_000_000,
        backupCount=5,
        encoding="utf-8",
    )
    file_handler.setFormatter(formatter)

    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    return logger


LOGGER = configurar_logger()


@contextmanager
def lock_execucao():
    if CONFIG.arquivo_lock.exists():
        raise RuntimeError(f"Robô de ponto já está em execução. Lock: {CONFIG.arquivo_lock}")

    CONFIG.arquivo_lock.write_text(
        f"pid={Path(sys.executable).name} | inicio={datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n",
        encoding="utf-8",
    )
    try:
        yield
    finally:
        try:
            if CONFIG.arquivo_lock.exists():
                CONFIG.arquivo_lock.unlink()
        except Exception as e:
            LOGGER.warning(f"Falha ao remover lock file: {e}")


def validar_configuracao() -> None:
    if str(CONFIG.modo_execucao).upper() not in {"PRODUCAO", "TESTE"}:
        raise ValueError("MODO_EXECUCAO inválido. Use 'PRODUCAO' ou 'TESTE'.")
    if str(CONFIG.modo_data_arquivo).strip().lower() not in {"ultima", "penultima", "d-1-util", "manual"}:
        raise ValueError("MODO_DATA inválido. Use 'ultima', 'penultima', 'd-1-util' ou 'manual'.")
    if str(CONFIG.log_nivel).upper() not in {"INFO", "DEBUG"}:
        raise ValueError("LOG_NIVEL inválido. Use 'INFO' ou 'DEBUG'.")
    if str(CONFIG.modo_data_arquivo).strip().lower() == "manual" and not str(CONFIG.data_manual_arquivo).strip():
        raise ValueError("Quando MODO_DATA = 'manual', informe DATA_MANUAL no formato dd/mm/aaaa.")
    if int(CONFIG.limite_he_alerta_minutos) < 0:
        raise ValueError("LIMITE_HE_ALERTA_MINUTOS não pode ser negativo.")


def modo_teste_ativo() -> bool:
    return str(CONFIG.modo_execucao).upper() == "TESTE"


def salvar_resultado_execucao(
    *,
    status: str,
    total_arquivos_pdf: int,
    total_registros: int,
    total_erros: int,
    caminho_excel: str = "",
    caminho_excel_data_referencia: str = "",
    data_referencia: str = "",
    modo_data: str = "",
    detalhe: str = "",
    erros: Optional[list] = None,
) -> None:
    payload = {
        "status": status,
        "total_arquivos_pdf": total_arquivos_pdf,
        "total_registros": total_registros,
        "total_erros": total_erros,
        "caminho_excel": caminho_excel,
        "caminho_excel_data_referencia": caminho_excel_data_referencia,
        "data_referencia": data_referencia,
        "modo_data": modo_data,
        "modo_execucao": CONFIG.modo_execucao,
        "detalhe": detalhe,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "erros": erros or [],
    }
    with open(CONFIG.arquivo_resultado, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def registrar_historico_execucao(
    *,
    status: str,
    total_arquivos_pdf: int,
    total_registros: int,
    total_erros: int,
    data_referencia: str,
    modo_data: str,
    caminho_excel: str,
) -> None:
    cabecalho = [
        "timestamp", "modo_execucao", "status", "total_arquivos_pdf", "total_registros",
        "total_erros", "data_referencia", "modo_data", "caminho_excel",
    ]
    linha = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "modo_execucao": CONFIG.modo_execucao,
        "status": status,
        "total_arquivos_pdf": total_arquivos_pdf,
        "total_registros": total_registros,
        "total_erros": total_erros,
        "data_referencia": data_referencia,
        "modo_data": modo_data,
        "caminho_excel": caminho_excel,
    }
    arquivo = CONFIG.arquivo_historico_execucoes
    novo = not arquivo.exists()
    with open(arquivo, "a", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=cabecalho)
        if novo:
            writer.writeheader()
        writer.writerow(linha)


def mover_arquivo(origem: Path, destino_dir: Path) -> Path:
    destino_dir.mkdir(parents=True, exist_ok=True)
    destino = destino_dir / origem.name
    if destino.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        destino = destino_dir / f"{origem.stem}_{timestamp}{origem.suffix}"
    shutil.move(str(origem), str(destino))
    return destino


def normalizar_texto(texto):
    if not texto:
        return ""
    texto = str(texto).upper().strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def sanitizar_nome_aba(nome):
    if not nome or not str(nome).strip():
        return "Sem Turno"
    nome = str(nome).strip()
    nome = re.sub(r'[\\/*?:\[\]]', "-", nome)
    nome = re.sub(r"\s+", " ", nome).strip()
    return nome[:31]


def hhmm_para_minutos(valor):
    valor = str(valor).strip()
    if not valor or valor.lower() == "nan":
        return None
    m = re.fullmatch(r"(-?)(\d{2}):(\d{2})", valor)
    if not m:
        return None
    sinal = -1 if m.group(1) == "-" else 1
    horas = int(m.group(2))
    minutos = int(m.group(3))
    return sinal * (horas * 60 + minutos)


def minutos_para_hhmm(total_minutos):
    if total_minutos is None:
        return ""

    try:
        if pd.isna(total_minutos):
            return ""
    except Exception:
        pass

    try:
        total_minutos = int(round(float(total_minutos)))
    except (TypeError, ValueError):
        return ""

    sinal = "-" if total_minutos < 0 else ""
    total_minutos = abs(total_minutos)
    horas = total_minutos // 60
    minutos = total_minutos % 60
    return f"{sinal}{horas:02d}:{minutos:02d}"


def converter_saldo_para_decimal(valor):
    """
    Converte saldo de banco de horas para número decimal apenas para análises
    auxiliares (ex.: saldo negativo).

    Aceita os formatos mais comuns do PDF:
    - "-41,29"  => -41h29
    - "-41:29"  => -41h29
    - "20,22"   => 20h22
    """
    minutos = saldo_banco_para_minutos(valor)
    if minutos is None:
        return None
    return minutos / 60


def saldo_banco_para_minutos(valor):
   
    valor = str(valor).strip()
    if not valor or valor.lower() == "nan":
        return None

    valor = valor.replace(".", "")
    m = re.fullmatch(r"(-?)(\d{1,5})[:,](\d{2})", valor)
    if not m:
        return None

    sinal = -1 if m.group(1) == "-" else 1
    horas = int(m.group(2))
    minutos = int(m.group(3))
    return sinal * (horas * 60 + minutos)


def minutos_para_saldo_banco(total_minutos):
    
    if total_minutos is None:
        return ""

    try:
        if pd.isna(total_minutos):
            return ""
    except Exception:
        pass

    try:
        total_minutos = int(round(float(total_minutos)))
    except (TypeError, ValueError):
        return ""

    sinal = "-" if total_minutos < 0 else ""
    total_minutos = abs(total_minutos)
    horas = total_minutos // 60
    minutos = total_minutos % 60
    return f"{sinal}{horas:02d},{minutos:02d}"


def limpar_cc(valor):
    valor = str(valor).strip()
    if not valor or valor.lower() == "nan":
        return ""
    return valor.split("-", 1)[0].strip()


def limpar_funcao(valor):
    valor = str(valor).strip()
    if not valor or valor.lower() == "nan":
        return ""
    valor = re.sub(r"^\s*\d+\s*-\s*", "", valor).strip()
    return valor


def normalizar_turno(turno_bruto):
    if not turno_bruto or not str(turno_bruto).strip():
        return "Sem Turno"
    turno_norm = normalizar_texto(turno_bruto)
    for chave, nome_amigavel in MAPEAMENTO_TURNOS.items():
        if normalizar_texto(chave) in turno_norm:
            return nome_amigavel
    codigo = re.search(r"\b(\d{3})\b", turno_norm)
    hora = re.search(r"\b(\d{2}:\d{2})\b", turno_norm)
    if codigo and hora:
        codigo_val = codigo.group(1)
        hora_val = hora.group(1)
        for chave, nome_amigavel in MAPEAMENTO_TURNOS.items():
            chave_norm = normalizar_texto(chave)
            if codigo_val in chave_norm and hora_val in chave_norm:
                return nome_amigavel
    return turno_bruto.strip()


def preparar_para_exportacao_excel(df):
    return df.copy().rename(columns=RENOMEAR_COLUNAS_EXCEL)


def exportar_log_erros(erros, caminho_saida_txt=None):
    caminho_saida_txt = Path(caminho_saida_txt or CONFIG.arquivo_erros)
    if not erros:
        if caminho_saida_txt.exists():
            caminho_saida_txt.unlink()
        return
    with open(caminho_saida_txt, "w", encoding="utf-8") as f:
        for arquivo, erro in erros:
            f.write(f"{arquivo} -> {erro}\n")


def extrair_texto_paginas(caminho_pdf):
    paginas = []
    with pdfplumber.open(caminho_pdf) as pdf:
        for idx, pagina in enumerate(pdf.pages, start=1):
            texto = pagina.extract_text() or ""
            palavras = pagina.extract_words() or []
            paginas.append({"numero_pagina": idx, "texto": texto, "palavras": palavras})
    return paginas


def extrair_matricula(texto):
    m = re.search(r"Matrícula:\s*\d+\s*-\s*(\d+)", texto)
    return m.group(1).strip() if m else ""


def extrair_nome(texto):
    m = re.search(r"Nome:\s*(.+?)\s+Chapa:", texto)
    return m.group(1).strip() if m else ""


def extrair_cabecalho(texto):
    matricula = extrair_matricula(texto)
    nome = extrair_nome(texto)
    cc = re.search(r"C\.C\.\:\s*(.+?)\s+CPF:", texto)
    funcao = re.search(r"Função:\s*(.+?)\s+C\.C\.", texto)
    return {
        "Matrícula": matricula,
        "Nome": nome,
        "cc": cc.group(1).strip() if cc else "",
        "Função": funcao.group(1).strip() if funcao else "",
    }


def chave_colaborador(texto, numero_pagina):
    matricula = extrair_matricula(texto)
    nome = extrair_nome(texto)
    if matricula or nome:
        return matricula, nome
    return f"SEM_MATRICULA_PAG_{numero_pagina}", f"SEM_NOME_PAG_{numero_pagina}"


def agrupar_paginas_por_colaborador(caminho_pdf):
    paginas = extrair_texto_paginas(caminho_pdf)
    grupos = []
    grupo_atual = None
    chave_atual = None
    for pagina in paginas:
        texto = pagina["texto"]
        numero = pagina["numero_pagina"]
        chave = chave_colaborador(texto, numero)
        if grupo_atual is None:
            grupo_atual = {"chave": chave, "paginas": [pagina]}
            chave_atual = chave
            continue
        if chave == chave_atual:
            grupo_atual["paginas"].append(pagina)
        else:
            grupos.append(grupo_atual)
            grupo_atual = {"chave": chave, "paginas": [pagina]}
            chave_atual = chave
    if grupo_atual is not None:
        grupos.append(grupo_atual)
    return grupos


def extrair_data_emissao_pdf(texto_completo):
   
    if not texto_completo:
        return ""
    m = re.search(r"Emiss(?:ã|a)o:\s*(\d{2}/\d{2}/\d{4})", str(texto_completo), flags=re.IGNORECASE)
    return m.group(1).strip() if m else ""


def extrair_resumo_banco_horas(texto_completo):
   
    linhas = [linha.strip() for linha in texto_completo.split("\n") if linha.strip()]
    resumo = {
        "Saldo Anterior": "",
        "Débito": "",
        "Crédito": "",
        "Saldo Atual": "",
    }

    for i, linha in enumerate(linhas):
        if linha.startswith("Banco de Horas"):
            for j in range(i + 1, min(i + 6, len(linhas))):
                valores = re.findall(r"-?\d+,\d{2}", linhas[j])
                if len(valores) >= 4:
                    resumo["Saldo Anterior"] = valores[0]
                    resumo["Débito"] = valores[1]
                    resumo["Crédito"] = valores[2]
                    resumo["Saldo Atual"] = valores[3]
                    return resumo
    return resumo


def extrair_saldo_atual(texto_completo):
    
    return extrair_resumo_banco_horas(texto_completo).get("Saldo Atual", "")



def calcular_saldo_atual_pdf_apurado(saldo_anterior_pdf, debito_pdf, credito_pdf, saldo_atual_pdf=""):
    
    saldo_anterior_min = saldo_banco_para_minutos(saldo_anterior_pdf)
    if saldo_anterior_min is None:
        return str(saldo_atual_pdf or "").strip()

    debito_min = saldo_banco_para_minutos(debito_pdf)
    credito_min = saldo_banco_para_minutos(credito_pdf)

    if debito_min is None:
        debito_min = 0
    if credito_min is None:
        credito_min = 0

    saldo_atual_min = saldo_anterior_min - debito_min + credito_min
    return minutos_para_saldo_banco(saldo_atual_min)


def obter_chave_ciclo_medicao(data_valor):
   
    try:
        data = pd.to_datetime(str(data_valor).strip(), format="%d/%m/%Y", errors="coerce")
    except Exception:
        return "SEM_CICLO"

    if pd.isna(data):
        return "SEM_CICLO"

    ano = int(data.year)
    mes = int(data.month)

    if int(data.day) >= 15:
        return f"{ano:04d}-{mes:02d}-15"

    # Dias 01 a 14 pertencem ao ciclo iniciado no mês anterior.
    if mes == 1:
        ano -= 1
        mes = 12
    else:
        mes -= 1
    return f"{ano:04d}-{mes:02d}-15"


def linha_neutra_banco_horas(linha):
   
    observacao = normalizar_texto(linha.get("Observação", ""))
    data_txt = str(linha.get("Data", "") or "").strip()

    if "D.S.R" in observacao or "DSR" in observacao or "D S R" in observacao:
        return True
    if "COMPENSADO" in observacao:
        return True
    if "FERIADO" in observacao or "SEXTA-FEIRA SANTA" in observacao:
        return True

    try:
        data = pd.to_datetime(data_txt, format="%d/%m/%Y", errors="coerce")
    except Exception:
        data = None

    if data is not None and not pd.isna(data):
        data_fmt = data.strftime("%d/%m/%Y")
        dia_mes = data.strftime("%d/%m")
        if dia_mes in set(CONFIG.feriados_fixos):
            return True
        if data_fmt in set(CONFIG.feriados_especificos):
            return True

    return False


def _primeiro_saldo_valido(valores):
    
    for valor in valores:
        texto = str(valor or "").strip()
        if not texto or texto.lower() == "nan":
            continue
        minutos = saldo_banco_para_minutos(texto)
        if minutos is not None:
            return texto, minutos
    return "", None


def _primeiro_tempo_minutos(valores):
    
    for valor in valores:
        minutos = hhmm_para_minutos(valor)
        if minutos is not None:
            return minutos
    return 0


def _primeiro_tempo_linha(linha, colunas):
   
    for coluna in colunas:
        minutos = hhmm_para_minutos(linha.get(coluna, ""))
        if minutos is not None:
            return minutos
    return 0


def observacao_indica_banco_horas(linha):
   
    observacao = normalizar_texto(linha.get("Observação", ""))
    return (
        "BANCO DE HORAS" in observacao
        or "BANCO HORAS" in observacao
        or "B.H" in observacao
        or "BH" == observacao.strip()
    )


def observacao_indica_ausencia(linha):
    observacao = normalizar_texto(linha.get("Observação", ""))
    return "AUSENTE" in observacao or "AUSENCIA" in observacao or "AUSÊNCIA" in observacao


def movimento_banco_horas_linha(linha):
    
    if linha_neutra_banco_horas(linha):
        return 0, 0

    if observacao_indica_banco_horas(linha):
        debito = _primeiro_tempo_linha(linha, ["Absent.", "Jornada", "H.E.", "Abono"])
        return 0, debito

    if observacao_indica_ausencia(linha):
        debito = _primeiro_tempo_linha(linha, ["Absent.", "Jornada", "H.E.", "Abono"])
        return 0, debito

    credito = hhmm_para_minutos(linha.get("H.E.", "")) or 0
    debito = hhmm_para_minutos(linha.get("Absent.", "")) or 0
    return credito, debito


def movimento_banco_horas_do_dia(linhas_data):
    
    vistos = set()
    credito_total = 0
    debito_total = 0

    for _, linha in linhas_data.iterrows():
        if linha_neutra_banco_horas(linha):
            continue

        assinatura = (
            str(linha.get("Data", "")).strip(),
            str(linha.get("H.E.", "")).strip(),
            str(linha.get("Absent.", "")).strip(),
            str(linha.get("Jornada", "")).strip(),
            str(linha.get("Abono", "")).strip(),
            normalizar_texto(linha.get("Observação", "")),
        )
        if assinatura in vistos:
            continue
        vistos.add(assinatura)

        credito, debito = movimento_banco_horas_linha(linha)
        credito_total += credito
        debito_total += debito

    return credito_total, debito_total


def aplicar_saldo_calculado_por_ciclo(linhas_tabela, resumo_banco):
   
    for linha in linhas_tabela:
        linha.setdefault("Saldo Calculado", "")
    return linhas_tabela

def aplicar_evolucao_banco_horas(linhas_tabela, resumo_banco, data_emissao_pdf=""):
    
    saldo_anterior_pdf = str(resumo_banco.get("Saldo Anterior", "") or "").strip()
    debito_pdf = str(resumo_banco.get("Débito", "") or "").strip()
    credito_pdf = str(resumo_banco.get("Crédito", "") or "").strip()
    saldo_atual_pdf_original = str(resumo_banco.get("Saldo Atual", "") or "").strip()

   
    saldo_atual_pdf = saldo_atual_pdf_original or calcular_saldo_atual_pdf_apurado(
        saldo_anterior_pdf,
        debito_pdf,
        credito_pdf,
        saldo_atual_pdf_original,
    )

    data_emissao_pdf = str(data_emissao_pdf or "").strip()

    resultado = []
    for linha in linhas_tabela:
        nova = linha.copy()
        data_linha = str(nova.get("Data", "") or "").strip()
        preencher_fotografia_nesta_linha = bool(saldo_atual_pdf and data_emissao_pdf and data_linha == data_emissao_pdf)

        nova["Data Emissão PDF"] = data_emissao_pdf
        nova["Saldo Anterior PDF"] = saldo_anterior_pdf
        nova["Debito PDF"] = debito_pdf
        nova["Credito PDF"] = credito_pdf

        
        nova["Saldo Atual"] = saldo_atual_pdf

       
        nova["Saldo Informado PDF"] = saldo_atual_pdf if preencher_fotografia_nesta_linha else ""

        resultado.append(nova)

    resultado = aplicar_saldo_calculado_por_ciclo(resultado, resumo_banco)
    return resultado

def extrair_turnos_do_bloco(texto_completo):
    linhas = [linha.strip() for linha in texto_completo.split("\n") if linha.strip()]
    dentro_horarios = False
    turnos_encontrados = []
    for linha in linhas:
        if linha.startswith("Horários"):
            dentro_horarios = True
            continue
        if not dentro_horarios:
            continue
        if linha.startswith("Data 1a E."):
            continue
        if "Assinatura do Funcionário" in linha:
            continue
        m = re.match(r"^\d{2}/\d{2}/\d{4}\s+(.+?)\s+(?:\d{2}:\d{2}\s*){1,8}$", linha)
        if m:
            turno = m.group(1).strip()
            if turno and turno not in turnos_encontrados:
                turnos_encontrados.append(turno)
            continue
        m2 = re.match(r"^\d{2}/\d{2}/\d{4}\s+(.+)$", linha)
        if m2:
            valor = m2.group(1).strip()
            valor = re.sub(r"(?:\s+\d{2}:\d{2}){1,8}\s*$", "", valor).strip()
            if valor and valor not in turnos_encontrados:
                turnos_encontrados.append(valor)
    return turnos_encontrados


def escolher_turno_principal(texto_completo):
    turnos = extrair_turnos_do_bloco(texto_completo)
    if not turnos:
        return "Sem Turno"
    return normalizar_turno(turnos[0])


def agrupar_palavras_por_linha(palavras, tolerancia_top=2):
    palavras_ordenadas = sorted(palavras, key=lambda w: (w["top"], w["x0"]))
    linhas = []
    for palavra in palavras_ordenadas:
        adicionada = False
        for linha in linhas:
            if abs(palavra["top"] - linha["top"]) <= tolerancia_top:
                linha["palavras"].append(palavra)
                linha["tops"].append(palavra["top"])
                adicionada = True
                break
        if not adicionada:
            linhas.append({"top": palavra["top"], "tops": [palavra["top"]], "palavras": [palavra]})
    for linha in linhas:
        linha["top"] = sum(linha["tops"]) / len(linha["tops"])
        linha["palavras"] = sorted(linha["palavras"], key=lambda w: w["x0"])
        linha["texto"] = " ".join(w["text"] for w in linha["palavras"])
    return sorted(linhas, key=lambda l: l["top"])


def obter_centros_colunas_tabela(palavras):
    linhas = agrupar_palavras_por_linha(palavras)
    cabecalho = None
    for linha in linhas:
        texto = linha["texto"]
        if "Data" in texto and "Dia" in texto and "Observação" in texto and "H.E." in texto and "Absent." in texto:
            cabecalho = linha
            break
    if cabecalho is None:
        return None, None
    palavras_cab = cabecalho["palavras"]
    centros = {}
    idx = 0
    for rotulo, qtd_palavras in LABEL_SPECS_TABELA:
        trecho = palavras_cab[idx: idx + qtd_palavras]
        idx += qtd_palavras
        if not trecho:
            return cabecalho["top"], None
        x0 = min(w["x0"] for w in trecho)
        x1 = max(w["x1"] for w in trecho)
        centros[rotulo] = (x0 + x1) / 2
    return cabecalho["top"], centros


def classificar_coluna_extra_por_posicao(x_centro, centros_colunas):
    colunas_extras = ["Abono", "H.E.", "Absent.", "Jornada", "Ad. Not."]
    return min(colunas_extras, key=lambda col: abs(x_centro - centros_colunas[col]))


def limpar_observacao_texto(tokens):
    tokens_limpos = [t for t in tokens if t not in {"**", ""}]
    texto = " ".join(tokens_limpos).strip()
    return re.sub(r"\s+", " ", texto).strip()


def montar_registro_vazio(data, dia):
    return {
        "Data": data, "Dia": dia,
        "1a E.": "", "1a S.": "", "2a E.": "", "2a S.": "",
        "3a E.": "", "3a S.": "", "4a E.": "", "4a S.": "",
        "Abono": "", "H.E.": "", "Absent.": "", "Jornada": "", "Ad. Not.": "", "Observação": "",
    }


def parse_linha_tabela_palavras(palavras_linha, centros_colunas):
    if not palavras_linha:
        return None
    palavras_linha = sorted(palavras_linha, key=lambda w: w["x0"])
    tokens = [w["text"] for w in palavras_linha]
    if not tokens or not re.fullmatch(r"\d{2}/\d{2}/\d{4}", tokens[0]):
        return None
    data = tokens[0]
    dia = tokens[1] if len(tokens) > 1 else ""
    registro = montar_registro_vazio(data, dia)

    batidas = []
    observacoes = []
    i = 2
    while i < len(palavras_linha):
        palavra = palavras_linha[i]
        token = palavra["text"]
        x_centro = (palavra["x0"] + palavra["x1"]) / 2
        proximo = palavras_linha[i + 1]["text"] if i + 1 < len(palavras_linha) else None
        if re.fullmatch(r"\d{2}:\d{2}", token):
            if proximo in {"O", "I", "P"}:
                batidas.append(token)
                i += 2
                continue
            coluna_extra = classificar_coluna_extra_por_posicao(x_centro, centros_colunas)
            registro[coluna_extra] = token if not registro[coluna_extra] else f'{registro[coluna_extra]} {token}'.strip()
        else:
            if token not in {"O", "I", "P", "**"}:
                observacoes.append(token)
        i += 1

    while len(batidas) < 8:
        batidas.append("")

    registro["1a E."] = batidas[0]
    registro["1a S."] = batidas[1]
    registro["2a E."] = batidas[2]
    registro["2a S."] = batidas[3]
    registro["3a E."] = batidas[4]
    registro["3a S."] = batidas[5]
    registro["4a E."] = batidas[6]
    registro["4a S."] = batidas[7]

    registro["Observação"] = limpar_observacao_texto(observacoes)
    obs_norm = normalizar_texto(registro["Observação"])

    if "SEXTA-FEIRA SANTA" in registro["Observação"].upper():
        registro["Observação"] = "SEXTA-FEIRA SANTA"
    elif obs_norm in {"D.S.R.", "D S R ."}:
        registro["Observação"] = "D.S.R."
    elif "COMPENSADO" in obs_norm:
        registro["Observação"] = "Compensado"
    elif "AUSENTE" in obs_norm:
        registro["Observação"] = "Ausente"
        if not registro["Absent."]:
            tempos = [w["text"] for w in palavras_linha if re.fullmatch(r"\d{2}:\d{2}", w["text"])]
            if tempos:
                registro["Absent."] = tempos[-1]
    elif "FERIAS" in obs_norm or "FÉRIAS" in registro["Observação"].upper():
        registro["Observação"] = "Férias"
    elif "NAO TRABALHADO" in obs_norm:
        registro["Observação"] = "Não Trabalhado"

    return registro


def extrair_tabela_de_uma_pagina(texto, palavras):
    if not palavras:
        return []
    top_cabecalho, centros_colunas = obter_centros_colunas_tabela(palavras)
    if top_cabecalho is None or not centros_colunas:
        return []
    linhas = agrupar_palavras_por_linha(palavras)
    dados = []
    for linha in linhas:
        texto_linha = linha["texto"].strip()
        if linha["top"] <= top_cabecalho:
            continue
        if texto_linha.startswith("Banco de Horas") or texto_linha.startswith("Marcações desconsideradas") or texto_linha.startswith("Horários"):
            break
        if re.match(r"^\d{2}/\d{2}/\d{4}", texto_linha):
            registro = parse_linha_tabela_palavras(linha["palavras"], centros_colunas)
            if registro:
                dados.append(registro)
    return dados


def chave_registro_unico(registro):
    return (
        str(registro.get("Matrícula", "")).strip(),
        str(registro.get("Nome", "")).strip(),
        str(registro.get("cc", "")).strip(),
        str(registro.get("Função", "")).strip(),
        str(registro.get("Data", "")).strip(),
        str(registro.get("Dia", "")).strip(),
        str(registro.get("1a E.", "")).strip(),
        str(registro.get("1a S.", "")).strip(),
        str(registro.get("2a E.", "")).strip(),
        str(registro.get("2a S.", "")).strip(),
        str(registro.get("3a E.", "")).strip(),
        str(registro.get("3a S.", "")).strip(),
        str(registro.get("4a E.", "")).strip(),
        str(registro.get("4a S.", "")).strip(),
        str(registro.get("Abono", "")).strip(),
        str(registro.get("H.E.", "")).strip(),
        str(registro.get("Absent.", "")).strip(),
        str(registro.get("Jornada", "")).strip(),
        str(registro.get("Ad. Not.", "")).strip(),
        str(registro.get("Observação", "")).strip(),
        str(registro.get("Turno", "")).strip(),
    )


def remover_registros_duplicados(registros):
    vistos = set()
    unicos = []
    for registro in registros:
        chave = chave_registro_unico(registro)
        if chave in vistos:
            continue
        vistos.add(chave)
        unicos.append(registro)
    return unicos


def processar_pdf(caminho_pdf):
    grupos = agrupar_paginas_por_colaborador(caminho_pdf)
    registros = []
    for grupo in grupos:
        paginas = grupo["paginas"]
        textos_paginas = [p["texto"] for p in paginas if p["texto"].strip()]
        if not textos_paginas:
            continue
        texto_completo = "\n".join(textos_paginas)
        cabecalho = extrair_cabecalho(texto_completo)
        resumo_banco = extrair_resumo_banco_horas(texto_completo)
        data_emissao_pdf = extrair_data_emissao_pdf(texto_completo)
        turno = escolher_turno_principal(texto_completo)
        linhas_tabela = []
        for pagina in paginas:
            linhas_tabela.extend(extrair_tabela_de_uma_pagina(pagina["texto"], pagina.get("palavras", [])))

        linhas_tabela = aplicar_evolucao_banco_horas(linhas_tabela, resumo_banco, data_emissao_pdf)

        for linha in linhas_tabela:
            registros.append({
                "Arquivo Origem": Path(caminho_pdf).name,
                "Matrícula": cabecalho["Matrícula"],
                "Nome": cabecalho["Nome"],
                "cc": cabecalho["cc"],
                "Função": cabecalho["Função"],
                "Data": linha["Data"],
                "Dia": linha["Dia"],
                "1a E.": linha["1a E."],
                "1a S.": linha["1a S."],
                "2a E.": linha["2a E."],
                "2a S.": linha["2a S."],
                "3a E.": linha["3a E."],
                "3a S.": linha["3a S."],
                "4a E.": linha["4a E."],
                "4a S.": linha["4a S."],
                "Abono": linha["Abono"],
                "H.E.": linha["H.E."],
                "Absent.": linha["Absent."],
                "Jornada": linha["Jornada"],
                "Ad. Not.": linha["Ad. Not."],
                "Observação": linha["Observação"],
                "Saldo Atual": linha.get("Saldo Atual", ""),
                "Saldo Calculado": linha.get("Saldo Calculado", ""),
                "Saldo Informado PDF": linha.get("Saldo Informado PDF", ""),
                "Data Emissão PDF": linha.get("Data Emissão PDF", data_emissao_pdf),
                "Saldo Anterior PDF": linha.get("Saldo Anterior PDF", resumo_banco.get("Saldo Anterior", "")),
                "Debito PDF": linha.get("Debito PDF", ""),
                "Credito PDF": linha.get("Credito PDF", ""),
                "Turno": turno,
            })
    return remover_registros_duplicados(registros)


def listar_pdfs_da_pasta(pasta):
    return sorted(Path(pasta).glob("*.pdf"))


def processar_pasta(pasta_pdf):
    arquivos_pdf = listar_pdfs_da_pasta(pasta_pdf)
    if not arquivos_pdf:
        raise ValueError(f"Nenhum arquivo PDF encontrado na pasta: {pasta_pdf}")

    todos_registros = []
    erros = []
    arquivos_ok = []
    arquivos_erro = []

    for arquivo in arquivos_pdf:
        try:
            LOGGER.info(f"Processando: {arquivo.name}")
            registros = processar_pdf(str(arquivo))
            todos_registros.extend(registros)

            if CONFIG.mover_arquivos_processados and not modo_teste_ativo():
                destino = mover_arquivo(arquivo, CONFIG.pasta_processados)
                arquivos_ok.append(str(destino))
                LOGGER.info(f"  -> registros extraídos: {len(registros)}")
                LOGGER.info(f"  -> arquivo movido para processados: {destino}")
            else:
                arquivos_ok.append(str(arquivo))
                LOGGER.info(f"  -> registros extraídos: {len(registros)}")
                LOGGER.info("  -> arquivo mantido na pasta de entrada (modo teste ou movimentação desabilitada)")
        except Exception as e:
            erros.append((arquivo.name, str(e)))
            LOGGER.exception(f"Erro em {arquivo.name}: {e}")
            try:
                if CONFIG.mover_arquivos_processados and not modo_teste_ativo():
                    destino_erro = mover_arquivo(arquivo, CONFIG.pasta_erro)
                    arquivos_erro.append(str(destino_erro))
                    LOGGER.error(f"  -> arquivo movido para erro: {destino_erro}")
                else:
                    arquivos_erro.append(str(arquivo))
                    LOGGER.error("  -> arquivo mantido na pasta de entrada após erro (modo teste ou movimentação desabilitada)")
            except Exception as mover_exc:
                LOGGER.error(f"Falha ao tratar '{arquivo.name}' após erro: {mover_exc}")
    return todos_registros, erros, arquivos_pdf, arquivos_ok, arquivos_erro


def calcular_jornada_batidas(row):
    colunas_batidas = ["1a E.", "1a S.", "2a E.", "2a S.", "3a E.", "3a S.", "4a E.", "4a S."]
    batidas = [str(row.get(col, "")).strip() for col in colunas_batidas]
    batidas = [b for b in batidas if b and b.lower() != "nan"]
    if len(batidas) < 2 or len(batidas) % 2 != 0:
        return None
    total = 0
    for i in range(0, len(batidas), 2):
        entrada = hhmm_para_minutos(batidas[i])
        saida = hhmm_para_minutos(batidas[i + 1])
        if entrada is None or saida is None:
            return None
        if saida < entrada:
            saida += 24 * 60
        total += saida - entrada
    return total


def validar_linha(row, tolerancia_minutos=5):
    inconsistencias = []
    observacao = str(row.get("Observação", "")).strip().upper()
    turno = str(row.get("Turno", "")).strip()
    saldo_atual = str(row.get("Saldo Atual", "")).strip()

    colunas_batidas = ["1a E.", "1a S.", "2a E.", "2a S.", "3a E.", "3a S.", "4a E.", "4a S."]
    batidas = [str(row.get(col, "")).strip() for col in colunas_batidas]
    batidas_validas = [b for b in batidas if b and b.lower() != "nan"]

    jornada_informada = hhmm_para_minutos(row.get("Jornada", ""))
    jornada_calculada = calcular_jornada_batidas(row)

    if not turno or turno == "Sem Turno":
        inconsistencias.append("Turno não identificado")
    if saldo_atual and converter_saldo_para_decimal(saldo_atual) is None:
        inconsistencias.append("Saldo Atual inválido")
    if len(batidas_validas) % 2 != 0:
        inconsistencias.append("Quantidade ímpar de batidas")
    if "FÉRIAS" in observacao or "FERIAS" in observacao:
        if batidas_validas:
            inconsistencias.append("Férias com batidas preenchidas")
    if "D.S.R." in observacao and batidas_validas:
        inconsistencias.append("D.S.R. com batidas preenchidas")
    if "COMPENSADO" in observacao and batidas_validas:
        inconsistencias.append("Compensado com batidas preenchidas")
    if "AUSENTE" in observacao and len(batidas_validas) >= 2:
        inconsistencias.append("Ausente com batidas preenchidas")
    if jornada_calculada is not None and jornada_informada is not None:
        if abs(jornada_calculada - jornada_informada) > tolerancia_minutos:
            inconsistencias.append(
                f"Jornada divergente (batidas={minutos_para_hhmm(jornada_calculada)} x informada={minutos_para_hhmm(jornada_informada)})"
            )
    if jornada_informada is not None and not batidas_validas and not observacao:
        inconsistencias.append("Jornada informada sem batidas")

    status = "OK" if not inconsistencias else "REVISAR"
    return {
        "Status Validação": status,
        "Inconsistências": "; ".join(inconsistencias),
        "Jornada Calculada": minutos_para_hhmm(jornada_calculada) if jornada_calculada is not None else "",
        "Qtd Batidas": len(batidas_validas),
    }


def preparar_dataframe_consultas(registros):
    df = pd.DataFrame(registros)
    for coluna in COLUNAS_BASE:
        if coluna not in df.columns:
            df[coluna] = ""
    df = df[COLUNAS_BASE].copy()

    for col in ["Arquivo Origem", "Matrícula", "Nome", "cc", "Função", "Dia", "Observação", "Turno"]:
        df[col] = df[col].fillna("").astype(str).str.strip()

    df["cc"] = df["cc"].apply(limpar_cc)
    df["Função"] = df["Função"].apply(limpar_funcao)

    df["Data_dt"] = pd.to_datetime(df["Data"], format="%d/%m/%Y", errors="coerce")
    df["H.E. Min"] = df["H.E."].apply(hhmm_para_minutos)
    df["Absent. Min"] = df["Absent."].apply(hhmm_para_minutos)
    df["Jornada Min"] = df["Jornada"].apply(hhmm_para_minutos)
    df["Ad. Not. Min"] = df["Ad. Not."].apply(hhmm_para_minutos)
    df["Saldo Atual Num"] = df["Saldo Atual"].apply(converter_saldo_para_decimal)

    obs_upper = df["Observação"].str.upper()
    df["Teve Ausência"] = obs_upper.str.contains("AUSENTE", na=False) | (df["Absent."].fillna("").astype(str).str.strip() != "")
    df["Tem HE"] = df["H.E."].fillna("").astype(str).str.strip() != ""
    df["Saldo Negativo"] = df["Saldo Atual Num"].fillna(0) < 0
    df["É Férias"] = obs_upper.str.contains("FÉRIAS|FERIAS", na=False, regex=True)
    df["É DSR"] = obs_upper.str.contains("D.S.R.", na=False, regex=False)
    df["É Compensado"] = obs_upper.str.contains("COMPENSADO", na=False)
    df["Tem Observação"] = df["Observação"].fillna("").astype(str).str.strip() != ""
    df["Acima Limite HE"] = df["H.E. Min"].fillna(0) >= int(CONFIG.limite_he_alerta_minutos)

    # Compatibilidade: a evolução já vem calculada da extração do PDF.
    df = recalcular_saldo_por_competencia(df)

    # Recalcula campos derivados após eventual ajuste de saldo.
    df["Saldo Atual Num"] = df["Saldo Atual"].apply(converter_saldo_para_decimal)
    df["Saldo Negativo"] = df["Saldo Atual Num"].fillna(0) < 0

    return df


def chave_colaborador_dataframe(row):
   
    matricula = str(row.get("Matrícula", "") or "").strip()
    nome = normalizar_texto(row.get("Nome", ""))
    cc = limpar_cc(row.get("cc", row.get("CC", "")))
    if matricula:
        return f"MAT::{matricula}"
    return f"NOME_CC::{nome}::{cc}"


def atualizar_campos_derivados_dataframe(df):
   
    if df is None or df.empty:
        return df
    df = df.copy()
    df["Data_dt"] = pd.to_datetime(df["Data"], format="%d/%m/%Y", errors="coerce")
    df["H.E. Min"] = df["H.E."].apply(hhmm_para_minutos)
    df["Absent. Min"] = df["Absent."].apply(hhmm_para_minutos)
    df["Jornada Min"] = df["Jornada"].apply(hhmm_para_minutos)
    df["Ad. Not. Min"] = df["Ad. Not."].apply(hhmm_para_minutos)
    df["Saldo Atual Num"] = df["Saldo Atual"].apply(converter_saldo_para_decimal)

    obs_upper = df["Observação"].fillna("").astype(str).str.upper()
    df["Teve Ausência"] = obs_upper.str.contains("AUSENTE", na=False) | (df["Absent."].fillna("").astype(str).str.strip() != "")
    df["Tem HE"] = df["H.E."].fillna("").astype(str).str.strip() != ""
    df["Saldo Negativo"] = df["Saldo Atual Num"].fillna(0) < 0
    df["É Férias"] = obs_upper.str.contains("FÉRIAS|FERIAS", na=False, regex=True)
    df["É DSR"] = obs_upper.str.contains("D.S.R.", na=False, regex=False)
    df["É Compensado"] = obs_upper.str.contains("COMPENSADO", na=False)
    df["Tem Observação"] = df["Observação"].fillna("").astype(str).str.strip() != ""
    df["Acima Limite HE"] = df["H.E. Min"].fillna(0) >= int(CONFIG.limite_he_alerta_minutos)
    return df


def consolidar_consolidado_oficial(df_historico):
    
    if df_historico is None or df_historico.empty:
        return df_historico

    df = df_historico.copy()
    for coluna in ["Matrícula", "Nome", "cc", "Data", "Data Emissão PDF", "Arquivo Origem"]:
        if coluna not in df.columns:
            df[coluna] = ""

    df["_chave_colab_oficial"] = df.apply(chave_colaborador_dataframe, axis=1)
    df["_data_dt_oficial"] = pd.to_datetime(df["Data"].astype(str).str.strip(), format="%d/%m/%Y", errors="coerce")
    df["_emissao_dt_oficial"] = pd.to_datetime(df["Data Emissão PDF"].astype(str).str.strip(), format="%d/%m/%Y", errors="coerce")

   
    df = df.sort_values(
        by=["_chave_colab_oficial", "_data_dt_oficial", "_emissao_dt_oficial", "Arquivo Origem"],
        na_position="first",
        kind="mergesort",
    )
    df_oficial = (
        df.groupby(["_chave_colab_oficial", "Data"], dropna=False, sort=False)
        .tail(1)
        .copy()
    )

    df_oficial = df_oficial.drop(
        columns=["_chave_colab_oficial", "_data_dt_oficial", "_emissao_dt_oficial"],
        errors="ignore",
    )

 
    df_oficial = recalcular_saldo_por_competencia(df_oficial)
    df_oficial = atualizar_campos_derivados_dataframe(df_oficial)
    df_oficial = aplicar_validacoes(df_oficial)
    return df_oficial



def montar_saldos_disponiveis_bh(df_historico, df_oficial=None):
    
    colunas_saida = [
        "Visao Origem", "Matrícula", "Nome", "CC", "Função", "Turno",
        "Data", "Data Emissão PDF", "Origem Saldo", "Saldo Disponível",
        "Saldo Min", "Recomendado", "Arquivo Origem", "Observação",
    ]

    fontes = [
        ("Consolidado", df_oficial),
        ("Consolidado Historico", df_historico),
    ]
    colunas_saldo = [
        ("Saldo Informado PDF", "Saldo Informado PDF"),
        ("Saldo Atual", "Saldo Atual"),
        ("Saldo Anterior PDF", "Saldo Anterior PDF"),
        ("Saldo Calculado", "Saldo Calculado"),
    ]

    linhas = []
    vistos = set()

    for visao, df_base in fontes:
        if df_base is None or df_base.empty:
            continue
        base = df_base.copy()
        for coluna in [
            "Matrícula", "Nome", "cc", "Função", "Turno", "Data",
            "Data Emissão PDF", "Arquivo Origem", "Observação",
        ]:
            if coluna not in base.columns:
                base[coluna] = ""

        for _, row in base.iterrows():
            data_linha = str(row.get("Data", "") or "").strip()
            data_emissao = str(row.get("Data Emissão PDF", "") or "").strip()
            for origem_rotulo, coluna_saldo in colunas_saldo:
                if coluna_saldo not in base.columns:
                    continue
                saldo_txt = str(row.get(coluna_saldo, "") or "").strip()
                if not saldo_txt or saldo_txt.lower() == "nan":
                    continue
                saldo_min = saldo_banco_para_minutos(saldo_txt)
                if saldo_min is None:
                    continue

                chave = (
                    visao,
                    str(row.get("Matrícula", "") or "").strip(),
                    normalizar_texto(row.get("Nome", "")),
                    limpar_cc(row.get("cc", "")),
                    data_linha,
                    data_emissao,
                    origem_rotulo,
                    saldo_txt,
                    str(row.get("Arquivo Origem", "") or "").strip(),
                )
                if chave in vistos:
                    continue
                vistos.add(chave)

                recomendado = "Sim" if (
                    visao == "Consolidado"
                    and origem_rotulo in {"Saldo Informado PDF", "Saldo Atual"}
                    and (not data_emissao or data_linha == data_emissao)
                ) else ""

                linhas.append({
                    "Visao Origem": visao,
                    "Matrícula": str(row.get("Matrícula", "") or "").strip(),
                    "Nome": str(row.get("Nome", "") or "").strip(),
                    "CC": limpar_cc(row.get("cc", "")),
                    "Função": limpar_funcao(row.get("Função", "")),
                    "Turno": str(row.get("Turno", "") or "").strip(),
                    "Data": data_linha,
                    "Data Emissão PDF": data_emissao,
                    "Origem Saldo": origem_rotulo,
                    "Saldo Disponível": saldo_txt,
                    "Saldo Min": int(saldo_min),
                    "Recomendado": recomendado,
                    "Arquivo Origem": str(row.get("Arquivo Origem", "") or "").strip(),
                    "Observação": str(row.get("Observação", "") or "").strip(),
                })

    resultado = pd.DataFrame(linhas, columns=colunas_saida)
    if resultado.empty:
        return resultado

    resultado["_data_ord"] = pd.to_datetime(resultado["Data"], format="%d/%m/%Y", errors="coerce")
    resultado["_emissao_ord"] = pd.to_datetime(resultado["Data Emissão PDF"], format="%d/%m/%Y", errors="coerce")
    resultado["_prio"] = resultado["Recomendado"].eq("Sim").astype(int)
    resultado = resultado.sort_values(
        by=["Nome", "CC", "_data_ord", "_prio", "_emissao_ord", "Origem Saldo"],
        ascending=[True, True, True, False, False, True],
        na_position="last",
        kind="mergesort",
    ).drop(columns=["_data_ord", "_emissao_ord", "_prio"], errors="ignore")
    return resultado


def ordenar_colunas_para_exportacao(df):
   
    df_export = df.copy()
    if "cc" in df_export.columns:
        df_export["cc"] = df_export["cc"].apply(limpar_cc)
    if "Função" in df_export.columns:
        df_export["Função"] = df_export["Função"].apply(limpar_funcao)
    for coluna in COLUNAS_EXPORTACAO:
        if coluna not in df_export.columns:
            df_export[coluna] = ""
    colunas_preferidas = [c for c in COLUNAS_EXPORTACAO if c in df_export.columns]
    colunas_restantes = [c for c in df_export.columns if c not in colunas_preferidas]
    return preparar_para_exportacao_excel(df_export[colunas_preferidas + colunas_restantes])





def _data_fechamento_ciclo_anterior(ciclo):
    
    try:
        inicio = pd.to_datetime(str(ciclo).strip(), format="%Y-%m-%d", errors="coerce")
    except Exception:
        return ""
    if pd.isna(inicio):
        return ""
    fechamento = inicio - pd.Timedelta(days=1)
    return fechamento.strftime("%d/%m/%Y")


def _saldo_abertura_ciclo_por_fechamento(sub_colab, ciclo):
   
    data_fechamento = _data_fechamento_ciclo_anterior(ciclo)
    if not data_fechamento or sub_colab is None or sub_colab.empty:
        return "", None

    candidatos = sub_colab[
        sub_colab["Data"].astype(str).str.strip() == data_fechamento
    ].copy()
    if candidatos.empty:
        return "", None

    if "Data Emissão PDF" not in candidatos.columns:
        candidatos["Data Emissão PDF"] = ""

    # Prioriza a fotografia cuja emissão também seja o fechamento do ciclo.
    candidatos["_prio_fechamento"] = (
        candidatos["Data Emissão PDF"].astype(str).str.strip() == data_fechamento
    ).astype(int)
    candidatos = candidatos.sort_values(
        by=["_prio_fechamento"],
        ascending=[False],
        kind="mergesort",
    )

    for coluna in ["Saldo Atual", "Saldo Informado PDF"]:
        if coluna not in candidatos.columns:
            continue
        saldo_txt, saldo_min = _primeiro_saldo_valido(candidatos[coluna].tolist())
        if saldo_min is not None:
            return saldo_txt, saldo_min

    return "", None

def recalcular_saldo_por_competencia(df):
   
    if df is None or df.empty:
        return df

    df = df.copy()

    for coluna in ["Matrícula", "Nome", "cc", "Data", "H.E.", "Absent.", "Observação", "Saldo Anterior PDF", "Saldo Calculado"]:
        if coluna not in df.columns:
            df[coluna] = ""

    df["_data_dt_calc"] = pd.to_datetime(df["Data"].astype(str).str.strip(), format="%d/%m/%Y", errors="coerce")
    df["_ciclo_calc"] = df["Data"].apply(obter_chave_ciclo_medicao)
    df["Saldo Calculado"] = ""

    def chave_colab(row):
        matricula = str(row.get("Matrícula", "") or "").strip()
        nome = str(row.get("Nome", "") or "").strip()
        cc = str(row.get("cc", "") or "").strip()
        if matricula:
            return ("MAT", matricula)
        return ("NOME_CC", normalizar_texto(nome), limpar_cc(cc))

    df["_chave_colab_calc"] = df.apply(chave_colab, axis=1)

 
    saldo_final_anterior_por_colab = {}

    for chave, idx_colab in df.groupby("_chave_colab_calc", sort=False).groups.items():
        sub_colab = df.loc[list(idx_colab)].copy()
        ciclos = sorted(
            [c for c in sub_colab["_ciclo_calc"].dropna().unique().tolist()],
            key=lambda c: str(c),
        )

        for ciclo in ciclos:
            idx_ciclo = sub_colab.index[sub_colab["_ciclo_calc"] == ciclo].tolist()
            if not idx_ciclo:
                continue

            grupo = df.loc[idx_ciclo].copy().sort_values(
                by=["_data_dt_calc", "Data"],
                na_position="last",
                kind="mergesort",
            )

           
            saldo_txt, saldo_corrente = _saldo_abertura_ciclo_por_fechamento(sub_colab, ciclo)

            
            if saldo_corrente is None:
                saldo_txt, saldo_corrente = _primeiro_saldo_valido(grupo.get("Saldo Anterior PDF", []))

           
            if saldo_corrente is None:
                saldo_corrente = saldo_final_anterior_por_colab.get(chave, 0)

            datas_ordenadas = []
            for data_val in grupo["Data"].astype(str).str.strip().tolist():
                if data_val not in datas_ordenadas:
                    datas_ordenadas.append(data_val)

            for data_txt in datas_ordenadas:
                idx_data = grupo.index[grupo["Data"].astype(str).str.strip() == data_txt].tolist()
                if not idx_data:
                    continue

                # Todas as fotografias/linhas do mesmo dia exibem a mesma posição de saldo.
                df.loc[idx_data, "Saldo Calculado"] = minutos_para_saldo_banco(saldo_corrente)

                linhas_data = df.loc[idx_data].copy()

                # Se todas as linhas daquele dia forem neutras, não há movimento.
                if all(linha_neutra_banco_horas(row) for _, row in linhas_data.iterrows()):
                    continue

                # Evita dupla contagem em caso de duplicidade do mesmo dia em PDFs diferentes.
                # Também corrige linhas com Observação = Banco de Horas:
                # nesses casos, o tempo é débito do banco, ainda que tenha sido
                # capturado visualmente em H.E. pelo layout do PDF.
                credito_he, debito_abs = movimento_banco_horas_do_dia(linhas_data)

                saldo_corrente = saldo_corrente + credito_he - debito_abs

            saldo_final_anterior_por_colab[chave] = saldo_corrente

    df = df.drop(columns=["_data_dt_calc", "_ciclo_calc", "_chave_colab_calc"], errors="ignore")
    return df

def aplicar_validacoes(df):
    df = df.copy()
    resultados = df.apply(validar_linha, axis=1, result_type="expand")
    for col in resultados.columns:
        df[col] = resultados[col]
    return df


def listar_ausentes(df):
    return df[df["Teve Ausência"]].copy()


def listar_hora_extra(df):
    return df[df["Tem HE"]].copy()


def listar_saldo_negativo(df):
    return df[df["Saldo Negativo"]].copy()


def listar_ferias(df):
    return df[df["É Férias"]].copy()


def listar_inconsistencias(df):
    return df[df["Status Validação"] == "REVISAR"].copy()


def resumo_validacao(df):
    return (
        df.groupby("Status Validação", dropna=False)
        .size()
        .reset_index(name="Quantidade")
        .sort_values(by="Status Validação")
    )


def resumo_por_turno(df):
    return (
        df.groupby("Turno", dropna=False)
        .agg(
            Registros=("Nome", "size"),
            Colaboradores=("Nome", "nunique"),
            Ausências=("Teve Ausência", "sum"),
            Hora_Extra=("Tem HE", "sum"),
            Férias=("É Férias", "sum"),
            Saldo_Negativo=("Saldo Negativo", "sum"),
            Inconsistencias=("Status Validação", lambda s: (s == "REVISAR").sum()),
        )
        .reset_index()
        .sort_values(by=["Turno"], ascending=True)
    )


def resumo_por_colaborador(df):
    return (
        df.groupby(["Nome", "cc", "Função", "Turno"], dropna=False)
        .agg(
            Registros=("Data", "size"),
            Dias_Com_HE=("Tem HE", "sum"),
            Dias_Com_Ausência=("Teve Ausência", "sum"),
            Dias_Férias=("É Férias", "sum"),
            Dias_Com_Inconsistência=("Status Validação", lambda s: (s == "REVISAR").sum()),
            Último_Saldo=("Saldo Atual", "last"),
        )
        .reset_index()
        .sort_values(by=["Nome", "Turno"], ascending=True)
    )


def resumo_por_data(df):
    base = (
        df.groupby("Data", dropna=False)
        .agg(
            Registros=("Nome", "size"),
            Colaboradores=("Nome", "nunique"),
            Ausências=("Teve Ausência", "sum"),
            Hora_Extra=("Tem HE", "sum"),
            Férias=("É Férias", "sum"),
            Inconsistencias=("Status Validação", lambda s: (s == "REVISAR").sum()),
        )
        .reset_index()
    )
    base["Data_dt"] = pd.to_datetime(base["Data"], format="%d/%m/%Y", errors="coerce")
    return base.sort_values(by="Data_dt").drop(columns=["Data_dt"])


def ranking_colaboradores_he(df):
    base = df[df["H.E. Min"].fillna(0) > 0].copy()
    if base.empty:
        return pd.DataFrame(columns=["Nome", "CC", "Função", "Turno", "Dias_Com_HE", "HE_Total", "HE_Total_Min"])
    resumo = (
        base.groupby(["Nome", "cc", "Função", "Turno"], dropna=False)
        .agg(
            Dias_Com_HE=("Tem HE", "sum"),
            HE_Total_Min=("H.E. Min", "sum"),
        )
        .reset_index()
        .sort_values(by=["HE_Total_Min", "Dias_Com_HE", "Nome"], ascending=[False, False, True])
        .head(int(CONFIG.top_n_colaboradores_he))
    )
    resumo["HE_Total"] = resumo["HE_Total_Min"].apply(minutos_para_hhmm)
    resumo = resumo.rename(columns={"cc": "CC"})
    return resumo[["Nome", "CC", "Função", "Turno", "Dias_Com_HE", "HE_Total", "HE_Total_Min"]]


def ranking_dias_he(df):
    base = df[df["H.E. Min"].fillna(0) > 0].copy()
    if base.empty:
        return pd.DataFrame(columns=["Data", "Colaboradores_Com_HE", "HE_Total", "HE_Total_Min"])
    resumo = (
        base.groupby("Data", dropna=False)
        .agg(
            Colaboradores_Com_HE=("Nome", "nunique"),
            HE_Total_Min=("H.E. Min", "sum"),
        )
        .reset_index()
    )
    resumo["Data_dt"] = pd.to_datetime(resumo["Data"], format="%d/%m/%Y", errors="coerce")
    resumo = resumo.sort_values(by=["HE_Total_Min", "Data_dt"], ascending=[False, False]).head(int(CONFIG.top_n_dias_he))
    resumo["HE_Total"] = resumo["HE_Total_Min"].apply(minutos_para_hhmm)
    return resumo[["Data", "Colaboradores_Com_HE", "HE_Total", "HE_Total_Min"]]


def alertas_hora_extra(df):
    base = df[df["Acima Limite HE"]].copy()
    if base.empty:
        return pd.DataFrame(columns=[
            "Data", "Nome", "cc", "Função", "Turno", "H.E.", "H.E. Min", "Observação", "Status Validação"
        ])
    base = base.sort_values(by=["H.E. Min", "Nome"], ascending=[False, True])
    return base[[
        "Data", "Nome", "cc", "Função", "Turno", "H.E.", "H.E. Min", "Observação", "Status Validação"
    ]].rename(columns={"cc": "CC"})


def top_consultas(df):
    return {
        "Ausentes": listar_ausentes(df),
        "Hora Extra": listar_hora_extra(df),
        "Saldo Negativo": listar_saldo_negativo(df),
        "Férias": listar_ferias(df),
        "Inconsistencias": listar_inconsistencias(df),
        "Resumo Validacao": resumo_validacao(df),
        "Resumo Turno": resumo_por_turno(df),
        "Resumo Colaborador": resumo_por_colaborador(df),
        "Por Data": resumo_por_data(df),
        "Ranking HE Colab": ranking_colaboradores_he(df),
        "Ranking HE Dias": ranking_dias_he(df),
        "Alertas HE": alertas_hora_extra(df),
    }


def montar_controle_operacional_resumo(df, total_pdfs, total_erros):
    total_registros = len(df)
    total_colaboradores = df["Nome"].nunique() if "Nome" in df.columns else 0
    total_turnos = df["Turno"].nunique() if "Turno" in df.columns else 0
    total_com_he = int(df["Tem HE"].sum()) if "Tem HE" in df.columns else 0
    total_com_ausencia = int(df["Teve Ausência"].sum()) if "Teve Ausência" in df.columns else 0
    total_com_inconsistencia = int((df["Status Validação"] == "REVISAR").sum()) if "Status Validação" in df.columns else 0
    total_ferias = int(df["É Férias"].sum()) if "É Férias" in df.columns else 0
    total_saldo_negativo = int(df["Saldo Negativo"].sum()) if "Saldo Negativo" in df.columns else 0
    total_alertas_he = int(df["Acima Limite HE"].sum()) if "Acima Limite HE" in df.columns else 0

    taxa_inconsistencia = (total_com_inconsistencia / total_registros * 100) if total_registros else 0
    taxa_erro_pdf = (total_erros / total_pdfs * 100) if total_pdfs else 0

    dados = [
        {"Indicador": "Modo Execução", "Valor": CONFIG.modo_execucao},
        {"Indicador": "Modo Data", "Valor": CONFIG.modo_data_arquivo},
        {"Indicador": "Limite Alerta HE", "Valor": minutos_para_hhmm(CONFIG.limite_he_alerta_minutos)},
        {"Indicador": "Total PDFs Lidos", "Valor": total_pdfs},
        {"Indicador": "Total PDFs com Erro", "Valor": total_erros},
        {"Indicador": "Taxa Erro PDF (%)", "Valor": round(taxa_erro_pdf, 2)},
        {"Indicador": "Total Registros", "Valor": total_registros},
        {"Indicador": "Total Colaboradores", "Valor": total_colaboradores},
        {"Indicador": "Total Turnos", "Valor": total_turnos},
        {"Indicador": "Dias com Hora Extra", "Valor": total_com_he},
        {"Indicador": "Dias com Ausência", "Valor": total_com_ausencia},
        {"Indicador": "Dias com Inconsistência", "Valor": total_com_inconsistencia},
        {"Indicador": "Taxa Inconsistência (%)", "Valor": round(taxa_inconsistencia, 2)},
        {"Indicador": "Dias em Férias", "Valor": total_ferias},
        {"Indicador": "Dias com Saldo Negativo", "Valor": total_saldo_negativo},
        {"Indicador": "Alertas de HE", "Valor": total_alertas_he},
    ]
    return pd.DataFrame(dados)


def montar_controle_operacional_por_data(df):
    base = (
        df.groupby("Data", dropna=False)
        .agg(
            Registros=("Nome", "size"),
            Colaboradores=("Nome", "nunique"),
            Dias_Com_HE=("Tem HE", "sum"),
            HE_Total_Min=("H.E. Min", "sum"),
            Dias_Com_Ausencia=("Teve Ausência", "sum"),
            Dias_Com_Inconsistencia=("Status Validação", lambda s: (s == "REVISAR").sum()),
            Dias_Saldo_Negativo=("Saldo Negativo", "sum"),
            Alertas_HE=("Acima Limite HE", "sum"),
        )
        .reset_index()
    )
    if base.empty:
        return pd.DataFrame(columns=[
            "Data", "Registros", "Colaboradores", "Dias_Com_HE", "HE_Total", "HE_Total_Min",
            "Dias_Com_Ausencia", "Dias_Com_Inconsistencia", "Dias_Saldo_Negativo", "Alertas_HE"
        ])
    base["HE_Total_Min"] = base["HE_Total_Min"].fillna(0).astype(int)
    base["HE_Total"] = base["HE_Total_Min"].apply(minutos_para_hhmm)
    base["Data_dt"] = pd.to_datetime(base["Data"], format="%d/%m/%Y", errors="coerce")
    base = base.sort_values(by="Data_dt").drop(columns=["Data_dt"])
    return base[[
        "Data", "Registros", "Colaboradores", "Dias_Com_HE", "HE_Total", "HE_Total_Min",
        "Dias_Com_Ausencia", "Dias_Com_Inconsistencia", "Dias_Saldo_Negativo", "Alertas_HE"
    ]]


def montar_dashboard_banco_horas_he(df):
    
    colunas = [
        "Nome", "CC", "Função", "Turno", "Primeira Data", "Última Data",
        "Dias no Período", "Dias com HE", "HE Período", "HE Período Min",
        "Débito/Absent. Período", "Débito/Absent. Min", "Saldo Inicial Estimado",
        "Saldo Atual Banco", "Variação Banco", "Variação Banco Min", "Status Banco",
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=colunas)

    base = df.copy()
    base["Data_dt"] = pd.to_datetime(base["Data"], format="%d/%m/%Y", errors="coerce")
    base["H.E. Min"] = base.get("H.E. Min", base["H.E."].apply(hhmm_para_minutos)).fillna(0)
    base["Absent. Min"] = base.get("Absent. Min", base["Absent."].apply(hhmm_para_minutos)).fillna(0)
    base["Saldo Atual Min"] = base["Saldo Atual"].apply(hhmm_para_minutos)

   
    base = base.sort_values(["Nome", "Matrícula", "Data_dt"])

    linhas = []
    chaves_grupo = ["Matrícula", "Nome", "cc", "Função", "Turno"]
    for (matricula, nome, cc, funcao, turno), grupo in base.groupby(chaves_grupo, dropna=False):
        grupo = grupo.sort_values("Data_dt").copy()
        grupo_saldo = grupo.dropna(subset=["Saldo Atual Min"])

        he_total = int(round(grupo["H.E. Min"].fillna(0).sum()))
        absent_total = int(round(grupo["Absent. Min"].fillna(0).sum()))
        variacao = he_total - absent_total

        saldo_final_min = None
        if not grupo_saldo.empty:
            saldo_final_min = int(round(grupo_saldo.iloc[-1]["Saldo Atual Min"]))

        saldo_inicial_min = saldo_final_min - variacao if saldo_final_min is not None else None

        if saldo_final_min is None:
            status_banco = "Sem saldo"
        elif saldo_final_min < 0:
            status_banco = "Negativo"
        elif saldo_final_min == 0:
            status_banco = "Zerado"
        else:
            status_banco = "Positivo"

        primeira_data = grupo["Data_dt"].min()
        ultima_data = grupo["Data_dt"].max()

        linhas.append({
            "Nome": str(nome).strip(),
            "CC": limpar_cc(cc),
            "Função": limpar_funcao(funcao),
            "Turno": str(turno).strip(),
            "Primeira Data": primeira_data.strftime("%d/%m/%Y") if pd.notna(primeira_data) else "",
            "Última Data": ultima_data.strftime("%d/%m/%Y") if pd.notna(ultima_data) else "",
            "Dias no Período": int(grupo["Data"].nunique()),
            "Dias com HE": int((grupo["H.E. Min"].fillna(0) > 0).sum()),
            "HE Período": minutos_para_hhmm(he_total),
            "HE Período Min": he_total,
            "Débito/Absent. Período": minutos_para_hhmm(absent_total),
            "Débito/Absent. Min": absent_total,
            "Saldo Inicial Estimado": minutos_para_hhmm(saldo_inicial_min),
            "Saldo Atual Banco": minutos_para_hhmm(saldo_final_min),
            "Variação Banco": minutos_para_hhmm(variacao),
            "Variação Banco Min": variacao,
            "Status Banco": status_banco,
        })

    resultado = pd.DataFrame(linhas, columns=colunas)
    if resultado.empty:
        return resultado
    return resultado.sort_values(
        by=["HE Período Min", "Variação Banco Min", "Nome"],
        ascending=[False, False, True],
    )


def escrever_dashboard_banco_horas_he(writer, df_dashboard):
    nome_aba = "Dashboard Banco HE"
    df_export = df_dashboard.copy()
    df_export.to_excel(writer, sheet_name=nome_aba, index=False, startrow=0)

    ws = writer.sheets[nome_aba]
    ajustar_largura_colunas(writer, nome_aba, df_export)
    aplicar_formatacao_colunas(writer, nome_aba, df_export)

    if df_export.empty:
        return

   
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Destaque visual simples por status do banco.
    fill_negativo = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
    fill_positivo = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fill_zerado = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    status_col = list(df_export.columns).index("Status Banco") + 1
    for row_idx in range(2, len(df_export) + 2):
        status = str(ws.cell(row=row_idx, column=status_col).value or "")
        if status == "Negativo":
            ws.cell(row=row_idx, column=status_col).fill = fill_negativo
        elif status == "Positivo":
            ws.cell(row=row_idx, column=status_col).fill = fill_positivo
        elif status == "Zerado":
            ws.cell(row=row_idx, column=status_col).fill = fill_zerado

  
    try:
        nome_col = list(df_export.columns).index("Nome") + 1
        he_col = list(df_export.columns).index("HE Período Min") + 1
        adicionar_grafico_coluna(
            ws=ws,
            min_row=1,
            max_row=len(df_export) + 1,
            categoria_col=nome_col,
            valor_col=he_col,
            titulo="Horas Extras no Período por Colaborador (min)",
            ancoragem="S2",
        )
    except Exception as e:
        LOGGER.warning(f"Não foi possível criar gráfico de HE do dashboard: {e}")

   
    try:
        saldo_col = list(df_export.columns).index("Variação Banco Min") + 1
        adicionar_grafico_coluna(
            ws=ws,
            min_row=1,
            max_row=len(df_export) + 1,
            categoria_col=nome_col,
            valor_col=saldo_col,
            titulo="Variação Líquida do Banco no Período (min)",
            ancoragem="S18",
        )
    except Exception as e:
        LOGGER.warning(f"Não foi possível criar gráfico de variação do dashboard: {e}")


def ajustar_largura_colunas(writer, sheet_name, df):
    worksheet = writer.sheets[sheet_name]
    for idx, col in enumerate(df.columns, start=1):
        valores = []
        for valor in df[col].tolist():
            if valor is None:
                valores.append("")
            else:
                try:
                    valores.append("" if pd.isna(valor) else str(valor))
                except Exception:
                    valores.append(str(valor))
        max_len = max([len(str(col))] + [len(v) for v in valores])
        largura = min(max_len + 2, 50)
        col_letter = worksheet.cell(row=1, column=idx).column_letter
        worksheet.column_dimensions[col_letter].width = largura
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def aplicar_formatacao_colunas(writer, sheet_name, df):
    worksheet = writer.sheets[sheet_name]
    azul = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    rosa = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    amarela = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fonte_negrito = Font(bold=True)
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        if col_name in COLUNAS_PDF:
            cell.fill = azul
        elif col_name in COLUNAS_CALCULADAS or col_name == "Valor":
            cell.fill = rosa
        elif "HE" in str(col_name):
            cell.fill = amarela
        cell.font = fonte_negrito


def adicionar_grafico_coluna(ws, min_row, max_row, categoria_col, valor_col, titulo, ancoragem):
    if max_row <= min_row:
        return
    chart = BarChart()
    chart.title = titulo
    chart.y_axis.title = "Valor"
    chart.x_axis.title = ""
    data = Reference(ws, min_col=valor_col, min_row=min_row, max_row=max_row)
    cats = Reference(ws, min_col=categoria_col, min_row=min_row + 1, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 8
    chart.width = 16
    ws.add_chart(chart, ancoragem)


def escrever_secao_dataframe(ws, titulo, df, start_row):
    ws.cell(row=start_row, column=1, value=titulo)
    ws.cell(row=start_row, column=1).font = Font(bold=True)
    header_row = start_row + 1

    if df is None or df.empty:
        ws.cell(row=header_row, column=1, value="Mensagem")
        ws.cell(row=header_row, column=1).font = Font(bold=True)
        ws.cell(row=header_row + 1, column=1, value=f"Nenhum registro encontrado para '{titulo}'.")
        return header_row, header_row + 1

    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=header_row, column=col_idx, value=col_name)
        ws.cell(row=header_row, column=col_idx).font = Font(bold=True)

    for row_idx, row in enumerate(df.itertuples(index=False), start=header_row + 1):
        for col_idx, valor in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    return header_row, header_row + len(df)


def formatar_secao_worksheet(ws, start_row, end_row, total_cols):
    for col_idx in range(1, total_cols + 1):
        valores = []
        for row_idx in range(start_row, end_row + 1):
            valor = ws.cell(row=row_idx, column=col_idx).value
            valores.append("" if valor is None else str(valor))
        max_len = max(len(v) for v in valores) if valores else 0
        letra = ws.cell(row=start_row, column=col_idx).column_letter
        ws.column_dimensions[letra].width = min(max_len + 2, 45)


def exportar_excel_com_consultas(registros, caminho_saida, total_pdfs=0, total_erros=0):
    if not registros:
        raise ValueError("Nenhum registro foi extraído dos PDFs.")

  
    df_historico = preparar_dataframe_consultas(registros)
    df_historico = aplicar_validacoes(df_historico)

   
    df = consolidar_consolidado_oficial(df_historico)

    grupos_turno = defaultdict(list)
    nomes_abas_usados = set()

    for _, row in df.iterrows():
        turno = row.get("Turno", "Sem Turno") or "Sem Turno"
        grupos_turno[turno].append(row.to_dict())

    consultas = top_consultas(df)
    df_controle_resumo = montar_controle_operacional_resumo(df, total_pdfs=total_pdfs, total_erros=total_erros)
    df_controle_por_data = montar_controle_operacional_por_data(df)
    df_dashboard_banco_he = montar_dashboard_banco_horas_he(df)

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        for turno, dados in grupos_turno.items():
            df_turno = pd.DataFrame(dados)
            df_turno = ordenar_colunas_para_exportacao(df_turno)

            nome_base = sanitizar_nome_aba(turno)
            nome_aba = nome_base
            contador = 1
            while nome_aba in nomes_abas_usados:
                sufixo = f"_{contador}"
                nome_aba = f"{nome_base[:31 - len(sufixo)]}{sufixo}"
                contador += 1
            nomes_abas_usados.add(nome_aba)

            df_turno.to_excel(writer, sheet_name=nome_aba, index=False)
            ajustar_largura_colunas(writer, nome_aba, df_turno)
            aplicar_formatacao_colunas(writer, nome_aba, df_turno)

        df_consolidado = ordenar_colunas_para_exportacao(df)
        df_consolidado.to_excel(writer, sheet_name="Consolidado", index=False)
        ajustar_largura_colunas(writer, "Consolidado", df_consolidado)
        aplicar_formatacao_colunas(writer, "Consolidado", df_consolidado)

        
        df_consolidado_historico = ordenar_colunas_para_exportacao(df_historico)
        df_consolidado_historico.to_excel(writer, sheet_name="Consolidado Historico", index=False)
        ajustar_largura_colunas(writer, "Consolidado Historico", df_consolidado_historico)
        aplicar_formatacao_colunas(writer, "Consolidado Historico", df_consolidado_historico)

        
        df_saldos_bh = montar_saldos_disponiveis_bh(df_historico=df_historico, df_oficial=df)
        df_saldos_bh.to_excel(writer, sheet_name="Saldos Disponiveis BH", index=False)
        ajustar_largura_colunas(writer, "Saldos Disponiveis BH", df_saldos_bh)
        aplicar_formatacao_colunas(writer, "Saldos Disponiveis BH", df_saldos_bh)

        escrever_dashboard_banco_horas_he(writer, df_dashboard_banco_he)

        nome_aba_controle = "Controle Operacional"
        df_controle_resumo.to_excel(writer, sheet_name=nome_aba_controle, index=False, startrow=0)
        ws_ctrl = writer.sheets[nome_aba_controle]
        aplicar_formatacao_colunas(writer, nome_aba_controle, df_controle_resumo)
        ajustar_largura_colunas(writer, nome_aba_controle, df_controle_resumo)

        start_row_data = len(df_controle_resumo) + 4
        ws_ctrl.cell(row=start_row_data, column=1, value="Controle Operacional por Data")
        ws_ctrl.cell(row=start_row_data, column=1).font = Font(bold=True)

        if df_controle_por_data.empty:
            ws_ctrl.cell(row=start_row_data + 1, column=1, value="Mensagem")
            ws_ctrl.cell(row=start_row_data + 1, column=1).font = Font(bold=True)
            ws_ctrl.cell(row=start_row_data + 2, column=1, value="Nenhuma linha disponível para análise por data.")
            data_header_row = start_row_data + 1
            data_end_row = start_row_data + 2
        else:
            df_controle_por_data.to_excel(writer, sheet_name=nome_aba_controle, index=False, startrow=start_row_data)
            data_header_row = start_row_data + 1
            data_end_row = data_header_row + len(df_controle_por_data)
            ws_ctrl.auto_filter.ref = f"A{data_header_row}:J{data_end_row}"
            ws_ctrl.freeze_panes = f"A{data_header_row + 1}"
            formatar_secao_worksheet(ws_ctrl, data_header_row, data_end_row, total_cols=len(df_controle_por_data.columns))
            for col_idx in range(1, len(df_controle_por_data.columns) + 1):
                ws_ctrl.cell(row=data_header_row, column=col_idx).font = Font(bold=True)

        consultas_graficos = {
            "Ranking HE Colab": ("Nome", "HE_Total_Min", "Top Colaboradores por Hora Extra"),
            "Ranking HE Dias": ("Data", "HE_Total_Min", "Top Dias por Hora Extra"),
        }

        for nome, df_consulta in consultas.items():
            nome_aba = sanitizar_nome_aba(nome)
            if df_consulta.empty:
                df_vazio = pd.DataFrame({"Mensagem": [f"Nenhum registro encontrado para '{nome}'."]})
                df_vazio = preparar_para_exportacao_excel(df_vazio)
                df_vazio.to_excel(writer, sheet_name=nome_aba, index=False)
                ajustar_largura_colunas(writer, nome_aba, df_vazio)
                aplicar_formatacao_colunas(writer, nome_aba, df_vazio)
            else:
                df_export = df_consulta.copy()
                if "cc" in df_export.columns:
                    df_export["cc"] = df_export["cc"].apply(limpar_cc)
                if "Função" in df_export.columns:
                    df_export["Função"] = df_export["Função"].apply(limpar_funcao)

                colunas_tecnicas = [
                    "Arquivo Origem", "Matrícula", "Data_dt", "H.E. Min",
                    "Absent. Min", "Jornada Min", "Ad. Not. Min", "Saldo Atual Num",
                ]
                for col in colunas_tecnicas:
                    if col in df_export.columns and nome not in {"Ranking HE Colab", "Ranking HE Dias", "Alertas HE"}:
                        df_export = df_export.drop(columns=[col])

                colunas_preferidas = [c for c in COLUNAS_EXPORTACAO if c in df_export.columns]
                colunas_restantes = [c for c in df_export.columns if c not in colunas_preferidas]
                df_export = df_export[colunas_preferidas + colunas_restantes]
                df_export = preparar_para_exportacao_excel(df_export)

                df_export.to_excel(writer, sheet_name=nome_aba, index=False)
                ajustar_largura_colunas(writer, nome_aba, df_export)
                aplicar_formatacao_colunas(writer, nome_aba, df_export)

                if nome in consultas_graficos:
                    ws = writer.sheets[nome_aba]
                    categoria_nome, valor_nome, titulo = consultas_graficos[nome]
                    categoria_col = list(df_export.columns).index(categoria_nome) + 1
                    valor_col = list(df_export.columns).index(valor_nome) + 1
                    adicionar_grafico_coluna(
                        ws=ws,
                        min_row=1,
                        max_row=len(df_export) + 1,
                        categoria_col=categoria_col,
                        valor_col=valor_col,
                        titulo=titulo,
                        ancoragem=f"{chr(65 + min(len(df_export.columns) + 1, 12))}2",
                    )


def normalizar_feriados_fixos(feriados):
    resultado = set()
    for valor in feriados or []:
        texto = str(valor).strip()
        if not texto:
            continue
        try:
            dt = datetime.strptime(texto, "%d/%m")
            resultado.add((dt.month, dt.day))
        except ValueError:
            LOGGER.warning(f"Feriado fixo inválido ignorado: {texto}. Use o formato dd/mm.")
    return resultado


def normalizar_feriados_especificos(feriados):
    resultado = set()
    for valor in feriados or []:
        texto = str(valor).strip()
        if not texto:
            continue
        try:
            dt = pd.to_datetime(texto, format="%d/%m/%Y", errors="raise")
            resultado.add(pd.Timestamp(dt).normalize())
        except Exception:
            LOGGER.warning(f"Feriado específico inválido ignorado: {texto}. Use o formato dd/mm/aaaa.")
    return resultado


def calcular_pascoa(ano):
    a = ano % 19
    b = ano // 100
    c = ano % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    mes = (h + l - 7 * m + 114) // 31
    dia = ((h + l - 7 * m + 114) % 31) + 1
    return pd.Timestamp(year=ano, month=mes, day=dia)


def obter_feriados_moveis(ano):
    pascoa = calcular_pascoa(ano)
    return {
        (pascoa - pd.Timedelta(days=48)).normalize(),
        (pascoa - pd.Timedelta(days=47)).normalize(),
        (pascoa - pd.Timedelta(days=2)).normalize(),
        (pascoa + pd.Timedelta(days=60)).normalize(),
    }


def eh_feriado(data, feriados_fixos=None, feriados_especificos=None):
    data = pd.Timestamp(data).normalize()
    fixos = normalizar_feriados_fixos(feriados_fixos)
    especificos = normalizar_feriados_especificos(feriados_especificos)
    moveis = obter_feriados_moveis(data.year)
    return (data.month, data.day) in fixos or data in especificos or data in moveis


def eh_dia_util(data, feriados_fixos=None, feriados_especificos=None):
    data = pd.Timestamp(data).normalize()
    return data.weekday() < 5 and not eh_feriado(
        data,
        feriados_fixos=feriados_fixos,
        feriados_especificos=feriados_especificos,
    )


def obter_dia_util_anterior(datas_unicas):
    hoje = pd.Timestamp.now().normalize()
    datas_validas = [
        pd.Timestamp(data).normalize()
        for data in datas_unicas
        if pd.Timestamp(data).normalize() < hoje
        and eh_dia_util(
            data,
            feriados_fixos=CONFIG.feriados_fixos,
            feriados_especificos=CONFIG.feriados_especificos,
        )
    ]
    if not datas_validas:
        raise ValueError("Não foi encontrada nenhuma data útil anterior a hoje no consolidado.")
    return max(datas_validas)


def obter_datas_validas_dataframe(df):
    if "Data" not in df.columns:
        raise ValueError("O dataframe não possui a coluna 'Data'.")
    datas = pd.to_datetime(df["Data"], format="%d/%m/%Y", errors="coerce")
    datas_unicas = sorted(pd.Series(datas.dropna().unique()).tolist())
    if not datas_unicas:
        raise ValueError("Não foi possível identificar datas válidas no consolidado.")
    return [pd.Timestamp(data).normalize() for data in datas_unicas]


def obter_data_referencia_dataframe(df, modo_selecao="penultima", data_manual=None):
    datas_unicas = obter_datas_validas_dataframe(df)
    modo_normalizado = str(modo_selecao).strip().lower()

    if modo_normalizado == "ultima":
        return pd.Timestamp(datas_unicas[-1]).normalize(), "ultima"
    if modo_normalizado == "penultima":
        if len(datas_unicas) < 2:
            raise ValueError("Não foi possível identificar a penúltima data do consolidado, pois existe apenas uma data válida.")
        return pd.Timestamp(datas_unicas[-2]).normalize(), "penultima"
    if modo_normalizado == "d-1-util":
        return obter_dia_util_anterior(datas_unicas), "d-1-util"
    if modo_normalizado == "manual":
        if not data_manual or not str(data_manual).strip():
            raise ValueError("No modo manual, informe a data desejada em DATA_MANUAL_ARQUIVO no formato dd/mm/aaaa.")
        data_manual_dt = pd.to_datetime(str(data_manual).strip(), format="%d/%m/%Y", errors="coerce")
        if pd.isna(data_manual_dt):
            raise ValueError("A data informada em DATA_MANUAL_ARQUIVO é inválida. Use o formato dd/mm/aaaa.")
        data_manual_ts = pd.Timestamp(data_manual_dt).normalize()
        datas_normalizadas = [pd.Timestamp(d).normalize() for d in datas_unicas]
        if data_manual_ts not in datas_normalizadas:
            datas_disponiveis = ", ".join(pd.Timestamp(data).strftime("%d/%m/%Y") for data in datas_unicas)
            raise ValueError(
                f"A data manual informada ({data_manual_ts.strftime('%d/%m/%Y')}) não existe no consolidado. "
                f"Datas disponíveis: {datas_disponiveis}"
            )
        return data_manual_ts, "manual"
    raise ValueError("Modo de seleção de data inválido. Use 'ultima', 'penultima', 'd-1-util' ou 'manual'.")


def obter_rotulo_modo_data(modo_selecao):
    modo_normalizado = str(modo_selecao).strip().lower()
    if modo_normalizado == "ultima":
        return "ultima_data", "Última Data", "última data disponível"
    if modo_normalizado == "penultima":
        return "penultima_data", "Penultima Data", "penúltima data disponível"
    if modo_normalizado == "d-1-util":
        return "d_1_util", "D-1 Util", "último dia útil anterior disponível"
    if modo_normalizado == "manual":
        return "data_manual", "Data Manual", "data manual informada"
    raise ValueError("Modo de seleção de data inválido para geração de rótulo.")


def exportar_consolidado_data_referencia(caminho_excel_origem, caminho_excel_destino, modo_selecao="penultima", data_manual=None):
    df_consolidado = pd.read_excel(caminho_excel_origem, sheet_name="Consolidado")
    if df_consolidado.empty:
        raise ValueError("A aba 'Consolidado' está vazia.")
    data_referencia, modo_resolvido = obter_data_referencia_dataframe(
        df_consolidado,
        modo_selecao=modo_selecao,
        data_manual=data_manual,
    )
    mascara = pd.to_datetime(df_consolidado["Data"], format="%d/%m/%Y", errors="coerce").dt.normalize() == data_referencia
    df_filtrado = df_consolidado.loc[mascara].copy()
    if df_filtrado.empty:
        raise ValueError("Nenhum registro encontrado para a data de referência selecionada no consolidado.")

    _, nome_aba, descricao_modo = obter_rotulo_modo_data(modo_resolvido)
    with pd.ExcelWriter(caminho_excel_destino, engine="openpyxl") as writer:
        df_filtrado.to_excel(writer, sheet_name=nome_aba, index=False)
        ajustar_largura_colunas(writer, nome_aba, df_filtrado)
        aplicar_formatacao_colunas(writer, nome_aba, df_filtrado)
    return data_referencia.strftime("%d/%m/%Y"), len(df_filtrado), modo_resolvido, descricao_modo


def enviar_email_outlook(destinatarios, assunto, corpo, anexos=None, cc=None, cco=None, exibir_antes_de_enviar=False):
    if not destinatarios:
        raise ValueError("Informe ao menos um destinatário para o envio por Outlook.")
    if win32 is None:
        raise ImportError("pywin32 não está instalado neste ambiente. Instale com: pip install pywin32")
    outlook = win32.Dispatch("Outlook.Application")
    email = outlook.CreateItem(0)

    if isinstance(destinatarios, (list, tuple, set)):
        email.To = "; ".join(str(x).strip() for x in destinatarios if str(x).strip())
    else:
        email.To = str(destinatarios).strip()

    if cc:
        email.CC = "; ".join(str(x).strip() for x in cc) if isinstance(cc, (list, tuple, set)) else str(cc).strip()
    if cco:
        email.BCC = "; ".join(str(x).strip() for x in cco) if isinstance(cco, (list, tuple, set)) else str(cco).strip()

    email.Subject = assunto
    email.Body = corpo

    for anexo in anexos or []:
        caminho_anexo = Path(anexo).resolve()
        if not caminho_anexo.exists():
            raise FileNotFoundError(f"Anexo não encontrado: {caminho_anexo}")
        email.Attachments.Add(str(caminho_anexo))

    if exibir_antes_de_enviar:
        email.Display()
    else:
        email.Send()


def montar_assunto_email(data_referencia, descricao_modo, assunto_base=None):
    assunto_base = str(assunto_base or "Consolidado").strip()
    descricao_formatada = str(descricao_modo or "data de referência").strip()
    descricao_formatada = descricao_formatada[:1].upper() + descricao_formatada[1:]
    return f"{assunto_base} - {descricao_formatada} - {data_referencia}"


def montar_corpo_email(data_referencia, quantidade_registros, nome_arquivo, descricao_modo):
    return (
        f"Olá,\n\n"
        f"Segue em anexo o registo de ponto referente à {descricao_modo}: {data_referencia}.\n"
        f"Quantidade de registros no arquivo: {quantidade_registros}.\n"
        f"Arquivo: {nome_arquivo}.\n\n"
        f"E-mail enviado automaticamente pelo robô de tratamento de ponto.\n"
    )


def main() -> int:
    validar_configuracao()

    LOGGER.info("============================================================")
    LOGGER.info("Execução do robô de ponto iniciada")
    LOGGER.info(f"Modo de execução: {CONFIG.modo_execucao}")
    LOGGER.info(f"Modo de seleção de data: {CONFIG.modo_data_arquivo}")
    LOGGER.info(f"Pasta de entrada: {CONFIG.pasta_entrada}")
    LOGGER.info(f"Pasta de processados: {CONFIG.pasta_processados}")
    LOGGER.info(f"Pasta de erro: {CONFIG.pasta_erro}")
    LOGGER.info(f"Pasta de saída: {CONFIG.pasta_saida}")

    if modo_teste_ativo():
        LOGGER.warning("MODO TESTE ATIVO: envio de e-mail será ignorado e arquivos podem permanecer na entrada.")

    caminho_saida = CONFIG.caminho_saida_excel
    registros, erros, arquivos_pdf, arquivos_ok, arquivos_erro = processar_pasta(CONFIG.pasta_entrada)

    if not registros:
        detalhe = "Nenhum registro extraído dos PDFs processados."
        LOGGER.warning(detalhe)
        exportar_log_erros(erros)
        salvar_resultado_execucao(
            status="sem_registros",
            total_arquivos_pdf=len(arquivos_pdf),
            total_registros=0,
            total_erros=len(erros),
            detalhe=detalhe,
            erros=[{"arquivo": arq, "erro": erro} for arq, erro in erros],
        )
        LOGGER.info("============================================================")
        return 1

    exportar_excel_com_consultas(
        registros,
        caminho_saida,
        total_pdfs=len(arquivos_pdf),
        total_erros=len(erros),
    )
    exportar_log_erros(erros)

    sufixo_arquivo_data, _, _ = obter_rotulo_modo_data(CONFIG.modo_data_arquivo)
    caminho_saida_data_referencia = CONFIG.pasta_saida / f"resultado_consolidado_{sufixo_arquivo_data}.xlsx"

    data_referencia, qtd_registros_data_referencia, modo_resolvido, descricao_modo_data = exportar_consolidado_data_referencia(
        caminho_saida,
        caminho_saida_data_referencia,
        modo_selecao=CONFIG.modo_data_arquivo,
        data_manual=CONFIG.data_manual_arquivo,
    )

    LOGGER.info(f"Excel gerado com sucesso: {caminho_saida}")
    LOGGER.info(f"Arquivo da data de referência gerado com sucesso: {caminho_saida_data_referencia}")
    LOGGER.info(f"Critério de seleção de data utilizado: {modo_resolvido}")
    LOGGER.info(f"Data de referência identificada no consolidado: {data_referencia}")
    LOGGER.info(f"Registros no arquivo da data de referência: {qtd_registros_data_referencia}")
    LOGGER.info(f"Total de registros extraídos: {len(registros)}")
    LOGGER.info(f"Arquivos processados com sucesso: {len(arquivos_ok)}")
    LOGGER.info(f"Arquivos com erro: {len(arquivos_erro)}")

    if erros:
        LOGGER.warning(f"Foi gerado o arquivo de erros: {CONFIG.arquivo_erros}")
    else:
        LOGGER.info("Nenhum erro encontrado no processamento.")

    deve_enviar_email = CONFIG.enviar_email_automaticamente and not modo_teste_ativo()
    aviso_email = ""
    if deve_enviar_email:
        assunto_email = montar_assunto_email(
            data_referencia=data_referencia,
            descricao_modo=descricao_modo_data,
            assunto_base=CONFIG.assunto_email_base,
        )
        corpo_email = montar_corpo_email(
            data_referencia=data_referencia,
            quantidade_registros=qtd_registros_data_referencia,
            nome_arquivo=caminho_saida_data_referencia.name,
            descricao_modo=descricao_modo_data,
        )
        try:
            enviar_email_outlook(
                destinatarios=CONFIG.destinatarios_email,
                assunto=assunto_email,
                corpo=corpo_email,
                anexos=[caminho_saida_data_referencia],
                cc=CONFIG.cc_email,
                cco=CONFIG.cco_email,
                exibir_antes_de_enviar=CONFIG.exibir_email_antes_de_enviar,
            )
            modo_envio = "aberto para revisão no Outlook" if CONFIG.exibir_email_antes_de_enviar else "enviado automaticamente"
            LOGGER.info(f"E-mail {modo_envio} com o anexo: {caminho_saida_data_referencia}")
        except Exception as e:
            aviso_email = f"Excel gerado, mas o envio por Outlook foi ignorado: {e}"
            LOGGER.warning(aviso_email)
    else:
        if modo_teste_ativo():
            LOGGER.info("Envio por Outlook ignorado porque o robô está em MODO TESTE.")
        else:
            LOGGER.info("Envio por Outlook desabilitado.")

    status_final = "sucesso" if not erros else "parcial"
    detalhe_final = "Processamento concluído."
    if aviso_email:
        detalhe_final += " " + aviso_email
    salvar_resultado_execucao(
        status=status_final,
        total_arquivos_pdf=len(arquivos_pdf),
        total_registros=len(registros),
        total_erros=len(erros),
        caminho_excel=str(caminho_saida),
        caminho_excel_data_referencia=str(caminho_saida_data_referencia),
        data_referencia=data_referencia,
        modo_data=modo_resolvido,
        detalhe=detalhe_final,
        erros=[{"arquivo": arq, "erro": erro} for arq, erro in erros],
    )
    registrar_historico_execucao(
        status=status_final,
        total_arquivos_pdf=len(arquivos_pdf),
        total_registros=len(registros),
        total_erros=len(erros),
        data_referencia=data_referencia,
        modo_data=modo_resolvido,
        caminho_excel=str(caminho_saida),
    )

    LOGGER.info("Execução do robô de ponto encerrada")
    LOGGER.info("============================================================")
    return 0 if not erros else 1


if __name__ == "__main__":
    try:
        with lock_execucao():
            sys.exit(main())
    except Exception as e:
        LOGGER.exception(f"Falha fatal no robô de ponto: {e}")
        try:
            salvar_resultado_execucao(
                status="falha",
                total_arquivos_pdf=0,
                total_registros=0,
                total_erros=1,
                detalhe=str(e),
                erros=[{"arquivo": "execucao", "erro": str(e)}],
            )
        except Exception:
            pass
        sys.exit(1)
