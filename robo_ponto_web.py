
from pathlib import Path
from flask import Flask, render_template, jsonify, request, send_file, redirect, url_for, session
import base64
import gzip
from io import BytesIO
import pandas as pd
import os
import shutil
import re
import json
import hashlib
import sqlite3
import hmac
import smtplib
import urllib.request
import time
import threading
import sys
import subprocess
import socket
import platform
import uuid
from email.message import EmailMessage
from functools import wraps
from werkzeug.utils import secure_filename

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

try:
    import pdfplumber
except Exception:
    pdfplumber = None
from datetime import datetime
from user_db import (
    inicializar_banco_usuarios,
    autenticar_usuario,
    listar_usuarios,
    obter_usuario,
    criar_ou_atualizar_usuario,
    alterar_senha_usuario,
    definir_status_usuario,
    excluir_usuario,
)

try:
    from usuarios_config import USUARIOS as USUARIOS_LEGADO
except Exception:
    USUARIOS_LEGADO = None


# v8.17.20 
def garantir_estrutura_ponto_pdfs_v81720():
    try:
        from pathlib import Path
        base = Path.home() / "ponto_pdfs"
        for nome in ["entrada", "saida", "processados", "logs", "backup", "erro", "config", "controle", "backup_atualizacao"]:
            (base / nome).mkdir(parents=True, exist_ok=True)
        return str(base)
    except Exception as e:
        try:
            LOGGER.warning(f"Falha ao criar ponto_pdfs: {e}")
        except Exception:
            pass
        return ""
PASTA_PONTO_PDFS_AUTO = garantir_estrutura_ponto_pdfs_v81720()

app = Flask(__name__)
app.secret_key = os.environ.get("TEMPO_FECHADO_SECRET_KEY") or "bbf4b37477c714f4df81447fcdfdc58e1be480cc0342885f"
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["MAX_CONTENT_LENGTH"] = 250 * 1024 * 1024

# v8.21.38 
APP_VERSION = "v8.21.38"
APP_RELEASE_NAME = "Upload Manual PDFs"
APP_FULL_NAME = "Tempo Fechado v8.21.38 - Upload Manual PDFs"
APP_STARTED_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
APP_BASE_DIR = str(Path(__file__).resolve().parent)


# v8.12.02 
@app.after_request
def aplicar_gzip_respostas(response):
    try:
        if request.path in ("/static/app.js", "/static/style.css"):
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
    except Exception:
        pass

    try:
        if (
            response.status_code < 200
            or response.status_code >= 300
            or response.direct_passthrough
            or response.headers.get("Content-Encoding")
            or "gzip" not in request.headers.get("Accept-Encoding", "").lower()
        ):
            return response

        content_type = (response.headers.get("Content-Type") or "").lower()
        if not any(tipo in content_type for tipo in ("application/json", "text/html", "text/css", "application/javascript", "text/javascript")):
            return response

        dados = response.get_data()
        if not dados or len(dados) < 1024:
            return response

        buffer = BytesIO()
        with gzip.GzipFile(mode="wb", fileobj=buffer, compresslevel=5) as gz:
            gz.write(dados)
        response.set_data(buffer.getvalue())
        response.headers["Content-Encoding"] = "gzip"
        response.headers["Vary"] = "Accept-Encoding"
        response.headers["Content-Length"] = str(len(response.get_data()))
    except Exception:
        return response
    return response

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
ARQUIVO_PADRAO = DATA_DIR / "resultado_lote_consultas_validado.xlsx"
ARQUIVO_BASE_ATIVA_FLAG = DATA_DIR / "base_ativa.flag"
ARQUIVO_SALDOS_BASE_MEDICAO = DATA_DIR / "saldos_base_medicao.json"
ARQUIVO_ALERTAS_JORNADA_ENVIADOS = DATA_DIR / "alertas_jornada_enviados.json"
ARQUIVO_LOG_NOTIFICACOES_JORNADA = DATA_DIR / "notificacoes_jornada.log"
# v8.17.21 - Configuração editável por usuário, fora da pasta do executável/Program Files.
USER_CONFIG_DIR = Path.home() / "ponto_pdfs" / "config"
ARQUIVO_CONFIG_NOTIFICACOES_JORNADA = USER_CONFIG_DIR / "alertas_notificacao_config.json"
ARQUIVO_AUDITORIA_MULTIUSUARIO = DATA_DIR / "auditoria_multiusuario.log"
ARQUIVO_USUARIOS_DB = DATA_DIR / "usuarios.db"
ARQUIVO_CONFIG_ROBO = DATA_DIR / "configuracoes_robo.json"
ARQUIVO_CONFIG_ROBO_USUARIO_V81743 = USER_CONFIG_DIR / "configuracoes_robo.json"
ARQUIVO_CONFIG_EMAIL_ROBO_V81727 = Path.home() / "ponto_pdfs" / "config" / "config_email_robo.json"
ARQUIVO_CACHE_PDFS_BANCO_HORAS = DATA_DIR / "catalogo_pdfs_banco_horas.json"
ARQUIVO_RESUMO_BANCO_HORAS_PDF = DATA_DIR / "resumo_banco_horas_pdf.json"
ARQUIVO_CONFIG_BH2 = DATA_DIR / "banco_horas_2_config.json"
ARQUIVO_CACHE_BH2_CALCULO = DATA_DIR / "banco_horas_2_calculo_cache.json"
_CACHE_BH2_CALCULO_MEM = {"key": None, "payload": None, "gerado_em_ts": 0}
_CACHE_SCORE_OPERACIONAL_MEM = {"key": None, "payload": None, "gerado_em_ts": 0.0}

# v8.10.43 
BH_CORRECAO_LOCK = threading.Lock()
BH_CORRECAO_JOB = {
    "status": "ocioso",
    "ok": True,
    "mensagem": "Nenhuma correção em andamento.",
    "iniciado_em": "",
    "finalizado_em": "",
    "segundos": 0,
    "resultado": {},
}



# v8.11.12 
ARQUIVO_LOG_PROCESSAMENTO_INTEGRADO = DATA_DIR / "processamento_integrado.log"
ARQUIVO_HISTORICO_OPERACAO_V8190 = DATA_DIR / "historico_operacao.jsonl"
ARQUIVO_PAINEL_IMPLANTACAO_V8210 = DATA_DIR / "painel_implantacao_notebooks.json"
PROCESSAMENTO_INTEGRADO_LOCK = threading.Lock()
PROCESSAMENTO_INTEGRADO_JOB = {
    "status": "ocioso",
    "ok": True,
    "fase": "Aguardando",
    "mensagem": "Nenhum processamento em andamento.",
    "iniciado_em": "",
    "finalizado_em": "",
    "segundos": 0,
    "resultado": {},
    "log_tail": "",
}

# v8.10.45 
_CACHE_PDFS_MEM = {"mtime": None, "payload": None}

# v8.10.57 - cache em memória da base Consolidado.
# v8.10.63 - clean-up da guia Configuração do Robô: topbar simplificada nessa guia.
# Evita reler o Excel e refazer a normalização completa a cada troca de guia/filtro.
# O cache é invalidado automaticamente quando o arquivo importado muda.
_CACHE_CONSOLIDADO_MEM = {
    "sig": None,
    "df_raw": None,
    "df_norm": None,
    "gerado_em": "",
}

# v8.11.21 - cache leve do Dashboard Executivo para evitar travamento em navegação/filtros.
_CACHE_OPCOES_FILTROS_MEM = {"sig": None, "ccs": [], "nomes": [], "turnos": [], "gerado_em": ""}

_CACHE_DASHBOARD_EXECUTIVO_MEM = {
    "key": None,
    "payload": None,
    "gerado_em_ts": 0.0,
}

_CACHE_DASHBOARD_BANCO_HORAS_MEM = {
    "key": None,
    "payload": None,
    "gerado_em_ts": 0.0,
}

# Cache leve de metadados/abas do Excel. Abrir um .xlsx grande apenas para
# listar abas ou descobrir o modelo custa tempo perceptivel em rede/disco lento.
_CACHE_EXCEL_META_MEM = {
    "sig": None,
    "abas": [],
    "aba_historico": None,
    "aba_saldos_bh": None,
    "modelo": None,
}
_CACHE_EXCEL_ABAS_MEM = {}
_CACHE_EXCEL_LOCK = threading.RLock()


def _assinatura_arquivo(path):
    try:
        st = Path(path).stat()
        return (int(st.st_mtime_ns), int(st.st_size))
    except Exception:
        return None


def _chave_cache_excel(path):
    caminho = Path(path)
    try:
        caminho_txt = str(caminho.resolve())
    except Exception:
        caminho_txt = str(caminho)
    return (caminho_txt, _assinatura_arquivo(caminho))


def invalidar_cache_consolidado():
    try:
        _CACHE_CONSOLIDADO_MEM["sig"] = None
        _CACHE_CONSOLIDADO_MEM["df_raw"] = None
        _CACHE_CONSOLIDADO_MEM["df_norm"] = None
        _CACHE_CONSOLIDADO_MEM["gerado_em"] = ""
        # v8.11.21: quando a base muda, o Dashboard Executivo também precisa
        # perder o cache. Manter payload antigo após importação podia gerar
        # tela pesada/inconsistente e sensação de travamento.
        _CACHE_DASHBOARD_EXECUTIVO_MEM["key"] = None
        _CACHE_DASHBOARD_EXECUTIVO_MEM["payload"] = None
        _CACHE_DASHBOARD_EXECUTIVO_MEM["gerado_em_ts"] = 0.0
        _CACHE_DASHBOARD_BANCO_HORAS_MEM["key"] = None
        _CACHE_DASHBOARD_BANCO_HORAS_MEM["payload"] = None
        _CACHE_DASHBOARD_BANCO_HORAS_MEM["gerado_em_ts"] = 0.0
        _CACHE_SCORE_OPERACIONAL_MEM["key"] = None
        _CACHE_SCORE_OPERACIONAL_MEM["payload"] = None
        _CACHE_SCORE_OPERACIONAL_MEM["gerado_em_ts"] = 0.0
        _CACHE_OPCOES_FILTROS_MEM["sig"] = None
        _CACHE_OPCOES_FILTROS_MEM["ccs"] = []
        _CACHE_OPCOES_FILTROS_MEM["nomes"] = []
        _CACHE_OPCOES_FILTROS_MEM["turnos"] = []
        _CACHE_OPCOES_FILTROS_MEM["gerado_em"] = ""
        _CACHE_CONSOLIDADO_MEM.pop("sig_historico", None)
        _CACHE_CONSOLIDADO_MEM.pop("df_historico_raw", None)
        _CACHE_CONSOLIDADO_MEM.pop("df_historico_norm", None)
        with _CACHE_EXCEL_LOCK:
            _CACHE_EXCEL_META_MEM["sig"] = None
            _CACHE_EXCEL_META_MEM["abas"] = []
            _CACHE_EXCEL_META_MEM["aba_historico"] = None
            _CACHE_EXCEL_META_MEM["aba_saldos_bh"] = None
            _CACHE_EXCEL_META_MEM["modelo"] = None
            _CACHE_EXCEL_ABAS_MEM.clear()
    except Exception:
        pass


def _preaquecer_base_excel_em_segundo_plano():
    """Carrega e normaliza o Excel recém-importado fora da requisição HTTP.

    v8.11.21: a importação deve responder rápido. A leitura/normalização do
    Excel pode levar alguns segundos em bases grandes; por isso ela é aquecida
    em thread separada, sem bloquear navegação nem botões.
    """
    def _worker():
        try:
            if not ARQUIVO_PADRAO.exists():
                return
            df = _ler_excel_consolidado_otimizado(ARQUIVO_PADRAO).fillna("")
            sig = _assinatura_arquivo(ARQUIVO_PADRAO)
            _CACHE_CONSOLIDADO_MEM["sig"] = sig
            _CACHE_CONSOLIDADO_MEM["df_raw"] = df
            _CACHE_CONSOLIDADO_MEM["df_norm"] = _normalizar_colunas_impl(df)
            _CACHE_CONSOLIDADO_MEM["gerado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            pass
    try:
        threading.Thread(target=_worker, daemon=True).start()
    except Exception:
        pass


inicializar_banco_usuarios(ARQUIVO_USUARIOS_DB, USUARIOS_LEGADO)

# ============================================================
# INTEGRAÇÃO COM O ROBÔ DE PONTO
# ============================================================
# v8.10.23: importação rápida + cache dos PDFs. Estes caminhos são configuráveis pela interface
# Administração > Configuração do Robô. Mantemos os valores abaixo como padrão
# para instalações antigas e para o primeiro carregamento do sistema.
ARQUIVO_EXCEL_ROBO_PADRAO = Path.home() / "ponto_pdfs" / "saida" / "resultado_lote_consultas_validado.xlsx"
PASTA_RAIZ_PDFS_ROBO_PADRAO = ARQUIVO_EXCEL_ROBO_PADRAO.parent.parent

ARQUIVO_EXCEL_ROBO = ARQUIVO_EXCEL_ROBO_PADRAO
PASTA_RAIZ_PDFS_ROBO = PASTA_RAIZ_PDFS_ROBO_PADRAO
PASTAS_PDFS_ROBO = (
    PASTA_RAIZ_PDFS_ROBO / "entrada",
    PASTA_RAIZ_PDFS_ROBO / "processados",
    PASTA_RAIZ_PDFS_ROBO / "erro",
    PASTA_RAIZ_PDFS_ROBO,
)


def _normalizar_path_windows(valor, fallback):
    texto = str(valor or "").strip().strip('"').strip("'")
    if not texto:
        return Path(fallback)
    try:
        return Path(texto).expanduser()
    except Exception:
        return Path(fallback)


def _montar_pastas_pdfs_robo(pasta_raiz):
    raiz = Path(pasta_raiz)
    # Mantém compatibilidade com a organização clássica do robô
    # e permite que o usuário aponte diretamente para uma pasta diferente.
    return (
        raiz / "entrada",
        raiz / "processados",
        raiz / "erro",
        raiz,
    )


def _config_robo_padrao():
    return {
        "arquivo_excel_robo": str(ARQUIVO_EXCEL_ROBO_PADRAO),
        "pasta_pdfs_robo": str(PASTA_RAIZ_PDFS_ROBO_PADRAO),
        "script_launcher_robo": "",
        "auto_importar_excel_robo": True,
        "corrigir_saldo_atual_por_pdf_ao_importar": False,
        "max_pdfs_correcao_saldo": 1500,
        "max_segundos_correcao_saldo": 180,
        # v8.17.43: politica configuravel na Administracao do Robo.
        # False = captura apenas e-mails nao lidos; True = permite incluir lidos recentes.
        "capturar_emails_lidos": False,
        "limite_emails_outlook": 500,
        "remetente_filtro": os.environ.get("TEMPO_FECHADO_REMETENTE_FILTRO") or "",
        "assunto_filtro": "Análise de controle de marcações",
    }


def carregar_config_robo():
    cfg = _config_robo_padrao()
    # v8.17.43: le tambem a copia editavel em ~/ponto_pdfs/config,
    # usada pelo launcher quando o app roda empacotado. O arquivo local em data/
    # permanece compatível com as versoes anteriores.
    for arquivo_cfg in (ARQUIVO_CONFIG_ROBO, ARQUIVO_CONFIG_ROBO_USUARIO_V81743):
        try:
            if arquivo_cfg.exists():
                with arquivo_cfg.open("r", encoding="utf-8") as f:
                    salvo = json.load(f)
                if isinstance(salvo, dict):
                    cfg.update({k: v for k, v in salvo.items() if v is not None})
        except Exception:
            pass
    return cfg


def _script_launcher_persistido_invalido_v82123(caminho):
    texto = str(caminho or "").strip().strip('"').strip("'")
    if not texto:
        return True
    baixo = texto.replace("/", "\\").lower()
    if "\\dist\\tempofechado\\" in baixo or "\\dist\\tempo fechado\\" in baixo:
        return True
    if baixo.startswith("c:\\projetos\\"):
        return True
    try:
        p = Path(texto).expanduser()
        return not (p.exists() and p.is_file())
    except Exception:
        return True


def _script_launcher_padrao_runtime_v82123():
    try:
        return str(_localizar_script_launcher({}))
    except Exception:
        return str(BASE_DIR / "Script_Launcher_v5_filtrado.py")


def normalizar_config_robo_runtime_v82123(cfg):
    cfg = dict(cfg or {})
    if _script_launcher_persistido_invalido_v82123(cfg.get("script_launcher_robo")):
        cfg["script_launcher_robo"] = _script_launcher_padrao_runtime_v82123()
        cfg["_script_launcher_robo_autocorrigido"] = True
    else:
        cfg["_script_launcher_robo_autocorrigido"] = False
    return cfg


def _normalizar_limite_emails_outlook_v8216(valor):
    try:
        limite = int(valor or 500)
    except Exception:
        limite = 500
    return max(50, min(limite, 5000))


def salvar_config_robo(cfg):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    USER_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    atual = carregar_config_robo()
    atual.update(cfg or {})
    atual = normalizar_config_robo_runtime_v82123(atual)
    for arquivo_cfg in (ARQUIVO_CONFIG_ROBO, ARQUIVO_CONFIG_ROBO_USUARIO_V81743):
        try:
            arquivo_cfg.parent.mkdir(parents=True, exist_ok=True)
            with arquivo_cfg.open("w", encoding="utf-8") as f:
                json.dump(atual, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    aplicar_config_robo(atual)
    return atual


def aplicar_config_robo(cfg=None):
    global ARQUIVO_EXCEL_ROBO, PASTA_RAIZ_PDFS_ROBO, PASTAS_PDFS_ROBO
    global AUTO_IMPORTAR_EXCEL_ROBO, CORRIGIR_SALDO_ATUAL_POR_PDF_AO_IMPORTAR
    global MAX_PDFS_CORRECAO_SALDO, MAX_SEGUNDOS_CORRECAO_SALDO, CAPTURAR_EMAILS_LIDOS_ROBO

    cfg = normalizar_config_robo_runtime_v82123(cfg or carregar_config_robo())
    ARQUIVO_EXCEL_ROBO = _normalizar_path_windows(cfg.get("arquivo_excel_robo"), ARQUIVO_EXCEL_ROBO_PADRAO)
    PASTA_RAIZ_PDFS_ROBO = _normalizar_path_windows(cfg.get("pasta_pdfs_robo"), PASTA_RAIZ_PDFS_ROBO_PADRAO)
    PASTAS_PDFS_ROBO = _montar_pastas_pdfs_robo(PASTA_RAIZ_PDFS_ROBO)
    AUTO_IMPORTAR_EXCEL_ROBO = bool(cfg.get("auto_importar_excel_robo", True))
    CORRIGIR_SALDO_ATUAL_POR_PDF_AO_IMPORTAR = bool(cfg.get("corrigir_saldo_atual_por_pdf_ao_importar", True))
    CAPTURAR_EMAILS_LIDOS_ROBO = bool(cfg.get("capturar_emails_lidos", False))
    try:
        MAX_PDFS_CORRECAO_SALDO = int(cfg.get("max_pdfs_correcao_saldo", 1500) or 1500)
    except Exception:
        MAX_PDFS_CORRECAO_SALDO = 1500
    try:
        MAX_SEGUNDOS_CORRECAO_SALDO = int(cfg.get("max_segundos_correcao_saldo", 180) or 180)
    except Exception:
        MAX_SEGUNDOS_CORRECAO_SALDO = 180
    return cfg

# Carrega configurações persistidas em data/configuracoes_robo.json.
aplicar_config_robo()




def _canon_nome_banco(valor):
    texto = str(valor or "").upper().strip()
    texto = re.sub(r"\s+", " ", texto)
    texto = texto.replace("Á", "A").replace("À", "A").replace("Ã", "A").replace("Â", "A")
    texto = texto.replace("É", "E").replace("Ê", "E")
    texto = texto.replace("Í", "I")
    texto = texto.replace("Ó", "O").replace("Ô", "O").replace("Õ", "O")
    texto = texto.replace("Ú", "U").replace("Ç", "C")
    return texto




def _normalizar_matricula_bh(valor):
    """Normaliza matrícula/chapa para casar Excel x PDF.

    O PDF costuma trazer algo como "02 - 004858" enquanto algumas bases
    podem trazer apenas "004858" ou "4858". A comparação usa os últimos
    6 dígitos, com zero à esquerda quando necessário.
    """
    texto = str(valor or "")
    digitos = re.sub(r"\D", "", texto)
    if not digitos:
        return ""
    if len(digitos) >= 6:
        return digitos[-6:]
    return digitos.zfill(6)


def _matricula_da_linha_bh(row):
    if row is None:
        return ""
    candidatos = [
        "Matrícula", "Matricula", "MATRÍCULA", "MATRICULA",
        "Chapa", "CHAPA", "matricula", "chapa", "matrícula",
        "Matrícula PDF", "Matricula PDF",
    ]
    for col in candidatos:
        try:
            if col in row.index:
                mat = _normalizar_matricula_bh(row.get(col, ""))
                if mat:
                    return mat
        except Exception:
            pass

    # v8.10.21: fallback genérico. Alguns arquivos do robô chegam com
    # nomes de coluna levemente diferentes (ex.: "Matricula Colaborador",
    # "CHAPA_FUNC", etc.). Antes, a matrícula do Excel podia ficar vazia
    # mesmo existindo na linha, quebrando o match com o PDF.
    try:
        for col in getattr(row, "index", []):
            c = str(col or "").strip().lower()
            c = c.translate(str.maketrans("áàãâéêíóôõúç", "aaaaeeiooouc"))
            if "matric" in c or "chapa" in c:
                mat = _normalizar_matricula_bh(row.get(col, ""))
                if mat:
                    return mat
    except Exception:
        pass
    return ""


def _tokens_nome_bh(valor):
    texto = _canon_nome_banco(valor)
    ignorar = {"DA", "DE", "DO", "DAS", "DOS", "E"}
    return [t for t in re.findall(r"[A-Z0-9]+", texto) if len(t) >= 3 and t not in ignorar]


def _nomes_compativeis_bh(nome_a, nome_b):
    a = _canon_nome_banco(nome_a)
    b = _canon_nome_banco(nome_b)
    if not a or not b:
        return False
    if a == b or a.startswith(b) or b.startswith(a):
        return True
    ta, tb = set(_tokens_nome_bh(a)), set(_tokens_nome_bh(b))
    if not ta or not tb:
        return False
    inter = ta & tb
    # Casamento tolerante para nomes truncados ou invertidos pelo extrator.
    return len(inter) >= 3 and (len(inter) / max(1, min(len(ta), len(tb)))) >= 0.60

def _parse_data_pdf(valor):
    """Converte datas vindas do PDF, Excel ou pandas para Timestamp.

    v8.10.17:
    a regra anterior só aceitava texto no formato dd/mm/aaaa. Quando a
    coluna Data chegava do Excel como Timestamp/ISO ou como dd/m/aaaa,
    o match PDF x linha falhava e o Extrato caía na coluna antiga
    `saldo_atual`.
    """
    try:
        if valor is None:
            return pd.NaT
        if isinstance(valor, pd.Timestamp):
            return valor.normalize()
        # datetime/date nativos: pandas resolve sem depender do formato.
        if hasattr(valor, "year") and hasattr(valor, "month") and hasattr(valor, "day"):
            return pd.to_datetime(valor, errors="coerce").normalize()
        texto = str(valor).strip()
        if not texto or texto.lower() in {"nan", "none", "nat"}:
            return pd.NaT
        # Datas brasileiras do PDF, inclusive 15/5/2026.
        # IMPORTANTE: tentar dayfirst=True antes de ISO evita que 03/06/2026
        # seja interpretado como 06/03/2026 e que 01/06/2026 vire 06/01/2026.
        if re.match(r"^\d{1,2}/\d{1,2}/\d{4}", texto):
            dt_br = pd.to_datetime(texto, errors="coerce", dayfirst=True)
            if not pd.isna(dt_br):
                return dt_br.normalize()
        # Excel pode trazer ISO: 2026-05-15 00:00:00
        dt_iso = pd.to_datetime(texto, errors="coerce", dayfirst=False)
        if not pd.isna(dt_iso):
            return dt_iso.normalize()
        # Fallback final para textos ambíguos.
        dt_br = pd.to_datetime(texto, errors="coerce", dayfirst=True)
        if not pd.isna(dt_br):
            return dt_br.normalize()
        return pd.NaT
    except Exception:
        return pd.NaT


def _extrair_espelhos_banco_horas_pdf(caminho_pdf):
    """
    Extrai do PDF o quadro real do Banco de Horas:
    Saldo Anterior, Débito, Crédito e Saldo Atual.

    v8.10.13:
    - usa texto e também tabelas do pdfplumber;
    - pega somente datas de linhas diárias do espelho, não datas soltas do cabeçalho;
    - aceita nomes e layouts com pequenas variações;
    - registra metadados suficientes para diagnosticar quando o PDF não foi usado.
    """
    resultados = []
    if pdfplumber is None:
        return resultados

    def _datas_operacionais(texto):
        datas = []
        for linha in (texto or '').splitlines():
            linha = linha.strip()
            # Linhas operacionais do espelho começam com data + dia da semana.
            if re.match(r"^\d{2}/\d{2}/\d{4}\s+", linha):
                dt = _parse_data_pdf(linha[:10])
                if not pd.isna(dt):
                    datas.append(dt)
        return sorted(set(datas))

    def _resumo_por_texto(texto):
        linhas = [linha.strip() for linha in (texto or '').splitlines() if linha.strip()]
        for i, linha in enumerate(linhas):
            linha_norm = _canon_nome_banco(linha)
            if "BANCO DE HORAS" not in linha_norm:
                continue
            janela = linhas[i:i + 12]
            blob = "\n".join(janela)
            valores = re.findall(r"-?\d{1,5}[,.]\d{2}", blob)
            # Evita capturar valores da tabela diária acima: só aceita quando no bloco
            # também existem os rótulos do quadro de Banco de Horas.
            blob_norm = _canon_nome_banco(blob)
            if len(valores) >= 4 and "SALDO ANTERIOR" in blob_norm and "SALDO ATUAL" in blob_norm:
                return {
                    "saldo_anterior_pdf": valores[0].replace('.', ','),
                    "debito_bh_pdf": valores[1].replace('.', ','),
                    "credito_bh_pdf": valores[2].replace('.', ','),
                    "saldo_atual_pdf": valores[3].replace('.', ','),
                }
        return None

    def _resumo_por_tabela(pagina):
        try:
            tabelas = pagina.extract_tables() or []
        except Exception:
            tabelas = []
        for tabela in tabelas:
            linhas = []
            for row in tabela or []:
                celulas = [str(c or '').strip() for c in row]
                linhas.append(celulas)
            for i, celulas in enumerate(linhas):
                joined = " ".join(celulas)
                joined_norm = _canon_nome_banco(joined)
                if "BANCO DE HORAS" in joined_norm and "SALDO ANTERIOR" in joined_norm and "SALDO ATUAL" in joined_norm:
                    # Busca os valores na própria linha e nas próximas 3 linhas.
                    blob = " ".join(" ".join(x) for x in linhas[i:i+4])
                    valores = re.findall(r"-?\d{1,5}[,.]\d{2}", blob)
                    if len(valores) >= 4:
                        return {
                            "saldo_anterior_pdf": valores[0].replace('.', ','),
                            "debito_bh_pdf": valores[1].replace('.', ','),
                            "credito_bh_pdf": valores[2].replace('.', ','),
                            "saldo_atual_pdf": valores[3].replace('.', ','),
                        }
        return None

    try:
        with pdfplumber.open(str(caminho_pdf)) as pdf:
            for pagina in pdf.pages:
                texto = pagina.extract_text() or ""
                texto_norm = _canon_nome_banco(texto)
                if "BANCO DE HORAS" not in texto_norm or "SALDO ATUAL" not in texto_norm:
                    continue

                periodo = re.search(r"Espelho do Ponto\s+(\d{1,2}/\d{1,2}/\d{4})\s*-\s*(\d{1,2}/\d{1,2}/\d{4})", texto, re.I)
                nome = re.search(r"Nome:\s*(.+?)\s+Chapa:", texto, re.I | re.S)
                if not nome:
                    nome = re.search(r"Nome:\s*([^\n]+)", texto, re.I)
                matricula = re.search(r"Matr[ií]cula:\s*([^\n]+?)\s+Nome:", texto, re.I | re.S)
                if not matricula:
                    matricula = re.search(r"Matr[ií]cula:\s*([^\n]+)", texto, re.I)
                emissao = re.search(r"Emiss[aã]o:\s*(\d{1,2}/\d{1,2}/\d{4})", texto, re.I)

                resumo = _resumo_por_tabela(pagina) or _resumo_por_texto(texto)
                if not periodo or not nome or not resumo:
                    continue

                datas_unicas = _datas_operacionais(texto)
                datas_reais_pdf = bool(datas_unicas)
                if not datas_unicas:
                    # Fallback conservador: se o texto não trouxe as linhas diárias,
                    # usa início do período. A partir da v8.10.14 esse fallback não
                    # tem a mesma prioridade de um PDF em que as datas reais foram
                    # lidas no corpo do espelho, justamente para não abrir o ciclo
                    # com saldo de um PDF errado/incompleto.
                    dt_ini = _parse_data_pdf(periodo.group(1))
                    datas_unicas = [] if pd.isna(dt_ini) else [dt_ini]

                resultados.append({
                    "arquivo_pdf": Path(caminho_pdf).name,
                    "caminho_pdf": str(caminho_pdf),
                    "nome": nome.group(1).strip(),
                    "nome_norm": _canon_nome_banco(nome.group(1)),
                    "matricula": matricula.group(1).strip() if matricula else "",
                    "matricula_norm": _normalizar_matricula_bh(matricula.group(1)) if matricula else "",
                    "inicio": _parse_data_pdf(periodo.group(1)),
                    "fim": _parse_data_pdf(periodo.group(2)),
                    "emissao": _parse_data_pdf(emissao.group(1)) if emissao else pd.NaT,
                    "datas_lancadas": datas_unicas,
                    "datas_reais_pdf": datas_reais_pdf,
                    "data_max_lancada": max(datas_unicas) if datas_unicas else pd.NaT,
                    **resumo,
                })
    except Exception:
        return resultados

    return resultados



def _listar_pdfs_robo_limitado(max_pdfs=MAX_PDFS_CORRECAO_SALDO):
    """
    Lista PDFs do robô de forma controlada, mas agora com busca recursiva.

    v8.10.13:
    - a versão anterior olhava só a pasta raiz e um nível abaixo;
    - em algumas instalações os PDFs ficam em subpastas por data/lote;
    - isso fazia a correção do Saldo Atual não encontrar o PDF correto e o
      Extrato continuava caindo na coluna antiga do Excel.
    """
    arquivos = []
    vistos = set()
    for pasta in PASTAS_PDFS_ROBO:
        try:
            if not pasta.exists():
                continue
            for pdf in pasta.rglob("*.pdf"):
                try:
                    chave = str(pdf.resolve()).lower()
                    if chave in vistos:
                        continue
                    vistos.add(chave)
                    arquivos.append(pdf)
                except Exception:
                    continue
        except Exception:
            continue

    arquivos.sort(key=lambda x: x.stat().st_mtime if x.exists() else 0, reverse=True)
    return arquivos[:max_pdfs]

def _catalogar_saldos_pdf_robo(max_pdfs=MAX_PDFS_CORRECAO_SALDO, max_segundos=MAX_SEGUNDOS_CORRECAO_SALDO):
    entradas = []
    inicio = time.monotonic()
    for pdf in _listar_pdfs_robo_limitado(max_pdfs=max_pdfs):
        if time.monotonic() - inicio > max_segundos:
            break
        try:
            entradas.extend(_extrair_espelhos_banco_horas_pdf(pdf))
        except Exception:
            continue
    return entradas


def _serializar_valor_cache_pdf(valor):
    """Converte objetos pandas/Timestamp para JSON, preservando datas."""
    try:
        if valor is None or pd.isna(valor):
            return None
    except Exception:
        pass
    if isinstance(valor, (pd.Timestamp, datetime)):
        try:
            return valor.strftime("%Y-%m-%d")
        except Exception:
            return str(valor)
    if isinstance(valor, list):
        return [_serializar_valor_cache_pdf(v) for v in valor]
    if isinstance(valor, dict):
        return {str(k): _serializar_valor_cache_pdf(v) for k, v in valor.items()}
    return valor


def _normalizar_item_cache_pdf(item):
    """Reidrata datas do cache para Timestamp, para manter o match histórico funcionando."""
    if not isinstance(item, dict):
        return {}
    item = dict(item)
    for chave in ["inicio", "fim", "emissao", "data_min_lancada", "data_max_lancada"]:
        item[chave] = _parse_data_pdf(item.get(chave))
    datas = []
    for d in item.get("datas_lancadas") or []:
        dt = _parse_data_pdf(d)
        if not pd.isna(dt):
            datas.append(dt)
    item["datas_lancadas"] = datas
    return item


def salvar_cache_pdfs_banco_horas(catalogo, pdfs_encontrados=0, segundos=0):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "versao": "v8.10.23",
        "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "pasta_pdfs_robo": str(PASTA_RAIZ_PDFS_ROBO),
        "pastas_verificadas": [str(p) for p in PASTAS_PDFS_ROBO],
        "pdfs_encontrados": int(pdfs_encontrados or 0),
        "quadros_lidos": len(catalogo or []),
        "segundos": segundos,
        "catalogo": [_serializar_valor_cache_pdf(i) for i in (catalogo or [])],
    }
    with ARQUIVO_CACHE_PDFS_BANCO_HORAS.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    # força recarga em memória na próxima consulta
    try:
        _CACHE_PDFS_MEM.update({"mtime": None, "payload": None})
    except Exception:
        pass
    return payload


def carregar_cache_pdfs_banco_horas():
    """Carrega o catálogo dos PDFs com cache em memória.

    v8.10.45: várias telas chamavam esta função muitas vezes numa mesma
    navegação. Cada chamada relia o JSON e normalizava centenas/milhares de
    itens, deixando o sistema instável. Agora só recarrega se o arquivo mudar.
    """
    if not ARQUIVO_CACHE_PDFS_BANCO_HORAS.exists():
        return None
    try:
        mtime = ARQUIVO_CACHE_PDFS_BANCO_HORAS.stat().st_mtime
        if _CACHE_PDFS_MEM.get("payload") is not None and _CACHE_PDFS_MEM.get("mtime") == mtime:
            return _CACHE_PDFS_MEM.get("payload")

        with ARQUIVO_CACHE_PDFS_BANCO_HORAS.open("r", encoding="utf-8") as f:
            payload = json.load(f)
        catalogo = [_normalizar_item_cache_pdf(i) for i in payload.get("catalogo", []) if isinstance(i, dict)]
        payload["catalogo"] = catalogo
        _CACHE_PDFS_MEM.update({"mtime": mtime, "payload": payload})
        return payload
    except Exception:
        return None


def atualizar_cache_pdfs_banco_horas():
    resultado = {
        "ok": True,
        "pdfplumber_instalado": pdfplumber is not None,
        "cache": str(ARQUIVO_CACHE_PDFS_BANCO_HORAS),
        "pastas_verificadas": [str(p) for p in PASTAS_PDFS_ROBO],
    }
    if pdfplumber is None:
        resultado.update({"ok": False, "mensagem": "pdfplumber não está instalado; não foi possível atualizar o histórico de PDFs."})
        return resultado

    t0 = time.monotonic()
    pdfs = _listar_pdfs_robo_limitado(max_pdfs=MAX_PDFS_CORRECAO_SALDO)
    catalogo = []
    for pdf in pdfs:
        if time.monotonic() - t0 > MAX_SEGUNDOS_CORRECAO_SALDO:
            break
        try:
            catalogo.extend(_extrair_espelhos_banco_horas_pdf(pdf))
        except Exception:
            continue

    segundos = round(time.monotonic() - t0, 1)
    cache = salvar_cache_pdfs_banco_horas(catalogo, pdfs_encontrados=len(pdfs), segundos=segundos)
    resultado.update({
        "ok": True,
        "mensagem": f"Histórico de PDFs atualizado. PDFs encontrados: {len(pdfs)}; quadros BH lidos: {len(catalogo)}; tempo: {segundos}s.",
        "pdfs_encontrados": len(pdfs),
        "quadros_lidos": len(catalogo),
        "segundos": segundos,
        "gerado_em": cache.get("gerado_em"),
    })
    return resultado



def _data_versao_pdf_bh(item):
    """Data que representa a versão do espelho de ponto/BH.

    v8.10.21: o mesmo PDF pode conter o dia 15/05, mas ter sido emitido
    somente em 05/06; nesse caso, o Saldo Atual do quadro Banco de Horas
    representa a fotografia de 05/06, não a abertura do ciclo. Por isso,
    a escolha do saldo informado passa a tratar cada PDF como uma versão
    histórica e usa, preferencialmente, a última data real lançada no corpo
    do espelho.
    """
    for chave in ("data_max_lancada", "emissao", "fim"):
        try:
            dt = item.get(chave)
            if dt is not None and not pd.isna(dt):
                return pd.to_datetime(dt).normalize()
        except Exception:
            continue
    return pd.NaT


def _chave_historica_versao_pdf(item, data_dt):
    """Ordena PDFs como histórico de versões.

    Para uma linha de 15/05, um PDF de 05/06 não deve vencer um PDF de
    15/05 só porque também contém a data 15/05 no corpo. A versão ideal é
    o PDF cuja data de versão seja igual à data da linha; se não existir,
    usa a primeira versão posterior. Apenas se não houver versão posterior
    usa a última versão anterior como fallback, com penalização.
    """
    versao = _data_versao_pdf_bh(item)
    if pd.isna(data_dt) or pd.isna(versao):
        return (2, 999999)
    try:
        delta = int((versao - data_dt).days)
    except Exception:
        return (2, 999999)
    if delta >= 0:
        return (0, delta)
    # Fallback: a versão é anterior à data da linha. Mantém possível, mas
    # sempre perde para uma versão da mesma data ou posterior.
    return (1, abs(delta))


def _escolher_saldo_pdf_para_linha(nome, data_linha, catalogo, matricula=""):
    """Escolhe o PDF de Banco de Horas para uma linha do Excel.

    v8.10.21 - histórico de versões:
    - o Saldo Atual do PDF é uma fotografia da versão/emissão do espelho;
    - um PDF de 05/06 pode conter 15/05 no corpo, mas não representa o
      saldo de abertura de 15/05;
    - para cada data do extrato, escolhe a versão mais próxima daquela data,
      preferindo a própria data ou a primeira versão posterior;
    - só usa versão anterior como fallback quando não existir versão posterior.
    """
    nome_norm = _canon_nome_banco(nome)
    mat_norm = _normalizar_matricula_bh(matricula)
    data_dt = _parse_data_pdf(data_linha)
    if not nome_norm and not mat_norm:
        return None

    candidatos = []
    for item in catalogo:
        item_mat = item.get("matricula_norm", "")
        item_nome = item.get("nome_norm", "")
        match_matricula = bool(mat_norm and item_mat and mat_norm == item_mat)
        match_nome = _nomes_compativeis_bh(nome_norm, item_nome)
        if not (match_matricula or match_nome):
            continue

        ini, fim = item.get("inicio"), item.get("fim")
        datas_lancadas = item.get("datas_lancadas") or []
        dentro_periodo = False
        contem_data_real = False
        data_entre_datas_reais = False

        if not pd.isna(data_dt):
            if not pd.isna(ini) and not pd.isna(fim):
                dentro_periodo = bool(ini <= data_dt <= fim)
            if datas_lancadas:
                contem_data_real = bool(any(_parse_data_pdf(d) == data_dt for d in datas_lancadas))
                dmin = item.get("data_min_lancada")
                dmax = item.get("data_max_lancada")
                if not pd.isna(dmin) and not pd.isna(dmax):
                    data_entre_datas_reais = bool(dmin <= data_dt <= dmax)

            # Mantém somente PDFs do ciclo/espelho compatível com a data da linha.
            if not (dentro_periodo or contem_data_real or data_entre_datas_reais):
                continue

        score_identidade = 0
        if match_matricula:
            score_identidade += 120
        if match_nome:
            score_identidade += 80
        if item.get("saldo_atual_pdf"):
            score_identidade += 10
        if item.get("datas_reais_pdf"):
            score_identidade += 5

        item2 = dict(item)
        item2["_score_match_v21"] = score_identidade
        item2["_match_v20"] = "matricula" if match_matricula else "nome"
        item2["_match_v21"] = "matricula" if match_matricula else "nome"
        item2["_dentro_periodo_v20"] = dentro_periodo
        item2["_contem_data_real_v20"] = contem_data_real
        item2["_data_versao_pdf_bh"] = _data_versao_pdf_bh(item)
        janela_tipo, janela_delta = _chave_historica_versao_pdf(item2, data_dt)
        item2["_historico_tipo_v21"] = janela_tipo
        item2["_historico_delta_v21"] = janela_delta
        candidatos.append(item2)

    if not candidatos:
        return None

    # Primeiro separa identidade forte; depois escolhe a versão histórica
    # correta. Isso impede que o PDF mais recente do ciclo contamine o saldo
    # inicial quando existir PDF mais antigo para a mesma pessoa.
    candidatos.sort(key=lambda x: (
        x.get("_historico_tipo_v21", 2),
        x.get("_historico_delta_v21", 999999),
        -int(x.get("_score_match_v21", 0)),
        x.get("data_max_lancada") if not pd.isna(x.get("data_max_lancada")) else pd.Timestamp.max,
        x.get("emissao") if not pd.isna(x.get("emissao")) else pd.Timestamp.max,
        x.get("arquivo_pdf", ""),
    ))
    return candidatos[0]



# ============================================================
# v8.10.27 - Banco de Horas 100% PDF no Extrato
# Base do ciclo = Saldo Atual do PDF de fechamento do ciclo anterior
# O Excel não é mais fallback silencioso para saldo de Banco de Horas
# ============================================================
def _dt_norm(valor):
    dt = _parse_data_pdf(valor)
    if pd.isna(dt):
        return pd.NaT
    return pd.to_datetime(dt).normalize()


def _mesmo_ciclo_pdf(item, inicio, fim):
    try:
        ini = _dt_norm(item.get("inicio"))
        f = _dt_norm(item.get("fim"))
        return (not pd.isna(ini)) and (not pd.isna(f)) and ini == pd.to_datetime(inicio).normalize() and f == pd.to_datetime(fim).normalize()
    except Exception:
        return False


def _identidade_pdf_compativel(nome, matricula, item):
    nome_norm = _canon_nome_banco(nome)
    mat_norm = _normalizar_matricula_bh(matricula)
    item_nome = item.get("nome_norm") or _canon_nome_banco(item.get("nome", ""))
    item_mat = item.get("matricula_norm") or _normalizar_matricula_bh(item.get("matricula", ""))
    match_matricula = bool(mat_norm and item_mat and mat_norm == item_mat)
    match_nome = _nomes_compativeis_bh(nome_norm, item_nome)
    return match_matricula or match_nome, ("matricula" if match_matricula else ("nome" if match_nome else ""))


def _arquivo_pdf_parece_ciclo(arquivo_pdf, inicio, fim):
    """Heurística leve para priorizar PDFs cujo nome sugere o ciclo desejado.

    Aceita padrões reais encontrados no projeto, como:
    - 008005000 15-04 a 14-05.pdf
    - 008005000 15 a 05-06.pdf
    - arquivos com barras já extraídas do texto do PDF.
    """
    try:
        nome = _canon_nome_banco(Path(str(arquivo_pdf)).name)
        ini = pd.to_datetime(inicio).normalize()
        f = pd.to_datetime(fim).normalize()
        padroes = [
            f"{ini.day:02d}-{ini.month:02d}",
            f"{f.day:02d}-{f.month:02d}",
            f"{ini.day:02d}/{ini.month:02d}",
            f"{f.day:02d}/{f.month:02d}",
            f"{ini.day:02d} A {f.day:02d}-{f.month:02d}",
            f"{ini.day:02d}-{ini.month:02d} A {f.day:02d}-{f.month:02d}",
        ]
        return any(_canon_nome_banco(x) in nome for x in padroes[:2]) or _canon_nome_banco(padroes[-1]) in nome
    except Exception:
        return False


def _buscar_pdf_fechamento_ciclo_anterior_direto(nome, matricula, inicio_anterior, fim_anterior, max_segundos=45):
    """Busca emergencial fora do cache para o PDF de fechamento do ciclo anterior.

    v8.10.28:
    O diagnóstico mostrou que o PDF correto pode existir na pasta, mas não entrar
    no cache por limite, cache antigo ou por não estar entre os PDFs mais recentes.
    Esta busca varre a pasta configurada dando prioridade a nomes de arquivo que
    contenham o período anterior (ex.: 15-04 a 14-05), e só então tenta os demais.
    """
    inicio_t = time.monotonic()
    try:
        todos = []
        vistos = set()
        for pasta in PASTAS_PDFS_ROBO:
            try:
                if not pasta.exists():
                    continue
                for pdf in pasta.rglob('*.pdf'):
                    try:
                        chave = str(pdf.resolve()).lower()
                        if chave in vistos:
                            continue
                        vistos.add(chave)
                        todos.append(pdf)
                    except Exception:
                        continue
            except Exception:
                continue

        # Prioriza nomes com o período exato, mas mantém todos como fallback.
        todos.sort(key=lambda x: (
            0 if _arquivo_pdf_parece_ciclo(x, inicio_anterior, fim_anterior) else 1,
            x.stat().st_mtime if x.exists() else 0,
        ), reverse=False)

        melhor = None
        melhor_chave = None
        for pdf in todos:
            if time.monotonic() - inicio_t > max_segundos:
                break
            try:
                for item in _extrair_espelhos_banco_horas_pdf(pdf):
                    ok_ident, tipo_match = _identidade_pdf_compativel(nome, matricula, item)
                    if not ok_ident:
                        continue
                    if not _mesmo_ciclo_pdf(item, inicio_anterior, fim_anterior):
                        continue
                    saldo = str(item.get('saldo_atual_pdf', '') or '').strip()
                    if not saldo:
                        continue
                    emissao = _dt_norm(item.get('emissao'))
                    # O critério principal é o PERÍODO do espelho. Emissão é só desempate.
                    delta_emissao = 999999
                    if not pd.isna(emissao):
                        delta_emissao = abs(int((emissao - pd.to_datetime(fim_anterior).normalize()).days))
                    contem_fim = False
                    try:
                        datas = item.get('datas_lancadas') or []
                        contem_fim = any(_dt_norm(d) == pd.to_datetime(fim_anterior).normalize() for d in datas)
                    except Exception:
                        contem_fim = False
                    chave = (
                        0 if tipo_match == 'matricula' else 1,
                        0 if contem_fim else 1,
                        delta_emissao,
                        str(pdf.name),
                    )
                    if melhor is None or chave < melhor_chave:
                        melhor_chave = chave
                        melhor = {**item, '_match_base_pdf_v25': tipo_match, '_match_base_pdf_v28': tipo_match, '_inicio_ciclo_anterior_v25': inicio_anterior, '_fim_ciclo_anterior_v25': fim_anterior, '_busca_direta_v28': True}
            except Exception:
                continue
        return melhor
    except Exception:
        return None


def _buscar_pdf_fechamento_ciclo_anterior(nome, matricula, inicio_ciclo, catalogo=None, permitir_busca_direta=False):
    """
    Retorna o espelho que fecha o ciclo anterior ao ciclo em análise.

    v8.10.28 - Correção conceitual definitiva:
    Para abrir o ciclo atual, a âncora do Banco de Horas é o Saldo Atual
    do PDF cujo PERÍODO seja exatamente o ciclo anterior, independentemente
    da data de emissão. Ex.: para 15/05/2026 a 14/06/2026, buscar o espelho
    15/04/2026 a 14/05/2026 e usar o Saldo Atual dele.
    """
    try:
        inicio_ciclo = pd.to_datetime(inicio_ciclo).normalize()
    except Exception:
        return None

    fim_anterior = inicio_ciclo - pd.Timedelta(days=1)
    inicio_anterior = _data_inicio_medicao(fim_anterior)
    fim_anterior = _data_fim_medicao(inicio_anterior) or fim_anterior
    inicio_anterior = pd.to_datetime(inicio_anterior).normalize()
    fim_anterior = pd.to_datetime(fim_anterior).normalize()

    if catalogo is None:
        try:
            cache = carregar_cache_pdfs_banco_horas()
            catalogo = cache.get('catalogo', []) if cache else []
        except Exception:
            catalogo = []

    candidatos = []
    for item in catalogo or []:
        ok_ident, tipo_match = _identidade_pdf_compativel(nome, matricula, item)
        if not ok_ident:
            continue
        # Regra dura: período do espelho precisa ser o ciclo anterior.
        if not _mesmo_ciclo_pdf(item, inicio_anterior, fim_anterior):
            continue
        saldo = str(item.get('saldo_atual_pdf', '') or '').strip()
        if not saldo:
            continue
        emissao = _dt_norm(item.get('emissao'))
        contem_fim = False
        try:
            datas = item.get('datas_lancadas') or []
            contem_fim = any(_dt_norm(d) == fim_anterior for d in datas)
        except Exception:
            contem_fim = False
        delta_emissao = 999999
        if not pd.isna(emissao):
            delta_emissao = abs(int((emissao - fim_anterior).days))

        candidatos.append((
            0 if tipo_match == 'matricula' else 1,
            0 if contem_fim else 1,
            delta_emissao,
            emissao if not pd.isna(emissao) else pd.Timestamp.max,
            item.get('arquivo_pdf', ''),
            {**item, '_match_base_pdf_v25': tipo_match, '_match_base_pdf_v28': tipo_match, '_inicio_ciclo_anterior_v25': inicio_anterior, '_fim_ciclo_anterior_v25': fim_anterior, '_busca_direta_v28': False}
        ))

    if candidatos:
        candidatos.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4]))
        return candidatos[0][-1]

    # v8.10.31: permitimos uma busca direta CIRÚRGICA apenas quando o cache não
    # contém o PDF do fechamento do ciclo anterior. Essa busca prioriza arquivos
    # cujo nome indica o período anterior, ex.: 15-04 a 14-05, evitando varrer
    # pesado em uso normal e corrigindo cache incompleto/limitado.
    if permitir_busca_direta:
        return _buscar_pdf_fechamento_ciclo_anterior_direto(nome, matricula, inicio_anterior, fim_anterior)
    return None


def _saldo_pdf_para_data_exata(nome, matricula, data_dt, catalogo=None):
    """Busca Saldo Atual informado por fotografia de PDF que contenha a data.

    v8.10.34:
    O Saldo Atual de um PDF não vale para o ciclo inteiro, mas vale para a
    fotografia/versão daquele espelho. Na prática operacional, um PDF emitido
    em 20/05 pode conter as linhas de 15/05 a 20/05; nesse caso, o Saldo Atual
    daquele PDF deve alimentar o Saldo Informado das datas que aparecem nesse
    PDF, até que exista outra fotografia mais adequada.

    Regra por data do extrato:
    - procurar PDFs do mesmo colaborador/matrícula cujo período/datas reais
      contenham a data da linha;
    - escolher a fotografia mais próxima da data, preferindo a menor
      data_max_lancada/emissão que ainda contenha aquela data;
    - isso evita repetir automaticamente o último PDF do ciclo inteiro em todas
      as linhas, mas também evita marcar como "PDF diário não encontrado" quando
      existe uma fotografia parcial contendo aquela data.
    """
    try:
        data_dt = pd.to_datetime(data_dt).normalize()
    except Exception:
        return None
    if catalogo is None:
        try:
            cache = carregar_cache_pdfs_banco_horas()
            catalogo = cache.get("catalogo", []) if cache else []
        except Exception:
            catalogo = []

    candidatos = []
    for item in catalogo or []:
        ok_ident, tipo_match = _identidade_pdf_compativel(nome, matricula, item)
        if not ok_ident:
            continue
        saldo = str(item.get("saldo_atual_pdf", "") or "").strip()
        if not saldo:
            continue

        ini = _dt_norm(item.get("inicio"))
        fim = _dt_norm(item.get("fim"))
        if not pd.isna(ini) and not pd.isna(fim) and not (ini <= data_dt <= fim):
            continue

        datas_lancadas = item.get("datas_lancadas") or []
        contem_data_real = False
        if datas_lancadas:
            for d in datas_lancadas:
                dd = _dt_norm(d)
                if not pd.isna(dd) and dd == data_dt:
                    contem_data_real = True
                    break
        else:
            # Sem datas reais, aceita apenas se o período do espelho contiver a data.
            contem_data_real = bool(not pd.isna(ini) and not pd.isna(fim) and ini <= data_dt <= fim)

        if not contem_data_real:
            continue

        data_max = _dt_norm(item.get("data_max_lancada"))
        emissao = _dt_norm(item.get("emissao"))
        versao = _data_versao_pdf_bh(item)
        if pd.isna(versao):
            versao = data_max if not pd.isna(data_max) else emissao
        if pd.isna(versao):
            versao = data_dt

        try:
            delta = int((versao - data_dt).days)
        except Exception:
            delta = 999999

        # Preferir a primeira fotografia posterior/igual que contém a data.
        # Fotografia anterior só entra como fallback e perde prioridade.
        tipo_janela = 0 if delta >= 0 else 1
        prioridade_ident = 0 if tipo_match == "matricula" else 1
        item2 = {**item, "_match_diario_pdf_v25": tipo_match, "_saldo_info_v34": "data_contida_na_fotografia_pdf"}
        candidatos.append((
            tipo_janela,
            abs(delta),
            prioridade_ident,
            versao if not pd.isna(versao) else pd.Timestamp.max,
            item.get("arquivo_pdf", ""),
            item2,
        ))

    if not candidatos:
        return None
    candidatos.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[4]))
    return candidatos[0][-1]



# v8.10.38 - formatação segura do período do PDF base
# Evita quebrar /api/extrato-banco-horas quando algum candidato de PDF vem sem
# inicio/fim ou quando o caminho de fallback monta o item com chaves diferentes.
def _periodo_pdf_base_seguro(pdf_base):
    if not pdf_base:
        return ""

    def _fmt_data_segura(valor):
        if valor is None:
            return ""
        try:
            if pd.isna(valor):
                return ""
        except Exception:
            pass
        try:
            return pd.to_datetime(valor).strftime("%d/%m/%Y")
        except Exception:
            return str(valor or "")

    ini = (
        pdf_base.get("_inicio_ciclo_anterior_v25")
        or pdf_base.get("inicio")
        or pdf_base.get("periodo_inicio")
        or pdf_base.get("inicio_ciclo")
    )
    fim = (
        pdf_base.get("_fim_ciclo_anterior_v25")
        or pdf_base.get("fim")
        or pdf_base.get("periodo_fim")
        or pdf_base.get("fim_ciclo")
    )

    ini_s = _fmt_data_segura(ini)
    fim_s = _fmt_data_segura(fim)
    if ini_s and fim_s:
        return f"{ini_s} a {fim_s}"

    # Alguns itens antigos do cache podem ter o período já salvo como string.
    periodo = str(pdf_base.get("periodo", "") or pdf_base.get("periodo_pdf", "") or "").strip()
    return periodo


def _catalogo_tem_identidade_bh(catalogo, nome, matricula=""):
    """Verifica se o cache possui ao menos um PDF compatível com colaborador/matrícula."""
    try:
        for item in catalogo or []:
            ok, _tipo = _identidade_pdf_compativel(nome, matricula, item)
            if ok:
                return True
    except Exception:
        pass
    return False


def _catalogo_emergencial_colaborador_bh(nome, matricula, inicio_ciclo, fim_ciclo=None, max_segundos=50):
    """Monta um mini-catálogo em disco apenas para o colaborador selecionado.

    v8.11.12:
    A lógica do Banco/Extrato depende do catálogo JSON dos PDFs. Quando esse
    catálogo fica vazio após troca de release, o extrato inicia em zero. Esta
    busca é acionada só para o colaborador selecionado e evita varredura pesada
    em lote.
    """
    if pdfplumber is None:
        return []
    try:
        inicio_ciclo = pd.to_datetime(inicio_ciclo).normalize()
    except Exception:
        return []
    try:
        fim_ciclo = pd.to_datetime(fim_ciclo).normalize() if fim_ciclo is not None else (_data_fim_medicao(inicio_ciclo) or inicio_ciclo + pd.Timedelta(days=30))
    except Exception:
        fim_ciclo = inicio_ciclo + pd.Timedelta(days=30)

    fim_anterior = inicio_ciclo - pd.Timedelta(days=1)
    inicio_anterior = _data_inicio_medicao(fim_anterior)
    fim_anterior = _data_fim_medicao(inicio_anterior) or fim_anterior

    inicio_t = time.monotonic()
    pdfs = []
    vistos = set()
    try:
        for pasta in PASTAS_PDFS_ROBO:
            try:
                if not pasta.exists():
                    continue
                for pdf in pasta.rglob("*.pdf"):
                    try:
                        chave = str(pdf.resolve()).lower()
                        if chave in vistos:
                            continue
                        vistos.add(chave)
                        pdfs.append(pdf)
                    except Exception:
                        continue
            except Exception:
                continue
    except Exception:
        return []

    def prioridade(pdf):
        try:
            p = 0
            if _arquivo_pdf_parece_ciclo(pdf, inicio_anterior, fim_anterior):
                p -= 100
            if _arquivo_pdf_parece_ciclo(pdf, inicio_ciclo, fim_ciclo):
                p -= 80
            return (p, -(pdf.stat().st_mtime if pdf.exists() else 0))
        except Exception:
            return (999, 0)

    pdfs.sort(key=prioridade)

    saida = []
    for pdf in pdfs:
        if time.monotonic() - inicio_t > max_segundos:
            break
        try:
            itens = _extrair_espelhos_banco_horas_pdf(pdf)
        except Exception:
            continue
        for item in itens or []:
            try:
                ok, _tipo = _identidade_pdf_compativel(nome, matricula, item)
                if not ok:
                    continue
                ini = _dt_norm(item.get("inicio"))
                fim = _dt_norm(item.get("fim"))
                mesmo_atual = (not pd.isna(ini)) and (not pd.isna(fim)) and ini == pd.to_datetime(inicio_ciclo).normalize() and fim == pd.to_datetime(fim_ciclo).normalize()
                mesmo_anterior = (not pd.isna(ini)) and (not pd.isna(fim)) and ini == pd.to_datetime(inicio_anterior).normalize() and fim == pd.to_datetime(fim_anterior).normalize()
                contem_ciclo = False
                for d in item.get("datas_lancadas") or []:
                    dd = _dt_norm(d)
                    if not pd.isna(dd) and inicio_ciclo <= dd <= fim_ciclo:
                        contem_ciclo = True
                        break
                if mesmo_atual or mesmo_anterior or contem_ciclo:
                    saida.append(item)
            except Exception:
                continue
    return saida

def _corrigir_saldo_atual_excel_com_catalogo(caminho_excel, catalogo, origem_catalogo="cache"):
    resultado = {
        "executado": True,
        "ok": True,
        "mensagem": "Correção de Banco de Horas não executada.",
        "pdfs_lidos": len(catalogo or []),
        "linhas_corrigidas": 0,
        "origem_catalogo": origem_catalogo,
    }
    caminho_excel = Path(caminho_excel)
    if not caminho_excel.exists():
        resultado.update({"ok": False, "mensagem": f"Excel não encontrado: {caminho_excel}"})
        return resultado
    if not catalogo:
        resultado.update({"ok": False, "mensagem": "Catálogo de PDFs vazio. Atualize o Histórico de PDFs antes de corrigir o Banco de Horas."})
        return resultado

    try:
        abas = pd.read_excel(caminho_excel, sheet_name=None)
        if "Consolidado" not in abas:
            resultado.update({"ok": False, "mensagem": "Aba Consolidado não encontrada no Excel."})
            return resultado

        df = abas["Consolidado"].copy()
        if "Nome" not in df.columns or "Data" not in df.columns:
            resultado.update({"ok": False, "mensagem": "Colunas Nome/Data não encontradas no Consolidado."})
            return resultado

        for col in ["Saldo Anterior PDF", "Débito BH PDF", "Crédito BH PDF", "Saldo Atual PDF", "Origem Saldo PDF", "Saldo Informado"]:
            if col not in df.columns:
                df[col] = ""

        corrigidas = 0
        sem_match = 0
        for idx, row in df.iterrows():
            item = _escolher_saldo_pdf_para_linha(row.get("Nome", ""), row.get("Data", ""), catalogo, _matricula_da_linha_bh(row))
            if not item:
                sem_match += 1
                continue
            saldo_atual_pdf = item.get("saldo_atual_pdf", "")
            df.at[idx, "Saldo Anterior PDF"] = item.get("saldo_anterior_pdf", "")
            df.at[idx, "Débito BH PDF"] = item.get("debito_bh_pdf", "")
            df.at[idx, "Crédito BH PDF"] = item.get("credito_bh_pdf", "")
            df.at[idx, "Saldo Atual PDF"] = saldo_atual_pdf
            origem_match = item.get("_match_v20", "pdf")
            df.at[idx, "Origem Saldo PDF"] = f"saldo_atual_pdf_cache_{origem_match}: {item.get('arquivo_pdf', '')}"
            if "Saldo Atual" in df.columns and saldo_atual_pdf not in (None, ""):
                df.at[idx, "Saldo Atual"] = saldo_atual_pdf
            if saldo_atual_pdf not in (None, ""):
                df.at[idx, "Saldo Informado"] = saldo_atual_pdf
            corrigidas += 1

        abas["Consolidado"] = df
        with pd.ExcelWriter(caminho_excel, engine="openpyxl") as writer:
            for nome_aba, dados in abas.items():
                dados.to_excel(writer, sheet_name=str(nome_aba)[:31], index=False)

        resultado.update({
            "ok": True,
            "mensagem": f"Banco de Horas corrigido pelo histórico/cache dos PDFs em {corrigidas} linha(s). Linhas sem match: {sem_match}.",
            "linhas_corrigidas": int(corrigidas),
            "linhas_sem_match": int(sem_match),
        })
        return resultado
    except Exception as e:
        resultado.update({"ok": False, "mensagem": f"Falha ao corrigir Banco de Horas pelo catálogo de PDFs: {e}"})
        return resultado


def corrigir_saldo_atual_excel_pelo_pdf(caminho_excel, usar_cache=True, atualizar_cache_se_necessario=False, forcar_execucao=False):
    """
    Corrige a aba Consolidado usando o catálogo de PDFs.

    v8.10.23: esta função não precisa mais rodar durante a importação rápida.
    Ela pode usar o cache em data/catalogo_pdfs_banco_horas.json ou, se pedido,
    atualizar o histórico de PDFs antes da correção.
    """
    if not forcar_execucao and not CORRIGIR_SALDO_ATUAL_POR_PDF_AO_IMPORTAR and not atualizar_cache_se_necessario:
        return {"executado": False, "ok": True, "mensagem": "Correção automática por PDF desativada por configuração; Excel importado normalmente. Para executar manualmente, use o botão Corrigir Banco de Horas pelos PDFs."}
    if pdfplumber is None:
        return {"executado": False, "ok": True, "mensagem": "pdfplumber não instalado; correção por PDF ignorada."}

    cache = carregar_cache_pdfs_banco_horas() if usar_cache else None
    if (not cache or not cache.get("catalogo")) and atualizar_cache_se_necessario:
        atualizacao = atualizar_cache_pdfs_banco_horas()
        if not atualizacao.get("ok"):
            return atualizacao
        cache = carregar_cache_pdfs_banco_horas()

    if cache and cache.get("catalogo"):
        resultado = _corrigir_saldo_atual_excel_com_catalogo(caminho_excel, cache.get("catalogo"), origem_catalogo="cache")
        resultado["cache_gerado_em"] = cache.get("gerado_em")
        resultado["pdfs_encontrados"] = cache.get("pdfs_encontrados")
        return resultado

    # Fallback compatível: se não houver cache e a correção automática estiver ligada,
    # monta catálogo em memória. O uso diário deve preferir o botão separado.
    t0 = time.monotonic()
    pdfs_candidatos = _listar_pdfs_robo_limitado()
    catalogo = _catalogar_saldos_pdf_robo()
    resultado = _corrigir_saldo_atual_excel_com_catalogo(caminho_excel, catalogo, origem_catalogo="memoria")
    resultado["pdfs_encontrados"] = len(pdfs_candidatos)
    resultado["segundos"] = round(time.monotonic() - t0, 1)
    return resultado


def sincronizar_excel_robo(forcar=False, corrigir_saldo_pdf=False):
    """
    Copia automaticamente o Excel gerado pelo robô para a pasta data do sistema web.

    Origem padrão:
    %USERPROFILE%\\ponto_pdfs\\saida\\resultado_lote_consultas_validado.xlsx

    Destino:
    data\\resultado_lote_consultas_validado.xlsx
    """
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    if not ARQUIVO_EXCEL_ROBO.exists():
        return {
            "ok": False,
            "mensagem": f"Arquivo de origem do robô não encontrado: {ARQUIVO_EXCEL_ROBO}",
            "origem": str(ARQUIVO_EXCEL_ROBO),
            "destino": str(ARQUIVO_PADRAO),
            "copiado": False,
        }

    precisa_copiar = forcar or not ARQUIVO_PADRAO.exists()

    if ARQUIVO_PADRAO.exists() and not forcar:
        origem_mtime = ARQUIVO_EXCEL_ROBO.stat().st_mtime
        destino_mtime = ARQUIVO_PADRAO.stat().st_mtime
        precisa_copiar = origem_mtime > destino_mtime

    if precisa_copiar:
        shutil.copy2(ARQUIVO_EXCEL_ROBO, ARQUIVO_PADRAO)
        invalidar_cache_consolidado()
        # v8.11.17: ao importar novo Excel, descarta resumo antigo do Banco de Horas.
        # O Excel passou a trazer Saldo Calculado oficial; manter JSON antigo podia
        # fazer a tela exibir resultado de versões anteriores/cache PDF.
        try:
            for _arquivo_cache_bh in [ARQUIVO_RESUMO_BANCO_HORAS_PDF]:
                if _arquivo_cache_bh.exists():
                    _arquivo_cache_bh.unlink()
        except Exception:
            pass

        # v8.11.26: gera imediatamente o resumo leve do Banco de Horas a partir
        # do Excel oficial do Script_Robo_Ponto. Antes o cache era apagado na
        # importação rápida e a guia Banco de Horas ficava sem dados até uma
        # correção manual por PDFs, exibindo excel_oficial_indisponivel.
        resumo_bh_importacao = None
        try:
            resumo_bh_importacao = gerar_resumo_banco_horas_pdf()
        except Exception as e:
            resumo_bh_importacao = {"ok": False, "mensagem": f"Falha ao gerar resumo do Banco de Horas na importação: {e}"}

        retorno = {
            "ok": True,
            "mensagem": "Excel do robô importado com sucesso em modo rápido.",
            "origem": str(ARQUIVO_EXCEL_ROBO),
            "destino": str(ARQUIVO_PADRAO),
            "copiado": True,
            "importacao_rapida": True,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "resumo_banco_horas": resumo_bh_importacao,
        }
        if corrigir_saldo_pdf:
            retorno["correcao_saldo_pdf"] = corrigir_saldo_atual_excel_pelo_pdf(ARQUIVO_PADRAO, usar_cache=True, atualizar_cache_se_necessario=True, forcar_execucao=False)
            retorno["mensagem"] = "Excel do robô importado com sucesso, com correção de Banco de Horas solicitada."
            retorno["importacao_rapida"] = False
        return retorno

    return {
        "ok": True,
        "mensagem": "Excel do sistema web já está atualizado.",
        "origem": str(ARQUIVO_EXCEL_ROBO),
        "destino": str(ARQUIVO_PADRAO),
        "copiado": False,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def normalizar_cc_web(valor):
    """
    Normaliza Centro de Custo para comparação web.
    Retorna somente os dígitos principais do CC.
    """
    texto = str(valor or "").strip()
    if not texto or texto.lower() in {"nan", "none"}:
        return ""

    texto = texto.split("-", 1)[0].strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    digitos = re.sub(r"\D", "", texto)
    if not digitos:
        return ""

    return digitos


def rotulo_cc_web(valor):
    cc = normalizar_cc_web(valor)
    return cc or str(valor or "").strip()


def normalizar_texto_filtro(valor):
    texto = str(valor or "").strip()
    if texto.lower() in {"nan", "none"}:
        return ""
    return re.sub(r"\s+", " ", texto).strip().lower()


def valor_cc_display(valor):
    texto = str(valor or "").strip()
    if texto.lower() in {"nan", "none"}:
        return ""
    return re.sub(r"\s+", " ", texto).strip()



def normalizar_cc_filtro_backend(valor):
    """
    Normalização simples e rastreável do CC:
    - converte para texto;
    - remove .0 quando Excel trata como número;
    - pega apenas a parte antes de hífen;
    - remove caracteres não numéricos.
    """
    texto = str(valor or "").strip()
    if not texto or texto.lower() in {"nan", "none"}:
        return ""

    texto = texto.split("-", 1)[0].strip()

    if texto.endswith(".0"):
        texto = texto[:-2]

    return re.sub(r"\D", "", texto)


def aplicar_filtro_cc_backend(df, cc_filtro):
    cc_norm = normalizar_cc_filtro_backend(cc_filtro)
    if not cc_norm or "cc" not in df.columns:
        return df

    df = df.copy()
    df["_cc_norm_web"] = df["cc"].apply(normalizar_cc_filtro_backend)
    df = df[df["_cc_norm_web"] == cc_norm].copy()
    df.drop(columns=["_cc_norm_web"], inplace=True, errors="ignore")
    return df

def hhmm_para_minutos(valor):
    if pd.isna(valor):
        return 0
    valor = str(valor).strip()
    if not valor or valor.lower() == "nan":
        return 0
    sinal = -1 if valor.startswith("-") else 1
    valor = valor.replace("-", "").strip()
    if ":" in valor:
        try:
            h, m = valor.split(":")[:2]
            return sinal * (int(h) * 60 + int(m))
        except Exception:
            return 0
    if "," in valor:
        try:
            h, m = valor.split(",", 1)
            return sinal * (int(h) * 60 + int(m[:2]))
        except Exception:
            return 0
    return 0


def minutos_para_hhmm(total):
    try:
        total = int(total)
    except Exception:
        return "00:00"
    sinal = "-" if total < 0 else ""
    total = abs(total)
    return f"{sinal}{total // 60:02d}:{total % 60:02d}"




def _base_operacional_ativa():
    """
    A base operacional só deve ser considerada ativa após importação manual.
    Isso impede que o sistema carregue automaticamente um Excel antigo existente na pasta data/.
    """
    try:
        return ARQUIVO_BASE_ATIVA_FLAG.exists()
    except Exception:
        return False


def _ativar_base_operacional(nome_arquivo=""):
    try:
        ARQUIVO_BASE_ATIVA_FLAG.write_text(str(nome_arquivo or "base_importada"), encoding="utf-8")
    except Exception:
        pass


def _desativar_base_operacional():
    try:
        if ARQUIVO_BASE_ATIVA_FLAG.exists():
            ARQUIVO_BASE_ATIVA_FLAG.unlink()
    except Exception:
        pass



# ============================================================
# v8.11.22 - Leitura otimizada do Excel oficial
# ============================================================
_COLUNAS_CONSOLIDADO_WEB_CANON = {
    "arquivo origem", "matricula", "matrícula", "chapa", "cc", "nome", "funcao", "função", "turno",
    "data", "dia", "1a e.", "1a s.", "2a e.", "2a s.", "3a e.", "3a s.", "4a e.", "4a s.",
    "abono", "h.e.", "absent.", "jornada", "ad. not.", "observacao", "observação",
    "saldo atual", "saldo atual pdf", "saldo informado pdf", "saldo informado",
    "saldo anterior pdf", "saldo anterior", "debito pdf", "débito pdf", "debito bh pdf", "débito bh pdf",
    "credito pdf", "crédito pdf", "credito bh pdf", "crédito bh pdf", "saldo calculado",
    "data emissao pdf", "data emissão pdf", "data de emissao pdf", "data de emissão pdf", "emissao", "emissão",
    "status validação", "status validacao", "inconsistências", "inconsistencias", "jornada calculada", "qtd batidas",
}

def _canon_header_excel_web(valor):
    texto = str(valor or "").strip().lower()
    texto = texto.replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
    texto = texto.replace("é", "e").replace("ê", "e")
    texto = texto.replace("í", "i")
    texto = texto.replace("ó", "o").replace("ô", "o").replace("õ", "o")
    texto = texto.replace("ú", "u").replace("ç", "c")
    texto = " ".join(texto.replace("_", " ").split())
    return texto


def _listar_abas_excel(caminho_excel):
    """Lista abas do Excel ativo sem carregar os dados."""
    caminho_excel = Path(caminho_excel)
    cache_id = _chave_cache_excel(caminho_excel)
    try:
        with _CACHE_EXCEL_LOCK:
            if (
                cache_id[1]
                and _CACHE_EXCEL_META_MEM.get("sig") == cache_id
                and _CACHE_EXCEL_META_MEM.get("abas")
            ):
                return list(_CACHE_EXCEL_META_MEM.get("abas") or [])
    except Exception:
        pass

    try:
        if load_workbook is not None:
            wb = load_workbook(filename=str(caminho_excel), read_only=True, data_only=True)
            try:
                abas = list(wb.sheetnames)
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
        else:
            abas = list(pd.ExcelFile(caminho_excel).sheet_names)
        try:
            with _CACHE_EXCEL_LOCK:
                if _CACHE_EXCEL_META_MEM.get("sig") != cache_id:
                    _CACHE_EXCEL_META_MEM.update({
                        "sig": cache_id,
                        "abas": abas,
                        "aba_historico": None,
                        "aba_saldos_bh": None,
                        "modelo": None,
                    })
                else:
                    _CACHE_EXCEL_META_MEM["abas"] = abas
        except Exception:
            pass
        return abas
    except Exception:
        return []


def _nome_aba_historico_v6(caminho_excel=None):
    """Retorna o nome da aba historica do modelo v6, se existir.

    O Script_Robo_Ponto_v6_PRO passou a entregar duas visoes:
    - Consolidado: base oficial operacional, uma linha por colaborador + data.
    - Consolidado Historico/Historico: fotografias completas para auditoria.
    """
    caminho_excel = Path(caminho_excel or ARQUIVO_PADRAO)
    cache_id = _chave_cache_excel(caminho_excel)
    try:
        with _CACHE_EXCEL_LOCK:
            if _CACHE_EXCEL_META_MEM.get("sig") == cache_id and _CACHE_EXCEL_META_MEM.get("aba_historico") is not None:
                return _CACHE_EXCEL_META_MEM.get("aba_historico") or ""
    except Exception:
        pass
    abas = _listar_abas_excel(caminho_excel)
    candidatos = {
        "consolidado historico",
        "consolidado_historico",
        "consolidado historico pdf",
        "consolidado historico v6",
        "consolidado histórico",
        "historico",
        "historico pdf",
        "histórico",
        "auditoria",
        "auditoria fechamentos",
        "auditoria premium",
        "fechamentos",
        "fotografias",
        "fotografias pdf",
    }
    candidatos_norm = {_canon_header_excel_web(c) for c in candidatos}
    resultado = ""
    for aba in abas:
        if _canon_header_excel_web(aba) in candidatos_norm:
            resultado = aba
            break
    if not resultado:
        for aba in abas:
            aba_norm = _canon_header_excel_web(aba)
            if (
                ("historico" in aba_norm or "auditoria" in aba_norm or "fotografia" in aba_norm)
                and ("consolidado" in aba_norm or "fechamento" in aba_norm or "pdf" in aba_norm)
            ):
                resultado = aba
                break
    try:
        with _CACHE_EXCEL_LOCK:
            if _CACHE_EXCEL_META_MEM.get("sig") == cache_id:
                _CACHE_EXCEL_META_MEM["aba_historico"] = resultado
    except Exception:
        pass
    return resultado


def _modelo_excel_ativo(caminho_excel=None):
    """Identifica o modelo de Excel ativo para mensagens e compatibilidade."""
    caminho_excel = Path(caminho_excel or ARQUIVO_PADRAO)
    cache_id = _chave_cache_excel(caminho_excel)
    try:
        with _CACHE_EXCEL_LOCK:
            modelo_cache = _CACHE_EXCEL_META_MEM.get("modelo") if _CACHE_EXCEL_META_MEM.get("sig") == cache_id else None
            if modelo_cache is not None:
                modelo = dict(modelo_cache)
                modelo["abas"] = list(modelo.get("abas") or [])
                return modelo
    except Exception:
        pass

    abas = _listar_abas_excel(caminho_excel)
    tem_consolidado = "Consolidado" in abas
    aba_historico = _nome_aba_historico_v6(caminho_excel)
    if tem_consolidado and aba_historico:
        modelo = {"modelo": "v6", "aba_operacional": "Consolidado", "aba_historico": aba_historico, "abas": abas}
    elif tem_consolidado:
        modelo = {"modelo": "v5_compativel", "aba_operacional": "Consolidado", "aba_historico": "", "abas": abas}
    else:
        modelo = {"modelo": "desconhecido", "aba_operacional": "", "aba_historico": "", "abas": abas}
    try:
        with _CACHE_EXCEL_LOCK:
            if _CACHE_EXCEL_META_MEM.get("sig") == cache_id:
                _CACHE_EXCEL_META_MEM["modelo"] = dict(modelo)
    except Exception:
        pass
    return modelo

def _ler_excel_consolidado_otimizado(caminho_excel):
    """Lê apenas a aba Consolidado e apenas as colunas usadas pelo Tempo Fechado.

    A leitura via pandas/openpyxl carregava a planilha inteira e deixava a
    navegação pesada após a importação. Aqui usamos openpyxl em read_only e
    montamos um DataFrame enxuto, preservando os nomes originais das colunas.
    """
    caminho_excel = Path(caminho_excel)
    if load_workbook is None:
        return pd.read_excel(caminho_excel, sheet_name="Consolidado", dtype=str).fillna("")

    wb = load_workbook(filename=str(caminho_excel), read_only=True, data_only=True)
    try:
        if "Consolidado" not in wb.sheetnames:
            raise ValueError("Aba Consolidado não encontrada no Excel importado.")
        ws = wb["Consolidado"]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return pd.DataFrame()
        header = ["" if h is None else str(h).strip() for h in header]
        indices = []
        nomes = []
        for idx, nome in enumerate(header):
            if not nome:
                continue
            if _canon_header_excel_web(nome) in _COLUNAS_CONSOLIDADO_WEB_CANON:
                indices.append(idx)
                nomes.append(nome)
        # Fallback seguro: se por alguma razão o cabeçalho vier diferente, lê tudo.
        if not indices:
            indices = list(range(len(header)))
            nomes = header
        dados = []
        for row in rows:
            if row is None:
                continue
            linha = []
            tem_valor = False
            for idx in indices:
                val = row[idx] if idx < len(row) else ""
                if val is None:
                    val = ""
                elif hasattr(val, "strftime"):
                    # Datas do Excel chegam como datetime/date: preserva dd/mm/aaaa para filtros e relatórios.
                    try:
                        val = val.strftime("%d/%m/%Y")
                    except Exception:
                        val = str(val)
                else:
                    val = str(val)
                if str(val).strip():
                    tem_valor = True
                linha.append(val)
            if tem_valor:
                dados.append(linha)
        return pd.DataFrame(dados, columns=nomes).fillna("")
    finally:
        try:
            wb.close()
        except Exception:
            pass

def ler_consolidado():
    if not _base_operacional_ativa():
        return pd.DataFrame(), "Nenhuma base carregada. Importe o Excel gerado pelo Tempo Fechado para iniciar a análise."

    if AUTO_IMPORTAR_EXCEL_ROBO:
        sync = sincronizar_excel_robo(forcar=False)
        if sync.get("copiado"):
            invalidar_cache_consolidado()

    if not ARQUIVO_PADRAO.exists():
        return pd.DataFrame(), (
            "Arquivo Excel não encontrado. Copie 'resultado_lote_consultas_validado.xlsx' para a pasta data "
            "ou ajuste ARQUIVO_EXCEL_ROBO em robo_ponto_web.py."
        )

    sig = _assinatura_arquivo(ARQUIVO_PADRAO)
    try:
        if sig and _CACHE_CONSOLIDADO_MEM.get("sig") == sig and _CACHE_CONSOLIDADO_MEM.get("df_raw") is not None:
            # Cópia defensiva: filtros e ordenações das rotas não contaminam o cache.
            return _CACHE_CONSOLIDADO_MEM["df_raw"].copy(deep=False), ""

        t0 = time.monotonic()
        df = _ler_excel_consolidado_otimizado(ARQUIVO_PADRAO)
        df = df.fillna("")
        _CACHE_CONSOLIDADO_MEM["sig"] = sig
        _CACHE_CONSOLIDADO_MEM["df_raw"] = df
        _CACHE_CONSOLIDADO_MEM["df_norm"] = None
        _CACHE_CONSOLIDADO_MEM["gerado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _CACHE_CONSOLIDADO_MEM["segundos_leitura"] = round(time.monotonic() - t0, 3)
        return df.copy(deep=False), ""
    except Exception as e:
        return pd.DataFrame(), f"Erro ao ler a aba Consolidado: {e}"


def _ler_excel_aba_otimizado(caminho_excel, nome_aba):
    """Le uma aba especifica do Excel usando o mesmo modo leve do Consolidado."""
    caminho_excel = Path(caminho_excel)
    nome_aba = str(nome_aba or '').strip()
    if not nome_aba:
        return pd.DataFrame()
    cache_key = (_chave_cache_excel(caminho_excel), "aba_otimizada", nome_aba)
    try:
        with _CACHE_EXCEL_LOCK:
            df_cache = _CACHE_EXCEL_ABAS_MEM.get(cache_key)
            if df_cache is not None:
                return df_cache.copy(deep=False)
    except Exception:
        pass

    if load_workbook is None:
        df = pd.read_excel(caminho_excel, sheet_name=nome_aba, dtype=str).fillna('')
        try:
            with _CACHE_EXCEL_LOCK:
                _CACHE_EXCEL_ABAS_MEM[cache_key] = df.copy(deep=False)
        except Exception:
            pass
        return df.copy(deep=False)

    wb = load_workbook(filename=str(caminho_excel), read_only=True, data_only=True)
    try:
        if nome_aba not in wb.sheetnames:
            raise ValueError(f"Aba {nome_aba} nao encontrada no Excel importado.")
        ws = wb[nome_aba]
        rows = ws.iter_rows(values_only=True)
        try:
            header = next(rows)
        except StopIteration:
            return pd.DataFrame()
        header = ['' if h is None else str(h).strip() for h in header]
        indices = []
        nomes = []
        for idx, nome in enumerate(header):
            if not nome:
                continue
            if _canon_header_excel_web(nome) in _COLUNAS_CONSOLIDADO_WEB_CANON:
                indices.append(idx)
                nomes.append(nome)
        if not indices:
            indices = list(range(len(header)))
            nomes = header
        dados = []
        for row in rows:
            if row is None:
                continue
            linha = []
            tem_valor = False
            for idx in indices:
                val = row[idx] if idx < len(row) else ''
                if val is None:
                    val = ''
                elif hasattr(val, 'strftime'):
                    try:
                        val = val.strftime('%d/%m/%Y')
                    except Exception:
                        val = str(val)
                else:
                    val = str(val)
                if str(val).strip():
                    tem_valor = True
                linha.append(val)
            if tem_valor:
                dados.append(linha)
        df = pd.DataFrame(dados, columns=nomes).fillna('')
        try:
            with _CACHE_EXCEL_LOCK:
                _CACHE_EXCEL_ABAS_MEM[cache_key] = df.copy(deep=False)
        except Exception:
            pass
        return df.copy(deep=False)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def ler_consolidado_historico_v6():
    """Le a aba historica do Script_Robo_Ponto_v6_PRO para auditoria premium."""
    if not _base_operacional_ativa():
        return pd.DataFrame(), 'Nenhuma base carregada. Importe o Excel gerado pelo robô para iniciar a auditoria.'
    if not ARQUIVO_PADRAO.exists():
        return pd.DataFrame(), 'Arquivo Excel ativo não encontrado.'

    aba_historico = _nome_aba_historico_v6(ARQUIVO_PADRAO)
    if not aba_historico:
        abas = _listar_abas_excel(ARQUIVO_PADRAO)
        if "Consolidado" in abas:
            try:
                df_fallback = _ler_excel_aba_otimizado(ARQUIVO_PADRAO, "Consolidado").fillna('')
                df_fallback.attrs["auditoria_aviso"] = (
                    "Aba Consolidado Historico nao encontrada. Exibindo a aba Consolidado em modo compativel; "
                    "para rastrear varias fotografias por fechamento, gere o Excel com historico pelo Script_Robo_Ponto_v6_PRO."
                )
                df_saida = df_fallback.copy(deep=False)
                df_saida.attrs["auditoria_aviso"] = df_fallback.attrs["auditoria_aviso"]
                return df_saida, ''
            except Exception as e:
                return pd.DataFrame(), f'Erro ao ler a aba Consolidado em modo compativel: {e}'
        abas_txt = ", ".join(abas) if abas else "nenhuma aba encontrada"
        return pd.DataFrame(), f'Aba Consolidado Historico nao encontrada. Abas disponiveis: {abas_txt}.'

    sig = (str(_assinatura_arquivo(ARQUIVO_PADRAO)), aba_historico)
    chave_raw = 'df_historico_raw'
    chave_norm = 'df_historico_norm'
    try:
        if _CACHE_CONSOLIDADO_MEM.get('sig_historico') == sig and _CACHE_CONSOLIDADO_MEM.get(chave_raw) is not None:
            return _CACHE_CONSOLIDADO_MEM[chave_raw].copy(deep=False), ''
        df = _ler_excel_aba_otimizado(ARQUIVO_PADRAO, aba_historico).fillna('')
        _CACHE_CONSOLIDADO_MEM['sig_historico'] = sig
        _CACHE_CONSOLIDADO_MEM[chave_raw] = df
        _CACHE_CONSOLIDADO_MEM[chave_norm] = None
        return df.copy(deep=False), ''
    except Exception as e:
        return pd.DataFrame(), f'Erro ao ler a aba histórica: {e}'


def normalizar_colunas_historico(df):
    sig = (str(_assinatura_arquivo(ARQUIVO_PADRAO)), _nome_aba_historico_v6(ARQUIVO_PADRAO))
    try:
        if _CACHE_CONSOLIDADO_MEM.get('sig_historico') == sig and _CACHE_CONSOLIDADO_MEM.get('df_historico_norm') is not None:
            return _CACHE_CONSOLIDADO_MEM['df_historico_norm'].copy(deep=False)
    except Exception:
        pass
    df_norm = _normalizar_colunas_impl(df)
    try:
        if _CACHE_CONSOLIDADO_MEM.get('sig_historico') == sig:
            _CACHE_CONSOLIDADO_MEM['df_historico_norm'] = df_norm.copy(deep=False)
    except Exception:
        pass
    return df_norm.copy(deep=False)




# ============================================================
# v8.15.19 - primeira fotografia do dia para o Extrato BH
# ============================================================
def _primeira_fotografia_por_dia_extrato_bh(g_operacional):
    """Retorna, por Data, a primeira Data Emissao PDF conhecida no historico.

    O Consolidado operacional do Script_Robo_Ponto_v7_PRO normalmente guarda a
    visao final/oficial do dia. Para auditoria do Extrato Banco de Horas, a
    coluna Data Emissao PDF deve mostrar quando aquele dia apareceu pela
    primeira vez nos PDFs importados. Ex.: se 15/05 apareceu pela primeira vez
    na emissao 16/05, a linha do extrato de 15/05 deve exibir 16/05, mesmo que
    exista uma emissao posterior contendo a mesma data.
    """
    mapa = {}
    try:
        if g_operacional is None or g_operacional.empty:
            return mapa
        g0 = g_operacional.copy()
        nome_ref = str(g0.get('nome', pd.Series([''])).iloc[0] if 'nome' in g0.columns else '').strip()
        mat_ref = _normalizar_matricula_bh(str(g0.get('matricula', pd.Series([''])).iloc[0] if 'matricula' in g0.columns else ''))
        cc_ref = ''.join(re.findall(r'\d+', str(g0.get('cc', pd.Series([''])).iloc[0] if 'cc' in g0.columns else '')))
        datas_ref = set(str(x).strip() for x in g0.get('data', pd.Series([], dtype='object')).astype(str).tolist() if str(x).strip())
        if not datas_ref:
            return mapa

        df_hist, erro = ler_consolidado_historico_v6()
        if erro or df_hist is None or df_hist.empty:
            return mapa
        h = normalizar_colunas_historico(df_hist)
        if h is None or h.empty or 'data' not in h.columns:
            return mapa

        h = h[h['data'].astype(str).str.strip().isin(datas_ref)].copy()
        if h.empty:
            return mapa

        if mat_ref and 'matricula' in h.columns:
            h['_mat_norm_bh19'] = h['matricula'].apply(_normalizar_matricula_bh)
            h = h[h['_mat_norm_bh19'] == mat_ref].copy()
        else:
            if nome_ref and 'nome' in h.columns:
                nome_norm = normalizar_nome(nome_ref)
                h = h[h['nome'].astype(str).apply(normalizar_nome) == nome_norm].copy()
            if cc_ref and 'cc' in h.columns and not h.empty:
                h['_cc_norm_bh19'] = h['cc'].astype(str).str.replace(r'\D+', '', regex=True)
                h = h[h['_cc_norm_bh19'].str.endswith(cc_ref[-6:] if len(cc_ref) >= 6 else cc_ref)].copy()
        if h.empty:
            return mapa

        h['_data_dt_bh19'] = pd.to_datetime(h['data'].astype(str).str.strip(), format='%d/%m/%Y', errors='coerce')
        if 'data_emissao_pdf' in h.columns:
            h['_emissao_dt_bh19'] = pd.to_datetime(h['data_emissao_pdf'].astype(str).str.strip(), format='%d/%m/%Y', errors='coerce')
        else:
            h['_emissao_dt_bh19'] = pd.NaT
        h = h.sort_values(['_data_dt_bh19', '_emissao_dt_bh19'], na_position='last', kind='mergesort')

        for data_txt, grupo in h.groupby(h['data'].astype(str).str.strip(), sort=False):
            if not data_txt:
                continue
            escolhido = grupo.iloc[0]
            saldo_info = ''
            for col in ('saldo_informado_pdf', 'saldo_informado', 'saldo_atual', 'saldo'):
                if col in escolhido.index:
                    val = str(escolhido.get(col, '') or '').strip()
                    if val and val.lower() not in {'nan', 'none', 'nat'}:
                        saldo_info = val
                        break
            mapa[data_txt] = {
                'data_emissao_pdf': str(escolhido.get('data_emissao_pdf', '') or '').strip(),
                'saldo_informado': saldo_info,
                'arquivo_origem': str(escolhido.get('arquivo_origem', '') or '').strip(),
                'origem': 'primeira_fotografia_historico',
            }
    except Exception:
        return mapa
    return mapa


def _primeira_linha_do_dia_por_emissao(linhas_data):
    """Fallback local: escolhe a primeira Data Emissao PDF dentro do dataframe recebido."""
    try:
        if linhas_data is None or linhas_data.empty:
            return None
        temp = linhas_data.copy()
        if 'data_emissao_pdf' in temp.columns:
            temp['_emissao_dt_bh19_local'] = pd.to_datetime(temp['data_emissao_pdf'].astype(str).str.strip(), format='%d/%m/%Y', errors='coerce')
            temp = temp.sort_values('_emissao_dt_bh19_local', na_position='last', kind='mergesort')
        return temp.iloc[0]
    except Exception:
        try:
            return linhas_data.iloc[0]
        except Exception:
            return None


# ============================================================
# v8.11.23 - payloads leves para preservar navegação
# ============================================================
LIMITE_LINHAS_RESPOSTA_TABELA = 800

def _df_para_resposta_leve(df, limite=LIMITE_LINHAS_RESPOSTA_TABELA):
    """Retorna linhas limitadas para a tela sem alterar KPIs.

    O Excel oficial pode conter muitos registros. Enviar/renderizar tudo em
    uma única resposta deixa o navegador sem resposta. A tabela exibe uma
    amostra operacional; os KPIs continuam calculados sobre o conjunto filtrado.
    """
    try:
        total = int(len(df))
    except Exception:
        total = 0
    if df is None or total == 0:
        return [], {"limitado": False, "total_sem_limite": 0, "limite_linhas": int(limite)}
    try:
        limite = int(limite)
    except Exception:
        limite = LIMITE_LINHAS_RESPOSTA_TABELA
    limitado = total > limite
    df_saida = df.head(limite).copy() if limitado else df
    return df_saida.to_dict(orient="records"), {
        "limitado": bool(limitado),
        "total_sem_limite": total,
        "limite_linhas": int(limite),
    }


def _init_anotacoes_db() -> sqlite3.Connection:
    db_path = Path.home() / "ponto_pdfs" / "config" / "anotacoes.db"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    conn.execute("""
        CREATE TABLE IF NOT EXISTS anotacoes (
            id              TEXT PRIMARY KEY,
            nome            TEXT NOT NULL,
            data_ref        TEXT NOT NULL DEFAULT '',
            data_anterior   TEXT DEFAULT '',
            data_retorno    TEXT DEFAULT '',
            tipo_violacao   TEXT DEFAULT '',
            anotacao        TEXT NOT NULL DEFAULT '',
            usuario_criacao TEXT NOT NULL DEFAULT '',
            criado_em       TEXT NOT NULL DEFAULT '',
            atualizado_em   TEXT NOT NULL DEFAULT ''
        )
    """)
    conn.commit()
    return conn


def _gerar_anotacao_id(nome: str, data_ref: str, tipo_violacao: str = "", extra: str = "") -> str:
    chave = f"{nome}|{data_ref}|{tipo_violacao}|{extra}"
    return hashlib.md5(chave.encode("utf-8")).hexdigest()[:16]


def _normalizar_colunas_impl(df):
    """
    Normaliza o Consolidado preservando, quando existir, a diferença entre:
    - Saldo Anterior
    - Saldo Atual

    Correção v8.10.8:
    A guia Extrato Banco de Horas deve usar somente o campo Saldo Atual da folha.
    Em versões anteriores, se o Excel trouxesse uma coluna genérica/antiga chamada
    "Saldo" preenchida pelo extrator com o primeiro número do quadro Banco de Horas,
    o sistema podia acabar usando Saldo Anterior como se fosse Saldo Atual.

    Regra de preferência:
    1) coluna explícita Saldo Atual / Banco de Horas - Saldo Atual / BH Saldo Atual;
    2) coluna genérica Saldo apenas como fallback;
    3) Saldo Anterior nunca abre o ciclo do Extrato/Banco de Horas.
    """
    mapa_basico = {
        "Arquivo Origem": "arquivo_origem", "CC": "cc", "cc": "cc", "Nome": "nome", "Matrícula": "matricula", "Matricula": "matricula", "Chapa": "matricula", "Função": "funcao", "Turno": "turno",
        "Data": "data", "Dia": "dia", "1a E.": "e1", "1a S.": "s1", "2a E.": "e2", "2a S.": "s2",
        "3a E.": "e3", "3a S.": "s3", "4a E.": "e4", "4a S.": "s4", "Abono": "abono",
        "H.E.": "he", "Absent.": "absent", "Jornada": "jornada", "Ad. Not.": "adnot",
        "Observação": "observacao", "Status Validação": "status",
        "Inconsistências": "inconsistencias", "Jornada Calculada": "jornada_calculada", "Qtd Batidas": "qtd_batidas",
    }

    # Variantes possíveis vindas do Excel do robô ou de futuras extrações.
    # Ordem importante: colunas criadas pela correção direta do PDF têm prioridade,
    # mas somente quando possuem valor. Se a correção por PDF não encontrou nada,
    # o sistema cai para a coluna explícita Saldo Atual do Excel.
    variantes_saldo_atual_prioridade = {
        "saldo atual pdf": 1, "saldo_atual_pdf": 1,
        "saldo informado pdf": 2, "saldo informado": 2, "saldo_informado": 2,
        "saldo atual": 3, "saldo_atual": 3, "saldo atual bh": 3, "bh saldo atual": 3,
        "banco de horas - saldo atual": 3, "banco de horas saldo atual": 3, "saldo atual banco de horas": 3,
    }
    variantes_saldo_atual = set(variantes_saldo_atual_prioridade.keys())
    variantes_saldo_anterior = {
        "saldo anterior pdf", "saldo_anterior_pdf",
        "saldo anterior", "saldo_anterior", "banco de horas - saldo anterior", "bh saldo anterior",
        "saldo anterior banco de horas",
    }
    variantes_debito_bh = {"debito pdf", "débito pdf", "débito bh pdf", "debito bh pdf", "débito", "debito", "débito bh", "debito bh", "banco de horas - débito", "banco de horas - debito"}
    variantes_credito_bh = {"credito pdf", "crédito pdf", "crédito bh pdf", "credito bh pdf", "crédito", "credito", "crédito bh", "credito bh", "banco de horas - crédito", "banco de horas - credito"}
    variantes_saldo_calculado = {"saldo calculado", "saldo_calculado", "saldo calculado bh", "bh saldo calculado", "banco de horas - saldo calculado"}

    def canon(col):
        texto = str(col or "").strip().lower()
        texto = texto.replace("á", "a").replace("à", "a").replace("ã", "a").replace("â", "a")
        texto = texto.replace("é", "e").replace("ê", "e")
        texto = texto.replace("í", "i")
        texto = texto.replace("ó", "o").replace("ô", "o").replace("õ", "o")
        texto = texto.replace("ú", "u").replace("ç", "c")
        texto = " ".join(texto.replace("_", " ").split())
        return texto

    saida = pd.DataFrame()

    for origem, destino in mapa_basico.items():
        if origem in df.columns and destino not in saida.columns:
            saida[destino] = df[origem].astype(str)

    # Preserva colunas de saldo por semântica, sem confundir Saldo Anterior com Saldo Atual.
    colunas_saldo_atual = []
    coluna_saldo_anterior = None
    coluna_saldo_generico = None
    coluna_debito_bh = None
    coluna_credito_bh = None
    coluna_saldo_calculado = None

    variantes_debito_bh_canon = {canon(x) for x in variantes_debito_bh}
    variantes_credito_bh_canon = {canon(x) for x in variantes_credito_bh}
    variantes_saldo_calculado_canon = {canon(x) for x in variantes_saldo_calculado}

    for col in df.columns:
        c = canon(col)
        if c in variantes_saldo_atual:
            colunas_saldo_atual.append((variantes_saldo_atual_prioridade.get(c, 99), col))
        elif c in variantes_saldo_anterior:
            coluna_saldo_anterior = col
        elif c == "saldo":
            coluna_saldo_generico = col
        elif c in variantes_debito_bh_canon:
            coluna_debito_bh = col
        elif c in variantes_credito_bh_canon:
            coluna_credito_bh = col
        elif c in variantes_saldo_calculado_canon:
            coluna_saldo_calculado = col

    colunas_saldo_atual = [col for _, col in sorted(colunas_saldo_atual, key=lambda item: item[0])]

    def _valor_preenchido(valor):
        texto = str(valor or "").strip()
        return texto and texto.lower() not in {"nan", "none", "nat"}

    def _coalescer_colunas(df_origem, colunas):
        if not colunas:
            return pd.Series([""] * len(df_origem), index=df_origem.index, dtype="object")
        resultado = pd.Series([""] * len(df_origem), index=df_origem.index, dtype="object")
        for col in colunas:
            if col not in df_origem.columns:
                continue
            serie = df_origem[col].astype(str).str.strip()
            preenchido = (serie != "") & (~serie.str.lower().isin({"nan", "none", "nat"}))
            pendente = resultado.astype(str).str.strip() == ""
            if preenchido.any() and pendente.any():
                mask = pendente & preenchido
                resultado.loc[mask] = serie.loc[mask]
        return resultado.astype("object")

    # v8.11.17: preserva explicitamente a coluna nova do Script_Robo_Ponto.
    # Ela representa o Saldo Atual informado no PDF, sem cálculo.
    colunas_saldo_informado_pdf = []
    for col in df.columns:
        c = canon(col)
        if c in {"saldo informado pdf", "saldo informado", "saldo_informado", "saldo atual pdf", "saldo_atual_pdf"}:
            colunas_saldo_informado_pdf.append(col)
    saida["saldo_informado_pdf"] = _coalescer_colunas(df, colunas_saldo_informado_pdf).astype(str)

    # v8.11.17: quando o Excel já traz a coluna Saldo Informado PDF, ela
    # passa a ser a fonte oficial por linha. Valores vazios são intencionais:
    # significam que aquele dia não é a fotografia/data de emissão do PDF.
    saida["saldo_informado_pdf_excel_presente"] = "1" if colunas_saldo_informado_pdf else ""

    colunas_data_emissao_pdf = []
    for col in df.columns:
        c = canon(col)
        if c in {"data emissao pdf", "data de emissao pdf", "emissao pdf", "emissao"}:
            colunas_data_emissao_pdf.append(col)
    saida["data_emissao_pdf"] = _coalescer_colunas(df, colunas_data_emissao_pdf).astype(str)

    if colunas_saldo_atual:
        # Coalesce por linha preservando a origem real do valor escolhido.
        saldo_atual_serie = pd.Series([""] * len(df), index=df.index, dtype="object")
        origem_serie = pd.Series(["saldo_atual_nao_encontrado"] * len(df), index=df.index, dtype="object")
        for col in colunas_saldo_atual:
            if col not in df.columns:
                continue
            serie = df[col].astype(str).str.strip()
            preenchido = (serie != "") & (~serie.str.lower().isin({"nan", "none", "nat"}))
            pendente = saldo_atual_serie.astype(str).str.strip() == ""
            if preenchido.any() and pendente.any():
                mask = pendente & preenchido
                saldo_atual_serie.loc[mask] = serie.loc[mask]
                origem_serie.loc[mask] = canon(col).replace(" ", "_")
        if coluna_saldo_generico is not None and coluna_saldo_generico in df.columns:
            serie = df[coluna_saldo_generico].astype(str).str.strip()
            preenchido = (serie != "") & (~serie.str.lower().isin({"nan", "none", "nat"}))
            pendente = saldo_atual_serie.astype(str).str.strip() == ""
            if preenchido.any() and pendente.any():
                mask = pendente & preenchido
                saldo_atual_serie.loc[mask] = serie.loc[mask]
                origem_serie.loc[mask] = "coluna_generica_saldo_fallback"
        saida["saldo_atual"] = saldo_atual_serie.astype(str)
        saida["saldo"] = saldo_atual_serie.astype(str)
        saida["origem_saldo_importado"] = origem_serie.astype("object")
    elif coluna_saldo_generico is not None:
        saida["saldo"] = df[coluna_saldo_generico].astype(str)
        saida["saldo_atual"] = df[coluna_saldo_generico].astype(str)
        saida["origem_saldo_importado"] = "coluna_generica_saldo_fallback"
    else:
        saida["saldo"] = ""
        saida["saldo_atual"] = ""
        saida["origem_saldo_importado"] = "saldo_nao_encontrado"

    if coluna_saldo_anterior is not None:
        saida["saldo_anterior"] = df[coluna_saldo_anterior].astype(str)
    else:
        saida["saldo_anterior"] = ""

    if coluna_debito_bh is not None:
        saida["debito_bh"] = df[coluna_debito_bh].astype(str)
    else:
        saida["debito_bh"] = ""

    if coluna_credito_bh is not None:
        saida["credito_bh"] = df[coluna_credito_bh].astype(str)
    else:
        saida["credito_bh"] = ""

    # v8.11.17: Saldo Calculado passa a vir pronto do Excel gerado pelo robô.
    # O Tempo Fechado deve consumir essa coluna como fonte oficial do Extrato
    # e da guia Banco de Horas, sem recalcular nem buscar cache PDF.
    if coluna_saldo_calculado is not None:
        saida["saldo_calculado"] = df[coluna_saldo_calculado].astype(str)
    else:
        saida["saldo_calculado"] = ""

    obrigatorias = set(mapa_basico.values()) | {"saldo", "saldo_atual", "saldo_anterior", "debito_bh", "credito_bh", "saldo_calculado", "origem_saldo_importado", "matricula", "arquivo_origem", "saldo_informado_pdf", "saldo_informado_pdf_excel_presente", "data_emissao_pdf"}
    for destino in obrigatorias:
        if destino not in saida.columns:
            saida[destino] = ""

    saida["status"] = saida["status"].replace({"": "OK", "nan": "OK", "None": "OK"})
    return saida




def _garantir_colunas_calculadas(df):
    """Garante colunas derivadas usadas em múltiplas APIs sem recalcular à toa.

    v8.13.10: centraliza data_dt, he_min, absent_min e tem_revisao.
    Antes, cada rota recalculava essas colunas isoladamente, criando trabalho
    repetido na navegação entre Dashboard, Alertas e Banco de Horas.
    A função é idempotente e preserva a regra de negócio existente.
    """
    if df is None or df.empty:
        return df
    base = df.copy()

    if "data_dt" not in base.columns and "data" in base.columns:
        base["data_dt"] = pd.to_datetime(base["data"].astype(str).str.strip(), format="%d/%m/%Y", errors="coerce")

    if "he_min" not in base.columns and "he" in base.columns:
        base["he_min"] = base["he"].apply(hhmm_para_minutos)
    if "absent_min" not in base.columns and "absent" in base.columns:
        base["absent_min"] = base["absent"].apply(hhmm_para_minutos)

    if "he_min" in base.columns:
        base["he_min"] = pd.to_numeric(base["he_min"], errors="coerce").fillna(0).astype(int)
    if "absent_min" in base.columns:
        base["absent_min"] = pd.to_numeric(base["absent_min"], errors="coerce").fillna(0).astype(int)

    if "tem_revisao" not in base.columns:
        status = base["status"].astype(str) if "status" in base.columns else pd.Series(["OK"] * len(base), index=base.index)
        inconsistencias = base["inconsistencias"].astype(str) if "inconsistencias" in base.columns else pd.Series([""] * len(base), index=base.index)
        base["tem_revisao"] = (status.str.upper() != "OK") | (inconsistencias.str.strip() != "")

    return base


def normalizar_colunas(df):
    """Normaliza o Consolidado com cache em memória.

    v8.10.57: cada endpoint continuava lendo o mesmo Excel e refazendo a
    normalização completa. Agora, enquanto o arquivo importado não muda, a
    normalização é reaproveitada em memória.
    """
    sig = _assinatura_arquivo(ARQUIVO_PADRAO)
    try:
        if sig and _CACHE_CONSOLIDADO_MEM.get("sig") == sig and _CACHE_CONSOLIDADO_MEM.get("df_norm") is not None:
            return _CACHE_CONSOLIDADO_MEM["df_norm"].copy(deep=False)
    except Exception:
        pass

    t0 = time.monotonic()
    df_norm = _normalizar_colunas_impl(df)
    try:
        if sig and _CACHE_CONSOLIDADO_MEM.get("sig") == sig:
            _CACHE_CONSOLIDADO_MEM["df_norm"] = df_norm.copy(deep=False)
            _CACHE_CONSOLIDADO_MEM["segundos_normalizacao"] = round(time.monotonic() - t0, 3)
    except Exception:
        pass
    return df_norm.copy(deep=False)



# ============================================================
# AUTENTICAÇÃO - MULTIUSUÁRIO V1
# ============================================================

JWT_COOKIE_NAME = "tf_access_token"
JWT_ALGORITHM = "HS256"
JWT_ISSUER = "tempo-fechado"
JWT_EXP_SECONDS = int(os.environ.get("TEMPO_FECHADO_JWT_EXP_SECONDS", str(12 * 60 * 60)))


def _jwt_secret():
    return (os.environ.get("TEMPO_FECHADO_JWT_SECRET") or app.secret_key).encode("utf-8")


def _b64url_encode(data):
    raw = data if isinstance(data, bytes) else str(data).encode("utf-8")
    return base64.urlsafe_b64encode(raw).rstrip(b"=").decode("ascii")


def _b64url_decode(texto):
    texto = str(texto or "")
    padding = "=" * (-len(texto) % 4)
    return base64.urlsafe_b64decode((texto + padding).encode("ascii"))


def _json_canonico(obj):
    return json.dumps(obj, separators=(",", ":"), sort_keys=True, ensure_ascii=False).encode("utf-8")


def criar_jwt_usuario(dados_usuario):
    agora = int(time.time())
    usuario = (dados_usuario or {}).get("usuario", "")
    payload = {
        "iss": JWT_ISSUER,
        "sub": usuario,
        "nome": (dados_usuario or {}).get("nome", usuario),
        "perfil": (dados_usuario or {}).get("perfil", "consulta"),
        "trocar_senha": bool((dados_usuario or {}).get("trocar_senha", False)),
        "iat": agora,
        "exp": agora + JWT_EXP_SECONDS,
        "jti": uuid.uuid4().hex,
    }
    header = {"typ": "JWT", "alg": JWT_ALGORITHM}
    cabecalho = _b64url_encode(_json_canonico(header))
    corpo = _b64url_encode(_json_canonico(payload))
    assinatura = hmac.new(_jwt_secret(), f"{cabecalho}.{corpo}".encode("ascii"), hashlib.sha256).digest()
    return f"{cabecalho}.{corpo}.{_b64url_encode(assinatura)}", payload


def _token_da_requisicao():
    auth = request.headers.get("Authorization", "")
    if auth.lower().startswith("bearer "):
        return auth.split(" ", 1)[1].strip()
    return request.cookies.get(JWT_COOKIE_NAME, "")


def validar_jwt_usuario(token):
    try:
        partes = str(token or "").split(".")
        if len(partes) != 3:
            return None, "Token ausente ou malformado."
        cabecalho_b64, corpo_b64, assinatura_b64 = partes
        assinatura_esperada = hmac.new(_jwt_secret(), f"{cabecalho_b64}.{corpo_b64}".encode("ascii"), hashlib.sha256).digest()
        assinatura_recebida = _b64url_decode(assinatura_b64)
        if not hmac.compare_digest(assinatura_esperada, assinatura_recebida):
            return None, "Token com assinatura invalida."
        cabecalho = json.loads(_b64url_decode(cabecalho_b64).decode("utf-8"))
        payload = json.loads(_b64url_decode(corpo_b64).decode("utf-8"))
        if cabecalho.get("alg") != JWT_ALGORITHM:
            return None, "Algoritmo JWT nao permitido."
        if payload.get("iss") != JWT_ISSUER:
            return None, "Emissor JWT invalido."
        if int(payload.get("exp") or 0) < int(time.time()):
            return None, "Token expirado."
        usuario = str(payload.get("sub") or "").strip().lower()
        dados = obter_usuario(ARQUIVO_USUARIOS_DB, usuario)
        if not dados or not dados.get("ativo"):
            return None, "Usuario do token nao encontrado ou desativado."
        return {
            "usuario": dados.get("usuario"),
            "nome": dados.get("nome"),
            "perfil": dados.get("perfil", "consulta"),
            "trocar_senha": bool(dados.get("trocar_senha", False)),
            "jwt_payload": payload,
        }, ""
    except Exception as e:
        return None, f"Token invalido: {e}"


def _aplicar_usuario_na_sessao(dados_usuario):
    session["usuario"] = dados_usuario.get("usuario")
    session["nome"] = dados_usuario.get("nome") or dados_usuario.get("usuario")
    session["perfil"] = dados_usuario.get("perfil", "consulta")
    session["trocar_senha"] = bool(dados_usuario.get("trocar_senha", False))


def autenticar_requisicao_atual():
    token = _token_da_requisicao()
    if token:
        dados, erro = validar_jwt_usuario(token)
        if dados:
            _aplicar_usuario_na_sessao(dados)
            return dados, ""
        if request.path.startswith("/api/"):
            return None, erro
    if session.get("usuario"):
        return {
            "usuario": session.get("usuario"),
            "nome": session.get("nome"),
            "perfil": session.get("perfil", ""),
            "trocar_senha": bool(session.get("trocar_senha", False)),
        }, ""
    return None, "Sessao expirada. Faca login novamente."


def _resposta_com_cookie_jwt(response, dados_usuario):
    token, payload = criar_jwt_usuario(dados_usuario)
    response.set_cookie(
        JWT_COOKIE_NAME,
        token,
        max_age=JWT_EXP_SECONDS,
        httponly=True,
        samesite="Lax",
        secure=bool(request.is_secure),
    )
    try:
        response.headers["X-Tempo-Fechado-Auth"] = "jwt"
        response.headers["X-Tempo-Fechado-JWT-Exp"] = str(payload.get("exp", ""))
    except Exception:
        pass
    return response


def _limpar_cookie_jwt(response):
    response.delete_cookie(JWT_COOKIE_NAME, samesite="Lax")
    return response

def usuario_logado():
    dados, _ = autenticar_requisicao_atual()
    return dados.get("usuario") if dados else None


def perfil_usuario():
    dados, _ = autenticar_requisicao_atual()
    return dados.get("perfil", "") if dados else ""


def usuario_eh_admin():
    return perfil_usuario() == "admin"


def registrar_auditoria(acao, detalhe="", status="OK"):
    """Registra ações relevantes do modo multiusuário em arquivo texto simples."""
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        usuario = session.get("usuario", "sem_login")
        nome = session.get("nome", "")
        perfil = session.get("perfil", "")
        ip = request.headers.get("X-Forwarded-For", request.remote_addr or "")
        quando = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        linha = f"{quando}\t{usuario}\t{nome}\t{perfil}\t{ip}\t{acao}\t{status}\t{detalhe}\n"
        with ARQUIVO_AUDITORIA_MULTIUSUARIO.open("a", encoding="utf-8") as f:
            f.write(linha)
    except Exception:
        pass


def login_obrigatorio(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        dados_auth, erro_auth = autenticar_requisicao_atual()
        if not dados_auth:
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "erro": erro_auth or "Sessao expirada. Faca login novamente."}), 401
            return redirect(url_for("login"))
        return func(*args, **kwargs)
    return wrapper


def admin_obrigatorio(func):
    @wraps(func)
    def wrapper(*args, **kwargs):
        dados_auth, erro_auth = autenticar_requisicao_atual()
        if not dados_auth:
            if request.path.startswith("/api/"):
                return jsonify({"ok": False, "erro": erro_auth or "Sessao expirada. Faca login novamente."}), 401
            return redirect(url_for("login"))
        if not usuario_eh_admin():
            registrar_auditoria("ACESSO_NEGADO", request.path, "NEGADO")
            if request.path.startswith("/api/"):
                return jsonify({
                    "ok": False,
                    "erro": "Acesso negado. Esta ação exige perfil Administrador.",
                    "perfil_atual": perfil_usuario(),
                }), 403
            return "Acesso negado. Perfil administrador necessário.", 403
        return func(*args, **kwargs)
    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    erro = ""

    if request.method == "POST":
        usuario = request.form.get("usuario", "").strip()
        senha = request.form.get("senha", "").strip()

        ok, dados_usuario, mensagem = autenticar_usuario(ARQUIVO_USUARIOS_DB, usuario, senha)

        if ok and dados_usuario:
            _aplicar_usuario_na_sessao(dados_usuario)
            registrar_auditoria("LOGIN", "Entrada no sistema", "OK")
            return _resposta_com_cookie_jwt(redirect(url_for("index")), dados_usuario)

        registrar_auditoria("LOGIN_FALHA", f"Usuário informado: {usuario}", "NEGADO")
        erro = mensagem or "Usuário ou senha inválidos."

    return render_template("login.html", erro=erro)


@app.route("/logout")
def logout():
    registrar_auditoria("LOGOUT", "Saida do sistema", "OK")
    session.clear()
    return _limpar_cookie_jwt(redirect(url_for("login")))


@app.route("/api/auth/login", methods=["POST"])
def api_auth_login_jwt():
    payload = request.get_json(silent=True) or {}
    usuario = str(payload.get("usuario", "")).strip()
    senha = str(payload.get("senha", "")).strip()
    ok, dados_usuario, mensagem = autenticar_usuario(ARQUIVO_USUARIOS_DB, usuario, senha)
    if not ok or not dados_usuario:
        registrar_auditoria("LOGIN_API_FALHA", f"Usuario informado: {usuario}", "NEGADO")
        return jsonify({"ok": False, "erro": mensagem or "Usuario ou senha invalidos."}), 401
    _aplicar_usuario_na_sessao(dados_usuario)
    token, jwt_payload = criar_jwt_usuario(dados_usuario)
    registrar_auditoria("LOGIN_API", "Token JWT emitido", "OK")
    resp = jsonify({
        "ok": True,
        "token_type": "Bearer",
        "access_token": token,
        "expires_in": JWT_EXP_SECONDS,
        "exp": jwt_payload.get("exp"),
        "usuario": {
            "usuario": dados_usuario.get("usuario"),
            "nome": dados_usuario.get("nome"),
            "perfil": dados_usuario.get("perfil"),
            "trocar_senha": bool(dados_usuario.get("trocar_senha", False)),
        },
    })
    return _resposta_com_cookie_jwt(resp, dados_usuario)


@app.route("/api/auth/refresh", methods=["POST"])
@login_obrigatorio
def api_auth_refresh_jwt():
    dados_usuario, _ = autenticar_requisicao_atual()
    token, jwt_payload = criar_jwt_usuario(dados_usuario)
    resp = jsonify({
        "ok": True,
        "token_type": "Bearer",
        "access_token": token,
        "expires_in": JWT_EXP_SECONDS,
        "exp": jwt_payload.get("exp"),
    })
    return _resposta_com_cookie_jwt(resp, dados_usuario)




@app.errorhandler(404)
def tratar_404_api(e):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "erro": f"Endpoint não encontrado: {request.path}"}), 404
    return e


@app.errorhandler(405)
def tratar_405_api(e):
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "erro": f"Método não permitido para {request.path}."}), 405
    return e


@app.errorhandler(Exception)
def tratar_erro_api(e):
    # Evita que endpoints /api devolvam HTML. Assim o JavaScript recebe JSON
    # e mostra o erro real, em vez de "Unexpected token '<'".
    if request.path.startswith("/api/"):
        try:
            registrar_auditoria("ERRO_API", f"{request.path}: {e}", "ERRO")
        except Exception:
            pass
        return jsonify({"ok": False, "erro": f"Falha interna em {request.path}: {e}"}), 500
    raise e


@app.route("/api/versao-publica")
def api_versao_publica_v8215():
    return jsonify({
        "ok": True,
        "versao": APP_VERSION,
        "nome": APP_RELEASE_NAME,
        "titulo": APP_FULL_NAME,
        "iniciado_em": APP_STARTED_AT,
    })


@app.route("/api/versao")
@login_obrigatorio
def api_versao_ativa_v8215():
    try:
        caminho_executavel = str(Path(sys.executable).resolve())
    except Exception:
        caminho_executavel = sys.executable
    try:
        arquivo_app = str(Path(__file__).resolve())
    except Exception:
        arquivo_app = __file__
    return jsonify({
        "ok": True,
        "versao": APP_VERSION,
        "nome": APP_RELEASE_NAME,
        "titulo": APP_FULL_NAME,
        "iniciado_em": APP_STARTED_AT,
        "host": request.host,
        "url": request.url_root,
        "pasta_app": APP_BASE_DIR,
        "arquivo_backend": arquivo_app,
        "executavel": caminho_executavel,
        "python_frozen": bool(getattr(sys, "frozen", False)),
        "porta_padrao": 5050,
        "usuario_windows": os.environ.get("USERNAME") or os.environ.get("USER") or "",
        "maquina": platform.node(),
    })


@app.route("/api/usuario-atual")
@login_obrigatorio
def api_usuario_atual():
    perfil = session.get("perfil")
    return jsonify({
        "usuario": session.get("usuario"),
        "nome": session.get("nome"),
        "perfil": perfil,
        "is_admin": perfil == "admin",
        "trocar_senha": bool(session.get("trocar_senha", False)),
        "permissoes": {
            "importar_excel": perfil == "admin",
            "enviar_alertas": perfil == "admin",
            "gerenciar_usuarios": perfil == "admin",
            "baixar_excel": True,
            "consultar_dashboards": True,
        }
    })

@app.route("/")
@login_obrigatorio
def index():
    return render_template("index.html")




@app.route("/api/nomes-por-cc")
@login_obrigatorio
def api_nomes_por_cc():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "nomes": []})

    df = normalizar_colunas(df)

    cc_param = request.args.get("cc", "").strip()

    if cc_param and "cc" in df.columns:
        df = aplicar_filtro_cc_backend(df, cc_param)

    nomes = []
    if "nome" in df.columns:
        nomes = sorted(
            set(
                str(v).strip()
                for v in df["nome"].dropna().tolist()
                if str(v).strip() and str(v).strip().lower() != "nan"
            ),
            key=lambda x: x.lower()
        )

    return jsonify({
        "erro": "",
        "cc": cc_param,
        "total_nomes": len(nomes),
        "nomes": nomes,
    })



@app.route("/api/opcoes-filtros")
@login_obrigatorio
def api_opcoes_filtros():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "ccs": [], "nomes": [], "turnos": []})

    sig = str(_assinatura_arquivo(ARQUIVO_PADRAO))
    cc_param = request.args.get("cc", "").strip()
    refresh_param = str(request.args.get("refresh", "")).strip().lower() in {"1", "true", "sim", "yes"}

    # Sem filtro de CC, devolve cache pronto das opções. Isso evita normalizar
    # e varrer a base completa a cada importação/troca de guia.
    if not cc_param and not refresh_param and _CACHE_OPCOES_FILTROS_MEM.get("sig") == sig:
        return jsonify({
            "erro": "",
            "ccs": _CACHE_OPCOES_FILTROS_MEM.get("ccs", []),
            "nomes": _CACHE_OPCOES_FILTROS_MEM.get("nomes", []),
            "turnos": _CACHE_OPCOES_FILTROS_MEM.get("turnos", []),
        })

    df = normalizar_colunas(df)
    if df.empty:
        return jsonify({"erro": "", "ccs": [], "nomes": [], "turnos": []})

    ccs = []
    nomes = []
    turnos = []

    if "cc" in df.columns:
        ccs = sorted(set(valor_cc_display(v) for v in df["cc"].dropna().tolist() if valor_cc_display(v)), key=lambda x: x.lower())

    df_nomes = df
    if cc_param and "cc" in df_nomes.columns:
        df_nomes = aplicar_filtro_cc_backend(df_nomes, cc_param)

    if "nome" in df_nomes.columns:
        nomes = sorted(set(str(v).strip() for v in df_nomes["nome"].dropna().tolist() if str(v).strip() and str(v).strip().lower() != "nan"), key=lambda x: x.lower())

    df_turnos = df
    if cc_param and "cc" in df_turnos.columns:
        df_turnos = aplicar_filtro_cc_backend(df_turnos, cc_param)

    if "turno" in df_turnos.columns:
        turnos = sorted(set(str(v).strip() for v in df_turnos["turno"].dropna().tolist() if str(v).strip() and str(v).strip().lower() != "nan"), key=lambda x: x.lower())

    if not cc_param:
        _CACHE_OPCOES_FILTROS_MEM.update({"sig": sig, "ccs": ccs, "nomes": nomes, "turnos": turnos, "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})

    return jsonify({"erro": "", "ccs": ccs, "nomes": nomes, "turnos": turnos})



def _parse_data_espelho_pdf_v55(valor):
    """Converte a data da linha do Espelho com tolerância a Excel/datetime/texto."""
    try:
        if valor is None:
            return pd.NaT
        if isinstance(valor, pd.Timestamp):
            return valor.normalize()
        texto = str(valor).strip()
        if not texto or texto.lower() in {"nan", "none", "nat"}:
            return pd.NaT
        # Primeiro tenta o padrão exibido no app.
        dt = pd.to_datetime(texto, format="%d/%m/%Y", errors="coerce")
        if pd.notna(dt):
            return pd.to_datetime(dt).normalize()
        # Depois aceita strings vindas do Excel como 2026-05-15 00:00:00.
        dt = pd.to_datetime(texto, dayfirst=True, errors="coerce")
        if pd.notna(dt):
            return pd.to_datetime(dt).normalize()
    except Exception:
        pass
    return pd.NaT


def aplicar_saldo_informado_pdf_no_espelho(df):
    """Aplica na coluna Saldo Atual do Espelho o mesmo conceito de
    Saldo Informado do Extrato Banco de Horas.

    v8.10.55 - correção cirúrgica sobre a base real v8.10.45/v8.10.50:
    - usa somente o cache/histórico de PDFs já atualizado;
    - não varre PDFs em disco;
    - atualiza todos os ciclos disponíveis, não apenas o ciclo mais recente;
    - trata datas em formato brasileiro e também datetime do Excel;
    - para a primeira data real de cada ciclo, usa o fechamento do ciclo anterior,
      como o Extrato; para as demais datas, usa a fotografia PDF da data.
    """
    try:
        if df is None or df.empty:
            return df
        cache = carregar_cache_pdfs_banco_horas()
        catalogo = cache.get("catalogo", []) if cache else []
        if not catalogo:
            return df
    except Exception:
        return df

    try:
        saida = df.copy()
        if "data" not in saida.columns or "nome" not in saida.columns:
            return saida

        saida["_data_dt_espelho_pdf_v55"] = saida["data"].apply(_parse_data_espelho_pdf_v55)
        saida["_inicio_ciclo_espelho_pdf_v55"] = saida["_data_dt_espelho_pdf_v55"].apply(
            lambda d: _data_inicio_medicao(d) if pd.notna(d) else pd.NaT
        )

        # Descobre a primeira data real de cada colaborador/ciclo dentro da base completa
        # já normalizada. Isso evita tratar uma data filtrada como se fosse abertura do ciclo.
        primeira_data_por_chave = {}
        try:
            tmp = saida[saida["_data_dt_espelho_pdf_v55"].notna()].copy()
            if not tmp.empty:
                for idx, row in tmp.iterrows():
                    nome_key = str(row.get("nome", "") or "").strip().lower()
                    mat_key = str(_matricula_da_linha_bh(row) or "").strip()
                    ciclo = row.get("_inicio_ciclo_espelho_pdf_v55")
                    if pd.isna(ciclo):
                        continue
                    ciclo_key = pd.to_datetime(ciclo).strftime("%Y-%m-%d")
                    chave = (nome_key, mat_key, ciclo_key)
                    data_dt = row.get("_data_dt_espelho_pdf_v55")
                    atual = primeira_data_por_chave.get(chave)
                    if atual is None or pd.to_datetime(data_dt) < pd.to_datetime(atual):
                        primeira_data_por_chave[chave] = pd.to_datetime(data_dt).normalize()
        except Exception:
            primeira_data_por_chave = {}

        memo_data = {}
        memo_base = {}

        for idx, row in saida[saida["_data_dt_espelho_pdf_v55"].notna()].iterrows():
            data_dt = pd.to_datetime(row.get("_data_dt_espelho_pdf_v55")).normalize()
            nome_pdf = str(row.get("nome", "") or "").strip()
            matricula_pdf = str(_matricula_da_linha_bh(row) or "").strip()
            inicio = row.get("_inicio_ciclo_espelho_pdf_v55")
            if pd.isna(inicio):
                inicio = _data_inicio_medicao(data_dt)
            if inicio is None or pd.isna(inicio):
                continue
            inicio = pd.to_datetime(inicio).normalize()

            nome_key = nome_pdf.lower()
            ciclo_key = inicio.strftime("%Y-%m-%d")
            chave_ciclo = (nome_key, matricula_pdf, ciclo_key)
            primeira_data = primeira_data_por_chave.get(chave_ciclo)

            saldo_pdf = ""
            origem_pdf = ""

            try:
                # Repete a regra do Extrato: a abertura do ciclo usa o PDF de fechamento
                # do ciclo anterior. Não faz busca direta em disco.
                if primeira_data is not None and data_dt == pd.to_datetime(primeira_data).normalize():
                    if chave_ciclo not in memo_base:
                        memo_base[chave_ciclo] = _buscar_pdf_fechamento_ciclo_anterior(
                            nome_pdf, matricula_pdf, inicio, catalogo, permitir_busca_direta=False
                        )
                    pdf_base = memo_base.get(chave_ciclo)
                    if pdf_base and str(pdf_base.get("saldo_atual_pdf", "") or "").strip():
                        saldo_pdf = str(pdf_base.get("saldo_atual_pdf", "") or "").strip()
                        origem_pdf = f"saldo_informado_extrato_pdf_base: {pdf_base.get('arquivo_pdf','')}"

                # Para qualquer data sem saldo-base encontrado, usa a fotografia PDF da data.
                if not saldo_pdf:
                    chave_data = (nome_key, matricula_pdf, data_dt.strftime("%Y-%m-%d"))
                    if chave_data not in memo_data:
                        memo_data[chave_data] = _saldo_pdf_para_data_exata(nome_pdf, matricula_pdf, data_dt, catalogo)
                    pdf_data = memo_data.get(chave_data)
                    if pdf_data and str(pdf_data.get("saldo_atual_pdf", "") or "").strip():
                        saldo_pdf = str(pdf_data.get("saldo_atual_pdf", "") or "").strip()
                        origem_pdf = f"saldo_informado_extrato_pdf_data: {pdf_data.get('arquivo_pdf','')}"
            except Exception:
                saldo_pdf = ""
                origem_pdf = ""

            if saldo_pdf:
                saida.at[idx, "saldo_atual"] = saldo_pdf
                saida.at[idx, "saldo"] = saldo_pdf
                saida.at[idx, "origem_saldo_importado"] = origem_pdf or "saldo_informado_extrato_pdf"

        for col in ["_data_dt_espelho_pdf_v55", "_inicio_ciclo_espelho_pdf_v55"]:
            if col in saida.columns:
                try:
                    saida = saida.drop(columns=[col])
                except Exception:
                    pass
        return saida
    except Exception:
        try:
            for col in ["_data_dt_espelho_pdf_v55", "_inicio_ciclo_espelho_pdf_v55"]:
                if col in df.columns:
                    df = df.drop(columns=[col])
        except Exception:
            pass
        return df

@app.route("/api/anotacoes")
@login_obrigatorio
def api_get_anotacoes():
    ids_param = request.args.get("ids", "").strip()
    conn = _init_anotacoes_db()
    try:
        if ids_param:
            ids_lista = [i.strip() for i in ids_param.split(",") if i.strip()]
            if not ids_lista:
                return jsonify({"anotacoes": {}})
            placeholders = ",".join("?" for _ in ids_lista)
            rows = conn.execute(f"SELECT id, anotacao FROM anotacoes WHERE id IN ({placeholders})", ids_lista).fetchall()
        else:
            rows = conn.execute("SELECT id, anotacao FROM anotacoes").fetchall()
        return jsonify({"anotacoes": {r["id"]: r["anotacao"] for r in rows}})
    finally:
        conn.close()


@app.route("/api/anotacoes", methods=["POST"])
@login_obrigatorio
def api_post_anotacoes():
    dados = request.get_json(silent=True) or {}
    id_ = str(dados.get("id", "")).strip()
    anotacao = str(dados.get("anotacao", "")).strip()
    if not id_:
        return jsonify({"erro": "ID obrigatório"}), 400

    usuario = autenticar_requisicao_atual()[0].get("usuario", "")
    agora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    conn = _init_anotacoes_db()
    try:
        existente = conn.execute("SELECT * FROM anotacoes WHERE id = ?", (id_,)).fetchone()
        if existente:
            conn.execute("UPDATE anotacoes SET anotacao = ?, usuario_criacao = ?, atualizado_em = ? WHERE id = ?",
                         (anotacao, usuario, agora, id_))
        else:
            conn.execute(
                """INSERT INTO anotacoes (id, nome, data_ref, data_anterior, data_retorno, tipo_violacao,
                   anotacao, usuario_criacao, criado_em, atualizado_em)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (id_,
                 str(dados.get("nome", "")),
                 str(dados.get("data_ref", "")),
                 str(dados.get("data_anterior", "")),
                 str(dados.get("data_retorno", "")),
                 str(dados.get("tipo_violacao", "")),
                 anotacao, usuario, agora, agora),
            )
        conn.commit()
        return jsonify({"ok": True})
    finally:
        conn.close()


@app.route("/api/consolidado")
@login_obrigatorio
def api_consolidado():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "dados": [], "kpis": {}})
    df = normalizar_colunas(df)
    # v8.10.57: a guia Espelho de Ponto não exibe mais Saldo Atual.
    # Portanto, removemos daqui a etapa de cruzamento com PDFs/cache,
    # que era pesada e não trazia ganho visual para a página.
    busca = request.args.get("busca", "").strip().lower()
    cc_filtro = request.args.get("cc", "").strip().lower()
    nome = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data = request.args.get("data", "").strip()
    datas = request.args.get("datas", "").strip()
    tipo = request.args.get("tipo", "").strip()

    if busca:
        texto = df["funcao"].str.lower() + " " + df["turno"].str.lower()
        df = df[texto.str.contains(busca, na=False)]

    if cc_filtro:
        df = aplicar_filtro_cc_backend(df, cc_filtro)

    if nome:
        df = df[df["nome"].str.lower() == nome]

    if turno_filtro:
        df = df[df["turno"].str.lower() == turno_filtro]

    # v8.11.24: respeita também o filtro múltiplo de datas no Consolidado,
    # Hora Extra e Inconsistências. Antes, a UI enviava ?datas=..., mas este
    # endpoint ignorava o parâmetro e voltava registros demais, prejudicando a
    # navegação e dando a sensação de filtro "travado".
    if datas:
        datas_lista = [d.strip() for d in datas.split(",") if d.strip()]
        if datas_lista:
            data_convertida = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce").dt.strftime("%Y-%m-%d")
            df = df[data_convertida.isin(datas_lista)]
    elif data:
        data_convertida = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce").dt.strftime("%Y-%m-%d")
        df = df[data_convertida == data]
    if tipo == "he":
        df = df[df["he"].astype(str).str.strip() != ""]
    elif tipo == "ausencia":
        df = df[df["absent"].astype(str).str.strip() != ""]
    elif tipo == "saldo_negativo":
        df = df[df["saldo"].astype(str).str.strip().str.startswith("-")]

    he_total_min = int(df["he"].apply(hhmm_para_minutos).sum()) if not df.empty else 0
    absent_total_min = int(df["absent"].apply(hhmm_para_minutos).sum()) if not df.empty else 0
    kpis = {
        "registros": int(len(df)),
        "colaboradores": int(df["nome"].nunique()) if not df.empty else 0,
        "he_total": minutos_para_hhmm(he_total_min),
        "absent_total": minutos_para_hhmm(absent_total_min),
        "saldo_negativo": int(df["saldo"].astype(str).str.strip().str.startswith("-").sum()) if not df.empty else 0,
        "revisar": int((df["status"].str.upper() == "REVISAR").sum()) if not df.empty else 0,
    }
    dados_resp, meta_resp = _df_para_resposta_leve(df)
    for r in dados_resp:
        nome = str(r.get("nome", "") or "")
        data = str(r.get("data", "") or "")
        r["_anotacao_id"] = _gerar_anotacao_id(nome, data)
    return jsonify({"erro": "", "dados": dados_resp, "kpis": kpis, **meta_resp})



def _texto_vazio_ponto(valor):
    texto = str(valor or "").strip()
    return texto == "" or texto.lower() in {"nan", "none", "nat"}


def _sem_batidas_ponto(row):
    cols = ["e1", "s1", "e2", "s2", "e3", "s3", "e4", "s4"]
    return all(_texto_vazio_ponto(row.get(c, "")) for c in cols)


def _obs_desconsidera_ponto_aberto(valor):
    texto = str(valor or "").strip().lower()
    termos = [
        "dsr",
        "d.s.r",
        "compensado",
        "afastamento temporário por doença",
        "afastamento temporario por doenca",
        "afastamento temporário por doenca",
        "afastamento temporario por doença",
        "não trabalhado",
        "nao trabalhado",
        "férias",
        "ferias",
    ]
    return any(t in texto for t in termos)


def _aplicar_filtros_ponto_aberto(df):
    busca = request.args.get("busca", "").strip().lower()
    cc_filtro = request.args.get("cc", "").strip()
    nome = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data = request.args.get("data", "").strip()
    datas = request.args.get("datas", "").strip()

    if busca:
        texto = df["funcao"].str.lower() + " " + df["turno"].str.lower()
        df = df[texto.str.contains(busca, na=False)]

    if cc_filtro:
        df = aplicar_filtro_cc_backend(df, cc_filtro)

    if nome:
        df = df[df["nome"].str.lower() == nome]

    if turno_filtro:
        df = df[df["turno"].str.lower() == turno_filtro]

    data_convertida = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce").dt.strftime("%Y-%m-%d")
    if datas:
        lista = [d.strip() for d in datas.split(",") if d.strip()]
        if lista:
            df = df[data_convertida.isin(lista)]
    elif data:
        df = df[data_convertida == data]

    return df


def _kpis_ponto_aberto(df):
    return {
        "registros": int(len(df)),
        "colaboradores": int(df["nome"].nunique()) if not df.empty else 0,
        "he_total": "00:00",
        "absent_total": "00:00",
        "saldo_negativo": 0,
        "revisar": int(len(df)),
    }



def _qtd_batidas_oficial(row):
    """
    Usa a coluna oficial 'Qtd Batidas' gerada pelo robô.
    Se por algum motivo ela vier vazia/indisponível, faz fallback contando
    as colunas de marcação.
    """
    valor = row.get("qtd_batidas", "")

    try:
        if str(valor).strip() != "":
            return int(float(str(valor).replace(",", ".").strip()))
    except Exception:
        pass

    cols = ["e1", "s1", "e2", "s2", "e3", "s3", "e4", "s4"]
    total = 0

    for col in cols:
        v = str(row.get(col, "") or "").strip()
        if v and v.lower() not in {"nan", "none", "nat"}:
            total += 1

    return total



def _qtd_batidas_direta(valor):
    """
    Converte a coluna normalizada qtd_batidas para inteiro.
    Essa é a fonte oficial do Tempo Fechado.
    """
    try:
        texto = str(valor or "").strip().replace(",", ".")
        if texto == "" or texto.lower() in {"nan", "none", "nat"}:
            return 0
        return int(float(texto))
    except Exception:
        return 0


@app.route("/api/ponto-em-aberto")
@login_obrigatorio
def api_ponto_em_aberto():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "dados": [], "kpis": {}})

    df = normalizar_colunas(df)
    if df.empty:
        return jsonify({"erro": "", "dados": [], "kpis": _kpis_ponto_aberto(df)})

    df = _aplicar_filtros_ponto_aberto(df)

    df = df.copy()
    df["_qtd_batidas_oficial"] = df.apply(_qtd_batidas_oficial, axis=1)

    mask_sem_batidas = df["_qtd_batidas_oficial"] == 0
    mask_obs_ok = ~df["observacao"].apply(_obs_desconsidera_ponto_aberto)

    df = df[mask_sem_batidas & mask_obs_ok].copy()
    df.drop(columns=["_qtd_batidas_oficial"], inplace=True, errors="ignore")

    dados_resp, meta_resp = _df_para_resposta_leve(df)
    for r in dados_resp:
        nome = str(r.get("nome", "") or "")
        data = str(r.get("data", "") or "")
        r["_anotacao_id"] = _gerar_anotacao_id(nome, data)
    return jsonify({
        "erro": "",
        "dados": dados_resp,
        "kpis": _kpis_ponto_aberto(df),
        **meta_resp,
    })


@app.route("/api/ponto-aberto")
@login_obrigatorio
def api_ponto_aberto():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "dados": [], "kpis": {}})

    df = normalizar_colunas(df)
    if df.empty:
        return jsonify({"erro": "", "dados": [], "kpis": _kpis_ponto_aberto(df)})

    df = _aplicar_filtros_ponto_aberto(df).copy()

    # Fonte oficial: coluna normalizada qtd_batidas.
    df["_qtd_batidas_web"] = df["qtd_batidas"].apply(_qtd_batidas_direta)

    # Regra final: somente ímpares maiores que zero.
    df = df[
        (df["_qtd_batidas_web"] > 0) &
        (df["_qtd_batidas_web"] % 2 == 1)
    ].copy()

    df.drop(columns=["_qtd_batidas_web"], inplace=True, errors="ignore")

    dados_resp, meta_resp = _df_para_resposta_leve(df)
    for r in dados_resp:
        nome = str(r.get("nome", "") or "")
        data = str(r.get("data", "") or "")
        r["_anotacao_id"] = _gerar_anotacao_id(nome, data)
    return jsonify({
        "erro": "",
        "dados": dados_resp,
        "kpis": {
            "registros": int(len(df)),
            "colaboradores": int(df["nome"].nunique()) if not df.empty else 0,
            "he_total": "00:00",
            "absent_total": "00:00",
            "saldo_negativo": 0,
            "revisar": int(len(df)),
        },
        **meta_resp,
    })


@app.route("/api/debug-ponto-aberto")
@login_obrigatorio
def api_debug_ponto_aberto():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro})

    df = normalizar_colunas(df)
    df = _aplicar_filtros_ponto_aberto(df).copy()
    df["_qtd_batidas_web"] = df["qtd_batidas"].apply(_qtd_batidas_direta)

    filtrado = df[
        (df["_qtd_batidas_web"] > 0) &
        (df["_qtd_batidas_web"] % 2 == 1)
    ].copy()

    resumo_filtrado = (
        filtrado.groupby("_qtd_batidas_web", dropna=False)
        .size()
        .reset_index(name="quantidade")
        .sort_values("_qtd_batidas_web")
    )

    return jsonify({
        "erro": "",
        "total_filtrado": int(len(filtrado)),
        "resumo_filtrado": [
            {
                "qtd_batidas": int(row["_qtd_batidas_web"]),
                "quantidade": int(row["quantidade"]),
            }
            for _, row in resumo_filtrado.iterrows()
        ],
        "amostra": filtrado.head(10).drop(columns=["_qtd_batidas_web"], errors="ignore").to_dict(orient="records"),
    })




def saldo_dashboard_para_minutos(valor):
    """
    Converte saldo do Banco de Horas para minutos.
    Aceita HH,MM e HH:MM.
    Ex.: 210,26 = 210h26min | -45,30 = -45h30min
    """
    texto = str(valor or "").strip()
    if not texto or texto.lower() in {"nan", "none"}:
        return 0

    negativo = texto.startswith("-")
    texto = texto.replace("+", "").replace("-", "").strip()
    texto = texto.replace(",", ":")

    partes = texto.split(":")
    if len(partes) < 2:
        return 0

    try:
        horas = int(partes[0] or 0)
        minutos = int(partes[1] or 0)
    except Exception:
        return 0

    total = horas * 60 + minutos
    return -total if negativo else total








# ============================================================
# BANCO DE HORAS 2.0 - v8.14.00
# ============================================================
def _norm_bh2(valor):
    texto = str(valor or "").strip().upper()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def _canon_bh2(valor):
    texto = str(valor or "").strip().lower()
    mapa = str.maketrans("áàãâéêíóôõúç", "aaaaeeiooouc")
    texto = texto.translate(mapa)
    texto = texto.replace("_", " ").replace("-", " ")
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def _ler_aba_excel_completa(nome_aba):
    """Le uma aba inteira do Excel ativo, sem filtrar colunas.

    Usada para a aba 'Saldos Disponiveis BH' do Script_Robo_Ponto_v7_PRO,
    onde o objetivo é preservar as opções de saldo para escolha do gestor.
    """
    if not ARQUIVO_PADRAO.exists():
        return pd.DataFrame(), "Excel oficial não encontrado. Importe o arquivo gerado pelo robô."
    nome_aba = str(nome_aba or "").strip()
    if not nome_aba:
        return pd.DataFrame(), "Nome de aba não informado."
    cache_key = (_chave_cache_excel(ARQUIVO_PADRAO), "aba_completa", nome_aba)
    try:
        with _CACHE_EXCEL_LOCK:
            df_cache = _CACHE_EXCEL_ABAS_MEM.get(cache_key)
            if df_cache is not None:
                return df_cache.copy(deep=False), ""
    except Exception:
        pass

    try:
        if load_workbook is None:
            df = pd.read_excel(ARQUIVO_PADRAO, sheet_name=nome_aba, dtype=str).fillna("")
            try:
                with _CACHE_EXCEL_LOCK:
                    _CACHE_EXCEL_ABAS_MEM[cache_key] = df.copy(deep=False)
            except Exception:
                pass
            return df.copy(deep=False), ""
        wb = load_workbook(filename=str(ARQUIVO_PADRAO), read_only=True, data_only=True)
        try:
            if nome_aba not in wb.sheetnames:
                return pd.DataFrame(), f"Aba '{nome_aba}' não encontrada no Excel importado."
            ws = wb[nome_aba]
            rows = ws.iter_rows(values_only=True)
            try:
                header = next(rows)
            except StopIteration:
                return pd.DataFrame(), "Aba vazia."
            header = ["" if h is None else str(h).strip() for h in header]
            dados = []
            for row in rows:
                linha = []
                tem_valor = False
                for val in row[:len(header)]:
                    if val is None:
                        val = ""
                    elif hasattr(val, "strftime"):
                        try:
                            val = val.strftime("%d/%m/%Y")
                        except Exception:
                            val = str(val)
                    else:
                        val = str(val)
                    if str(val).strip():
                        tem_valor = True
                    linha.append(val)
                while len(linha) < len(header):
                    linha.append("")
                if tem_valor:
                    dados.append(linha)
            df = pd.DataFrame(dados, columns=header).fillna("")
            try:
                with _CACHE_EXCEL_LOCK:
                    _CACHE_EXCEL_ABAS_MEM[cache_key] = df.copy(deep=False)
            except Exception:
                pass
            return df.copy(deep=False), ""
        finally:
            try:
                wb.close()
            except Exception:
                pass
    except Exception as e:
        return pd.DataFrame(), f"Falha ao ler aba '{nome_aba}': {e}"


def _nome_aba_saldos_disponiveis_bh(caminho_excel=None):
    caminho_excel = Path(caminho_excel or ARQUIVO_PADRAO)
    cache_id = _chave_cache_excel(caminho_excel)
    try:
        with _CACHE_EXCEL_LOCK:
            if _CACHE_EXCEL_META_MEM.get("sig") == cache_id and _CACHE_EXCEL_META_MEM.get("aba_saldos_bh") is not None:
                return _CACHE_EXCEL_META_MEM.get("aba_saldos_bh") or ""
    except Exception:
        pass

    abas = _listar_abas_excel(caminho_excel)
    candidatos = {
        "saldos disponiveis bh", "saldos disponíveis bh", "saldos disponiveis banco horas",
        "saldos disponíveis banco horas", "saldos bh", "opcoes saldo bh", "opções saldo bh",
    }
    candidatos_norm = {_canon_bh2(c) for c in candidatos}
    resultado = ""
    for aba in abas:
        if _canon_bh2(aba) in candidatos_norm:
            resultado = aba
            break
    try:
        with _CACHE_EXCEL_LOCK:
            if _CACHE_EXCEL_META_MEM.get("sig") == cache_id:
                _CACHE_EXCEL_META_MEM["aba_saldos_bh"] = resultado
    except Exception:
        pass
    return resultado


def _coluna_por_alias_bh2(df, aliases):
    if df is None or df.empty:
        return ""
    aliases_norm = {_canon_bh2(a) for a in aliases}
    for col in df.columns:
        if _canon_bh2(col) in aliases_norm:
            return col
    for col in df.columns:
        c = _canon_bh2(col)
        if any(a in c for a in aliases_norm):
            return col
    return ""


def _normalizar_saldo_texto_bh2(valor):
    texto = str(valor or "").strip()
    if not texto or texto.lower() in {"nan", "none", "nat"}:
        return ""
    texto = texto.replace(".", "").replace(":", ",")
    m = re.match(r"^(-?)(\d{1,5}),?(\d{2})$", texto)
    if m and "," not in texto:
        return f"{m.group(1)}{int(m.group(2)):02d},{m.group(3)}"
    if re.match(r"^-?\d{1,5},\d{2}$", texto):
        sinal = "-" if texto.startswith("-") else ""
        base = texto[1:] if sinal else texto
        h, mm = base.split(",", 1)
        return f"{sinal}{int(h):02d},{mm[:2]}"
    return texto


def _carregar_config_bh2():
    try:
        if not ARQUIVO_CONFIG_BH2.exists():
            return {"ativo": True, "data_inicio": "", "saldos": {}, "origem_padrao": "saldo_atual", "saldo_manual": "", "atualizado_em": "", "versao": "8.15.10"}
        with open(ARQUIVO_CONFIG_BH2, "r", encoding="utf-8") as f:
            cfg = json.load(f) or {}
        cfg["ativo"] = True  # v8.15.00: BH 2.0 e o motor padrao; sem checkbox na interface
        cfg.setdefault("data_inicio", "")
        cfg.setdefault("saldos", {})
        cfg.setdefault("origem_padrao", "saldo_atual")
        cfg.setdefault("saldo_manual", "")
        if not isinstance(cfg.get("saldos", {}), dict):
            cfg["saldos"] = {}
        return cfg
    except Exception:
        return {"ativo": True, "data_inicio": "", "saldos": {}, "origem_padrao": "saldo_atual", "saldo_manual": "", "atualizado_em": "", "versao": "8.15.10"}


def _invalidar_cache_bh2_calculo():
    """Remove caches do motor BH 2.0.

    v8.15.23: o cálculo pesado do Banco de Horas passa a ser cacheado e só
    deve ser refeito quando mudar o Excel, a Data Inicial ou algum saldo
    manual por colaborador.
    """
    try:
        _CACHE_BH2_CALCULO_MEM["key"] = None
        _CACHE_BH2_CALCULO_MEM["payload"] = None
        _CACHE_BH2_CALCULO_MEM["gerado_em_ts"] = 0
    except Exception:
        pass
    for arq in (ARQUIVO_CACHE_BH2_CALCULO, ARQUIVO_RESUMO_BANCO_HORAS_PDF):
        try:
            if arq.exists():
                arq.unlink()
        except Exception:
            pass


def _salvar_config_bh2(cfg):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    cfg = dict(cfg or {})
    cfg["versao"] = "8.15.23"
    cfg["atualizado_em"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(ARQUIVO_CONFIG_BH2, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    _invalidar_cache_bh2_calculo()
    return cfg


def _saldo_config_bh2_para_colaborador(nome, matricula=""):
    cfg = _carregar_config_bh2()
    # v8.15.00: nao existe mais chave operacional de ativacao; data_inicio controla o motor.
    saldos = cfg.get("saldos", {}) or {}
    nome_norm = _norm_bh2(nome)
    mat = str(matricula or "").strip()
    for chave, item in saldos.items():
        if not isinstance(item, dict):
            continue
        item_nome = _norm_bh2(item.get("nome", chave))
        item_mat = str(item.get("matricula", "") or "").strip()
        if mat and item_mat and mat == item_mat:
            return item
        if nome_norm and item_nome and (nome_norm == item_nome or nome_norm.startswith(item_nome) or item_nome.startswith(nome_norm)):
            return item
    return None


def _opcoes_saldos_disponiveis_bh():
    aba = _nome_aba_saldos_disponiveis_bh()
    if not aba:
        return [], "Aba 'Saldos Disponiveis BH' não encontrada. Gere o Excel pelo Script_Robo_Ponto_v7_PRO."
    df, erro = _ler_aba_excel_completa(aba)
    if erro:
        return [], erro
    if df.empty:
        return [], "Aba 'Saldos Disponiveis BH' está vazia."

    col_nome = _coluna_por_alias_bh2(df, ["Nome", "Colaborador", "Funcionario", "Funcionário"])
    col_mat = _coluna_por_alias_bh2(df, ["Matrícula", "Matricula", "Chapa"])
    col_cc = _coluna_por_alias_bh2(df, ["CC", "C.C.", "Centro de Custo", "cc"])
    col_data = _coluna_por_alias_bh2(df, ["Data", "Data Base", "Data Inicio", "Data Início", "Data Referencia", "Data Referência"])
    col_emissao = _coluna_por_alias_bh2(df, ["Data Emissão PDF", "Data Emissao PDF", "Emissão", "Emissao"])
    col_origem = _coluna_por_alias_bh2(df, ["Origem", "Campo", "Tipo", "Fonte", "Origem Saldo"])
    col_saldo = _coluna_por_alias_bh2(df, ["Saldo", "Saldo Inicial", "Valor", "Saldo Disponível", "Saldo Disponivel", "Saldo Atual", "Saldo Informado PDF", "Saldo Calculado", "Saldo Anterior PDF"])
    col_recomendado = _coluna_por_alias_bh2(df, ["Recomendado", "Recomendacao", "Recomendação", "Preferencial", "Principal"])

    # Se não houver coluna explícita de saldo, procura a primeira coluna com "saldo".
    if not col_saldo:
        for col in df.columns:
            if "saldo" in _canon_bh2(col):
                col_saldo = col
                break

    opcoes = []
    for _, row in df.iterrows():
        nome = str(row.get(col_nome, "") if col_nome else "").strip()
        saldo = _normalizar_saldo_texto_bh2(row.get(col_saldo, "") if col_saldo else "")
        if not nome or not saldo:
            continue
        data_ref = str(row.get(col_data, "") if col_data else "").strip()
        data_emissao = str(row.get(col_emissao, "") if col_emissao else "").strip()
        origem = str(row.get(col_origem, "") if col_origem else col_saldo).strip() or col_saldo
        recomendado_txt = str(row.get(col_recomendado, "") if col_recomendado else "").strip().lower()
        recomendado = recomendado_txt in {"1", "sim", "s", "true", "verdadeiro", "yes", "y", "recomendado"}
        opcoes.append({
            "nome": nome,
            "matricula": str(row.get(col_mat, "") if col_mat else "").strip(),
            "cc": str(row.get(col_cc, "") if col_cc else "").strip(),
            "data": data_ref,
            "data_emissao_pdf": data_emissao,
            "origem": origem,
            "saldo": saldo,
            "recomendado": recomendado,
            "label": f"{saldo} | {origem}" + (" | recomendado" if recomendado else "") + (f" | emissão {data_emissao}" if data_emissao else ""),
        })
    return opcoes, ""


def _calcular_extrato_bh2_colaborador(g_ciclo):
    """Motor Banco de Horas 2.0.

    Regras v8.15.22:
    - Data Inicial e parametro global de calculo, nao filtro de tela.
    - Se Data Inicial estiver vazia, o motor considera todo o periodo do colaborador
      usando a primeira data disponivel como marco zero.
    - Filtros de data nas guias apenas mudam a visualizacao/posicao do saldo.
    - Origem/saldo manual continuam sendo regra individual por colaborador.
    """
    cfg = _carregar_config_bh2()
    if g_ciclo is None or g_ciclo.empty:
        return None

    g = g_ciclo.copy()
    g["data_dt"] = pd.to_datetime(g.get("data", ""), format="%d/%m/%Y", errors="coerce")
    g = g[g["data_dt"].notna()].sort_values(["data_dt", "data"], kind="mergesort").copy()
    if g.empty:
        return None

    data_inicio_txt = str(cfg.get("data_inicio", "") or "").strip()
    if data_inicio_txt:
        data_inicio = pd.to_datetime(data_inicio_txt, errors="coerce")
        if pd.isna(data_inicio):
            data_inicio = pd.to_datetime(data_inicio_txt, errors="coerce", dayfirst=True)
        if pd.isna(data_inicio):
            return None
        data_inicio = pd.Timestamp(data_inicio).normalize()
        g = g[g["data_dt"].dt.normalize() >= data_inicio].copy()
        origem_periodo = "data_inicial_gestor"
    else:
        data_inicio = pd.Timestamp(g["data_dt"].min()).normalize()
        origem_periodo = "todo_periodo_sem_data_inicial"

    if g.empty:
        return None

    nome = str(g.get("nome", pd.Series([""])).iloc[0] if "nome" in g.columns else "")
    matricula = str(g.get("matricula", pd.Series([""])).iloc[0] if "matricula" in g.columns else "")
    origem_padrao = str(cfg.get("origem_padrao", "saldo_atual") or "saldo_atual").strip().lower()
    saldo_cfg_individual = _saldo_config_bh2_para_colaborador(nome, matricula)
    if saldo_cfg_individual and isinstance(saldo_cfg_individual, dict):
        origem_padrao = str(saldo_cfg_individual.get("origem", saldo_cfg_individual.get("origem_padrao", origem_padrao)) or origem_padrao).strip().lower()
        if origem_padrao not in {"saldo_atual", "manual"}:
            origem_padrao = "saldo_atual"
        saldo_manual = _normalizar_saldo_texto_bh2(saldo_cfg_individual.get("saldo", saldo_cfg_individual.get("saldo_manual", "")))
    else:
        saldo_manual = ""

    origem_saldo_base = ""
    saldo_base_txt = ""
    if origem_padrao == "manual" and saldo_manual:
        saldo_base_txt = saldo_manual
        origem_saldo_base = "valor_manual_gestor_colaborador"
    else:
        candidatos_exatos = g[g["data_dt"].dt.normalize() == data_inicio.normalize()].copy()
        if candidatos_exatos.empty:
            candidatos_exatos = g.head(1).copy()
            origem_saldo_base = "saldo_atual_primeira_linha_posterior_ou_igual_data_inicial"
        else:
            origem_saldo_base = "saldo_atual_data_inicial_gestor"

        row_base = candidatos_exatos.iloc[-1] if not candidatos_exatos.empty else None
        if row_base is not None:
            for col in ("saldo_atual", "saldo", "saldo_informado_pdf", "saldo_calculado", "saldo_anterior_pdf"):
                if col in row_base.index:
                    saldo_base_txt = _normalizar_saldo_texto_bh2(row_base.get(col, ""))
                    if saldo_base_txt:
                        origem_saldo_base += f": {col}"
                        break

    if not saldo_base_txt:
        saldo_base_txt = "00,00"
        origem_saldo_base = (origem_saldo_base or origem_padrao or "saldo_atual") + "_nao_encontrado_usando_zero"

    saldo_corrente_min = _saldo_virgula_para_minutos(saldo_base_txt)
    credito_total_min = 0
    debito_total_min = 0
    linhas = []
    primeira_fotografia_por_data = _primeira_fotografia_por_dia_extrato_bh(g)

    datas_ordenadas = []
    for d in g["data"].astype(str).str.strip().tolist():
        if d and d not in datas_ordenadas:
            datas_ordenadas.append(d)

    for idx, data_txt in enumerate(datas_ordenadas):
        linhas_data = g[g["data"].astype(str).str.strip() == data_txt].copy()
        if linhas_data.empty:
            continue
        row = linhas_data.iloc[-1]
        row_primeira = _primeira_linha_do_dia_por_emissao(linhas_data)
        info_primeira = primeira_fotografia_por_data.get(data_txt, {}) if isinstance(primeira_fotografia_por_data, dict) else {}
        emissao_primeira = str(info_primeira.get("data_emissao_pdf", "") or "").strip()
        if not emissao_primeira and row_primeira is not None:
            emissao_primeira = str(row_primeira.get("data_emissao_pdf", "") or "").strip()
        saldo_info_primeira = str(info_primeira.get("saldo_informado", "") or "").strip()
        if not saldo_info_primeira and row_primeira is not None:
            for _col_info in ("saldo_informado_pdf", "saldo_informado", "saldo_atual", "saldo"):
                if _col_info in row_primeira.index:
                    _v_info = str(row_primeira.get(_col_info, "") or "").strip()
                    if _v_info and _v_info.lower() not in {"nan", "none", "nat"}:
                        saldo_info_primeira = _v_info
                        break
        he_min = int(sum(hhmm_para_minutos(v) for v in linhas_data.get("he", pd.Series([], dtype="object")).tolist()))
        absent_min = int(sum(hhmm_para_minutos(v) for v in linhas_data.get("absent", pd.Series([], dtype="object")).tolist()))
        if idx > 0:
            saldo_corrente_min += he_min
            saldo_corrente_min -= absent_min
            credito_total_min += he_min
            debito_total_min += absent_min
        linhas.append({
            "data": data_txt,
            "data_emissao_pdf": emissao_primeira or row.get("data_emissao_pdf", ""),
            "dia": row.get("dia", ""),
            "evento": "Saldo Base" if idx == 0 else "Movimento",
            "he": minutos_para_hhmm(he_min) if he_min else "",
            "absent": minutos_para_hhmm(absent_min) if absent_min else "",
            "saldo_calculado": _minutos_para_saldo_virgula(saldo_corrente_min),
            "saldo_informado": saldo_info_primeira or row.get("saldo_informado_pdf", "") or row.get("saldo_atual", "") or row.get("saldo", ""),
            "saldo_pdf": saldo_info_primeira or row.get("saldo_informado_pdf", "") or row.get("saldo_atual", "") or row.get("saldo", ""),
            "observacao": str(row.get("observacao", "") or ""),
            "status": str(row.get("status", "") or ""),
            "inconsistencias": str(row.get("inconsistencias", "") or ""),
        })

    resumo = {
        "nome": nome,
        "matricula": matricula,
        "cc": str(g.get("cc", pd.Series([""])).iloc[-1] if "cc" in g.columns else ""),
        "funcao": str(g.get("funcao", pd.Series([""])).iloc[-1] if "funcao" in g.columns else ""),
        "turno": str(g.get("turno", pd.Series([""])).iloc[-1] if "turno" in g.columns else ""),
        "data_base": data_inicio.strftime("%d/%m/%Y"),
        "inicio_ciclo": data_inicio.strftime("%d/%m/%Y"),
        "fim_ciclo": pd.to_datetime(g["data_dt"].max()).strftime("%d/%m/%Y"),
        "saldo_base": saldo_base_txt,
        "credito_total": minutos_para_hhmm(int(credito_total_min)),
        "debito_total": minutos_para_hhmm(int(debito_total_min)),
        "saldo_final": _minutos_para_saldo_virgula(saldo_corrente_min),
        "saldo_informado_final": linhas[-1].get("saldo_informado", "") if linhas else "",
        "saldo_pdf_final": linhas[-1].get("saldo_informado", "") if linhas else "",
        "origem_saldo_base": origem_saldo_base,
        "fonte_calculo": "banco_horas_2_0_simplificado_gestor",
        "origem_periodo": origem_periodo,
        "justificativa": cfg.get("justificativa", ""),
    }
    return {"resumo": resumo, "linhas": linhas}

def _normalizar_nome_saldo_base(nome):
    texto = str(nome or "").strip().upper()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def _carregar_saldos_base_medicao():
    """
    Carrega saldos oficiais de início de medição.

    Arquivo:
      data/saldos_base_medicao.json

    Formato:
    {
      "2026-05-15": {
        "NOME DO COLABORADOR": "263,48"
      }
    }

    Esses valores prevalecem sobre saldos importados do PDF/Excel.
    """
    try:
        if not ARQUIVO_SALDOS_BASE_MEDICAO.exists():
            return {}
        with open(ARQUIVO_SALDOS_BASE_MEDICAO, "r", encoding="utf-8") as f:
            return json.load(f) or {}
    except Exception:
        return {}


def _obter_saldo_base_medicao_oficial(nome, data_base):
    try:
        data_key = pd.to_datetime(data_base).strftime("%Y-%m-%d")
    except Exception:
        data_key = str(data_base or "").strip()

    nome_norm = _normalizar_nome_saldo_base(nome)
    dados = _carregar_saldos_base_medicao()

    mapa = dados.get(data_key, {})
    if not isinstance(mapa, dict):
        return None

    # Match exato
    for nome_json, saldo in mapa.items():
        if _normalizar_nome_saldo_base(nome_json) == nome_norm:
            return str(saldo).strip()

    # Match tolerante para nomes truncados no PDF/Excel
    for nome_json, saldo in mapa.items():
        nj = _normalizar_nome_saldo_base(nome_json)
        if nome_norm.startswith(nj) or nj.startswith(nome_norm):
            return str(saldo).strip()

    return None


def _data_inicio_medicao(data_dt):
    """
    Período de medição:
    - dia 15 do mês corrente até dia 14 do mês seguinte.

    Se a data estiver entre 01 e 14, o início da medição é dia 15 do mês anterior.
    Se a data estiver entre 15 e fim do mês, o início da medição é dia 15 do próprio mês.
    """
    try:
        if pd.isna(data_dt):
            return None

        data_dt = pd.to_datetime(data_dt)

        if data_dt.day >= 15:
            return pd.Timestamp(year=data_dt.year, month=data_dt.month, day=15)

        mes_anterior = data_dt - pd.DateOffset(months=1)
        return pd.Timestamp(year=mes_anterior.year, month=mes_anterior.month, day=15)
    except Exception:
        return None




def _data_fim_medicao(inicio_dt):
    """
    Retorna o fim oficial do ciclo de apuração do Banco de Horas.

    Regra do ciclo:
    - início: dia 15;
    - fim: dia 14 do mês seguinte.

    Evita usar +30 dias, porque meses com 28, 29, 30 ou 31 dias
    podem incluir indevidamente o dia 15 do ciclo seguinte.
    """
    try:
        if inicio_dt is None or pd.isna(inicio_dt):
            return None
        inicio_dt = pd.to_datetime(inicio_dt)
        proximo_mes = inicio_dt + pd.DateOffset(months=1)
        return pd.Timestamp(year=proximo_mes.year, month=proximo_mes.month, day=14)
    except Exception:
        return None


def _saldo_para_minutos_banco(valor):
    """
    Converte saldo no padrão do espelho/robô para minutos.
    Aceita HH,MM e HH:MM.
    Ex.:
    255,00 = 255h00
    -15,52 = -15h52
    """
    texto = str(valor or "").strip()
    if not texto or texto.lower() in {"nan", "none", "nat"}:
        return 0

    negativo = texto.startswith("-")
    texto = texto.replace("+", "").replace("-", "").strip()
    texto = texto.replace(",", ":")

    partes = texto.split(":")
    if len(partes) < 2:
        return 0

    try:
        horas = int(partes[0] or 0)
        minutos = int(partes[1][:2] or 0)
    except Exception:
        return 0

    total = horas * 60 + minutos
    return -total if negativo else total


def _minutos_para_saldo_virgula(total_minutos):
    try:
        total_minutos = int(total_minutos)
    except Exception:
        total_minutos = 0

    sinal = "-" if total_minutos < 0 else ""
    total_minutos = abs(total_minutos)

    return f"{sinal}{total_minutos // 60:02d},{total_minutos % 60:02d}"


def _saldo_virgula_para_minutos(valor):
    return _saldo_para_minutos_banco(valor)


def _carregar_resumo_banco_horas_pdf():
    if not ARQUIVO_RESUMO_BANCO_HORAS_PDF.exists():
        return None
    try:
        with ARQUIVO_RESUMO_BANCO_HORAS_PDF.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _salvar_resumo_banco_horas_pdf(resumo):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    try:
        with ARQUIVO_RESUMO_BANCO_HORAS_PDF.open("w", encoding="utf-8") as f:
            json.dump(resumo, f, ensure_ascii=False, indent=2, default=str)
        return True
    except Exception:
        return False


def _resumo_banco_horas_parece_valido(resumo):
    # v8.10.44: para a guia Banco de Horas carregar rápido, um resumo com
    # lista vazia também é considerado estruturalmente válido. Isso evita que
    # o endpoint caia no cálculo pesado em tempo real quando o cache existe,
    # mas ainda não tem linhas ou foi gerado parcialmente.
    return bool(resumo and isinstance(resumo.get("linhas"), list))


def _filtrar_linhas_resumo_banco_horas(linhas, busca="", cc="", nome="", turno=""):
    busca = str(busca or "").strip().lower()
    cc_norm = normalizar_cc_filtro_backend(cc)
    nome = str(nome or "").strip().lower()
    turno = str(turno or "").strip().lower()
    filtradas = []
    for item in linhas or []:
        cc_item_norm = normalizar_cc_filtro_backend(item.get("cc", ""))
        nome_item = str(item.get("nome", "") or "")
        turno_item = str(item.get("turno", "") or "")
        if cc_norm and cc_item_norm != cc_norm:
            continue
        if nome and nome not in nome_item.lower():
            continue
        if turno and turno_item.lower() != turno:
            continue
        if busca:
            texto = f"{item.get('cc','')} {item.get('funcao','')} {item.get('turno','')} {item.get('nome','')}".lower()
            if busca not in texto:
                continue
        filtradas.append(item)
    return filtradas


def _montar_kpis_banco_horas_de_linhas(linhas, registros_base=0, revisar=0):
    he_total_min = sum(hhmm_para_minutos(x.get("he_total", "")) for x in linhas or [])
    absent_total_min = sum(hhmm_para_minutos(x.get("absent_total", "")) for x in linhas or [])
    saldo_negativo = 0
    for item in linhas or []:
        if _saldo_virgula_para_minutos(item.get("saldo_atual", "")) < 0:
            saldo_negativo += 1
    return {
        "registros": int(registros_base or 0),
        "colaboradores": int(len(linhas or [])),
        "he_total": minutos_para_hhmm(int(he_total_min)),
        "absent_total": minutos_para_hhmm(int(absent_total_min)),
        "saldo_negativo": int(saldo_negativo),
        "revisar": int(revisar or 0),
    }




def _valor_hora_tem_movimento_bh(valor):
    """Retorna True quando um campo de H.E./ausência representa movimento real.

    Usado para calcular Dias com HE e Dias com ausência sem transformar dado
    ausente em zero visual. Aceita HH:MM, HH,MM e campos vazios.
    """
    try:
        return abs(int(hhmm_para_minutos(valor))) > 0
    except Exception:
        return False


def _contar_dias_movimento_bh(linhas, campo):
    return int(sum(1 for item in (linhas or []) if _valor_hora_tem_movimento_bh((item or {}).get(campo, ""))))


def _contar_dias_movimento_df_bh(df, col_min, col_texto):
    if df is None or df.empty:
        return 0
    try:
        if col_min in df.columns:
            vals = pd.to_numeric(df[col_min], errors="coerce").fillna(0)
            return int((vals.abs() > 0).sum())
    except Exception:
        pass
    try:
        if col_texto in df.columns:
            return int(df[col_texto].apply(_valor_hora_tem_movimento_bh).sum())
    except Exception:
        pass
    return 0


def _enriquecer_linha_dias_bh(linha, linhas_extrato=None):
    """Completa dias_com_he/dias_com_ausencia com métrica real quando possível."""
    base = dict(linha or {})
    if linhas_extrato is not None:
        base["dias_com_he"] = _contar_dias_movimento_bh(linhas_extrato, "he")
        base["dias_com_ausencia"] = _contar_dias_movimento_bh(linhas_extrato, "absent")
        base["metricas_dias_fonte"] = "extrato_bh2"
    else:
        base.setdefault("dias_com_he", None)
        base.setdefault("dias_com_ausencia", None)
        base.setdefault("metricas_dias_fonte", "nao_calculado")
    return base


def _assinatura_config_bh2(cfg=None):
    cfg = dict(cfg or _carregar_config_bh2() or {})
    # Campos transitórios não entram na assinatura, senão o cache nunca acerta.
    cfg.pop("atualizado_em", None)
    cfg.pop("versao", None)
    try:
        return json.dumps(cfg, ensure_ascii=False, sort_keys=True, default=str)
    except Exception:
        return str(cfg)


def _chave_bh2_colaborador(nome="", matricula=""):
    mat = str(matricula or "").strip()
    if mat:
        return f"MAT::{mat}"
    return f"NOME::{_norm_bh2(nome)}"


def _carregar_cache_bh2_calculo_disco():
    try:
        if not ARQUIVO_CACHE_BH2_CALCULO.exists():
            return None
        with open(ARQUIVO_CACHE_BH2_CALCULO, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _salvar_cache_bh2_calculo_disco(payload):
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(ARQUIVO_CACHE_BH2_CALCULO, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
    except Exception as e:
        LOGGER.warning(f"Não foi possível salvar cache BH 2.0: {e}")


def _cache_bh2_valido(payload, excel_sig, config_sig):
    if not isinstance(payload, dict):
        return False
    linhas = payload.get("linhas", [])
    metricas_ok = isinstance(linhas, list) and all(
        isinstance(item, dict) and "dias_com_he" in item and "dias_com_ausencia" in item
        for item in linhas
    )
    return (
        str(payload.get("assinatura_excel", "")) == str(excel_sig)
        and str(payload.get("assinatura_config", "")) == str(config_sig)
        and isinstance(payload.get("extratos", {}), dict)
        and isinstance(linhas, list)
        and metricas_ok
    )


def _obter_cache_bh2_calculo(forcar=False):
    """Retorna cálculo BH 2.0 cacheado para navegação rápida.

    O cálculo completo por colaborador é a parte cara da navegação. A partir da
    v8.15.23, Banco de Horas e Extrato leem este cache e apenas filtram a visão.
    O cache é invalidado quando o Excel muda ou quando a configuração BH 2.0 é
    salva.
    """
    cfg = _carregar_config_bh2()
    excel_sig = str(_assinatura_arquivo(ARQUIVO_PADRAO))
    config_sig = _assinatura_config_bh2(cfg)
    cache_key = (excel_sig, config_sig)

    if not forcar:
        try:
            if _CACHE_BH2_CALCULO_MEM.get("key") == cache_key and _CACHE_BH2_CALCULO_MEM.get("payload"):
                return _CACHE_BH2_CALCULO_MEM["payload"]
        except Exception:
            pass

        payload = _carregar_cache_bh2_calculo_disco()
        if _cache_bh2_valido(payload, excel_sig, config_sig):
            _CACHE_BH2_CALCULO_MEM["key"] = cache_key
            _CACHE_BH2_CALCULO_MEM["payload"] = payload
            _CACHE_BH2_CALCULO_MEM["gerado_em_ts"] = time.monotonic()
            return payload

    df, erro = ler_consolidado()
    if erro:
        return {"ok": False, "erro": erro, "linhas": [], "extratos": {}, "kpis": _montar_kpis_banco_horas_de_linhas([])}
    df = normalizar_colunas(df)
    if df is None or df.empty:
        return {"ok": True, "erro": "", "linhas": [], "extratos": {}, "kpis": _montar_kpis_banco_horas_de_linhas([])}
    df = _garantir_colunas_calculadas(df)

    # v8.16.10 - Fase 02:
    # A Central de Alertas passa a usar a mesma fonte de saldo do Dashboard,
    # Banco de Horas e Extrato: o cache do motor BH 2.0. Quando o cache não
    # estiver disponível, mantém o fallback do consolidado para não bloquear a tela.
    # v8.21.4: a Central de Alertas deve abrir rápido. O saldo BH 2.0 completo
    # pode recalcular extratos em bases grandes; aqui usamos o saldo operacional
    # já disponível no consolidado e deixamos o motor BH 2.0 para Banco/Extrato.
    mapa_bh2_nome, mapa_bh2_mat = {}, {}


    linhas = []
    extratos = {}
    chaves_grupo = [c for c in ["matricula", "nome"] if c in df.columns] or ["nome"]
    for _, g in df.sort_values("data_dt").groupby(chaves_grupo, dropna=False):
        extr = _calcular_extrato_bh2_colaborador(g.copy())
        if not extr:
            continue
        r = extr.get("resumo", {}) or {}
        nome_colab = str(r.get("nome", "") or (str(g["nome"].iloc[-1]) if "nome" in g.columns and not g.empty else "")).strip()
        mat_colab = str(r.get("matricula", "") or (str(g["matricula"].iloc[-1]) if "matricula" in g.columns and not g.empty else "")).strip()
        chave = _chave_bh2_colaborador(nome_colab, mat_colab)

        linhas_extrato = []
        for item in extr.get("linhas", []) or []:
            linha = dict(item)
            linha.setdefault("nome", nome_colab)
            linha.setdefault("matricula", mat_colab)
            linha.setdefault("cc", r.get("cc", ""))
            linha.setdefault("funcao", r.get("funcao", ""))
            linha.setdefault("turno", r.get("turno", ""))
            linhas_extrato.append(linha)

        extratos[chave] = {"resumo": r, "linhas": linhas_extrato}
        linhas.append({
            "nome": nome_colab,
            "matricula": mat_colab,
            "cc": r.get("cc", ""),
            "funcao": r.get("funcao", ""),
            "turno": r.get("turno", ""),
            "he_total": r.get("credito_total", "00:00"),
            "absent_total": r.get("debito_total", "00:00"),
            "dias_com_he": _contar_dias_movimento_bh(linhas_extrato, "he"),
            "dias_com_ausencia": _contar_dias_movimento_bh(linhas_extrato, "absent"),
            "metricas_dias_fonte": "extrato_bh2",
            "saldo_atual": r.get("saldo_final", ""),
            "saldo_pdf_final": r.get("saldo_informado_final", ""),
            "saldo_base": r.get("saldo_base", ""),
            "origem_saldo_base": r.get("origem_saldo_base", ""),
        })

    revisar = int((df.get("status", pd.Series([], dtype="object")).astype(str).str.upper() == "REVISAR").sum()) if "status" in df.columns else 0
    payload = {
        "ok": True,
        "erro": "",
        "versao": f"{APP_VERSION}_bh2_cacheado_metricas_reais",
        "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "assinatura_excel": excel_sig,
        "assinatura_config": config_sig,
        "data_inicio": cfg.get("data_inicio", ""),
        "periodo_calculo": cfg.get("data_inicio", "") or "todo_periodo",
        "fonte": "banco_horas_2_0_cacheado_metricas_reais_v8_21_20",
        "registros_base": int(len(df)),
        "revisar": revisar,
        "linhas": linhas,
        "extratos": extratos,
        "kpis": _montar_kpis_banco_horas_de_linhas(linhas, registros_base=len(df), revisar=revisar),
    }
    _salvar_cache_bh2_calculo_disco(payload)
    _CACHE_BH2_CALCULO_MEM["key"] = cache_key
    _CACHE_BH2_CALCULO_MEM["payload"] = payload
    _CACHE_BH2_CALCULO_MEM["gerado_em_ts"] = time.monotonic()
    return payload


def _linhas_bh2_ate_data_cache(data_limite):
    cache = _obter_cache_bh2_calculo()
    if not cache.get("ok") or data_limite is None or pd.isna(data_limite):
        return []
    limite = pd.Timestamp(data_limite).normalize()
    linhas = []
    for extr in (cache.get("extratos", {}) or {}).values():
        r = extr.get("resumo", {}) or {}
        linhas_validas = []
        for item in extr.get("linhas", []) or []:
            dt = pd.to_datetime(str(item.get("data", "")), format="%d/%m/%Y", errors="coerce")
            if pd.notna(dt) and pd.Timestamp(dt).normalize() <= limite:
                linhas_validas.append(item)
        if not linhas_validas:
            continue
        ultima = linhas_validas[-1]
        credito_total = sum(hhmm_para_minutos(x.get("he", "")) for x in linhas_validas)
        debito_total = sum(hhmm_para_minutos(x.get("absent", "")) for x in linhas_validas)
        linhas.append({
            "nome": r.get("nome", ultima.get("nome", "")),
            "matricula": r.get("matricula", ultima.get("matricula", "")),
            "cc": r.get("cc", ultima.get("cc", "")),
            "funcao": r.get("funcao", ultima.get("funcao", "")),
            "turno": r.get("turno", ultima.get("turno", "")),
            "he_total": minutos_para_hhmm(int(credito_total)),
            "absent_total": minutos_para_hhmm(int(debito_total)),
            "dias_com_he": _contar_dias_movimento_bh(linhas_validas, "he"),
            "dias_com_ausencia": _contar_dias_movimento_bh(linhas_validas, "absent"),
            "metricas_dias_fonte": "extrato_bh2_filtrado_por_data",
            "saldo_atual": ultima.get("saldo_calculado", ""),
            "saldo_pdf_final": ultima.get("saldo_informado", ""),
            "saldo_base": r.get("saldo_base", ""),
            "origem_saldo_base": r.get("origem_saldo_base", ""),
        })
    return linhas


def gerar_resumo_banco_horas_pdf():
    """Gera o painel-resumo usado pela guia Banco de Horas.

    v8.10.39: a guia Banco de Horas não recalcula mais 100 colaboradores
    em tempo real. O cálculo pesado roda no botão manual Corrigir Banco de
    Horas pelos PDFs e salva este JSON.
    """
    df, erro = ler_consolidado()
    if erro:
        return {"ok": False, "mensagem": erro, "linhas": []}
    df = normalizar_colunas(df)
    if df.empty:
        return {"ok": True, "mensagem": "Base vazia; resumo não gerado.", "linhas": []}

    df = _garantir_colunas_calculadas(df)

    def _valor_excel_ok_bh(valor):
        texto = str(valor or "").strip()
        return texto and texto.lower() not in {"nan", "none", "nat"}

    # v8.14.00: MODO BANCO DE HORAS 2.0.
    # Quando há data/saldos definidos pelo gestor, monta o resumo a partir do
    # motor configurável. A fonte oficial passa a ser a configuração auditável
    # do Tempo Fechado, não a tentativa de escolher automaticamente um PDF.
    cfg_bh2 = _carregar_config_bh2()
    if cfg_bh2.get("ativo", True):
        # v8.15.23: usa cache do motor BH 2.0 em vez de recalcular todos os
        # colaboradores a cada troca de guia/filtro.
        resumo = _obter_cache_bh2_calculo()
        if resumo.get("ok"):
            _salvar_resumo_banco_horas_pdf(resumo)
        return resumo

    # v8.11.21: MODO EXCEL OFICIAL.
    # Se o arquivo importado já traz Saldo Calculado, o Tempo Fechado não deve
    # misturar cache de PDFs nem reprocessar saldos. O Banco de Horas passa a ser
    # montado exclusivamente com as colunas oficiais geradas pelo Script_Robo_Ponto.
    if "saldo_calculado" in df.columns and df["saldo_calculado"].apply(_valor_excel_ok_bh).any():
        linhas = []
        chaves_grupo = [c for c in ["matricula", "nome"] if c in df.columns]
        if not chaves_grupo:
            chaves_grupo = ["nome"]

        for chave, g in df.sort_values("data_dt").groupby(chaves_grupo, dropna=False):
            g_ord = g.sort_values("data_dt").copy()
            saldos_validos = [str(v).strip() for v in g_ord.get("saldo_calculado", pd.Series([], dtype="object")).tolist() if _valor_excel_ok_bh(v)]
            saldo_final = saldos_validos[-1] if saldos_validos else ""
            saldo_base = saldos_validos[0] if saldos_validos else ""

            saldos_pdf = [str(v).strip() for v in g_ord.get("saldo_atual", pd.Series([], dtype="object")).tolist() if _valor_excel_ok_bh(v)]
            saldo_pdf_final = saldos_pdf[-1] if saldos_pdf else ""

            nome_colab = str(g_ord["nome"].iloc[-1]) if "nome" in g_ord.columns and not g_ord.empty else str(chave)
            linhas.append({
                "nome": nome_colab,
                "matricula": str(g_ord["matricula"].iloc[-1]) if "matricula" in g_ord.columns and not g_ord.empty else "",
                "cc": str(g_ord["cc"].iloc[-1]) if "cc" in g_ord.columns and not g_ord.empty else "",
                "funcao": str(g_ord["funcao"].iloc[-1]) if "funcao" in g_ord.columns and not g_ord.empty else "",
                "turno": str(g_ord["turno"].iloc[-1]) if "turno" in g_ord.columns and not g_ord.empty else "",
                "he_total": minutos_para_hhmm(int(g_ord["he_min"].sum())),
                "absent_total": minutos_para_hhmm(int(g_ord["absent_min"].sum())),
                "saldo_atual": saldo_final,
                "saldo_pdf_final": saldo_pdf_final,
                "saldo_base": saldo_base,
                "data_base_saldo": str(g_ord["data"].iloc[0]) if "data" in g_ord.columns and not g_ord.empty else "",
                "credito_periodo": minutos_para_hhmm(int(g_ord["he_min"].sum())),
                "debito_periodo": minutos_para_hhmm(int(g_ord["absent_min"].sum())),
                "saldo_pdf_ignorado": "",
                "origem_saldo_base": "excel_oficial_saldo_calculado",
                "saldo_base_oficial": saldo_base,
                "dias_com_he": _contar_dias_movimento_df_bh(g_ord, "he_min", "he"),
                "dias_com_ausencia": _contar_dias_movimento_df_bh(g_ord, "absent_min", "absent"),
                "metricas_dias_fonte": "excel_oficial",
            })
        linhas = sorted(linhas, key=lambda x: _saldo_virgula_para_minutos(x.get("saldo_atual", "")), reverse=True)
        revisar = int((df["status"].str.upper() == "REVISAR").sum()) if "status" in df.columns else 0
        resumo = {
            "ok": True,
            "versao": "v8.11.36_excel_oficial_sem_consolidacao_dashboard",
            "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "registros_base": int(len(df)),
            "revisar": revisar,
            "fonte": "excel_oficial_sem_consolidacao_dashboard",
            "assinatura_excel": str(_assinatura_arquivo(ARQUIVO_PADRAO)),
            "linhas": linhas,
        }
        resumo["kpis"] = _montar_kpis_banco_horas_de_linhas(linhas, registros_base=len(df), revisar=revisar)
        _salvar_resumo_banco_horas_pdf(resumo)
        return {"ok": True, "mensagem": f"Resumo do Banco de Horas gerado exclusivamente pelo Excel com {len(linhas)} colaborador(es).", "linhas": linhas, "arquivo": str(ARQUIVO_RESUMO_BANCO_HORAS_PDF), "fonte": "excel_oficial_sem_consolidacao_dashboard"}

    # v8.11.12: o resumo em lote depende do catálogo de PDFs. Se o usuário
    # trocou de release e o JSON de cache ficou vazio, tenta atualizar o
    # histórico uma única vez dentro do processamento manual de correção.
    try:
        cache_tmp = carregar_cache_pdfs_banco_horas()
        if not cache_tmp or not cache_tmp.get("catalogo"):
            atualizar_cache_pdfs_banco_horas()
    except Exception:
        pass

    linhas = []
    for nome_colab, g in df.sort_values("data_dt").groupby("nome", dropna=False):
        saldo_info = _calcular_saldo_base_medicao_por_colaborador(g)
        linhas.append({
            "nome": str(nome_colab),
            "cc": str(g["cc"].iloc[-1]) if "cc" in g.columns and not g.empty else "",
            "funcao": str(g["funcao"].iloc[-1]) if "funcao" in g.columns and not g.empty else "",
            "turno": str(g["turno"].iloc[-1]) if "turno" in g.columns and not g.empty else "",
            "he_total": minutos_para_hhmm(int(g["he_min"].sum())),
            "absent_total": minutos_para_hhmm(int(g["absent_min"].sum())),
            "saldo_atual": saldo_info.get("saldo_calculado", "00,00"),
            "saldo_base": saldo_info.get("saldo_base", ""),
            "data_base_saldo": saldo_info.get("data_base", ""),
            "credito_periodo": minutos_para_hhmm(int(saldo_info.get("credito_total_min", 0) or 0)),
            "debito_periodo": minutos_para_hhmm(int(saldo_info.get("debito_total_min", 0) or 0)),
            "saldo_pdf_ignorado": saldo_info.get("saldo_pdf_ignorado", ""),
            "origem_saldo_base": saldo_info.get("origem_saldo_base", ""),
            "saldo_base_oficial": saldo_info.get("saldo_base_oficial", ""),
            "dias_com_he": _contar_dias_movimento_df_bh(g, "he_min", "he"),
            "dias_com_ausencia": _contar_dias_movimento_df_bh(g, "absent_min", "absent"),
            "metricas_dias_fonte": "excel_pdf_legado",
        })

    linhas = sorted(linhas, key=lambda x: _saldo_virgula_para_minutos(x.get("saldo_atual", "")), reverse=True)
    revisar = int((df["status"].str.upper() == "REVISAR").sum()) if "status" in df.columns else 0
    resumo = {
        "ok": True,
        "versao": "v8.10.41",
        "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "registros_base": int(len(df)),
        "revisar": revisar,
        "linhas": linhas,
    }
    resumo["kpis"] = _montar_kpis_banco_horas_de_linhas(linhas, registros_base=len(df), revisar=revisar)
    _salvar_resumo_banco_horas_pdf(resumo)
    return {"ok": True, "mensagem": f"Resumo do Banco de Horas gerado com {len(linhas)} colaborador(es).", "linhas": linhas, "arquivo": str(ARQUIVO_RESUMO_BANCO_HORAS_PDF)}


def _calcular_saldo_base_medicao_por_colaborador(df_colab):
    """
    Calcula o saldo consolidado da guia Banco de Horas usando a mesma regra
    da guia Extrato Banco de Horas.

    v8.10.36:
    - A guia Banco de Horas não deve mais usar o saldo antigo do Excel.
    - O saldo base vem do PDF de fechamento do ciclo anterior.
    - Se o PDF de fechamento anterior não estiver disponível, usa como fallback
      o PDF/fotografia que contém a data-base do ciclo.
    - A movimentação posterior vem do Excel: crédito H.E. - débito ausência.
    """
    vazio = {
        "saldo_base": "PDF não encontrado",
        "saldo_base_min": 0,
        "saldo_calculado": "00,00",
        "saldo_calculado_min": 0,
        "debito_total_min": 0,
        "credito_total_min": 0,
        "data_base": "",
        "saldo_pdf_ignorado": "",
        "origem_saldo_base": "pdf_nao_encontrado_sem_fallback_excel",
        "saldo_base_oficial": "",
    }

    if df_colab.empty:
        return vazio

    g = df_colab.copy()
    g["data_dt"] = pd.to_datetime(g["data"], format="%d/%m/%Y", errors="coerce")
    g = g[g["data_dt"].notna()].sort_values("data_dt").copy()

    if g.empty:
        return vazio

    data_ref = g["data_dt"].max()
    inicio = _data_inicio_medicao(data_ref)

    if inicio is None:
        inicio = g["data_dt"].min()

    fim = _data_fim_medicao(inicio) or (inicio + pd.Timedelta(days=30))
    g_ciclo = g[(g["data_dt"] >= inicio) & (g["data_dt"] <= fim)].sort_values("data_dt").copy()

    if g_ciclo.empty:
        g_ciclo = g.sort_values("data_dt").copy()
        inicio = g_ciclo["data_dt"].min()
        fim = g_ciclo["data_dt"].max()

    base_rows = g_ciclo[g_ciclo["data_dt"] == inicio].copy()
    base = base_rows.iloc[0] if not base_rows.empty else g_ciclo.iloc[0]
    data_base = base["data_dt"]

    nome_base = str(g_ciclo["nome"].iloc[0]) if "nome" in g_ciclo.columns and not g_ciclo.empty else ""
    matricula_base = _matricula_da_linha_bh(base)
    saldo_base_oficial = _obter_saldo_base_medicao_oficial(nome_base, data_base)

    try:
        cache_bh = carregar_cache_pdfs_banco_horas()
        catalogo_bh = cache_bh.get("catalogo", []) if cache_bh else []
    except Exception:
        catalogo_bh = []

    # v8.10.41: na guia Banco de Horas/resumo em lote, NÃO fazer busca
    # direta nos PDFs para cada colaborador. Isso podia varrer a pasta 100+
    # vezes e deixar o botão Corrigir Banco de Horas preso por muitos minutos.
    # A busca direta fica restrita ao Extrato individual; aqui usamos somente
    # o histórico/cache já gerado pelo botão Atualizar Histórico de PDFs.
    pdf_base = _buscar_pdf_fechamento_ciclo_anterior(
        nome_base,
        matricula_base,
        inicio,
        catalogo_bh,
        permitir_busca_direta=False,
    )

    saldo_base_txt = ""
    origem_saldo_base = "pdf_nao_encontrado_sem_fallback_excel"
    saldo_pdf_ignorado = ""

    if pdf_base and str(pdf_base.get("saldo_atual_pdf", "") or "").strip():
        saldo_base_txt = str(pdf_base.get("saldo_atual_pdf", "") or "").strip()
        origem_saldo_base = f"pdf_fechamento_ciclo_anterior_{pdf_base.get('_match_base_pdf_v25','pdf')}: {pdf_base.get('arquivo_pdf','')}"
    else:
        # Mesmo fallback seguro usado no Extrato: se o fechamento anterior não
        # estiver catalogado, usar a fotografia em PDF que contém a data-base.
        pdf_base_por_data = _saldo_pdf_para_data_exata(nome_base, matricula_base, data_base, catalogo_bh)
        if pdf_base_por_data and str(pdf_base_por_data.get("saldo_atual_pdf", "") or "").strip():
            saldo_base_txt = str(pdf_base_por_data.get("saldo_atual_pdf", "") or "").strip()
            origem_saldo_base = f"pdf_data_base_sem_fallback_excel: {pdf_base_por_data.get('arquivo_pdf','')}"
            pdf_base = pdf_base_por_data

    saldo_base_min = _saldo_para_minutos_banco(saldo_base_txt) if saldo_base_txt else 0

    mov = g_ciclo[g_ciclo["data_dt"] > data_base].copy()

    credito_total_min = int(mov["he"].apply(hhmm_para_minutos).sum()) if "he" in mov.columns else 0
    debito_total_min = int(mov["absent"].apply(hhmm_para_minutos).sum()) if "absent" in mov.columns else 0

    saldo_calculado_min = saldo_base_min + credito_total_min - debito_total_min

    if pdf_base:
        saldo_pdf_ignorado = str(pdf_base.get("saldo_atual_pdf", "") or "")
    else:
        ultimo = g_ciclo.tail(1)
        saldo_pdf_ignorado = str(ultimo["saldo"].iloc[0]) if not ultimo.empty and "saldo" in ultimo.columns else ""

    return {
        "saldo_base": (saldo_base_txt if saldo_base_txt else "PDF não encontrado"),
        "saldo_base_min": int(saldo_base_min),
        "saldo_calculado": _minutos_para_saldo_virgula(saldo_calculado_min),
        "saldo_calculado_min": int(saldo_calculado_min),
        "debito_total_min": int(debito_total_min),
        "credito_total_min": int(credito_total_min),
        "data_base": pd.to_datetime(data_base).strftime("%d/%m/%Y"),
        "saldo_pdf_ignorado": saldo_pdf_ignorado,
        "origem_saldo_base": origem_saldo_base,
        "saldo_base_oficial": saldo_base_oficial or "",
    }






# ============================================================
# HELPERS GUIAS ESPECIAIS: INTER JORNADA / VIOLAÇÕES DE JORNADA
# ============================================================

def _hora_marcacao_para_time(valor):
    texto = str(valor or "").strip()
    if not texto or texto.lower() in {"nan", "none", "nat"}:
        return None

    m = re.search(r"(\d{1,2}):(\d{2})", texto)
    if not m:
        return None

    try:
        h = int(m.group(1))
        mi = int(m.group(2))
        if 0 <= h <= 23 and 0 <= mi <= 59:
            return h, mi
    except Exception:
        return None

    return None


def _primeira_ultima_marcacao(row):
    """
    Retorna primeira e última marcação válida do dia.

    Aceita colunas normalizadas e também os nomes originais do Excel/PDF.
    """
    grupos = [
        ["e1", "s1", "e2", "s2", "e3", "s3", "e4", "s4"],
        ["1a E.", "1a S.", "2a E.", "2a S.", "3a E.", "3a S.", "4a E.", "4a S."],
        ["1ª E.", "1ª S.", "2ª E.", "2ª S.", "3ª E.", "3ª S.", "4ª E.", "4ª S."],
    ]

    marcas = []

    for cols in grupos:
        marcas = []
        for col in cols:
            hm = _hora_marcacao_para_time(row.get(col, ""))
            if hm is not None:
                marcas.append((col, hm[0] * 60 + hm[1], f"{hm[0]:02d}:{hm[1]:02d}"))
        if marcas:
            break

    if not marcas:
        return None, None

    return marcas[0], marcas[-1]


def _minutos_para_hhmm_simples(total):
    try:
        total = int(total)
    except Exception:
        return "00:00"
    total = max(0, total)
    return f"{total // 60:02d}:{total % 60:02d}"


def _tem_trabalho_jornada(row):
    """
    Considera dia trabalhado quando há marcação válida.

    Para Violações de Jornada, batida prevalece sobre observação:
    se houve marcação, conta como trabalho mesmo que a observação esteja
    como DSR, Compensado ou Não Trabalhado.
    """
    # Fonte principal quando existir.
    for qtd_col in ["qtd_batidas", "Qtd Batidas", "Qtd. Batidas", "Quantidade Batidas"]:
        valor_qtd = row.get(qtd_col, "")
        try:
            texto = str(valor_qtd or "").strip().replace(",", ".")
            if texto and texto.lower() not in {"nan", "none", "nat"}:
                if int(float(texto)) > 0:
                    return True
        except Exception:
            pass

    primeira, ultima = _primeira_ultima_marcacao(row)
    return primeira is not None or ultima is not None


def _turno_regra_violacao_jornada(turno):
    texto = str(turno or "").strip().lower()

    if texto in {"12x36m7", "12x36m8", "12x36n19", "12x36n19.1"}:
        return "12x36"

    if texto in {"comercial7", "comercial8"}:
        return "comercial"

    if texto in {"noturno1", "noturno2"}:
        return "noturno"

    if "12x36" in texto or "12 x 36" in texto:
        return "12x36"
    if texto.startswith("comercial"):
        return "comercial"
    if texto.startswith("noturno"):
        return "noturno"

    return "outros"


def _semana_iso_label(data_dt):
    try:
        iso = data_dt.isocalendar()
        return f"{int(iso.year)}-S{int(iso.week):02d}"
    except Exception:
        return ""


def _primeira_data_semana(data_dt):
    try:
        return (data_dt - pd.Timedelta(days=int(data_dt.weekday()))).strftime("%d/%m/%Y")
    except Exception:
        return ""


def _ultima_data_semana(data_dt):
    try:
        return (data_dt + pd.Timedelta(days=6-int(data_dt.weekday()))).strftime("%d/%m/%Y")
    except Exception:
        return ""



# ============================================================
# NOTIFICAÇÕES AUTOMÁTICAS - INTERJORNADA / VIOLAÇÕES DE JORNADA
# ============================================================

def _log_notificacao_jornada(mensagem):
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        linha = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {mensagem}\n"
        with open(ARQUIVO_LOG_NOTIFICACOES_JORNADA, "a", encoding="utf-8") as f:
            f.write(linha)
    except Exception:
        pass


def _cfg_notificacoes_jornada():
    cfg = {
        "EMAIL_ATIVO": False,
        "SMTP_HOST": "",
        "SMTP_PORT": 587,
        "SMTP_USAR_TLS": True,
        "SMTP_USUARIO": "",
        "SMTP_SENHA": "",
        "EMAIL_REMETENTE": "",
        "EMAIL_DESTINATARIOS": [],
        "WHATSAPP_ATIVO": False,
        "WHATSAPP_API_URL": "",
        "WHATSAPP_TOKEN": "",
        "WHATSAPP_DESTINATARIOS": [],
        "ENVIAR_AO_IMPORTAR_EXCEL": True,
        "ENVIAR_APENAS_OCORRENCIAS_NOVAS": True,
        "LIMITE_OCORRENCIAS_POR_ENVIO": 50,
    }

    # 1) Configuração editável do usuário, gravada em C:\Users\<usuario>\ponto_pdfs\config.
    try:
        USER_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
        if ARQUIVO_CONFIG_NOTIFICACOES_JORNADA.exists():
            with open(ARQUIVO_CONFIG_NOTIFICACOES_JORNADA, "r", encoding="utf-8") as f:
                dados = json.load(f) or {}
            for k in cfg:
                if k in dados:
                    cfg[k] = dados[k]
    except Exception as e:
        _log_notificacao_jornada(f"Configuração JSON de notificações não carregada: {e}")

    # 2) Compatibilidade com arquivo Python antigo, se existir no pacote.
    try:
        import alertas_notificacao_config as c
        for k in cfg:
            if hasattr(c, k) and (cfg.get(k) in [None, "", [], False]):
                cfg[k] = getattr(c, k)

        if hasattr(c, "SMTP_SERVIDOR") and not cfg.get("SMTP_HOST"):
            cfg["SMTP_HOST"] = getattr(c, "SMTP_SERVIDOR")
        if hasattr(c, "SMTP_SERVER") and not cfg.get("SMTP_HOST"):
            cfg["SMTP_HOST"] = getattr(c, "SMTP_SERVER")
        if not cfg.get("EMAIL_REMETENTE") and cfg.get("SMTP_USUARIO"):
            cfg["EMAIL_REMETENTE"] = cfg.get("SMTP_USUARIO")
    except Exception:
        pass

    return cfg


def _salvar_cfg_notificacoes_jornada(dados):
    cfg = _cfg_notificacoes_jornada()
    permitidas = set(cfg.keys())
    for k, v in (dados or {}).items():
        if k in permitidas:
            cfg[k] = v

    def lista_limpa(valor):
        if isinstance(valor, list):
            base = valor
        else:
            base = str(valor or "").replace(";", ",").split(",")
        return [str(x).strip() for x in base if str(x).strip()]

    cfg["EMAIL_ATIVO"] = bool(cfg.get("EMAIL_ATIVO"))
    cfg["WHATSAPP_ATIVO"] = bool(cfg.get("WHATSAPP_ATIVO"))
    cfg["SMTP_USAR_TLS"] = bool(cfg.get("SMTP_USAR_TLS", True))
    cfg["ENVIAR_AO_IMPORTAR_EXCEL"] = bool(cfg.get("ENVIAR_AO_IMPORTAR_EXCEL", True))
    cfg["ENVIAR_APENAS_OCORRENCIAS_NOVAS"] = bool(cfg.get("ENVIAR_APENAS_OCORRENCIAS_NOVAS", True))
    cfg["EMAIL_DESTINATARIOS"] = lista_limpa(cfg.get("EMAIL_DESTINATARIOS"))
    cfg["WHATSAPP_DESTINATARIOS"] = lista_limpa(cfg.get("WHATSAPP_DESTINATARIOS"))
    try:
        cfg["SMTP_PORT"] = int(cfg.get("SMTP_PORT") or 587)
    except Exception:
        cfg["SMTP_PORT"] = 587
    try:
        cfg["LIMITE_OCORRENCIAS_POR_ENVIO"] = int(cfg.get("LIMITE_OCORRENCIAS_POR_ENVIO") or 50)
    except Exception:
        cfg["LIMITE_OCORRENCIAS_POR_ENVIO"] = 50

    USER_CONFIG_DIR.mkdir(parents=True, exist_ok=True)
    with open(ARQUIVO_CONFIG_NOTIFICACOES_JORNADA, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    return cfg


def _carregar_alertas_jornada_enviados():
    try:
        if ARQUIVO_ALERTAS_JORNADA_ENVIADOS.exists():
            return json.loads(ARQUIVO_ALERTAS_JORNADA_ENVIADOS.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {"ids": [], "historico": []}


def _salvar_alertas_jornada_enviados(controle):
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        # Mantém o arquivo leve mesmo após muitos ciclos.
        controle["ids"] = list(dict.fromkeys(controle.get("ids", [])))[-5000:]
        controle["historico"] = controle.get("historico", [])[-1000:]
        ARQUIVO_ALERTAS_JORNADA_ENVIADOS.write_text(
            json.dumps(controle, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception as e:
        _log_notificacao_jornada(f"Falha ao salvar controle de alertas enviados: {e}")


def _id_ocorrencia_jornada(ocorrencia):
    base = "|".join([
        str(ocorrencia.get("tipo_alerta", "")),
        str(ocorrencia.get("nome", "")),
        str(ocorrencia.get("cc", "")),
        str(ocorrencia.get("turno", "")),
        str(ocorrencia.get("data_referencia", "")),
        str(ocorrencia.get("data_retorno", "")),
        str(ocorrencia.get("violacao", "")),
        str(ocorrencia.get("detalhe", "")),
    ])
    return hashlib.sha256(base.encode("utf-8", errors="ignore")).hexdigest()


def _calcular_interjornada_notificacao(df):
    resultados = []
    if df.empty:
        return resultados

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")
    df = df[df["data_dt"].notna()].copy()
    if df.empty:
        return resultados

    for nome, g in df.sort_values("data_dt").groupby("nome", dropna=False):
        registros = []
        for _, row in g.sort_values("data_dt").iterrows():
            primeira, ultima = _primeira_ultima_marcacao(row)
            if primeira is None or ultima is None:
                continue
            registros.append({"row": row, "data_dt": row["data_dt"], "primeira": primeira, "ultima": ultima})

        for anterior, atual in zip(registros, registros[1:]):
            data_anterior = anterior["data_dt"]
            data_atual = atual["data_dt"]
            fim_abs = int(data_anterior.timestamp() // 60) + anterior["ultima"][1]
            inicio_abs = int(data_atual.timestamp() // 60) + atual["primeira"][1]
            intervalo = inicio_abs - fim_abs
            if intervalo < 0:
                continue
            if intervalo < 11 * 60:
                row_atual = atual["row"]
                resultados.append({
                    "tipo_alerta": "Inter Jornada",
                    "criticidade": "Crítico",
                    "cc": str(row_atual.get("cc", "")),
                    "nome": str(nome),
                    "funcao": str(row_atual.get("funcao", "")),
                    "turno": str(row_atual.get("turno", "")),
                    "data_referencia": data_atual.strftime("%d/%m/%Y"),
                    "data_anterior": data_anterior.strftime("%d/%m/%Y"),
                    "ultima_marcacao_anterior": anterior["ultima"][2],
                    "data_retorno": data_atual.strftime("%d/%m/%Y"),
                    "primeira_marcacao_retorno": atual["primeira"][2],
                    "violacao": "Interjornada inferior a 11h",
                    "detalhe": f"Intervalo de {_minutos_para_hhmm_simples(intervalo)} entre a saída de {data_anterior.strftime('%d/%m/%Y')} e a entrada de {data_atual.strftime('%d/%m/%Y')}.",
                    "intervalo": _minutos_para_hhmm_simples(intervalo),
                    "deficit": _minutos_para_hhmm_simples((11 * 60) - intervalo),
                })
    return resultados


def _calcular_violacoes_jornada_notificacao(df):
    resultados = []
    if df.empty:
        return resultados

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")
    df = df[df["data_dt"].notna()].copy()
    if df.empty:
        return resultados

    for nome_colab, g in df.sort_values("data_dt").groupby("nome", dropna=False):
        g = g.sort_values("data_dt").copy()
        g["_trabalhado"] = g.apply(_tem_trabalho_jornada, axis=1)
        g["_regra_turno"] = g["turno"].apply(_turno_regra_violacao_jornada)

        trabalhados_12x36 = g[(g["_trabalhado"]) & (g["_regra_turno"] == "12x36")].copy()
        registros_12x36 = trabalhados_12x36.to_dict(orient="records")
        for anterior, atual in zip(registros_12x36, registros_12x36[1:]):
            dias_diff = (atual["data_dt"].date() - anterior["data_dt"].date()).days
            if dias_diff == 1:
                resultados.append({
                    "tipo_alerta": "Violação de Jornada",
                    "criticidade": "Crítico",
                    "cc": str(atual.get("cc", "")),
                    "nome": str(nome_colab),
                    "funcao": str(atual.get("funcao", "")),
                    "turno": str(atual.get("turno", "")),
                    "tipo_jornada": "12x36",
                    "data_referencia": atual["data_dt"].strftime("%d/%m/%Y"),
                    "violacao": "Quebra da escala 12x36",
                    "detalhe": f"Trabalho em dias consecutivos: {anterior['data_dt'].strftime('%d/%m/%Y')} e {atual['data_dt'].strftime('%d/%m/%Y')}.",
                    "limite_esperado": "Sem dias consecutivos",
                })

        base_semana = g[(g["_trabalhado"]) & (g["_regra_turno"].isin(["comercial", "noturno"]))].copy()
        if not base_semana.empty:
            base_semana["_semana"] = base_semana["data_dt"].apply(_semana_iso_label)
            for (semana, tipo), gs in base_semana.groupby(["_semana", "_regra_turno"], dropna=False):
                dias_trabalhados = int(gs["data_dt"].dt.date.nunique())
                if dias_trabalhados <= 5:
                    continue
                ref_dt = gs["data_dt"].min()
                resultados.append({
                    "tipo_alerta": "Violação de Jornada",
                    "criticidade": "Atenção",
                    "cc": str(gs["cc"].iloc[-1]),
                    "nome": str(nome_colab),
                    "funcao": str(gs["funcao"].iloc[-1]),
                    "turno": str(gs["turno"].iloc[-1]),
                    "tipo_jornada": str(tipo).capitalize(),
                    "periodo": f"{_primeira_data_semana(ref_dt)} a {_ultima_data_semana(ref_dt)}",
                    "data_referencia": ref_dt.strftime("%d/%m/%Y"),
                    "violacao": "Excesso de dias trabalhados na semana",
                    "detalhe": f"Foram identificados {dias_trabalhados} dia(s) trabalhado(s) na semana de segunda a domingo.",
                    "dias_trabalhados": dias_trabalhados,
                    "limite_esperado": "Até 5 dias/semana",
                })
    return resultados


def _montar_resumo_notificacao_jornada(ocorrencias, motivo=""):
    total_inter = sum(1 for o in ocorrencias if o.get("tipo_alerta") == "Inter Jornada")
    total_viol = sum(1 for o in ocorrencias if o.get("tipo_alerta") == "Violação de Jornada")
    colaboradores = sorted(set(str(o.get("nome", "")).strip() for o in ocorrencias if str(o.get("nome", "")).strip()))
    linhas = [
        "ALERTA AUTOMÁTICO - ROBÔ DE PONTO",
        "",
        f"Motivo da verificação: {motivo or 'Verificação automática'}",
        f"Data/hora: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        "",
        f"Total de ocorrências novas: {len(ocorrencias)}",
        f"Interjornada inferior a 11h: {total_inter}",
        f"Violação de jornada: {total_viol}",
        f"Colaboradores impactados: {len(colaboradores)}",
        "",
        "Ocorrências:",
    ]
    for i, o in enumerate(ocorrencias, 1):
        linhas.extend([
            "",
            f"{i}. [{o.get('tipo_alerta', '')}] {o.get('nome', '')} - {o.get('data_referencia', o.get('data_retorno', ''))}",
            f"   CC: {o.get('cc', '')} | Turno: {o.get('turno', '')}",
            f"   Regra: {o.get('violacao', '')}",
            f"   Detalhe: {o.get('detalhe', '')}",
        ])
        if o.get("deficit"):
            linhas.append(f"   Déficit interjornada: {o.get('deficit')}")
    linhas.extend(["", "Ação recomendada: validar a ocorrência e registrar a tratativa operacional."])
    return "\n".join(linhas)


def _enviar_email_alerta_jornada(cfg, assunto, corpo):
    destinatarios = [d for d in cfg.get("EMAIL_DESTINATARIOS", []) if str(d).strip()]
    if not cfg.get("EMAIL_ATIVO"):
        return {"ok": False, "mensagem": "E-mail desativado em alertas_notificacao_config.py."}
    if not destinatarios:
        return {"ok": False, "mensagem": "Nenhum destinatário de e-mail configurado."}
    if not cfg.get("SMTP_HOST") or not cfg.get("EMAIL_REMETENTE"):
        return {"ok": False, "mensagem": "SMTP_HOST ou EMAIL_REMETENTE não configurado."}

    msg = EmailMessage()
    msg["Subject"] = assunto
    msg["From"] = cfg.get("EMAIL_REMETENTE")
    msg["To"] = ", ".join(destinatarios)
    msg.set_content(corpo)

    with smtplib.SMTP(cfg.get("SMTP_HOST"), int(cfg.get("SMTP_PORT", 587)), timeout=30) as smtp:
        if cfg.get("SMTP_USAR_TLS", True):
            smtp.starttls()
        usuario = cfg.get("SMTP_USUARIO")
        senha = cfg.get("SMTP_SENHA")
        if usuario and senha:
            smtp.login(usuario, senha)
        smtp.send_message(msg)
    return {"ok": True, "mensagem": f"E-mail enviado para {len(destinatarios)} destinatário(s)."}


def _enviar_whatsapp_alerta_jornada(cfg, corpo):
    if not cfg.get("WHATSAPP_ATIVO"):
        return {"ok": False, "mensagem": "WhatsApp desativado em alertas_notificacao_config.py."}
    url = str(cfg.get("WHATSAPP_API_URL") or "").strip()
    token = str(cfg.get("WHATSAPP_TOKEN") or "").strip()
    destinatarios = [d for d in cfg.get("WHATSAPP_DESTINATARIOS", []) if str(d).strip()]
    if not url or not destinatarios:
        return {"ok": False, "mensagem": "WHATSAPP_API_URL ou destinatários não configurados."}

    enviados = 0
    erros = []
    for numero in destinatarios:
        payload = json.dumps({"to": numero, "message": corpo[:3500]}).encode("utf-8")
        req = urllib.request.Request(url, data=payload, headers={"Content-Type": "application/json"}, method="POST")
        if token:
            req.add_header("Authorization", f"Bearer {token}")
        try:
            with urllib.request.urlopen(req, timeout=30) as resp:
                if 200 <= int(resp.status) < 300:
                    enviados += 1
                else:
                    erros.append(f"{numero}: HTTP {resp.status}")
        except Exception as e:
            erros.append(f"{numero}: {e}")
    return {"ok": enviados > 0, "mensagem": f"WhatsApp enviado para {enviados} número(s)." + (" Erros: " + "; ".join(erros) if erros else "")}


def verificar_e_enviar_alertas_jornada(motivo="Verificação automática", forcar_reenvio=False):
    cfg = _cfg_notificacoes_jornada()
    df, erro = ler_consolidado()
    if erro:
        return {"ok": False, "mensagem": erro, "novas": 0, "total_detectadas": 0, "envios": []}

    df = normalizar_colunas(df)
    if df.empty:
        return {"ok": True, "mensagem": "Nenhuma base carregada para verificação.", "novas": 0, "total_detectadas": 0, "envios": []}

    ocorrencias = _calcular_interjornada_notificacao(df) + _calcular_violacoes_jornada_notificacao(df)
    controle = _carregar_alertas_jornada_enviados()
    enviados_set = set(controle.get("ids", []))

    novas = []
    for o in ocorrencias:
        oid = _id_ocorrencia_jornada(o)
        o["id_alerta"] = oid
        if forcar_reenvio or not cfg.get("ENVIAR_APENAS_OCORRENCIAS_NOVAS", True) or oid not in enviados_set:
            novas.append(o)

    limite = int(cfg.get("LIMITE_OCORRENCIAS_POR_ENVIO", 50) or 50)
    novas_envio = novas[:limite]
    if not novas_envio:
        _log_notificacao_jornada(f"Sem novas ocorrências. Detectadas: {len(ocorrencias)}. Motivo: {motivo}")
        return {"ok": True, "mensagem": "Nenhuma ocorrência nova para notificar.", "novas": 0, "total_detectadas": len(ocorrencias), "envios": []}

    assunto = f"Tempo Fechado - {len(novas_envio)} alerta(s) de jornada"
    corpo = _montar_resumo_notificacao_jornada(novas_envio, motivo=motivo)

    envios = []
    try:
        envios.append({"canal": "email", **_enviar_email_alerta_jornada(cfg, assunto, corpo)})
    except Exception as e:
        envios.append({"canal": "email", "ok": False, "mensagem": str(e)})
    try:
        envios.append({"canal": "whatsapp", **_enviar_whatsapp_alerta_jornada(cfg, corpo)})
    except Exception as e:
        envios.append({"canal": "whatsapp", "ok": False, "mensagem": str(e)})

    # Marca como enviado somente se ao menos um canal ativo tiver enviado com sucesso.
    houve_envio = any(e.get("ok") for e in envios)
    if houve_envio:
        for o in novas_envio:
            enviados_set.add(o["id_alerta"])
        controle["ids"] = list(enviados_set)
        controle.setdefault("historico", []).append({
            "quando": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "motivo": motivo,
            "quantidade": len(novas_envio),
            "envios": envios,
        })
        _salvar_alertas_jornada_enviados(controle)

    _log_notificacao_jornada(f"Verificação concluída. Novas={len(novas_envio)} Detectadas={len(ocorrencias)} Envios={envios}")
    return {
        "ok": houve_envio or not (cfg.get("EMAIL_ATIVO") or cfg.get("WHATSAPP_ATIVO")),
        "mensagem": "Verificação concluída.",
        "novas": len(novas_envio),
        "total_detectadas": len(ocorrencias),
        "envios": envios,
        "amostra": novas_envio[:10],
    }


@app.route("/api/debug-banco-horas")
@login_obrigatorio
def api_debug_banco_horas():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro})

    df = normalizar_colunas(df)
    nome = request.args.get("nome", "").strip().lower()

    if nome:
        df = df[df["nome"].str.lower().str.contains(nome, na=False)]

    if df.empty:
        return jsonify({"erro": "", "dados": [], "resumo": {}})

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")

    saida = []
    resumo = {}

    for nome_colab, g in df.sort_values("data_dt").groupby("nome", dropna=False):
        info = _calcular_saldo_base_medicao_por_colaborador(g)
        resumo[nome_colab] = info

        gg = g.sort_values("data_dt").copy()
        for _, r in gg.iterrows():
            saida.append({
                "nome": nome_colab,
                "data": r.get("data", ""),
                "he": r.get("he", ""),
                "absent": r.get("absent", ""),
                "saldo_pdf": r.get("saldo", ""),
                "saldo_base": info.get("saldo_base", ""),
                "saldo_calculado_final": info.get("saldo_calculado", ""),
                "data_base": info.get("data_base", ""),
                "origem_saldo_base": info.get("origem_saldo_base", ""),
                "saldo_base_oficial": info.get("saldo_base_oficial", ""),
            })

    return jsonify({"erro": "", "dados": saida, "resumo": resumo})




@app.route("/api/status-base")
@login_obrigatorio
def api_status_base():
    ativa = _base_operacional_ativa()
    arquivo = ""
    try:
        arquivo = ARQUIVO_BASE_ATIVA_FLAG.read_text(encoding="utf-8").strip() if ativa else ""
    except Exception:
        arquivo = ""

    existe_excel = False
    try:
        existe_excel = ARQUIVO_PADRAO.exists()
    except Exception:
        pass

    return jsonify({
        "erro": "",
        "base_ativa": bool(ativa),
        "arquivo": arquivo,
        "excel_existe": bool(existe_excel),
        "mensagem": "Base carregada" if ativa else "Nenhuma base carregada. Importe o Excel do Tempo Fechado."
    })




def _cleanup_remover_arquivo(path, removidos, erros, categoria="arquivo"):
    try:
        p = Path(path)
        if p.exists() and p.is_file():
            tamanho = p.stat().st_size
            p.unlink()
            removidos.append({"categoria": categoria, "caminho": str(p), "bytes": int(tamanho)})
    except Exception as e:
        erros.append({"categoria": categoria, "caminho": str(path), "erro": str(e)})


def _cleanup_remover_padroes(pasta, padroes, removidos, erros, categoria):
    try:
        p = Path(pasta)
        if not p.exists() or not p.is_dir():
            return
        for padrao in padroes:
            for item in p.glob(padrao):
                if item.is_file():
                    _cleanup_remover_arquivo(item, removidos, erros, categoria)
    except Exception as e:
        erros.append({"categoria": categoria, "caminho": str(pasta), "erro": str(e)})


def _executar_cleanup_operacional(payload=None, validar_job=True):
    payload = payload or {}
    limpar_base_app = bool(payload.get("limpar_base_app", True))
    limpar_pastas_pdf = bool(payload.get("limpar_pastas_pdf", True))
    limpar_cache_bh = bool(payload.get("limpar_cache_bh", True))
    limpar_pdfs_entrada = bool(payload.get("limpar_pdfs_entrada", limpar_pastas_pdf))
    limpar_pdfs_processados = bool(payload.get("limpar_pdfs_processados", limpar_pastas_pdf))
    limpar_pdfs_erro = bool(payload.get("limpar_pdfs_erro", limpar_pastas_pdf))
    limpar_saida_robo = bool(payload.get("limpar_saida_robo", limpar_pastas_pdf))
    limpar_controle_temporario = bool(payload.get("limpar_controle_temporario", limpar_pastas_pdf))
    limpar_logs_historico = bool(payload.get("limpar_logs_historico", False))

    if validar_job:
        with PROCESSAMENTO_INTEGRADO_LOCK:
            if PROCESSAMENTO_INTEGRADO_JOB.get("status") == "rodando":
                return {
                "ok": False,
                "erro": "Existe um processamento integrado em andamento. Aguarde concluir antes de executar o Cleanup.",
                    "status_code": 409,
                }

    removidos = []
    erros = []

    if limpar_base_app:
        for arquivo in [
            ARQUIVO_PADRAO,
            ARQUIVO_BASE_ATIVA_FLAG,
            ARQUIVO_SALDOS_BASE_MEDICAO,
        ]:
            _cleanup_remover_arquivo(arquivo, removidos, erros, "base_app")

    if limpar_base_app or limpar_cache_bh:
        for arquivo in [
            ARQUIVO_CACHE_PDFS_BANCO_HORAS,
            ARQUIVO_RESUMO_BANCO_HORAS_PDF,
            ARQUIVO_CACHE_BH2_CALCULO,
        ]:
            _cleanup_remover_arquivo(arquivo, removidos, erros, "cache_banco_horas")
        invalidar_cache_consolidado()
        try:
            _CACHE_PDFS_MEM["mtime"] = None
            _CACHE_PDFS_MEM["payload"] = None
        except Exception:
            pass
        try:
            _CACHE_BH2_CALCULO_MEM["key"] = None
            _CACHE_BH2_CALCULO_MEM["payload"] = None
            _CACHE_BH2_CALCULO_MEM["gerado_em_ts"] = 0
        except Exception:
            pass

    if any([limpar_pdfs_entrada, limpar_pdfs_processados, limpar_pdfs_erro, limpar_saida_robo, limpar_controle_temporario]):
        cfg = carregar_config_robo()
        aplicar_config_robo(cfg)
        raiz = Path(PASTA_RAIZ_PDFS_ROBO)
        if limpar_pdfs_entrada:
            _cleanup_remover_padroes(raiz / "entrada", ["*.pdf", "*.PDF"], removidos, erros, "pdfs_entrada")
        if limpar_pdfs_processados:
            _cleanup_remover_padroes(raiz / "processados", ["*.pdf", "*.PDF"], removidos, erros, "pdfs_processados")
        if limpar_pdfs_erro:
            _cleanup_remover_padroes(raiz / "erro", ["*.pdf", "*.PDF"], removidos, erros, "pdfs_erro")
        if limpar_saida_robo:
            _cleanup_remover_padroes(raiz / "saida", ["*.xlsx", "*.xls", "*.csv"], removidos, erros, "saida_robo")
        if limpar_controle_temporario:
            _cleanup_remover_padroes(raiz / "controle", ["download.lock", "robo_ponto_resultado.json"], removidos, erros, "controle_robo")

    if limpar_logs_historico:
        for arquivo in [
            ARQUIVO_LOG_PROCESSAMENTO_INTEGRADO,
            ARQUIVO_HISTORICO_OPERACAO_V8190,
        ]:
            _cleanup_remover_arquivo(arquivo, removidos, erros, "logs_historico_processamento")

        cfg = carregar_config_robo()
        aplicar_config_robo(cfg)
        controle = Path(PASTA_RAIZ_PDFS_ROBO) / "controle"
        for nome in [
            "processamento_manual_robo.log",
            "robo_ponto.log",
            "download_anexos.log",
        ]:
            _cleanup_remover_arquivo(controle / nome, removidos, erros, "logs_historico_processamento")

        try:
            PROCESSAMENTO_INTEGRADO_JOB["log_tail"] = ""
            PROCESSAMENTO_INTEGRADO_JOB["resultado"] = {}
        except Exception:
            pass

    total_bytes = sum(int(x.get("bytes") or 0) for x in removidos)
    mensagem = f"Cleanup concluido. {len(removidos)} arquivo(s) removido(s)."
    try:
        registrar_auditoria("CLEANUP_OPERACIONAL", mensagem, "OK" if not erros else "PARCIAL")
    except Exception:
        pass
    if not limpar_logs_historico:
        try:
            registrar_historico_operacao_v8190(
                "Cleanup Operacional",
                "OK" if not erros else "PARCIAL",
                mensagem,
                {"removidos": removidos[:200], "erros": erros[:50], "bytes": total_bytes},
                origem="Processamento"
            )
        except Exception:
            pass

    return {
        "ok": True,
        "parcial": bool(erros),
        "erro": "",
        "aviso": "" if not erros else "Cleanup concluido com alguns itens nao removidos.",
        "mensagem": mensagem,
        "removidos": removidos,
        "erros": erros,
        "total_removidos": len(removidos),
        "bytes": total_bytes,
    }


@app.route("/api/cleanup-operacional", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_cleanup_operacional_v82117():
    resultado = _executar_cleanup_operacional(request.get_json(silent=True) or {}, validar_job=True)
    status_code = int(resultado.pop("status_code", 200) or 200)
    return jsonify(resultado), status_code


def _calcular_extrato_banco_horas_colaborador(df_colab):
    """
    Gera extrato diário do Banco de Horas no ciclo 15 -> 14.

    Regra operacional consolidada:
    - O ciclo de apuração começa no dia 15 e termina no dia 14 do mês seguinte.
    - A coluna Saldo Informado mostra exatamente o campo Saldo Atual vindo da folha,
      dia a dia, sem cálculo.
    - A coluna Saldo Calculado parte do Saldo Atual informado no primeiro registro
      disponível dentro do ciclo e evolui diariamente por:
          Saldo Calculado = Saldo Base + Crédito H.E. - Débito Ausência
    - O campo Saldo Anterior da folha não é usado para abrir o ciclo.
    """
    resultado_vazio = {
        "resumo": {
            "nome": "",
            "cc": "",
            "funcao": "",
            "turno": "",
            "data_base": "",
            "inicio_ciclo": "",
            "fim_ciclo": "",
            "saldo_base": "00,00",
            "credito_total": "00:00",
            "debito_total": "00:00",
            "saldo_final": "00,00",
            "saldo_informado_final": "",
            "saldo_pdf_final": "",
            "origem_saldo_base": "saldo_atual_primeiro_dia_ciclo",
        },
        "linhas": []
    }

    if df_colab.empty:
        return resultado_vazio

    g = df_colab.copy()
    g["data_dt"] = pd.to_datetime(g["data"], format="%d/%m/%Y", errors="coerce")
    g = g[g["data_dt"].notna()].sort_values("data_dt").copy()
    if g.empty:
        return resultado_vazio

    inicio = _data_inicio_medicao(g["data_dt"].max())
    if inicio is None:
        inicio = g["data_dt"].min()

    fim = _data_fim_medicao(inicio) or (inicio + pd.Timedelta(days=30))

    g_ciclo = g[(g["data_dt"] >= inicio) & (g["data_dt"] <= fim)].sort_values("data_dt").copy()
    if g_ciclo.empty:
        g_ciclo = g.copy()
        inicio = g_ciclo["data_dt"].min()
        fim = g_ciclo["data_dt"].max()

    # v8.14.00: se o gestor configurou Banco de Horas 2.0, este motor prevalece
    # sobre o saldo calculado do Excel para Banco/Extrato, pois a data e o saldo
    # inicial passam a ser uma decisão operacional auditável.
    _bh2 = _calcular_extrato_bh2_colaborador(g_ciclo)
    if _bh2 is not None:
        return _bh2

    # v8.11.17: se o Excel já trouxe Saldo Calculado, ele é a fonte oficial.
    # Não recalcula, não usa cache PDF e não reinicia por fotografia intermediária.
    def _valor_excel_ok(valor):
        texto = str(valor or "").strip()
        return texto and texto.lower() not in {"nan", "none", "nat"}

    if "saldo_calculado" in g_ciclo.columns and g_ciclo["saldo_calculado"].apply(_valor_excel_ok).any():
        linhas = []
        credito_total_min = 0
        debito_total_min = 0
        saldo_final = ""
        saldo_base = ""
        saldo_informado_final = ""
        primeira_fotografia_por_data = _primeira_fotografia_por_dia_extrato_bh(g_ciclo)

        for idx, (_, row) in enumerate(g_ciclo.sort_values("data_dt").iterrows()):
            he_min = hhmm_para_minutos(row.get("he", ""))
            absent_min = hhmm_para_minutos(row.get("absent", ""))
            credito_total_min += he_min
            debito_total_min += absent_min

            saldo_calc = str(row.get("saldo_calculado", "") or "").strip()
            if _valor_excel_ok(saldo_calc):
                if not saldo_base:
                    saldo_base = saldo_calc
                saldo_final = saldo_calc

            saldo_info = ""
            # Para Saldo Informado, prioriza a fotografia por emissão quando houver;
            # na ausência dela, usa o Saldo Atual do PDF importado no Excel.
            for _col_info in ("saldo_informado_pdf", "saldo_atual", "saldo"):
                val = str(row.get(_col_info, "") or "").strip()
                if _valor_excel_ok(val):
                    saldo_info = val
                    break
            if _valor_excel_ok(saldo_info):
                saldo_informado_final = saldo_info

            data_txt_linha = str(row.get("data", "") or "").strip()
            info_primeira = primeira_fotografia_por_data.get(data_txt_linha, {}) if isinstance(primeira_fotografia_por_data, dict) else {}
            emissao_primeira = str(info_primeira.get("data_emissao_pdf", "") or "").strip()
            saldo_info_primeira = str(info_primeira.get("saldo_informado", "") or "").strip()
            if saldo_info_primeira:
                saldo_info = saldo_info_primeira
                saldo_informado_final = saldo_info

            linhas.append({
                "data": row.get("data", ""),
                "data_emissao_pdf": emissao_primeira or row.get("data_emissao_pdf", ""),
                "dia": row.get("dia", ""),
                "evento": "Saldo Base" if idx == 0 else "Movimento",
                "he": minutos_para_hhmm(he_min) if he_min else "",
                "absent": minutos_para_hhmm(absent_min) if absent_min else "",
                "saldo_calculado": saldo_calc,
                "saldo_informado": saldo_info,
                "saldo_pdf": saldo_info,
                "observacao": str(row.get("observacao", "") or ""),
                "status": str(row.get("status", "") or ""),
                "inconsistencias": str(row.get("inconsistencias", "") or ""),
            })

        resumo = {
            "nome": str(g_ciclo["nome"].iloc[-1]) if "nome" in g_ciclo.columns else "",
            "cc": str(g_ciclo["cc"].iloc[-1]) if "cc" in g_ciclo.columns else "",
            "funcao": str(g_ciclo["funcao"].iloc[-1]) if "funcao" in g_ciclo.columns else "",
            "turno": str(g_ciclo["turno"].iloc[-1]) if "turno" in g_ciclo.columns else "",
            "data_base": pd.to_datetime(g_ciclo["data_dt"].min()).strftime("%d/%m/%Y"),
            "inicio_ciclo": pd.to_datetime(inicio).strftime("%d/%m/%Y"),
            "fim_ciclo": pd.to_datetime(fim).strftime("%d/%m/%Y"),
            "saldo_base": saldo_base or "",
            "credito_total": minutos_para_hhmm(int(credito_total_min)),
            "debito_total": minutos_para_hhmm(int(debito_total_min)),
            "saldo_final": saldo_final or "",
            "saldo_informado_final": saldo_informado_final,
            "saldo_pdf_final": saldo_informado_final,
            "origem_saldo_base": "excel_saldo_calculado_oficial",
            "fonte_calculo": "excel_oficial_completo_sem_consolidacao_dashboard",
            "pdf_base_banco_horas": "",
            "periodo_base_banco_horas": "",
        }
        return {"resumo": resumo, "linhas": linhas}

    # Primeiro registro disponível dentro do ciclo.
    # v8.10.25: para Banco de Horas, a base confiável não deve vir do Excel
    # quando houver cache de PDFs. A base correta do ciclo atual é o Saldo Atual
    # do PDF de fechamento do ciclo anterior. Ex.: ciclo 15/05-14/06 abre com
    # o Saldo Atual do espelho 15/04-14/05.
    base = g_ciclo.iloc[0]
    data_base_dt = base["data_dt"]
    coluna_saldo_atual = "saldo_atual" if "saldo_atual" in g_ciclo.columns and str(base.get("saldo_atual", "")).strip() else "saldo"
    origem_saldo_importado = str(base.get("origem_saldo_importado", "") or "")
    nome_base_pdf = str(base.get("nome", "") or (g_ciclo["nome"].iloc[0] if "nome" in g_ciclo.columns else ""))
    matricula_base_pdf = _matricula_da_linha_bh(base)
    try:
        cache_bh = carregar_cache_pdfs_banco_horas()
        catalogo_bh = cache_bh.get("catalogo", []) if cache_bh else []
    except Exception:
        catalogo_bh = []

    # v8.11.12: se o cache estiver vazio/desatualizado para este colaborador,
    # monta um mini-catálogo emergencial apenas para o Extrato selecionado.
    if not _catalogo_tem_identidade_bh(catalogo_bh, nome_base_pdf, matricula_base_pdf):
        catalogo_extra = _catalogo_emergencial_colaborador_bh(nome_base_pdf, matricula_base_pdf, inicio, fim, max_segundos=55)
        if catalogo_extra:
            catalogo_bh = list(catalogo_bh or []) + catalogo_extra

    # v8.11.13: pode existir cache com o PDF do ciclo anterior suficiente
    # para abrir o Saldo Calculado, mas sem nenhuma fotografia do ciclo atual
    # para preencher o Saldo Informado diário. Nesse caso, a identidade existe
    # no catálogo e a v8.11.12 não disparava a busca emergencial. Agora fazemos
    # uma checagem específica: se nenhuma data do ciclo encontra fotografia
    # no catálogo atual, buscamos de forma cirúrgica os PDFs do colaborador.
    try:
        tem_fotografia_atual = False
        for _dt_check in list(g_ciclo["data_dt"].dropna().unique()):
            _pdf_check = _saldo_pdf_para_data_exata(nome_base_pdf, matricula_base_pdf, _dt_check, catalogo_bh)
            if _pdf_check and str(_pdf_check.get("saldo_atual_pdf", "") or "").strip():
                tem_fotografia_atual = True
                break
        if not tem_fotografia_atual:
            catalogo_extra = _catalogo_emergencial_colaborador_bh(nome_base_pdf, matricula_base_pdf, inicio, fim, max_segundos=60)
            if catalogo_extra:
                # Evita duplicar o mesmo PDF/quadro quando já existir no cache.
                vistos = set()
                catalogo_merged = []
                for _item in list(catalogo_bh or []) + list(catalogo_extra or []):
                    chave = (str(_item.get("arquivo_pdf", "")), str(_item.get("matricula_norm", "")), str(_item.get("nome_norm", "")), str(_item.get("inicio", "")), str(_item.get("fim", "")), str(_item.get("saldo_atual_pdf", "")))
                    if chave in vistos:
                        continue
                    vistos.add(chave)
                    catalogo_merged.append(_item)
                catalogo_bh = catalogo_merged
    except Exception:
        pass

    # v8.10.45: Extrato em modo estável usa somente o histórico/cache de PDFs.
    # A busca direta em disco ficava lenta e contaminava a navegação entre guias.
    # Para incluir PDFs novos, use primeiro Configuração do Robô > Atualizar Histórico de PDFs.
    pdf_base = _buscar_pdf_fechamento_ciclo_anterior(nome_base_pdf, matricula_base_pdf, inicio, catalogo_bh, permitir_busca_direta=False)
    if not pdf_base and not _catalogo_tem_identidade_bh(catalogo_bh, nome_base_pdf, matricula_base_pdf):
        pdf_base = _buscar_pdf_fechamento_ciclo_anterior(nome_base_pdf, matricula_base_pdf, inicio, catalogo_bh, permitir_busca_direta=True)
    # v8.10.27: regra dura para o Extrato Banco de Horas.
    # O saldo do Excel NÃO abre mais o ciclo de forma silenciosa.
    # Se o PDF de fechamento do ciclo anterior não for encontrado, o Extrato
    # sinaliza a ausência em origem_saldo_base e inicia o cálculo em 00,00
    # apenas para não quebrar a tela.
    #
    # v8.10.35: fallback seguro pelo próprio PDF da data-base.
    # Nos testes reais, o Saldo Informado já encontrou a fotografia que contém
    # 15/05/2026 com Saldo Atual 263,48, mas a busca do PDF de fechamento do
    # ciclo anterior pode falhar se o cache ainda estiver focado só no ciclo
    # atual. Como esse saldo vem do PDF e não do Excel, ele pode abrir também
    # o Saldo Calculado quando o PDF anterior não estiver catalogado.
    pdf_base_encontrado = bool(pdf_base and str(pdf_base.get("saldo_atual_pdf", "") or "").strip())
    pdf_base_por_data = None
    if pdf_base_encontrado:
        saldo_base_informado = str(pdf_base.get("saldo_atual_pdf", "") or "").strip()
        origem_saldo_importado = f"pdf_fechamento_ciclo_anterior_{pdf_base.get('_match_base_pdf_v25','pdf')}: {pdf_base.get('arquivo_pdf','')}"
    else:
        pdf_base_por_data = _saldo_pdf_para_data_exata(nome_base_pdf, matricula_base_pdf, data_base_dt, catalogo_bh)
        if pdf_base_por_data and str(pdf_base_por_data.get("saldo_atual_pdf", "") or "").strip():
            saldo_base_informado = str(pdf_base_por_data.get("saldo_atual_pdf", "") or "").strip()
            origem_saldo_importado = f"pdf_data_base_sem_fallback_excel: {pdf_base_por_data.get('arquivo_pdf','')}"
            # Mantém o objeto para diagnóstico e para exibir na linha de saldo base.
            pdf_base = {**pdf_base_por_data, "_match_base_pdf_v35": pdf_base_por_data.get("_match_diario_pdf_v25", "pdf_data_base")}
            pdf_base_encontrado = True
        else:
            saldo_base_informado = ""
            origem_saldo_importado = "pdf_fechamento_ciclo_anterior_nao_encontrado_sem_fallback_excel"

    saldo_corrente_min = _saldo_para_minutos_banco(saldo_base_informado) if saldo_base_informado else 0

    credito_total_min = 0
    debito_total_min = 0

    linhas = []
    for _, row in g_ciclo.iterrows():
        data_dt = row["data_dt"]
        he_min = hhmm_para_minutos(row.get("he", ""))
        absent_min = hhmm_para_minutos(row.get("absent", ""))
        # v8.10.32: Saldo Informado separado do Saldo Calculado.
        # - A linha de Saldo Base continua exibindo o fechamento do ciclo anterior
        #   (ex.: PDF 15/04 a 14/05, Saldo Atual 263,48).
        # - As demais linhas tentam mostrar a fotografia do ciclo atual a partir
        #   do cache de PDFs (ex.: PDF 15/05 a 05/06, Saldo Atual 431,00).
        # - Não há varredura direta de PDFs aqui; usa apenas o histórico/cache já
        #   atualizado na guia Configuração do Robô, mantendo a tela estável.
        saldo_informado = ""

        # v8.11.17: prioridade máxima para o saldo informado importado
        # do Excel, respeitando a Data de Emissão do PDF.
        # Importante: quando a coluna Saldo Informado PDF existe, valores
        # vazios são intencionais e não devem disparar fallback por cache,
        # pois isso voltaria a repetir/sobrescrever saldos em datas indevidas.
        saldo_pdf_excel_presente = str(row.get("saldo_informado_pdf_excel_presente", "") or "").strip() == "1"
        for _col_info in ("saldo_informado_pdf", "saldo_informado", "saldo_atual_pdf", "saldo atual pdf", "saldo informado"):
            try:
                _v = str(row.get(_col_info, "") or "").strip()
                if _v and _v.lower() not in ("nan", "none", "nat"):
                    saldo_informado = _v
                    break
            except Exception:
                pass

        if not saldo_informado:
            if saldo_pdf_excel_presente:
                saldo_informado = ""
            elif data_dt == data_base_dt and pdf_base_encontrado:
                saldo_informado = str(pdf_base.get("saldo_atual_pdf", "") or "").strip()
            else:
                # Fallback legado: usa fotografia do cache quando o Excel ainda
                # não trouxer Saldo Informado PDF.
                pdf_diario = _saldo_pdf_para_data_exata(nome_base_pdf, matricula_base_pdf, data_dt, catalogo_bh)
                if pdf_diario and str(pdf_diario.get("saldo_atual_pdf", "") or "").strip():
                    saldo_informado = str(pdf_diario.get("saldo_atual_pdf", "") or "").strip()
                else:
                    saldo_informado = "PDF da data não encontrado"

        if data_dt > data_base_dt:
            saldo_corrente_min += he_min
            saldo_corrente_min -= absent_min
            credito_total_min += he_min
            debito_total_min += absent_min
            evento = "Movimento"
        elif data_dt == data_base_dt:
            evento = "Saldo Base"
        else:
            evento = "Anterior à base"

        linhas.append({
            "data": row.get("data", ""),
            "data_emissao_pdf": row.get("data_emissao_pdf", ""),
            "dia": row.get("dia", ""),
            "evento": evento,
            "he": minutos_para_hhmm(he_min) if he_min else "",
            "absent": minutos_para_hhmm(absent_min) if absent_min else "",
            "saldo_calculado": _minutos_para_saldo_virgula(saldo_corrente_min),
            "saldo_informado": saldo_informado,
            # Mantido por compatibilidade com a tela, caso algum JS antigo ainda consulte saldo_pdf.
            "saldo_pdf": saldo_informado,
            "observacao": str(row.get("observacao", "") or ""),
            "status": str(row.get("status", "") or ""),
            "inconsistencias": str(row.get("inconsistencias", "") or ""),
        })

    # v8.10.27: saldo informado final deriva da última linha calculada pelo PDF,
    # não do Excel.
    saldo_informado_final = ""
    for _linha in reversed(linhas):
        val = str(_linha.get("saldo_informado", "") or "").strip()
        if val and val != "PDF não encontrado":
            saldo_informado_final = val
            break

    resumo = {
        "nome": str(g_ciclo["nome"].iloc[-1]) if "nome" in g_ciclo.columns else "",
        "cc": str(g_ciclo["cc"].iloc[-1]) if "cc" in g_ciclo.columns else "",
        "funcao": str(g_ciclo["funcao"].iloc[-1]) if "funcao" in g_ciclo.columns else "",
        "turno": str(g_ciclo["turno"].iloc[-1]) if "turno" in g_ciclo.columns else "",
        "data_base": pd.to_datetime(data_base_dt).strftime("%d/%m/%Y"),
        "inicio_ciclo": pd.to_datetime(inicio).strftime("%d/%m/%Y"),
        "fim_ciclo": pd.to_datetime(fim).strftime("%d/%m/%Y"),
        "saldo_base": (saldo_base_informado if saldo_base_informado else "PDF não encontrado"),
        "credito_total": minutos_para_hhmm(int(credito_total_min)),
        "debito_total": minutos_para_hhmm(int(debito_total_min)),
        "saldo_final": _minutos_para_saldo_virgula(saldo_corrente_min),
        "saldo_informado_final": saldo_informado_final,
        # Mantido por compatibilidade com cards antigos.
        "saldo_pdf_final": saldo_informado_final,
        "origem_saldo_base": origem_saldo_importado,
        "pdf_base_banco_horas": str(pdf_base.get("arquivo_pdf", "") if pdf_base else ""),
        # v8.10.37: alguns caminhos de fallback por PDF da data-base
        # montam pdf_base sem _inicio_ciclo_anterior_v25/_fim_ciclo_anterior_v25.
        # Antes isso gerava: 'NoneType' object has no attribute 'strftime'.
        "periodo_base_banco_horas": _periodo_pdf_base_seguro(pdf_base),
    }

    return {"resumo": resumo, "linhas": linhas}




def _selecionar_saldos_bh2_por_origem(opcoes, origem_padrao, saldo_manual=""):
    """Seleciona o saldo inicial global do Banco de Horas 2.0 simplificado.

    v8.14.02: a operação deixa de ter exceções por colaborador. O gestor
    escolhe apenas uma data inicial e uma origem global:
      - saldo_atual: usa o campo Saldo Atual disponível para cada colaborador;
      - manual: aplica o mesmo saldo informado a todos os colaboradores.

    A função materializa cfg["saldos"] para manter compatibilidade com o motor
    de cálculo já existente.
    """
    origem_padrao = str(origem_padrao or "saldo_atual").strip().lower()
    saldo_manual = _normalizar_saldo_texto_bh2(saldo_manual)

    grupos = {}
    for o in opcoes or []:
        nome = str(o.get("nome", "") or "").strip()
        if not nome:
            continue
        chave = _norm_bh2(o.get("matricula") or nome)
        grupos.setdefault(chave, []).append(o)

    resultado = {}

    for chave, lista in grupos.items():
        if not lista:
            continue
        base = lista[0]
        nome = str(base.get("nome", "") or "").strip()
        matricula = str(base.get("matricula", "") or "").strip()
        cc = str(base.get("cc", "") or "").strip()

        if origem_padrao == "manual":
            if not saldo_manual:
                continue
            resultado[chave] = {
                "nome": nome,
                "matricula": matricula,
                "cc": cc,
                "saldo": saldo_manual,
                "origem": "manual_global_gestor",
                "data_emissao_pdf": "",
                "justificativa": "saldo manual global definido pelo gestor",
            }
            continue

        # Origem operacional padrão: Saldo Atual. Filtra pela origem textual
        # entregue pela aba Saldos Disponiveis BH; se não houver correspondência,
        # usa o primeiro saldo disponível como fallback explícito.
        candidatos_saldo_atual = [
            o for o in lista
            if "saldoatual" in _canon_bh2(o.get("origem", "")) or _canon_bh2(o.get("origem", "")) == "atual"
        ]
        candidatos = candidatos_saldo_atual or lista
        escolhido = sorted(
            candidatos,
            key=lambda o: (bool(o.get("recomendado")), str(o.get("data_emissao_pdf", ""))),
            reverse=True,
        )[0]
        saldo = _normalizar_saldo_texto_bh2(escolhido.get("saldo", ""))
        if not saldo:
            continue
        resultado[chave] = {
            "nome": str(escolhido.get("nome", nome) or nome).strip(),
            "matricula": str(escolhido.get("matricula", matricula) or matricula).strip(),
            "cc": str(escolhido.get("cc", cc) or cc).strip(),
            "saldo": saldo,
            "origem": "saldo_atual" if candidatos_saldo_atual else str(escolhido.get("origem", "saldo_disponivel") or "saldo_disponivel"),
            "data_emissao_pdf": str(escolhido.get("data_emissao_pdf", "") or "").strip(),
            "justificativa": "origem global: Saldo Atual" if candidatos_saldo_atual else "fallback: primeiro saldo disponível",
        }
    return resultado

def _normalizar_data_config_bh2(valor):
    texto = str(valor or "").strip()
    if not texto:
        return ""
    dt = pd.to_datetime(texto, errors="coerce", dayfirst=False)
    if pd.isna(dt):
        dt = pd.to_datetime(texto, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        return texto
    return pd.Timestamp(dt).strftime("%Y-%m-%d")

@app.route("/api/banco-horas-2/config", methods=["GET", "POST"])
@login_obrigatorio
def api_banco_horas_2_config():
    """Configuracao do Banco de Horas 2.0 por colaborador.

    v8.15.10: a Data Inicial continua sendo global, mas a origem do saldo e o
    saldo manual passam a ser configurados somente para o colaborador filtrado.
    Sem configuracao individual, o motor usa Saldo Atual na data inicial.
    """
    if request.method == "GET":
        cfg = _carregar_config_bh2()
        nome_req = str(request.args.get("nome", "") or "").strip()
        matricula_req = str(request.args.get("matricula", "") or "").strip()
        total_colaboradores = 0
        try:
            sig_atual = str(_assinatura_arquivo(ARQUIVO_PADRAO))
            if _CACHE_OPCOES_FILTROS_MEM.get("sig") == sig_atual:
                total_colaboradores = int(len(_CACHE_OPCOES_FILTROS_MEM.get("nomes") or []))
        except Exception:
            total_colaboradores = 0

        cfg_individual = _saldo_config_bh2_para_colaborador(nome_req, matricula_req) if nome_req or matricula_req else None
        return jsonify({
            "erro": "",
            "config": cfg,
            "config_colaborador": cfg_individual or {},
            "colaborador_selecionado": nome_req,
            "total_opcoes": 0,
            "total_colaboradores": total_colaboradores,
            "modo": "bh2_por_colaborador",
        })

    try:
        payload = request.get_json(silent=True) or {}
        cfg = _carregar_config_bh2()
        data_inicio = _normalizar_data_config_bh2(payload.get("data_inicio", cfg.get("data_inicio", "")))
        nome = str(payload.get("nome", "") or "").strip()
        matricula = str(payload.get("matricula", "") or "").strip()
        origem = str(payload.get("origem", payload.get("origem_padrao", "saldo_atual")) or "saldo_atual").strip().lower()
        if origem not in {"saldo_atual", "manual"}:
            origem = "saldo_atual"
        saldo_manual = _normalizar_saldo_texto_bh2(payload.get("saldo_manual", ""))

        # v8.15.22: Data Inicial pode ficar em branco. Nesse caso, o motor BH 2.0
        # considera todo o periodo disponivel para cada colaborador.
        saldos = cfg.get("saldos", {}) if isinstance(cfg.get("saldos", {}), dict) else {}

        # v8.15.21: a Data Inicial e uma configuracao global do motor BH 2.0.
        # Ela pode ser salva sozinha, sem colaborador selecionado. A origem/saldo
        # manual continuam sendo regras individuais, aplicadas apenas quando um
        # colaborador for informado.
        if nome or matricula:
            if origem == "manual" and not saldo_manual:
                return jsonify({"ok": False, "erro": "Informe o saldo manual no formato HH,MM para este colaborador."}), 400
            chave = matricula or _norm_bh2(nome)
            if origem == "saldo_atual":
                saldo_manual = ""
            saldos[chave] = {
                "nome": nome,
                "matricula": matricula,
                "origem": origem,
                "saldo": saldo_manual,
                "atualizado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }

        cfg.update({
            "ativo": True,
            "data_inicio": data_inicio,
            "origem_padrao": "saldo_atual",
            "saldo_manual": "",
            "saldos": saldos,
        })
        cfg = _salvar_config_bh2(cfg)
        invalidar_cache_consolidado()
        try:
            if ARQUIVO_RESUMO_BANCO_HORAS_PDF.exists():
                ARQUIVO_RESUMO_BANCO_HORAS_PDF.unlink()
        except Exception:
            pass

        try:
            gerar_resumo_banco_horas_pdf()
        except Exception as e:
            LOGGER.warning(f"Configuração BH 2.0 por colaborador salva, mas o resumo imediato falhou: {e}")

        origem_label = "Saldo Atual" if origem == "saldo_atual" else "Valor Manual"
        if nome or matricula:
            periodo_label = data_inicio or "todo o periodo"
            mensagem = f"Banco de Horas atualizado para {nome or matricula}: origem {origem_label}, periodo {periodo_label}."
        else:
            mensagem = f"Data inicial do Banco de Horas atualizada para {data_inicio or 'todo o periodo'}."
        return jsonify({
            "ok": True,
            "mensagem": mensagem,
            "config": cfg,
        })
    except Exception as e:
        LOGGER.exception(f"Falha ao salvar Banco de Horas 2.0 por colaborador: {e}")
        return jsonify({"ok": False, "erro": f"Falha interna ao atualizar Banco de Horas 2.0: {e}"}), 500

@app.route("/api/banco-horas-2/opcoes-saldos")
@login_obrigatorio
def api_banco_horas_2_opcoes_saldos():
    opcoes, erro = _opcoes_saldos_disponiveis_bh()
    nome = str(request.args.get("nome", "") or "").strip().lower()
    cc = str(request.args.get("cc", "") or "").strip()
    if nome:
        opcoes = [o for o in opcoes if nome in str(o.get("nome", "")).lower()]
    if cc:
        cc_norm = normalizar_cc_filtro_backend(cc)
        opcoes = [o for o in opcoes if normalizar_cc_filtro_backend(o.get("cc", "")) == cc_norm]
    return jsonify({"erro": erro, "opcoes": opcoes[:2000], "total": len(opcoes)})


@app.route("/api/opcoes-extrato-banco-horas")
@login_obrigatorio
def api_opcoes_extrato_banco_horas():
    """Opções leves e sempre atualizadas para a guia Extrato Banco de Horas.

    Esta rota existe para não depender do cache global de filtros. Na prática,
    quando a base é reimportada ou a guia é aberta diretamente, o combo de
    colaboradores precisa ser recarregado a partir do Consolidado atual.
    """
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "ccs": [], "nomes": [], "turnos": []})

    df = normalizar_colunas(df)
    if df is None or df.empty:
        return jsonify({"erro": "", "ccs": [], "nomes": [], "turnos": []})

    cc_filtro = request.args.get("cc", "").strip()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data = request.args.get("data", "").strip()
    datas = request.args.get("datas", "").strip()

    # Lista de CCs vem da base inteira, para o usuário conseguir trocar o recorte.
    ccs = []
    if "cc" in df.columns:
        ccs = sorted(
            set(valor_cc_display(v) for v in df["cc"].dropna().tolist() if valor_cc_display(v)),
            key=lambda x: x.lower(),
        )

    df_op = df.copy()

    if cc_filtro and "cc" in df_op.columns:
        df_op = aplicar_filtro_cc_backend(df_op, cc_filtro)

    if turno_filtro and "turno" in df_op.columns:
        df_op = df_op[df_op["turno"].astype(str).str.lower() == turno_filtro]

    if "data" in df_op.columns:
        df_op = df_op.copy()
        df_op["_data_iso_extrato_opcoes"] = pd.to_datetime(
            df_op["data"].astype(str).str.strip(),
            format="%d/%m/%Y",
            errors="coerce",
        ).dt.strftime("%Y-%m-%d")
        if datas:
            lista_datas = [d.strip() for d in datas.split(",") if d.strip()]
            if lista_datas:
                df_op = df_op[df_op["_data_iso_extrato_opcoes"].isin(lista_datas)].copy()
        elif data:
            df_op = df_op[df_op["_data_iso_extrato_opcoes"] == data].copy()
        df_op.drop(columns=["_data_iso_extrato_opcoes"], inplace=True, errors="ignore")

    nomes = []
    if "nome" in df_op.columns:
        nomes = sorted(
            set(
                str(v).strip()
                for v in df_op["nome"].dropna().tolist()
                if str(v).strip() and str(v).strip().lower() != "nan"
            ),
            key=lambda x: x.lower(),
        )

    # Turnos respeitam CC e data, mas não precisam respeitar o turno atual.
    df_turnos = df.copy()
    if cc_filtro and "cc" in df_turnos.columns:
        df_turnos = aplicar_filtro_cc_backend(df_turnos, cc_filtro)
    if "data" in df_turnos.columns:
        df_turnos = df_turnos.copy()
        df_turnos["_data_iso_extrato_opcoes"] = pd.to_datetime(
            df_turnos["data"].astype(str).str.strip(),
            format="%d/%m/%Y",
            errors="coerce",
        ).dt.strftime("%Y-%m-%d")
        if datas:
            lista_datas = [d.strip() for d in datas.split(",") if d.strip()]
            if lista_datas:
                df_turnos = df_turnos[df_turnos["_data_iso_extrato_opcoes"].isin(lista_datas)].copy()
        elif data:
            df_turnos = df_turnos[df_turnos["_data_iso_extrato_opcoes"] == data].copy()
        df_turnos.drop(columns=["_data_iso_extrato_opcoes"], inplace=True, errors="ignore")

    turnos = []
    if "turno" in df_turnos.columns:
        turnos = sorted(
            set(
                str(v).strip()
                for v in df_turnos["turno"].dropna().tolist()
                if str(v).strip() and str(v).strip().lower() != "nan"
            ),
            key=lambda x: x.lower(),
        )

    return jsonify({
        "erro": "",
        "ccs": ccs,
        "nomes": nomes,
        "turnos": turnos,
        "total_nomes": len(nomes),
    })


@app.route("/api/extrato-banco-horas")
@login_obrigatorio
def api_extrato_banco_horas():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "resumo": {}, "dados": []})

    df = normalizar_colunas(df)
    if df.empty:
        return jsonify({"erro": "Nenhuma base carregada.", "resumo": {}, "dados": []})

    # v8.11.27: a guia Extrato Banco de Horas agora respeita os filtros
    # principais antes de calcular o extrato. Antes ela usava apenas nome,
    # então CC/turno/data pareciam não produzir efeito na guia.
    cc_filtro = request.args.get("cc", "").strip()
    nome = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data = request.args.get("data", "").strip()
    datas = request.args.get("datas", "").strip()

    if cc_filtro and "cc" in df.columns:
        df = aplicar_filtro_cc_backend(df, cc_filtro)

    if turno_filtro and "turno" in df.columns:
        df = df[df["turno"].astype(str).str.lower() == turno_filtro]

    # v8.11.28: os filtros de data vindos do HTML input[type=date] chegam
    # no formato ISO (yyyy-mm-dd), enquanto o Excel oficial traz Data em
    # dd/mm/aaaa. A v8.11.27 comparava texto com texto e, por isso, o filtro
    # de data/datas do Extrato Banco de Horas não encontrava linhas.
    cfg_bh2_extrato = _carregar_config_bh2()
    bh2_ativo_extrato = bool(cfg_bh2_extrato.get("data_inicio"))

    # Quando BH 2.0 está ativo, o filtro de data deve ser aplicado somente na
    # visualização final. O cálculo precisa enxergar todo o histórico desde a
    # data inicial do gestor, senão o saldo é aberto no dia filtrado e fica errado.
    if not bh2_ativo_extrato:
        if "data" in df.columns:
            df = df.copy()
            df["_data_iso_extrato_web"] = pd.to_datetime(
                df["data"].astype(str).str.strip(),
                format="%d/%m/%Y",
                errors="coerce",
            ).dt.strftime("%Y-%m-%d")

        if datas and "_data_iso_extrato_web" in df.columns:
            lista_datas = [d.strip() for d in datas.split(",") if d.strip()]
            if lista_datas:
                df = df[df["_data_iso_extrato_web"].isin(lista_datas)].copy()
        elif data and "_data_iso_extrato_web" in df.columns:
            df = df[df["_data_iso_extrato_web"] == data].copy()

        df.drop(columns=["_data_iso_extrato_web"], inplace=True, errors="ignore")

    if df.empty:
        return jsonify({"erro": "Nenhum registro encontrado para os filtros selecionados.", "resumo": {}, "dados": []})

    def _identidade_linha(row):
        return {
            "nome": str(row.get("nome", "") or "").strip(),
            "cc": str(row.get("cc", "") or "").strip(),
            "turno": str(row.get("turno", "") or "").strip(),
            "funcao": str(row.get("funcao", "") or "").strip(),
        }

    def _hhmm_total(valor):
        try:
            return int(hhmm_para_minutos(valor) or 0)
        except Exception:
            return 0

    def _extrato_para_grupo(df_grupo):
        extrato_local = _calcular_extrato_banco_horas_colaborador(df_grupo)
        identidade = _identidade_linha(df_grupo.iloc[0]) if not df_grupo.empty else {}
        resumo_local = extrato_local.get("resumo", {}) or {}
        for k, v in identidade.items():
            resumo_local.setdefault(k, v)
        linhas_local = []
        for linha in extrato_local.get("linhas", []) or []:
            nova = dict(linha)
            for k, v in identidade.items():
                nova.setdefault(k, v)
            linhas_local.append(nova)

        # BH 2.0: filtra as linhas exibidas por data somente depois do cálculo,
        # preservando o saldo acumulado correto desde a data inicial do gestor.
        if bh2_ativo_extrato and (data or datas):
            datas_alvo = set()
            if datas:
                datas_alvo.update(d.strip() for d in datas.split(",") if d.strip())
            elif data:
                datas_alvo.add(data)
            if datas_alvo:
                filtradas_data = []
                for item in linhas_local:
                    iso = pd.to_datetime(str(item.get("data", "")), format="%d/%m/%Y", errors="coerce")
                    iso_txt = iso.strftime("%Y-%m-%d") if pd.notna(iso) else ""
                    if iso_txt in datas_alvo:
                        filtradas_data.append(item)
                linhas_local = filtradas_data
        return resumo_local, linhas_local

    if nome:
        df_nome = df[df["nome"].astype(str).str.lower() == nome].copy()
        if df_nome.empty:
            # fallback tolerante
            df_nome = df[df["nome"].astype(str).str.lower().str.contains(nome, na=False)].copy()

        if df_nome.empty:
            return jsonify({"erro": "Colaborador não encontrado na base para os filtros selecionados.", "resumo": {}, "dados": []})

        # Se vierem vários por contains, escolhe o primeiro nome exato disponível
        # dentro da base já filtrada.
        nome_real = str(df_nome["nome"].iloc[0])
        df_nome = df[df["nome"] == nome_real].copy()
        if bh2_ativo_extrato:
            try:
                cache_bh2 = _obter_cache_bh2_calculo()
                mat_real = str(df_nome["matricula"].iloc[0]) if "matricula" in df_nome.columns and not df_nome.empty else ""
                chave_cache = _chave_bh2_colaborador(nome_real, mat_real)
                extr_cache = (cache_bh2.get("extratos", {}) or {}).get(chave_cache)
                if not extr_cache:
                    # fallback por nome para bases sem matrícula confiável
                    for _ch, _ex in (cache_bh2.get("extratos", {}) or {}).items():
                        _r = _ex.get("resumo", {}) or {}
                        if _norm_bh2(_r.get("nome", "")) == _norm_bh2(nome_real):
                            extr_cache = _ex
                            break
                if extr_cache:
                    resumo = dict(extr_cache.get("resumo", {}) or {})
                    linhas = [dict(x) for x in (extr_cache.get("linhas", []) or [])]
                    if data or datas:
                        datas_alvo = set()
                        if datas:
                            datas_alvo.update(d.strip() for d in datas.split(",") if d.strip())
                        elif data:
                            datas_alvo.add(data)
                        if datas_alvo:
                            linhas = [x for x in linhas if (pd.to_datetime(str(x.get("data", "")), format="%d/%m/%Y", errors="coerce").strftime("%Y-%m-%d") if pd.notna(pd.to_datetime(str(x.get("data", "")), format="%d/%m/%Y", errors="coerce")) else "") in datas_alvo]
                    return jsonify({"erro": "", "resumo": resumo, "dados": linhas, "fonte": "bh2_cacheado_v8_15_23"})
            except Exception as exc:
                LOGGER.warning(f"Falha ao usar cache BH 2.0 no Extrato: {exc}")
        resumo, linhas = _extrato_para_grupo(df_nome)
        return jsonify({"erro": "", "resumo": resumo, "dados": linhas})

    nomes_unicos = sorted(set(str(v).strip() for v in df["nome"].dropna().tolist() if str(v).strip()))

    # v8.11.30: blindagem de performance. Calcular o extrato detalhado de
    # dezenas/centenas de colaboradores de uma vez trava o navegador e também
    # pesa no backend. Sem colaborador selecionado, só abrimos automaticamente
    # quando a seleção ficou pequena. Caso contrário, devolvemos uma resposta
    # leve orientando o usuário a refinar o filtro.
    try:
        limite_colaboradores = int(request.args.get("max_colaboradores", "20") or 20)
    except Exception:
        limite_colaboradores = 20
    limite_colaboradores = max(1, min(limite_colaboradores, 50))

    try:
        limite_linhas = int(request.args.get("limit", "500") or 500)
    except Exception:
        limite_linhas = 500
    limite_linhas = max(50, min(limite_linhas, 2000))

    if len(nomes_unicos) == 1:
        df_nome = df[df["nome"] == nomes_unicos[0]].copy()
        resumo, linhas = _extrato_para_grupo(df_nome)
        total_linhas = len(linhas)
        linhas = linhas[:limite_linhas]
        return jsonify({
            "erro": "",
            "resumo": resumo,
            "dados": linhas,
            "total_linhas": total_linhas,
            "limit": limite_linhas,
            "truncado": total_linhas > len(linhas),
        })

    resumo_leve = {
        "nome": f"{len(nomes_unicos)} colaboradores",
        "cc": cc_filtro or "Todos os CCs",
        "funcao": "",
        "turno": turno_filtro or "Todos os turnos",
        "data_base": data or datas or "filtros aplicados",
        "inicio_ciclo": "",
        "fim_ciclo": "",
        "saldo_base": "por colaborador",
        "credito_total": "--",
        "debito_total": "--",
        "saldo_final": "por colaborador",
        "saldo_informado_final": "",
        "saldo_pdf_final": "",
        "origem_saldo_base": f"seleção ampla: {len(nomes_unicos)} colaborador(es)",
    }

    if len(nomes_unicos) > limite_colaboradores:
        return jsonify({
            "erro": "",
            "resumo": resumo_leve,
            "dados": [],
            "total_colaboradores": len(nomes_unicos),
            "limite_colaboradores": limite_colaboradores,
            "aviso": (
                f"Filtro amplo demais para detalhar sem travar: {len(nomes_unicos)} colaborador(es). "
                "Selecione um colaborador ou reduza por CC, turno e data. "
                f"A abertura automática detalhada está limitada a {limite_colaboradores} colaborador(es)."
            ),
        })

    linhas_todas = []
    credito_total = 0
    debito_total = 0
    chaves = ["nome"]
    for extra in ["cc", "turno"]:
        if extra in df.columns:
            chaves.append(extra)

    for _, grupo in df.groupby(chaves, dropna=False, sort=True):
        if grupo.empty:
            continue
        resumo_g, linhas_g = _extrato_para_grupo(grupo.copy())
        linhas_todas.extend(linhas_g)
        credito_total += _hhmm_total(resumo_g.get("credito_total", ""))
        debito_total += _hhmm_total(resumo_g.get("debito_total", ""))
        if len(linhas_todas) >= limite_linhas:
            break

    try:
        linhas_todas = sorted(
            linhas_todas,
            key=lambda r: (pd.to_datetime(str(r.get("data", "")), format="%d/%m/%Y", errors="coerce"), str(r.get("nome", ""))),
        )
    except Exception:
        pass

    total_linhas = len(linhas_todas)
    linhas_saida = linhas_todas[:limite_linhas]

    resumo_multi = dict(resumo_leve)
    resumo_multi.update({
        "credito_total": minutos_para_hhmm(int(credito_total)),
        "debito_total": minutos_para_hhmm(int(debito_total)),
        "origem_saldo_base": f"extrato consolidado limitado: {len(nomes_unicos)} colaborador(es)",
    })

    return jsonify({
        "erro": "",
        "resumo": resumo_multi,
        "dados": linhas_saida,
        "total_linhas": total_linhas,
        "limit": limite_linhas,
        "truncado": total_linhas > len(linhas_saida),
        "aviso": f"Exibindo extrato consolidado limitado para {len(nomes_unicos)} colaborador(es). Para auditoria completa, selecione um colaborador.",
    })






# ============================================================
# ALERTAS AUTOMÁTICOS - V8.7.0
# ============================================================
def _aplicar_filtros_executivos(df):
    """Reaproveita os filtros principais para Dashboard Executivo e Alertas."""
    cc_filtro = request.args.get("cc", "").strip()
    nome_filtro = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data = request.args.get("data", "").strip()
    datas = request.args.get("datas", "").strip()

    # v8.11.21: se o usuário clicar em outras guias durante o carregamento,
    # o Dashboard não deve recalcular tudo repetidas vezes.
    if cc_filtro:
        df = aplicar_filtro_cc_backend(df, cc_filtro)
    if nome_filtro and "nome" in df.columns:
        df = df[df["nome"].astype(str).str.lower() == nome_filtro]
    if turno_filtro and "turno" in df.columns:
        df = df[df["turno"].astype(str).str.lower() == turno_filtro]

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df.get("data", ""), format="%d/%m/%Y", errors="coerce")

    if datas:
        datas_lista = [d.strip() for d in datas.split(",") if d.strip()]
        if datas_lista:
            data_iso = df["data_dt"].dt.strftime("%Y-%m-%d")
            df = df[data_iso.isin(datas_lista)].copy()
    elif data:
        data_iso = df["data_dt"].dt.strftime("%Y-%m-%d")
        df = df[data_iso == data].copy()
    return df


def _nivel_alerta_por_score(score):
    if score >= 90:
        return "crítico"
    if score >= 55:
        return "atenção"
    return "monitoramento"


def _classe_alerta_por_score(score):
    if score >= 90:
        return "danger"
    if score >= 55:
        return "warning"
    return "info"


# v8.21.4 - cache leve para a Central de Alertas Gerenciais.
# A guia era recalculada integralmente a cada clique. Em bases maiores, isso
# deixava a navegação com cara de porta emperrada. O cache usa assinatura do
# Excel + filtros e expira rapidamente, preservando atualização sem travar a UX.
_CACHE_ALERTAS_AUTOMATICOS_V8212 = {"key": None, "payload": None, "ts": 0}

def _cache_key_alertas_automaticos_v8212():
    try:
        excel_sig = str(_assinatura_arquivo(ARQUIVO_PADRAO))
    except Exception:
        excel_sig = "sem_assinatura"
    return (
        excel_sig,
        request.args.get("cc", "").strip(),
        request.args.get("nome", "").strip().lower(),
        request.args.get("turno", "").strip().lower(),
        request.args.get("data", "").strip(),
        request.args.get("datas", "").strip(),
    )


@app.route("/api/alertas-automaticos")
@login_obrigatorio
def api_alertas_automaticos():
    cache_key = _cache_key_alertas_automaticos_v8212()
    try:
        if (_CACHE_ALERTAS_AUTOMATICOS_V8212.get("key") == cache_key
                and _CACHE_ALERTAS_AUTOMATICOS_V8212.get("payload")
                and (time.monotonic() - float(_CACHE_ALERTAS_AUTOMATICOS_V8212.get("ts") or 0)) < 180):
            return jsonify(_CACHE_ALERTAS_AUTOMATICOS_V8212["payload"])
    except Exception:
        pass

    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "kpis": {}, "categorias": [], "colaboradores": [], "ocorrencias": [], "recomendacoes": []})

    df = normalizar_colunas(df)
    if df.empty:
        return jsonify({"erro": "Nenhuma base carregada.", "kpis": {}, "categorias": [], "colaboradores": [], "ocorrencias": [], "recomendacoes": []})

    df = _aplicar_filtros_executivos(df)
    if df.empty:
        return jsonify({
            "erro": "",
            "kpis": {"alertas_criticos": 0, "alertas_atencao": 0, "colaboradores_risco": 0, "ocorrencias": 0, "score_medio": 0},
            "categorias": [], "colaboradores": [], "ocorrencias": [],
            "recomendacoes": ["Nenhum alerta gerencial encontrado para os filtros selecionados."],
        })

    df = _garantir_colunas_calculadas(df)
    df["texto_inconsistencia"] = (
        df["status"].astype(str).fillna("") + " " + df["inconsistencias"].astype(str).fillna("")
    ).str.strip()
    df["tem_revisao"] = (df["status"].astype(str).str.upper() != "OK") | (df["inconsistencias"].astype(str).str.strip() != "")

    try:
        mapa_bh2_nome, mapa_bh2_mat = _mapa_saldos_bh2_para_dashboard(
            data=request.args.get("data", "").strip(),
            datas=request.args.get("datas", "").strip(),
        )
    except Exception:
        mapa_bh2_nome, mapa_bh2_mat = {}, {}

    colaboradores = []
    categorias = {
        "Banco negativo crítico": {"qtd": 0, "nivel": "danger", "descricao": "Déficit elevado de banco de horas."},
        "Banco negativo em atenção": {"qtd": 0, "nivel": "warning", "descricao": "Déficit de banco de horas a acompanhar."},
        "Horas extras concentradas": {"qtd": 0, "nivel": "warning", "descricao": "Volume de H.E. acima do padrão esperado."},
        "Ausências recorrentes": {"qtd": 0, "nivel": "warning", "descricao": "Ausências que merecem validação operacional."},
        "Registros para revisão": {"qtd": 0, "nivel": "danger", "descricao": "Status ou inconsistência pendente de tratamento."},
        "Marcação incompleta": {"qtd": 0, "nivel": "danger", "descricao": "Possível falta de marcação ou ponto aberto."},
        "Jornada sensível": {"qtd": 0, "nivel": "warning", "descricao": "Indícios de interjornada ou violação de jornada."},
    }

    ocorrencias = []

    for nome_colab, g in df.groupby("nome", dropna=False):
        saldo_info = _calcular_saldo_base_medicao_por_colaborador(g)

        # Prioridade para saldo recalculado pelo BH 2.0, preservando fallback antigo.
        saldo_min = None
        try:
            matricula_ref = str(g["matricula"].iloc[-1]).strip() if "matricula" in g.columns else ""
            nome_ref = str(nome_colab or "").strip()
            item_bh2 = None
            if matricula_ref and matricula_ref in mapa_bh2_mat:
                item_bh2 = mapa_bh2_mat.get(matricula_ref)
            else:
                item_bh2 = mapa_bh2_nome.get(_norm_bh2(nome_ref))
            if item_bh2:
                saldo_min = int(item_bh2.get("saldo_min", 0) or 0)
        except Exception:
            saldo_min = None

        if saldo_min is None:
            saldo_min = int(saldo_info.get("saldo_calculado_min", 0) or 0)
        he_min = int(g["he_min"].sum())
        absent_min = int(g["absent_min"].sum())
        dias_he = int((g["he_min"] > 0).sum())
        dias_abs = int((g["absent_min"] > 0).sum())
        qtd_revisao = int(g["tem_revisao"].sum())
        texto = " ".join(g["texto_inconsistencia"].astype(str).tolist()).lower()

        sinais = []
        score = 0

        if saldo_min <= -600:
            sinais.append({"tipo": "Banco negativo crítico", "nivel": "danger", "detalhe": f"Saldo calculado de {_minutos_para_saldo_virgula(saldo_min)}."})
            categorias["Banco negativo crítico"]["qtd"] += 1
            score += 40
        elif saldo_min < 0:
            sinais.append({"tipo": "Banco negativo em atenção", "nivel": "warning", "detalhe": f"Saldo calculado de {_minutos_para_saldo_virgula(saldo_min)}."})
            categorias["Banco negativo em atenção"]["qtd"] += 1
            score += 22

        if he_min >= 600 or dias_he >= 3:
            sinais.append({"tipo": "Horas extras concentradas", "nivel": "warning", "detalhe": f"{minutos_para_hhmm(he_min)} de H.E. em {dias_he} dia(s)."})
            categorias["Horas extras concentradas"]["qtd"] += 1
            score += 25

        if absent_min >= 480 or dias_abs >= 2:
            sinais.append({"tipo": "Ausências recorrentes", "nivel": "warning", "detalhe": f"{minutos_para_hhmm(absent_min)} de ausência em {dias_abs} dia(s)."})
            categorias["Ausências recorrentes"]["qtd"] += 1
            score += 25

        if qtd_revisao > 0:
            sinais.append({"tipo": "Registros para revisão", "nivel": "danger" if qtd_revisao >= 3 else "warning", "detalhe": f"{qtd_revisao} registro(s) com status/inconsistência."})
            categorias["Registros para revisão"]["qtd"] += 1
            score += min(30, 10 + qtd_revisao * 5)

        if any(p in texto for p in ["sem marca", "sem marcação", "ponto aberto", "marcação incompleta", "batida impar", "batida ímpar"]):
            sinais.append({"tipo": "Marcação incompleta", "nivel": "danger", "detalhe": "Há indicação de marcação incompleta ou ponto aberto."})
            categorias["Marcação incompleta"]["qtd"] += 1
            score += 30

        if any(p in texto for p in ["inter jornada", "interjornada", "violação", "violacao", "jornada"]):
            sinais.append({"tipo": "Jornada sensível", "nivel": "warning", "detalhe": "Há indicação de jornada sensível para validação."})
            categorias["Jornada sensível"]["qtd"] += 1
            score += 20

        if sinais:
            score = min(100, score)
            colaboradores.append({
                "nome": str(nome_colab),
                "cc": valor_cc_display(g["cc"].iloc[-1]),
                "funcao": str(g["funcao"].iloc[-1]),
                "turno": str(g["turno"].iloc[-1]),
                "score": int(score),
                "nivel": _nivel_alerta_por_score(score),
                "classe": _classe_alerta_por_score(score),
                "saldo": _minutos_para_saldo_virgula(saldo_min),
                "he": minutos_para_hhmm(he_min),
                "ausencia": minutos_para_hhmm(absent_min),
                "qtd_revisao": qtd_revisao,
                "sinais": sinais,
            })

        # Ocorrências pontuais: as mais operacionais para tratativa do dia.
        for _, row in g.iterrows():
            motivos = []
            if int(row.get("he_min", 0) or 0) >= 120:
                motivos.append(f"H.E. {minutos_para_hhmm(int(row.get('he_min', 0) or 0))}")
            if int(row.get("absent_min", 0) or 0) > 0:
                motivos.append(f"Ausência {minutos_para_hhmm(int(row.get('absent_min', 0) or 0))}")
            inc = str(row.get("inconsistencias", "") or "").strip()
            status = str(row.get("status", "") or "").strip()
            if status.upper() != "OK" or inc:
                motivos.append(inc or status)
            if motivos:
                ocorrencias.append({
                    "data": row.get("data", ""),
                    "nome": str(nome_colab),
                    "cc": valor_cc_display(row.get("cc", "")),
                    "turno": row.get("turno", ""),
                    "motivo": " | ".join(motivos)[:220],
                })

    colaboradores = sorted(colaboradores, key=lambda x: x["score"], reverse=True)
    score_medio = round(sum(c["score"] for c in colaboradores) / max(len(colaboradores), 1), 1) if colaboradores else 0
    categorias_lista = [
        {"titulo": k, "qtd": v["qtd"], "nivel": v["nivel"], "descricao": v["descricao"]}
        for k, v in categorias.items() if v["qtd"] > 0
    ]

    criticos = sum(1 for c in colaboradores if c["score"] >= 90)
    atencao = sum(1 for c in colaboradores if 55 <= c["score"] < 90)
    ocorrencias = sorted(ocorrencias, key=lambda x: pd.to_datetime(x.get("data", ""), format="%d/%m/%Y", errors="coerce"), reverse=True)[:60]

    recomendacoes = []
    if criticos:
        recomendacoes.append("Priorizar os colaboradores classificados como críticos antes do fechamento da medição.")
    if categorias.get("Marcação incompleta", {}).get("qtd", 0):
        recomendacoes.append("Tratar marcações incompletas antes de discutir banco de horas, pois elas podem distorcer saldo, H.E. e ausência.")
    if categorias.get("Banco negativo crítico", {}).get("qtd", 0):
        recomendacoes.append("Montar plano de compensação para saldos negativos críticos, com acompanhamento semanal.")
    if categorias.get("Horas extras concentradas", {}).get("qtd", 0):
        recomendacoes.append("Avaliar escala, cobertura e distribuição de demanda onde houver concentração de H.E.")
    if not recomendacoes:
        recomendacoes.append("Não há concentração relevante de risco gerencial nos filtros atuais.")

    payload_resp = {
        "erro": "",
        "kpis": {
            "alertas_criticos": int(criticos),
            "alertas_atencao": int(atencao),
            "colaboradores_risco": int(len(colaboradores)),
            "ocorrencias": int(len(ocorrencias)),
            "score_medio": score_medio,
        },
        "categorias": categorias_lista,
        "colaboradores": colaboradores[:30],
        "ocorrencias": ocorrencias,
        "recomendacoes": recomendacoes,
        "cache_v8212": {
            "ativo": True,
            "ttl_segundos": 180,
            "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
    }
    try:
        _CACHE_ALERTAS_AUTOMATICOS_V8212.update({"key": cache_key, "payload": payload_resp, "ts": time.monotonic()})
    except Exception:
        pass
    return jsonify(payload_resp)




def _valor_saldo_valido_dashboard(v):
    texto = str(v or "").strip()
    return bool(texto) and texto.lower() not in {"nan", "none", "nat"}


def _saldo_final_excel_oficial_grupo(g):
    """
    v8.11.21: versão leve e segura para o Dashboard Executivo.

    Usa o último Saldo Calculado válido do Excel no grupo recebido, sem
    cópias pesadas e sem acessar cache/PDF. Isso evita travamento da página
    quando o Dashboard é recalculado após importação ou troca de filtros.
    """
    try:
        if g is None or g.empty:
            return {"saldo_min": 0, "saldo": "00,00", "origem": "grupo_vazio"}

        # O endpoint já cria data_dt e ordena o dataframe principal.
        # Aqui só procuramos a última célula válida nas colunas oficiais.
        for col, origem in [
            ("saldo_calculado", "excel_saldo_calculado"),
            ("saldo_atual", "excel_saldo_atual_fallback"),
            ("saldo", "excel_saldo_fallback"),
            ("saldo_informado_pdf", "excel_saldo_informado_pdf_fallback"),
        ]:
            if col not in g.columns:
                continue
            serie = g[col].astype(str).str.strip()
            mascara = ~serie.str.lower().isin(["", "nan", "none", "nat"])
            if mascara.any():
                saldo_txt = str(serie[mascara].iloc[-1]).strip()
                return {
                    "saldo_min": int(_saldo_para_minutos_banco(saldo_txt)),
                    "saldo": saldo_txt,
                    "origem": origem,
                }
    except Exception:
        pass
    return {"saldo_min": 0, "saldo": "00,00", "origem": "saldo_nao_encontrado"}






def _mapa_saldos_bh2_para_dashboard(data="", datas=""):
    """Retorna saldos do Banco de Horas 2.0 para o Dashboard Executivo.

    v8.15.28:
    O Dashboard deixa de usar o saldo original do Consolidado para indicadores
    de Banco de Horas e passa a consumir a mesma fonte das guias Banco de Horas
    e Extrato: o cache do motor BH 2.0.

    Quando houver filtro de data, o saldo usado e a posicao acumulada ate a
    data limite do filtro, preservando a regra de calculo global do BH 2.0.
    """
    mapa_nome = {}
    mapa_mat = {}
    try:
        cache = _obter_cache_bh2_calculo()
    except Exception as e:
        LOGGER.warning(f"Falha ao obter cache BH 2.0 para Dashboard: {e}")
        return mapa_nome, mapa_mat

    if not isinstance(cache, dict) or not cache.get("ok"):
        return mapa_nome, mapa_mat

    data_limite = None
    try:
        if datas:
            datas_lista = [d.strip() for d in str(datas).split(",") if d.strip()]
            datas_dt = [pd.to_datetime(d, errors="coerce") for d in datas_lista]
            datas_dt = [pd.Timestamp(d).normalize() for d in datas_dt if not pd.isna(d)]
            if datas_dt:
                data_limite = max(datas_dt)
        elif data:
            d = pd.to_datetime(str(data).strip(), errors="coerce")
            if not pd.isna(d):
                data_limite = pd.Timestamp(d).normalize()
    except Exception:
        data_limite = None

    def registrar(item):
        nome = str(item.get("nome", "") or "").strip()
        mat = str(item.get("matricula", "") or "").strip()
        saldo_txt = str(item.get("saldo", "") or item.get("saldo_atual", "") or "").strip()
        saldo_min = _saldo_virgula_para_minutos(saldo_txt)
        if saldo_min is None:
            saldo_min = 0
            saldo_txt = "00,00"
        saida = {
            "nome": nome,
            "matricula": mat,
            "cc": item.get("cc", ""),
            "funcao": item.get("funcao", ""),
            "turno": item.get("turno", ""),
            "saldo": saldo_txt,
            "saldo_min": int(saldo_min),
            "origem": item.get("origem", "banco_horas_2_0_cache"),
        }
        if nome:
            mapa_nome[_norm_bh2(nome)] = saida
        if mat:
            mapa_mat[str(mat)] = saida

    if data_limite is None:
        for item in cache.get("linhas", []) or []:
            it = dict(item)
            it["saldo"] = it.get("saldo_atual", it.get("saldo", ""))
            it["origem"] = "banco_horas_2_0_cache_saldo_final"
            registrar(it)
        return mapa_nome, mapa_mat

    for extr in (cache.get("extratos", {}) or {}).values():
        if not isinstance(extr, dict):
            continue
        r = extr.get("resumo", {}) or {}
        melhor = None
        melhor_dt = None
        for linha in extr.get("linhas", []) or []:
            dt = pd.to_datetime(str(linha.get("data", "") or ""), format="%d/%m/%Y", errors="coerce")
            if pd.isna(dt):
                continue
            dt = pd.Timestamp(dt).normalize()
            if dt <= data_limite and (melhor_dt is None or dt >= melhor_dt):
                melhor = linha
                melhor_dt = dt
        if melhor is None:
            continue
        registrar({
            "nome": r.get("nome", melhor.get("nome", "")),
            "matricula": r.get("matricula", melhor.get("matricula", "")),
            "cc": r.get("cc", melhor.get("cc", "")),
            "funcao": r.get("funcao", melhor.get("funcao", "")),
            "turno": r.get("turno", melhor.get("turno", "")),
            "saldo": melhor.get("saldo_calculado", ""),
            "origem": "banco_horas_2_0_cache_posicao_data",
        })
    return mapa_nome, mapa_mat


def _saldo_bh2_dashboard_para_grupo(g, mapa_nome, mapa_mat):
    """Localiza o saldo BH 2.0 correspondente a um grupo do Dashboard."""
    try:
        if g is not None and not g.empty and "matricula" in g.columns:
            for mat in g["matricula"].astype(str).str.strip().tolist():
                if mat and mat in mapa_mat:
                    return mapa_mat[mat]
        if g is not None and not g.empty and "nome" in g.columns:
            for nome in g["nome"].astype(str).str.strip().tolist():
                chave = _norm_bh2(nome)
                if chave and chave in mapa_nome:
                    return mapa_nome[chave]
    except Exception:
        pass
    return None

def _preparar_base_dashboard_executivo(df):
    """Prepara a base exclusiva para a guia Dashboard Executivo.

    v8.12.00:
    Com o Script_Robo_Ponto_v6_PRO, a aba Consolidado ja e a base oficial
    operacional, contendo uma unica linha por colaborador + data. Portanto o
    Tempo Fechado nao deve mais aplicar consolidacao por Data Emissao PDF no
    Dashboard. Essa consolidacao era um curativo para o modelo v5, mas podia
    contaminar leituras quando o Excel ja vinha corretamente tratado pelo robo.

    A aba Consolidado Historico permanece disponivel apenas para auditoria e
    diagnostico, nao para KPIs operacionais.
    """
    if df is None:
        return pd.DataFrame()
    return df.copy(deep=False)


@app.route("/api/modelo-excel")
@login_obrigatorio
def api_modelo_excel():
    """Informa se o Excel ativo segue o modelo v6 do robo de ponto."""
    if not ARQUIVO_PADRAO.exists():
        return jsonify({"ok": False, "erro": "Excel ativo não encontrado.", "modelo": {}}), 404
    info = _modelo_excel_ativo(ARQUIVO_PADRAO)
    return jsonify({"ok": True, "modelo": info})


def _variacao_percentual_texto(atual, anterior):
    try:
        atual = float(atual or 0)
        anterior = float(anterior or 0)
    except Exception:
        return "0,0%"
    if anterior == 0:
        if atual == 0:
            return "0,0%"
        return "+100,0%"
    variacao = ((atual - anterior) / abs(anterior)) * 100
    sinal = "+" if variacao > 0 else ""
    return f"{sinal}{variacao:.1f}%".replace(".", ",")


def _montar_tendencias_dashboard(df_dashboard):
    """Compara a metade final do período com a metade inicial.

    v8.13.00 - tendência operacional simples, sem alterar regras de negócio.
    Usa apenas o Consolidado oficial já filtrado.
    """
    vazio = {
        "modo": "periodo_indisponivel",
        "atual": {"inicio": "-", "fim": "-", "he": "00:00", "ausencias": "00:00", "registros_revisao": 0},
        "anterior": {"inicio": "-", "fim": "-", "he": "00:00", "ausencias": "00:00", "registros_revisao": 0},
        "comparativos": [],
    }
    if df_dashboard is None or df_dashboard.empty or "data_dt" not in df_dashboard.columns:
        return vazio
    datas = sorted([d for d in df_dashboard["data_dt"].dropna().dt.normalize().unique().tolist()])
    if len(datas) < 2:
        return vazio
    corte = max(1, len(datas) // 2)
    datas_ant = set(datas[:corte])
    datas_atu = set(datas[corte:]) or set(datas[-1:])
    anterior = df_dashboard[df_dashboard["data_dt"].dt.normalize().isin(datas_ant)].copy()
    atual = df_dashboard[df_dashboard["data_dt"].dt.normalize().isin(datas_atu)].copy()

    def resumo(g, datas_ref):
        if g.empty or not datas_ref:
            return {"inicio": "-", "fim": "-", "he_min": 0, "absent_min": 0, "registros_revisao": 0, "he": "00:00", "ausencias": "00:00"}
        inicio = min(datas_ref).strftime("%d/%m/%Y")
        fim = max(datas_ref).strftime("%d/%m/%Y")
        he_min = int(g.get("he_min", pd.Series(dtype=int)).fillna(0).sum())
        absent_min = int(g.get("absent_min", pd.Series(dtype=int)).fillna(0).sum())
        revisao = int(g.get("tem_revisao", pd.Series(dtype=bool)).fillna(False).sum())
        return {"inicio": inicio, "fim": fim, "he_min": he_min, "absent_min": absent_min, "registros_revisao": revisao, "he": minutos_para_hhmm(he_min), "ausencias": minutos_para_hhmm(absent_min)}

    r_ant = resumo(anterior, datas_ant)
    r_atu = resumo(atual, datas_atu)
    comparativos = [
        {"indicador": "Horas Extras", "atual": r_atu["he"], "anterior": r_ant["he"], "variacao": _variacao_percentual_texto(r_atu["he_min"], r_ant["he_min"]), "direcao": "alta" if r_atu["he_min"] > r_ant["he_min"] else "queda" if r_atu["he_min"] < r_ant["he_min"] else "estavel"},
        {"indicador": "Ausências", "atual": r_atu["ausencias"], "anterior": r_ant["ausencias"], "variacao": _variacao_percentual_texto(r_atu["absent_min"], r_ant["absent_min"]), "direcao": "alta" if r_atu["absent_min"] > r_ant["absent_min"] else "queda" if r_atu["absent_min"] < r_ant["absent_min"] else "estavel"},
        {"indicador": "Registros em Revisão", "atual": str(r_atu["registros_revisao"]), "anterior": str(r_ant["registros_revisao"]), "variacao": _variacao_percentual_texto(r_atu["registros_revisao"], r_ant["registros_revisao"]), "direcao": "alta" if r_atu["registros_revisao"] > r_ant["registros_revisao"] else "queda" if r_atu["registros_revisao"] < r_ant["registros_revisao"] else "estavel"},
    ]
    return {"modo": "metade_periodo", "atual": r_atu, "anterior": r_ant, "comparativos": comparativos}


def _montar_alertas_inteligentes(colabs, df_dashboard):
    """Classificação gerencial v8.13.00: crítico, alto, médio e baixo."""
    alertas = []
    for c in colabs or []:
        saldo = int(c.get("saldo_min", 0) or 0)
        he = int(c.get("he_min", 0) or 0)
        absent = int(c.get("absent_min", 0) or 0)
        nivel = None
        motivo = ""
        score = 0
        if saldo <= -100 * 60:
            nivel, score, motivo = "critico", 100, "Banco de horas abaixo de -100h."
        elif saldo <= -50 * 60:
            nivel, score, motivo = "alto", 75, "Banco de horas entre -50h e -100h."
        elif absent >= 24 * 60:
            nivel, score, motivo = "medio", 55, "Volume elevado de ausências no período."
        elif he >= 24 * 60:
            nivel, score, motivo = "medio", 50, "Volume elevado de horas extras no período."
        elif saldo < 0 or he > 0 or absent > 0:
            nivel, score, motivo = "baixo", 25, "Monitoramento operacional recomendado."
        if nivel:
            alertas.append({
                "nivel": nivel,
                "score": score,
                "nome": c.get("nome", ""),
                "cc": c.get("cc", ""),
                "turno": c.get("turno", ""),
                "saldo": c.get("saldo", "00,00"),
                "he": c.get("he", "00:00"),
                "ausencias": c.get("absent", "00:00"),
                "motivo": motivo,
            })
    ordem = {"critico": 0, "alto": 1, "medio": 2, "baixo": 3}
    return sorted(alertas, key=lambda x: (ordem.get(x["nivel"], 9), -int(x.get("score", 0))))[:20]



@app.route("/api/sistema/diagnostico-tecnico")
@login_obrigatorio
def api_sistema_diagnostico_tecnico():
    """Diagnóstico leve de cache e modelo ativo para suporte técnico.

    v8.13.10: ajuda a verificar se o Excel v6 está sendo lido pela aba
    operacional correta e se o cache em memória está aquecido.
    """
    modelo = _modelo_excel_ativo(ARQUIVO_PADRAO) if ARQUIVO_PADRAO.exists() else {"modelo": "sem_base", "aba_operacional": "", "aba_historico": "", "abas": []}
    return jsonify({
        "ok": True,
        "versao": "v8.13.10",
        "arquivo_base_existe": bool(ARQUIVO_PADRAO.exists()),
        "assinatura_excel": str(_assinatura_arquivo(ARQUIVO_PADRAO)),
        "modelo_excel": modelo,
        "cache_consolidado": {
            "aquecido_raw": _CACHE_CONSOLIDADO_MEM.get("df_raw") is not None,
            "aquecido_norm": _CACHE_CONSOLIDADO_MEM.get("df_norm") is not None,
            "gerado_em": _CACHE_CONSOLIDADO_MEM.get("gerado_em", ""),
            "segundos_leitura": _CACHE_CONSOLIDADO_MEM.get("segundos_leitura", ""),
            "segundos_normalizacao": _CACHE_CONSOLIDADO_MEM.get("segundos_normalizacao", ""),
        },
        "cache_dashboard": {
            "aquecido": _CACHE_DASHBOARD_EXECUTIVO_MEM.get("payload") is not None,
            "gerado_em_ts": _CACHE_DASHBOARD_EXECUTIVO_MEM.get("gerado_em_ts", 0.0),
        },
    })

@app.route("/api/dashboard-executivo")
@login_obrigatorio
def api_dashboard_executivo():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "kpis": {}, "linhas": [], "rankings": {}, "resumos": {}, "alertas": [], "alertas_inteligentes": [], "tendencias": {}})

    df = normalizar_colunas(df)
    if df.empty:
        return jsonify({"erro": "Nenhuma base carregada.", "kpis": {}, "linhas": [], "rankings": {}, "resumos": {}, "alertas": [], "alertas_inteligentes": [], "tendencias": {}})

    cc_filtro = request.args.get("cc", "").strip()
    nome_filtro = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data = request.args.get("data", "").strip()
    datas = request.args.get("datas", "").strip()

    # v8.11.21: se o usuário clicar em outras guias durante o carregamento,
    # o Dashboard não deve recalcular tudo repetidas vezes.
    cache_key = (str(_assinatura_arquivo(ARQUIVO_PADRAO)), _assinatura_config_bh2(), cc_filtro, nome_filtro, turno_filtro, data, datas)
    try:
        if (
            _CACHE_DASHBOARD_EXECUTIVO_MEM.get("key") == cache_key
            and _CACHE_DASHBOARD_EXECUTIVO_MEM.get("payload") is not None
            and time.monotonic() - float(_CACHE_DASHBOARD_EXECUTIVO_MEM.get("gerado_em_ts") or 0) < 30
        ):
            return jsonify(_CACHE_DASHBOARD_EXECUTIVO_MEM["payload"])
    except Exception:
        pass

    if cc_filtro:
        df = aplicar_filtro_cc_backend(df, cc_filtro)
    if nome_filtro:
        df = df[df["nome"].str.lower() == nome_filtro]
    if turno_filtro:
        df = df[df["turno"].str.lower() == turno_filtro]

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")

    if datas:
        datas_lista = [d.strip() for d in datas.split(",") if d.strip()]
        if datas_lista:
            data_iso = df["data_dt"].dt.strftime("%Y-%m-%d")
            df = df[data_iso.isin(datas_lista)].copy()
    elif data:
        data_iso = df["data_dt"].dt.strftime("%Y-%m-%d")
        df = df[data_iso == data].copy()

    if df.empty:
        return jsonify({
            "erro": "",
            "kpis": {
                "registros": 0,
                "colaboradores": 0,
                "banco_positivo": "00,00",
                "banco_negativo": "00,00",
                "colab_saldo_positivo": 0,
                "colab_saldo_negativo": 0,
                "he_total": "00:00",
                "ausencia_total": "00:00",
                "dias_com_he": 0,
                "dias_com_ausencia": 0,
                "registros_revisao": 0,
                "taxa_revisao": "0%",
            },
            "periodo": {"inicio": "-", "fim": "-"},
            "linhas": [],
            "rankings": {"banco_positivo": [], "banco_negativo": [], "horas_extras": [], "ausencias": []},
            "resumos": {"por_cc": [], "por_turno": []},
            "alertas": [],
            "alertas_inteligentes": [],
            "tendencias": {},
        })

    df = _garantir_colunas_calculadas(df)
    try:
        df = df.sort_values(["nome", "matricula", "data_dt"]).copy()
    except Exception:
        df = df.sort_values(["nome", "data_dt"]).copy()

    # v8.11.36: a consolidação por última Data Emissão PDF fica encapsulada
    # em função própria e é usada exclusivamente nesta rota. Banco de Horas e
    # Extrato continuam consumindo o Excel oficial completo.
    df_dashboard = _preparar_base_dashboard_executivo(df)

    # v8.15.28: saldos do Dashboard agora vêm do mesmo motor/cache BH 2.0
    # usado pelas guias Banco de Horas e Extrato. H.E., ausências e revisão
    # continuam respeitando os filtros do Consolidado operacional.
    mapa_bh2_nome, mapa_bh2_mat = _mapa_saldos_bh2_para_dashboard(data=data, datas=datas)

    colabs = []
    for nome_colab, g in df_dashboard.groupby("nome", dropna=False):
        saldo_info_bh2 = _saldo_bh2_dashboard_para_grupo(g, mapa_bh2_nome, mapa_bh2_mat)
        if saldo_info_bh2:
            saldo_min = int(saldo_info_bh2.get("saldo_min", 0) or 0)
            saldo_txt = str(saldo_info_bh2.get("saldo", "") or _minutos_para_saldo_virgula(saldo_min))
            origem_saldo = saldo_info_bh2.get("origem", "banco_horas_2_0_cache")
        else:
            # Fallback defensivo para cenários sem configuração/cache BH 2.0.
            saldo_info_excel = _saldo_final_excel_oficial_grupo(g)
            saldo_min = int(saldo_info_excel.get("saldo_min", 0) or 0)
            saldo_txt = str(saldo_info_excel.get("saldo", "") or _minutos_para_saldo_virgula(saldo_min))
            origem_saldo = saldo_info_excel.get("origem", "excel_saldo_calculado_fallback")
        he_min = int(g["he_min"].sum())
        absent_min = int(g["absent_min"].sum())
        colabs.append({
            "nome": str(nome_colab),
            "cc": valor_cc_display(g["cc"].iloc[-1]),
            "funcao": str(g["funcao"].iloc[-1]),
            "turno": str(g["turno"].iloc[-1]),
            "saldo_min": saldo_min,
            "saldo": saldo_txt,
            "origem_saldo": origem_saldo,
            "he_min": he_min,
            "he": minutos_para_hhmm(he_min),
            "absent_min": absent_min,
            "absent": minutos_para_hhmm(absent_min),
            "registros": int(len(g)),
            "dias_he": int((g["he_min"] > 0).sum()),
            "dias_absent": int((g["absent_min"] > 0).sum()),
        })

    banco_positivo = sum(x["saldo_min"] for x in colabs if x["saldo_min"] > 0)
    banco_negativo = sum(x["saldo_min"] for x in colabs if x["saldo_min"] < 0)
    registros_revisao = int(df_dashboard["tem_revisao"].sum())
    taxa_revisao = (registros_revisao / max(int(len(df_dashboard)), 1)) * 100

    datas_validas = df_dashboard["data_dt"].dropna()
    if not datas_validas.empty:
        periodo = {
            "inicio": datas_validas.min().strftime("%d/%m/%Y"),
            "fim": datas_validas.max().strftime("%d/%m/%Y"),
        }
    else:
        periodo = {"inicio": "-", "fim": "-"}

    kpis = {
        "registros": int(len(df_dashboard)),
        "colaboradores": int(len(colabs)),
        "banco_positivo": _minutos_para_saldo_virgula(banco_positivo),
        "banco_negativo": _minutos_para_saldo_virgula(banco_negativo),
        "colab_saldo_positivo": int(sum(1 for x in colabs if x["saldo_min"] > 0)),
        "colab_saldo_negativo": int(sum(1 for x in colabs if x["saldo_min"] < 0)),
        "he_total": minutos_para_hhmm(int(df_dashboard["he_min"].sum())),
        "ausencia_total": minutos_para_hhmm(int(df_dashboard["absent_min"].sum())),
        "dias_com_he": int((df_dashboard["he_min"] > 0).sum()),
        "dias_com_ausencia": int((df_dashboard["absent_min"] > 0).sum()),
        "registros_revisao": registros_revisao,
        "taxa_revisao": f"{taxa_revisao:.1f}%".replace(".", ","),
    }

    def _rank_item(x, valor_key, texto_key):
        return {
            "nome": x["nome"],
            "cc": x["cc"],
            "funcao": x["funcao"],
            "turno": x["turno"],
            "valor": x[texto_key],
            "minutos": int(x[valor_key]),
        }

    rankings = {
        "banco_positivo": [_rank_item(x, "saldo_min", "saldo") for x in sorted([c for c in colabs if c["saldo_min"] > 0], key=lambda i: i["saldo_min"], reverse=True)[:8]],
        "banco_negativo": [_rank_item(x, "saldo_min", "saldo") for x in sorted([c for c in colabs if c["saldo_min"] < 0], key=lambda i: i["saldo_min"])[:8]],
        "horas_extras": [_rank_item(x, "he_min", "he") for x in sorted([c for c in colabs if c["he_min"] > 0], key=lambda i: i["he_min"], reverse=True)[:8]],
        "ausencias": [_rank_item(x, "absent_min", "absent") for x in sorted([c for c in colabs if c["absent_min"] > 0], key=lambda i: i["absent_min"], reverse=True)[:8]],
    }

    def _resumo_grupo(campo, limite=8):
        linhas = []
        for rotulo, g in df_dashboard.groupby(campo, dropna=False):
            saldo_grupo = 0
            for _, gc in g.groupby("nome", dropna=False):
                saldo_info_bh2 = _saldo_bh2_dashboard_para_grupo(gc, mapa_bh2_nome, mapa_bh2_mat)
                if saldo_info_bh2:
                    saldo_grupo += int(saldo_info_bh2.get("saldo_min", 0) or 0)
                else:
                    saldo_grupo += int(_saldo_final_excel_oficial_grupo(gc).get("saldo_min", 0) or 0)
            linhas.append({
                "rotulo": valor_cc_display(rotulo) if campo == "cc" else str(rotulo or "Sem informação"),
                "colaboradores": int(g["nome"].nunique()),
                "registros": int(len(g)),
                "saldo_min": int(saldo_grupo),
                "saldo": _minutos_para_saldo_virgula(saldo_grupo),
                "he_min": int(g["he_min"].sum()),
                "he": minutos_para_hhmm(int(g["he_min"].sum())),
                "absent_min": int(g["absent_min"].sum()),
                "absent": minutos_para_hhmm(int(g["absent_min"].sum())),
            })
        return sorted(linhas, key=lambda x: abs(x["saldo_min"]) + x["he_min"] + x["absent_min"], reverse=True)[:limite]

    resumos = {
        "por_cc": _resumo_grupo("cc", limite=10),
        "por_turno": _resumo_grupo("turno", limite=10),
    }

    alertas = []
    if kpis["colab_saldo_negativo"]:
        alertas.append({
            "nivel": "danger",
            "titulo": "Banco negativo em aberto",
            "descricao": f"{kpis['colab_saldo_negativo']} colaborador(es) com saldo negativo. Priorizar análise de compensação e escala.",
        })
    if int(df_dashboard["he_min"].sum()) > 0:
        top_he = rankings["horas_extras"][0]["nome"] if rankings["horas_extras"] else "colaboradores com H.E."
        alertas.append({
            "nivel": "warning",
            "titulo": "Concentração de horas extras",
            "descricao": f"Há {kpis['he_total']} de H.E. no filtro atual. Maior concentração: {top_he}.",
        })
    if int(df_dashboard["absent_min"].sum()) > 0:
        alertas.append({
            "nivel": "warning",
            "titulo": "Ausências registradas",
            "descricao": f"Foram identificadas {kpis['ausencia_total']} de ausência. Avaliar justificativas, abonos e impacto operacional.",
        })
    if registros_revisao:
        alertas.append({
            "nivel": "danger" if taxa_revisao >= 10 else "warning",
            "titulo": "Registros para revisão",
            "descricao": f"{registros_revisao} registro(s), ou {kpis['taxa_revisao']}, possuem status ou inconsistência para tratamento.",
        })

    tendencias = _montar_tendencias_dashboard(df_dashboard)
    alertas_inteligentes = _montar_alertas_inteligentes(colabs, df_dashboard)

    linhas = []
    for grupo, lista in rankings.items():
        rotulo = {
            "banco_positivo": "Top Banco Positivo",
            "banco_negativo": "Top Banco Negativo",
            "horas_extras": "Top Horas Extras",
            "ausencias": "Top Ausências",
        }.get(grupo, grupo)
        for x in lista:
            linhas.append({"grupo": rotulo, "nome": x["nome"], "cc": x["cc"], "funcao": x["funcao"], "turno": x["turno"], "valor": x["valor"]})

    payload_dashboard = {
        "erro": "",
        "kpis": kpis,
        "periodo": periodo,
        "linhas": linhas,
        "rankings": rankings,
        "resumos": resumos,
        "alertas": alertas,
        "alertas_inteligentes": alertas_inteligentes,
        "tendencias": tendencias,
        "fonte_banco_horas": "banco_horas_2_0_cache_v8_15_28",
    }
    try:
        _CACHE_DASHBOARD_EXECUTIVO_MEM["key"] = cache_key
        _CACHE_DASHBOARD_EXECUTIVO_MEM["payload"] = payload_dashboard
        _CACHE_DASHBOARD_EXECUTIVO_MEM["gerado_em_ts"] = time.monotonic()
    except Exception:
        pass
    return jsonify(payload_dashboard)



@app.route("/api/auditoria-premium")
@login_obrigatorio
def api_auditoria_premium():
    """Guia Auditoria Premium v8.13.00.

    Usa a aba Consolidado Historico do Script_Robo_Ponto_v6_PRO para permitir
    rastrear fotografias por colaborador, data e Data Emissao PDF, sem mexer na
    base operacional exibida nas demais guias.
    """
    df, erro = ler_consolidado_historico_v6()
    if erro:
        return jsonify({"ok": False, "erro": erro, "dados": [], "kpis": {}, "opcoes": {}})
    aviso = ""
    try:
        aviso = str(getattr(df, "attrs", {}).get("auditoria_aviso", "") or "")
    except Exception:
        aviso = ""
    df = normalizar_colunas_historico(df)
    if df.empty:
        return jsonify({"ok": True, "erro": "", "aviso": aviso, "dados": [], "kpis": {"linhas": 0}, "opcoes": {}})

    cc_filtro = request.args.get("cc", "").strip()
    nome_filtro = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data = request.args.get("data", "").strip()
    emissao = request.args.get("emissao", "").strip()

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")
    df["emissao_dt"] = pd.to_datetime(df["data_emissao_pdf"], format="%d/%m/%Y", errors="coerce")

    if cc_filtro:
        df = aplicar_filtro_cc_backend(df, cc_filtro)
    if nome_filtro:
        df = df[df["nome"].astype(str).str.lower() == nome_filtro]
    if turno_filtro:
        df = df[df["turno"].astype(str).str.lower() == turno_filtro]
    if data:
        df = df[df["data_dt"].dt.strftime("%Y-%m-%d") == data]
    if emissao:
        df = df[df["emissao_dt"].dt.strftime("%Y-%m-%d") == emissao]

    try:
        df = df.sort_values(["nome", "data_dt", "emissao_dt"], na_position="last").copy()
    except Exception:
        pass

    total = int(len(df))
    limite = 600
    df_saida = df.head(limite).copy()

    def _linha(row):
        return {
            "data": str(row.get("data", "")),
            "data_emissao_pdf": str(row.get("data_emissao_pdf", "")),
            "nome": str(row.get("nome", "")),
            "cc": valor_cc_display(row.get("cc", "")),
            "turno": str(row.get("turno", "")),
            "he": str(row.get("he", "")),
            "absent": str(row.get("absent", "")),
            "saldo_calculado": str(row.get("saldo_calculado", "")),
            "saldo_informado_pdf": str(row.get("saldo_informado_pdf", "")),
            "saldo_atual": str(row.get("saldo_atual", "")),
            "observacao": str(row.get("observacao", "")),
            "inconsistencias": str(row.get("inconsistencias", "")),
            "arquivo_origem": str(row.get("arquivo_origem", "")),
        }

    kpis = {
        "linhas": total,
        "colaboradores": int(df["nome"].nunique()) if "nome" in df.columns else 0,
        "datas": int(df["data"].nunique()) if "data" in df.columns else 0,
        "emissoes": int(df["data_emissao_pdf"].nunique()) if "data_emissao_pdf" in df.columns else 0,
        "limitado": bool(total > limite),
        "limite": limite,
    }
    opcoes = {
        "ccs": sorted([valor_cc_display(x) for x in df.get("cc", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if str(x).strip()]),
        "nomes": sorted([str(x) for x in df.get("nome", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if str(x).strip()]),
        "turnos": sorted([str(x) for x in df.get("turno", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if str(x).strip()]),
        "emissoes": sorted([str(x) for x in df.get("data_emissao_pdf", pd.Series(dtype=str)).dropna().astype(str).unique().tolist() if str(x).strip()]),
    }
    return jsonify({"ok": True, "erro": "", "aviso": aviso, "dados": [_linha(r) for _, r in df_saida.iterrows()], "kpis": kpis, "opcoes": opcoes})



def _resolver_datas_filtro_banco_horas_request():
    """Resolve filtro de data da guia Banco de Horas.

    A UI envia data em yyyy-mm-dd. O Excel costuma estar em dd/mm/aaaa.
    Para o painel por colaborador, a data funciona como posicao de corte:
    mostra o saldo acumulado ate a data selecionada, preservando a evolucao
    desde a data inicial configurada do BH 2.0.
    """
    valores = []
    data_unica = request.args.get("data", "").strip()
    datas_mult = request.args.get("datas", "").strip()
    if datas_mult:
        valores.extend([x.strip() for x in datas_mult.split(",") if x.strip()])
    elif data_unica:
        valores.append(data_unica)

    resolvidas = []
    for valor in valores:
        dt = pd.to_datetime(valor, errors="coerce")
        if pd.isna(dt):
            dt = pd.to_datetime(valor, format="%d/%m/%Y", errors="coerce")
        if not pd.isna(dt):
            resolvidas.append(pd.Timestamp(dt).normalize())
    return resolvidas


def _montar_resumo_banco_horas_ate_data(df, data_limite):
    """Monta resumo do Banco de Horas respeitando filtro de data.

    v8.15.23: primeiro tenta resolver pelo cache do BH 2.0. Assim o filtro de
    data vira só uma projeção visual, sem recalcular o Banco inteiro.
    """
    if df is None or df.empty or data_limite is None or pd.isna(data_limite):
        return []
    try:
        linhas_cache = _linhas_bh2_ate_data_cache(data_limite)
        if linhas_cache:
            return linhas_cache
    except Exception as exc:
        LOGGER.warning(f"Falha ao usar cache BH 2.0 no filtro de data: {exc}")
    base = normalizar_colunas(df).copy()
    if base.empty or "data" not in base.columns:
        return []
    base = _garantir_colunas_calculadas(base)
    if "data_dt" not in base.columns:
        base["data_dt"] = pd.to_datetime(base["data"], format="%d/%m/%Y", errors="coerce")
    base = base[base["data_dt"].notna() & (base["data_dt"] <= pd.Timestamp(data_limite).normalize())].copy()
    if base.empty:
        return []

    linhas = []
    chaves_grupo = [c for c in ["matricula", "nome"] if c in base.columns] or ["nome"]
    for _, g in base.sort_values("data_dt").groupby(chaves_grupo, dropna=False):
        extr = _calcular_extrato_bh2_colaborador(g.copy())
        if not extr:
            # Fallback para Excel oficial sem BH2 ativo.
            g_ord = g.sort_values("data_dt").copy()
            if g_ord.empty:
                continue
            saldos_validos = [str(v).strip() for v in g_ord.get("saldo_calculado", pd.Series([], dtype="object")).tolist() if str(v).strip() and str(v).strip().lower() not in {"nan", "none", "nat"}]
            saldo_final = saldos_validos[-1] if saldos_validos else ""
            saldo_base = saldos_validos[0] if saldos_validos else ""
            linhas.append({
                "nome": str(g_ord["nome"].iloc[-1]) if "nome" in g_ord.columns and not g_ord.empty else "",
                "matricula": str(g_ord["matricula"].iloc[-1]) if "matricula" in g_ord.columns and not g_ord.empty else "",
                "cc": str(g_ord["cc"].iloc[-1]) if "cc" in g_ord.columns and not g_ord.empty else "",
                "funcao": str(g_ord["funcao"].iloc[-1]) if "funcao" in g_ord.columns and not g_ord.empty else "",
                "turno": str(g_ord["turno"].iloc[-1]) if "turno" in g_ord.columns and not g_ord.empty else "",
                "he_total": minutos_para_hhmm(int(g_ord.get("he_min", pd.Series([], dtype="float64")).fillna(0).sum())) if "he_min" in g_ord.columns else "00:00",
                "absent_total": minutos_para_hhmm(int(g_ord.get("absent_min", pd.Series([], dtype="float64")).fillna(0).sum())) if "absent_min" in g_ord.columns else "00:00",
                "dias_com_he": _contar_dias_movimento_df_bh(g_ord, "he_min", "he"),
                "dias_com_ausencia": _contar_dias_movimento_df_bh(g_ord, "absent_min", "absent"),
                "metricas_dias_fonte": "excel_oficial_filtrado_por_data",
                "saldo_atual": saldo_final,
                "saldo_pdf_final": "",
                "saldo_base": saldo_base,
                "origem_saldo_base": "excel_oficial_filtrado_por_data",
            })
            continue

        linhas_extrato = extr.get("linhas", []) or []
        linhas_extrato_filtradas = []
        for item in linhas_extrato:
            dt_item = pd.to_datetime(str(item.get("data", "")), format="%d/%m/%Y", errors="coerce")
            if pd.notna(dt_item) and pd.Timestamp(dt_item).normalize() <= pd.Timestamp(data_limite).normalize():
                linhas_extrato_filtradas.append(item)
        if not linhas_extrato_filtradas:
            continue

        r = extr.get("resumo", {}) or {}
        ultima = linhas_extrato_filtradas[-1]
        credito_total = sum(hhmm_para_minutos(x.get("he", "")) for x in linhas_extrato_filtradas)
        debito_total = sum(hhmm_para_minutos(x.get("absent", "")) for x in linhas_extrato_filtradas)
        linhas.append({
            "nome": r.get("nome", ""),
            "matricula": r.get("matricula", str(g["matricula"].iloc[-1]) if "matricula" in g.columns and not g.empty else ""),
            "cc": r.get("cc", ""),
            "funcao": r.get("funcao", ""),
            "turno": r.get("turno", ""),
            "he_total": minutos_para_hhmm(int(credito_total)),
            "absent_total": minutos_para_hhmm(int(debito_total)),
            "dias_com_he": _contar_dias_movimento_bh(linhas_extrato_filtradas, "he"),
            "dias_com_ausencia": _contar_dias_movimento_bh(linhas_extrato_filtradas, "absent"),
            "metricas_dias_fonte": "extrato_bh2_filtrado_por_data",
            "saldo_atual": ultima.get("saldo_calculado", ""),
            "saldo_pdf_final": ultima.get("saldo_informado", ""),
            "saldo_base": r.get("saldo_base", ""),
            "origem_saldo_base": r.get("origem_saldo_base", ""),
        })
    return linhas


@app.route("/api/dashboard-banco-he")
@login_obrigatorio
def api_dashboard_banco_he():
    """Carrega a guia Banco de Horas sempre em modo rápido.

    v8.10.44:
    - A guia Banco de Horas deixou de ter fallback pesado em tempo real.
    - Antes, se o resumo/cache não existisse ou estivesse incompleto, o endpoint
      recalculava saldo PDF/ciclo para todos os colaboradores e a tela ficava
      indefinidamente em "Carregando".
    - Agora a guia lê somente data/resumo_banco_horas_pdf.json. Se o resumo não
      existir, retorna imediatamente com aviso para usar o botão manual.
    """
    busca = request.args.get("busca", "").strip().lower()
    cc_filtro = request.args.get("cc", "").strip().lower()
    nome = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    order_saldo = request.args.get("order_saldo", "").strip().lower()
    datas_filtro_bh = _resolver_datas_filtro_banco_horas_request()
    data_limite_bh = max(datas_filtro_bh) if datas_filtro_bh else None
    cache_key_bh = (
        str(_assinatura_arquivo(ARQUIVO_PADRAO)),
        _assinatura_config_bh2(),
        busca,
        cc_filtro,
        nome,
        turno_filtro,
        order_saldo,
        tuple(str(d) for d in datas_filtro_bh),
    )
    try:
        if (
            _CACHE_DASHBOARD_BANCO_HORAS_MEM.get("key") == cache_key_bh
            and _CACHE_DASHBOARD_BANCO_HORAS_MEM.get("payload") is not None
            and time.monotonic() - float(_CACHE_DASHBOARD_BANCO_HORAS_MEM.get("gerado_em_ts") or 0) < 45
        ):
            return jsonify(_CACHE_DASHBOARD_BANCO_HORAS_MEM["payload"])
    except Exception:
        pass

    # v8.21.38: prioriza resposta imediata a partir do resumo ja salvo.
    # A versao anterior podia tentar gerar o resumo antes de usar o cache,
    # deixando a guia sem sinal claro de progresso em bases grandes.
    fonte_resumo = "resumo_banco_horas_cache_v8_21_14"
    resumo_cache = _carregar_resumo_banco_horas_pdf()
    sig_atual_bh = str(_assinatura_arquivo(ARQUIVO_PADRAO))
    resumo_valido = _resumo_banco_horas_parece_valido(resumo_cache)
    assinatura_resumo = str((resumo_cache or {}).get("assinatura_excel", "") or "")
    resumo_compativel = resumo_valido and (not assinatura_resumo or assinatura_resumo == sig_atual_bh)

    if resumo_valido and not resumo_compativel:
        try:
            gerado = gerar_resumo_banco_horas_pdf()
            if gerado.get("ok"):
                resumo_cache = _carregar_resumo_banco_horas_pdf()
                fonte_resumo = gerado.get("fonte", "excel_oficial_gerado_sob_demanda")
        except Exception:
            pass
    if not _resumo_banco_horas_parece_valido(resumo_cache):
        # Última tentativa segura: gerar o resumo diretamente do Excel oficial.
        try:
            gerado = gerar_resumo_banco_horas_pdf()
            if gerado.get("ok"):
                resumo_cache = _carregar_resumo_banco_horas_pdf()
                fonte_resumo = gerado.get("fonte", "excel_oficial_gerado_sob_demanda")
        except Exception:
            pass

    if not _resumo_banco_horas_parece_valido(resumo_cache):
        return jsonify({
            "erro": "",
            "dados": [],
            "kpis": _montar_kpis_banco_horas_de_linhas([], registros_base=0, revisar=0),
            "fonte": "excel_oficial_indisponivel",
            "resumo_gerado_em": "",
            "aviso": "Resumo do Banco de Horas ainda não foi gerado. Reimporte o Excel do robô ou use o botão Corrigir Banco de Horas pelos PDFs.",
        })

    linhas_base_resumo = resumo_cache.get("linhas", [])
    fonte_data = ""
    if data_limite_bh is not None:
        try:
            df_data_bh, erro_data_bh = ler_consolidado()
            if not erro_data_bh:
                linhas_data = _montar_resumo_banco_horas_ate_data(df_data_bh, data_limite_bh)
                if linhas_data:
                    linhas_base_resumo = linhas_data
                    fonte_data = f"; data={pd.Timestamp(data_limite_bh).strftime('%d/%m/%Y')}"
        except Exception as exc:
            LOGGER.warning(f"Falha ao aplicar filtro de data no Banco de Horas: {exc}")

    linhas = _filtrar_linhas_resumo_banco_horas(
        linhas_base_resumo,
        busca=busca,
        cc=cc_filtro,
        nome=nome,
        turno=turno_filtro,
    )

    if order_saldo in {"asc", "desc"}:
        linhas = sorted(
            linhas,
            key=lambda item: saldo_dashboard_para_minutos(item.get("saldo_atual", item.get("saldo_total", item.get("saldo", "")))),
            reverse=(order_saldo == "desc")
        )

    # v8.21.38: Dias com HE e Dias com ausência passam a ser métrica real.
    # Se algum cache antigo chegar sem a coluna, o front-end recebe None em vez
    # de um zero artificial. Caches BH 2.0 sem métricas são invalidados acima.
    def _normalizar_linha_banco_horas_saida(item):
        linha = dict(item or {})
        for campo in ("dias_com_he", "dias_com_ausencia"):
            valor = linha.get(campo, None)
            if valor is None or str(valor).strip().lower() in {"", "undefined", "nan", "none"}:
                linha[campo] = None
            else:
                try:
                    linha[campo] = int(float(str(valor).replace(",", ".")))
                except Exception:
                    linha[campo] = None
        linha.setdefault("metricas_dias_fonte", "nao_calculado")
        linha.setdefault("he_total", "00:00")
        linha.setdefault("absent_total", "00:00")
        linha.setdefault("cc", "")
        linha.setdefault("nome", "")
        linha.setdefault("funcao", "")
        linha.setdefault("turno", "")
        linha.setdefault("saldo_atual", "")
        return linha

    linhas = [_normalizar_linha_banco_horas_saida(item) for item in linhas]

    kpis = _montar_kpis_banco_horas_de_linhas(
        linhas,
        registros_base=resumo_cache.get("registros_base", len(linhas)),
        revisar=resumo_cache.get("revisar", 0),
    )
    payload_bh = {
        "erro": "",
        "dados": linhas,
        "kpis": kpis,
        "fonte": f"{resumo_cache.get('fonte', fonte_resumo)}{fonte_data}",
        "resumo_gerado_em": resumo_cache.get("gerado_em", ""),
        "aviso": "",
    }
    try:
        _CACHE_DASHBOARD_BANCO_HORAS_MEM["key"] = cache_key_bh
        _CACHE_DASHBOARD_BANCO_HORAS_MEM["payload"] = payload_bh
        _CACHE_DASHBOARD_BANCO_HORAS_MEM["gerado_em_ts"] = time.monotonic()
    except Exception:
        pass
    return jsonify(payload_bh)


@app.route("/api/debug-violacoes-jornada")
@login_obrigatorio
def api_debug_violacoes_jornada():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro})

    df_original, _ = ler_consolidado()
    df = normalizar_colunas(df_original)
    if df.empty:
        return jsonify({"erro": "", "total_linhas": 0, "turnos": [], "amostra": []})

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")
    df["_trabalhado"] = df.apply(_tem_trabalho_jornada, axis=1)
    df["_regra_turno"] = df["turno"].apply(_turno_regra_violacao_jornada)
    df["_semana"] = df["data_dt"].apply(lambda x: _semana_iso_label(x) if pd.notna(x) else "")

    resumo_turnos = []
    if "turno" in df.columns:
        resumo = (
            df.groupby(["turno", "_regra_turno"], dropna=False)
            .agg(
                linhas=("nome", "size"),
                trabalhados=("_trabalhado", "sum"),
                colaboradores=("nome", pd.Series.nunique)
            )
            .reset_index()
        )
        for _, row in resumo.iterrows():
            resumo_turnos.append({
                "turno": str(row["turno"]),
                "regra": str(row["_regra_turno"]),
                "linhas": int(row["linhas"]),
                "dias_trabalhados_detectados": int(row["trabalhados"]),
                "colaboradores": int(row["colaboradores"]),
            })

    excesso_semana = []
    base = df[(df["_trabalhado"]) & (df["_regra_turno"].isin(["comercial", "noturno"]))].copy()
    if not base.empty:
        tmp = (
            base.groupby(["nome", "turno", "_regra_turno", "_semana"], dropna=False)["data_dt"]
            .nunique()
            .reset_index(name="dias_trabalhados")
        )
        tmp = tmp[tmp["dias_trabalhados"] > 5].copy()
        for _, row in tmp.head(30).iterrows():
            excesso_semana.append({
                "nome": str(row["nome"]),
                "turno": str(row["turno"]),
                "regra": str(row["_regra_turno"]),
                "semana": str(row["_semana"]),
                "dias_trabalhados": int(row["dias_trabalhados"]),
            })

    consecutivos_12x36 = []
    for nome_colab, g in df[(df["_trabalhado"]) & (df["_regra_turno"] == "12x36")].sort_values("data_dt").groupby("nome", dropna=False):
        datas = sorted(pd.to_datetime(g["data_dt"], errors="coerce").dropna().dt.date.unique())
        for a, b in zip(datas, datas[1:]):
            if (b - a).days == 1:
                turno_ref = str(g["turno"].iloc[-1]) if not g.empty else ""
                consecutivos_12x36.append({
                    "nome": str(nome_colab),
                    "turno": turno_ref,
                    "data_anterior": a.strftime("%d/%m/%Y"),
                    "data_atual": b.strftime("%d/%m/%Y"),
                })
                if len(consecutivos_12x36) >= 30:
                    break
        if len(consecutivos_12x36) >= 30:
            break

    return jsonify({
        "erro": "",
        "total_linhas": int(len(df)),
        "linhas_trabalhadas_detectadas": int(df["_trabalhado"].sum()),
        "turnos": resumo_turnos,
        "excessos_semanais_amostra": excesso_semana,
        "consecutivos_12x36_amostra": consecutivos_12x36,
    })


@app.route("/api/violacoes-jornada")
@login_obrigatorio
def api_violacoes_jornada():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "dados": [], "kpis": {}})

    df = normalizar_colunas(df)
    if df.empty:
        return jsonify({"erro": "", "dados": [], "kpis": {
            "registros": 0, "colaboradores": 0, "he_total": "00:00",
            "absent_total": "00:00", "saldo_negativo": 0, "revisar": 0,
            "quebras_12x36": 0, "excessos_semanais": 0
        }})

    cc_filtro = request.args.get("cc", "").strip()
    nome = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data = request.args.get("data", "").strip()
    datas = request.args.get("datas", "").strip()

    # v8.11.21: se o usuário clicar em outras guias durante o carregamento,
    # o Dashboard não deve recalcular tudo repetidas vezes.
    cache_key = (str(_assinatura_arquivo(ARQUIVO_PADRAO)), cc_filtro, nome, turno_filtro, data, datas)
    try:
        if (
            _CACHE_DASHBOARD_EXECUTIVO_MEM.get("key") == cache_key
            and _CACHE_DASHBOARD_EXECUTIVO_MEM.get("payload") is not None
            and time.monotonic() - float(_CACHE_DASHBOARD_EXECUTIVO_MEM.get("gerado_em_ts") or 0) < 30
        ):
            return jsonify(_CACHE_DASHBOARD_EXECUTIVO_MEM["payload"])
    except Exception:
        pass

    if cc_filtro:
        df = aplicar_filtro_cc_backend(df, cc_filtro)

    if nome:
        df = df[df["nome"].str.lower() == nome]

    if turno_filtro:
        df = df[df["turno"].str.lower() == turno_filtro]

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")
    df = df[df["data_dt"].notna()].copy()

    if df.empty:
        return jsonify({"erro": "", "dados": [], "kpis": {
            "registros": 0, "colaboradores": 0, "he_total": "00:00",
            "absent_total": "00:00", "saldo_negativo": 0, "revisar": 0,
            "quebras_12x36": 0, "excessos_semanais": 0
        }})

    datas_filtro = []
    if datas:
        datas_filtro = [d.strip() for d in datas.split(",") if d.strip()]
    elif data:
        datas_filtro = [data]

    resultados = []

    for nome_colab, g in df.sort_values("data_dt").groupby("nome", dropna=False):
        g = g.sort_values("data_dt").copy()
        g["_trabalhado"] = g.apply(_tem_trabalho_jornada, axis=1)
        g["_regra_turno"] = g["turno"].apply(_turno_regra_violacao_jornada)

        # ============================================================
        # Regra 12x36:
        # Não pode trabalhar dias consecutivos.
        # ============================================================
        trabalhados_12x36 = g[(g["_trabalhado"]) & (g["_regra_turno"] == "12x36")].copy()
        registros_12x36 = trabalhados_12x36.to_dict(orient="records")

        for anterior, atual in zip(registros_12x36, registros_12x36[1:]):
            dias_diff = (atual["data_dt"].date() - anterior["data_dt"].date()).days

            if dias_diff == 1:
                data_iso = atual["data_dt"].strftime("%Y-%m-%d")
                if datas_filtro and data_iso not in datas_filtro:
                    continue

                resultados.append({
                    "cc": atual.get("cc", ""),
                    "nome": nome_colab,
                    "funcao": atual.get("funcao", ""),
                    "turno": atual.get("turno", ""),
                    "tipo_jornada": "12x36",
                    "periodo": "",
                    "data_referencia": atual["data_dt"].strftime("%d/%m/%Y"),
                    "violacao": "Quebra da escala 12x36",
                    "detalhe": f"Trabalho em dias consecutivos: {anterior['data_dt'].strftime('%d/%m/%Y')} e {atual['data_dt'].strftime('%d/%m/%Y')}.",
                    "dias_trabalhados": "",
                    "limite_esperado": "Sem dias consecutivos",
                    "status": "REVISAR",
                })

        # ============================================================
        # Regra Comercial / Noturno:
        # Não pode trabalhar MAIS de 5 dias na semana.
        # Semana: segunda a domingo.
        # ============================================================
        base_semana = g[
            (g["_trabalhado"]) &
            (g["_regra_turno"].isin(["comercial", "noturno"]))
        ].copy()

        if not base_semana.empty:
            base_semana["_semana"] = base_semana["data_dt"].apply(_semana_iso_label)

            for (semana, tipo), gs in base_semana.groupby(["_semana", "_regra_turno"], dropna=False):
                dias_trabalhados = int(gs["data_dt"].dt.date.nunique())

                # Apenas excesso é violação.
                if dias_trabalhados <= 5:
                    continue

                ref_dt = gs["data_dt"].min()
                data_iso = ref_dt.strftime("%Y-%m-%d")
                if datas_filtro and data_iso not in datas_filtro:
                    continue

                turno_ref = str(gs["turno"].iloc[-1])
                resultados.append({
                    "cc": str(gs["cc"].iloc[-1]),
                    "nome": nome_colab,
                    "funcao": str(gs["funcao"].iloc[-1]),
                    "turno": turno_ref,
                    "tipo_jornada": tipo.capitalize(),
                    "periodo": f"{_primeira_data_semana(ref_dt)} a {_ultima_data_semana(ref_dt)}",
                    "data_referencia": ref_dt.strftime("%d/%m/%Y"),
                    "violacao": "Excesso de dias trabalhados na semana",
                    "detalhe": f"Foram identificados {dias_trabalhados} dia(s) trabalhado(s) na semana de segunda a domingo.",
                    "dias_trabalhados": dias_trabalhados,
                    "limite_esperado": "Até 5 dias/semana",
                    "status": "REVISAR",
                })

    quebras_12x36 = sum(1 for r in resultados if "12x36" in str(r.get("tipo_jornada", "")).lower())
    excessos_semanais = sum(1 for r in resultados if "semana" in str(r.get("violacao", "")).lower())

    kpis = {
        "registros": int(len(resultados)),
        "colaboradores": int(len(set(str(r.get("nome", "")).strip() for r in resultados if str(r.get("nome", "")).strip()))),
        "he_total": "00:00",
        "absent_total": "00:00",
        "saldo_negativo": 0,
        "revisar": int(len(resultados)),
        "quebras_12x36": int(quebras_12x36),
        "excessos_semanais": int(excessos_semanais),
    }

    for r in resultados:
        nome = str(r.get("nome", "") or "")
        data_ref = str(r.get("data_referencia", "") or "")
        if "12x36" in str(r.get("tipo_jornada", "")):
            r["_anotacao_id"] = _gerar_anotacao_id(nome, data_ref, "12x36")
        elif "semana" in str(r.get("violacao", "")).lower():
            periodo = str(r.get("periodo", "") or "")
            r["_anotacao_id"] = _gerar_anotacao_id(nome, periodo, "excesso_semanal")
        else:
            r["_anotacao_id"] = _gerar_anotacao_id(nome, data_ref, "violacao")

    dados_resp = resultados[:LIMITE_LINHAS_RESPOSTA_TABELA]
    meta_resp = {"limitado": len(resultados) > LIMITE_LINHAS_RESPOSTA_TABELA, "total_sem_limite": len(resultados), "limite_linhas": LIMITE_LINHAS_RESPOSTA_TABELA}
    payload = {"erro": "", "dados": dados_resp, "kpis": kpis, **meta_resp}
    try:
        _CACHE_DASHBOARD_EXECUTIVO_MEM["key"] = cache_key
        _CACHE_DASHBOARD_EXECUTIVO_MEM["payload"] = payload
        _CACHE_DASHBOARD_EXECUTIVO_MEM["gerado_em_ts"] = time.monotonic()
    except Exception:
        pass
    return jsonify(payload)


@app.route("/api/inter-jornada")
@login_obrigatorio
def api_inter_jornada():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "dados": [], "kpis": {}})

    df = normalizar_colunas(df)
    if df.empty:
        return jsonify({"erro": "", "dados": [], "kpis": {
            "registros": 0, "colaboradores": 0, "he_total": "00:00",
            "absent_total": "00:00", "saldo_negativo": 0, "revisar": 0
        }})

    busca = request.args.get("busca", "").strip().lower()
    cc_filtro = request.args.get("cc", "").strip()
    nome_filtro = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data = request.args.get("data", "").strip()
    datas = request.args.get("datas", "").strip()

    if busca:
        texto = df["funcao"].str.lower() + " " + df["turno"].str.lower()
        df = df[texto.str.contains(busca, na=False)]

    if cc_filtro:
        df = aplicar_filtro_cc_backend(df, cc_filtro)

    if nome_filtro:
        df = df[df["nome"].str.lower() == nome_filtro]

    if turno_filtro:
        df = df[df["turno"].str.lower() == turno_filtro]

    df = df.copy()
    df["data_dt"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")
    df = df[df["data_dt"].notna()].copy()

    if df.empty:
        return jsonify({"erro": "", "dados": [], "kpis": {
            "registros": 0, "colaboradores": 0, "he_total": "00:00",
            "absent_total": "00:00", "saldo_negativo": 0, "revisar": 0
        }})

    resultados = []

    for nome, g in df.sort_values("data_dt").groupby("nome", dropna=False):
        registros = []

        for _, row in g.sort_values("data_dt").iterrows():
            primeira, ultima = _primeira_ultima_marcacao(row)
            if primeira is None or ultima is None:
                continue

            registros.append({
                "row": row,
                "data_dt": row["data_dt"],
                "primeira": primeira,
                "ultima": ultima,
            })

        for anterior, atual in zip(registros, registros[1:]):
            data_anterior = anterior["data_dt"]
            data_atual = atual["data_dt"]

            # Diferença entre última marcação do dia anterior e primeira marcação do próximo dia.
            fim_abs = int(data_anterior.timestamp() // 60) + anterior["ultima"][1]
            inicio_abs = int(data_atual.timestamp() // 60) + atual["primeira"][1]
            intervalo = inicio_abs - fim_abs

            if intervalo < 0:
                continue

            if intervalo < 11 * 60:
                row_atual = atual["row"]
                row_anterior = anterior["row"]

                # Filtra por data de retorno/primeira marcação quando aplicável.
                data_retorno_iso = data_atual.strftime("%Y-%m-%d")
                if datas:
                    lista = [d.strip() for d in datas.split(",") if d.strip()]
                    if lista and data_retorno_iso not in lista:
                        continue
                elif data and data_retorno_iso != data:
                    continue

                resultado = row_atual.to_dict()
                resultado.update({
                    "data_anterior": data_anterior.strftime("%d/%m/%Y"),
                    "ultima_marcacao_anterior": anterior["ultima"][2],
                    "data_retorno": data_atual.strftime("%d/%m/%Y"),
                    "primeira_marcacao_retorno": atual["primeira"][2],
                    "intervalo_interjornada": _minutos_para_hhmm_simples(intervalo),
                    "deficit_interjornada": _minutos_para_hhmm_simples((11 * 60) - intervalo),
                    "inconsistencias": "Interjornada inferior a 11h",
                })
                resultados.append(resultado)

    kpis = {
        "registros": int(len(resultados)),
        "colaboradores": int(len(set(str(r.get("nome", "")).strip() for r in resultados if str(r.get("nome", "")).strip()))),
        "he_total": "00:00",
        "absent_total": "00:00",
        "saldo_negativo": 0,
        "revisar": int(len(resultados)),
    }

    for r in resultados:
        nome = str(r.get("nome", "") or "")
        data_ant = str(r.get("data_anterior", "") or "")
        data_ret = str(r.get("data_retorno", "") or "")
        r["_anotacao_id"] = _gerar_anotacao_id(nome, f"{data_ant}|{data_ret}", "interjornada")

    dados_resp = resultados[:LIMITE_LINHAS_RESPOSTA_TABELA]
    meta_resp = {"limitado": len(resultados) > LIMITE_LINHAS_RESPOSTA_TABELA, "total_sem_limite": len(resultados), "limite_linhas": LIMITE_LINHAS_RESPOSTA_TABELA}
    return jsonify({"erro": "", "dados": dados_resp, "kpis": kpis, **meta_resp})


@app.route("/api/inconsistencias-opcoes")
@login_obrigatorio
def api_inconsistencias_opcoes():
    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "opcoes": []})

    df = normalizar_colunas(df)
    if df.empty or "inconsistencias" not in df.columns:
        return jsonify({"erro": "", "opcoes": []})

    valores = []
    for texto in df["inconsistencias"].dropna().astype(str):
        texto = texto.strip()
        if not texto or texto.lower() in {"nan", "none"}:
            continue

        partes = [p.strip() for p in texto.split(";") if p.strip()]
        valores.extend(partes if partes else [texto])

    opcoes = sorted(set(valores), key=lambda x: x.lower())
    return jsonify({"erro": "", "opcoes": opcoes})



@app.route("/api/importar-excel-robo", methods=["POST", "GET"])
@login_obrigatorio
@admin_obrigatorio
def api_importar_excel_robo():
    registrar_auditoria("IMPORTAR_EXCEL_ROBO", "Solicitação de importação manual", "INICIO")
    resultado = sincronizar_excel_robo(forcar=True, corrigir_saldo_pdf=str(request.args.get("corrigir_pdf", "")).lower() in {"1", "true", "sim"})
    status = 200 if resultado.get("ok") else 404

    if resultado.get("ok"):
        _ativar_base_operacional("resultado_lote_consultas_validado.xlsx")
        # v8.11.21: importação precisa ser atômica e leve. Não gerar resumo do
        # Banco de Horas nem disparar notificações dentro da requisição, pois isso
        # mantinha a chamada HTTP aberta e travava a navegação depois da importação.
        # Banco/Extrato/Dashboard passam a consumir o Excel sob demanda.
        # v8.11.22: não pré-aquece Excel imediatamente após importação.
        # A leitura pesada era disparada no mesmo momento em que o navegador
        # tentava atualizar filtros/guias, causando congelamento. A base será
        # carregada sob demanda com leitor otimizado.
        pass
        resultado["banco_horas_excel"] = {
            "ok": True,
            "executado": False,
            "mensagem": "Resumo será montado sob demanda a partir do Excel oficial."
        }
        resultado["notificacoes_jornada"] = {
            "ok": True,
            "executado": False,
            "mensagem": "Verificação automática ao importar foi adiada para preservar a navegação."
        }
        if "mensagem" not in resultado:
            resultado["mensagem"] = "Excel importado com sucesso em modo leve."

    registrar_auditoria(
        "IMPORTAR_EXCEL_ROBO",
        resultado.get("mensagem", "") + (" | copiado=" + str(resultado.get("copiado"))),
        "OK" if resultado.get("ok") else "ERRO",
    )
    registrar_historico_operacao_v8190(
        "Importação Excel",
        "OK" if resultado.get("ok") else "ERRO",
        resultado.get("mensagem", ""),
        {"importacao_excel": resultado},
        origem="Administração/Topbar",
    )
    return jsonify(resultado), status


@app.route("/api/pdfs-banco-horas/atualizar-historico", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_atualizar_historico_pdfs_banco_horas():
    registrar_auditoria("ATUALIZAR_HISTORICO_PDFS_BH", "Atualização manual do cache de PDFs", "INICIO")
    resultado = atualizar_cache_pdfs_banco_horas()
    registrar_auditoria(
        "ATUALIZAR_HISTORICO_PDFS_BH",
        resultado.get("mensagem", ""),
        "OK" if resultado.get("ok") else "ERRO",
    )
    return jsonify(resultado), (200 if resultado.get("ok") else 500)


@app.route("/api/banco-horas/corrigir-pelos-pdfs", methods=["POST", "GET"])
@app.route("/api/banco-horas/corrigir-pelos-pdfs/", methods=["POST", "GET"])
@login_obrigatorio
@admin_obrigatorio
def api_corrigir_banco_horas_pelos_pdfs():
    """Inicia a correção/resumo do Banco de Horas em segundo plano.

    v8.10.43:
    - A versão anterior ainda mantinha a requisição HTTP aberta enquanto gerava
      o resumo de todos os colaboradores; em bases maiores o navegador estourava
      com "Failed to fetch".
    - Agora o botão apenas dispara um job assíncrono e retorna imediatamente.
    - O front-end consulta /status até concluir.
    """
    try:
        payload = request.get_json(silent=True) or {}
        atualizar_cache = bool(payload.get("atualizar_cache", False))

        if not ARQUIVO_PADRAO.exists():
            return jsonify({
                "ok": False,
                "mensagem": "Excel do sistema web não encontrado. Importe o Excel do Robô antes de corrigir o Banco de Horas.",
            }), 500

        with BH_CORRECAO_LOCK:
            if BH_CORRECAO_JOB.get("status") == "rodando":
                return jsonify({
                    "ok": True,
                    "status": "rodando",
                    "em_processamento": True,
                    "mensagem": "Correção do Banco de Horas já está em andamento. Aguarde a conclusão.",
                    "job": dict(BH_CORRECAO_JOB),
                }), 202

            BH_CORRECAO_JOB.clear()
            BH_CORRECAO_JOB.update({
                "status": "rodando",
                "ok": True,
                "mensagem": "Correção do Banco de Horas iniciada em segundo plano.",
                "iniciado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "finalizado_em": "",
                "segundos": 0,
                "resultado": {},
            })

        def _worker_corrigir_bh_pdf(atualizar_cache_worker=False):
            t0 = time.monotonic()
            resultado = {"ok": False, "mensagem": "Correção não concluída."}
            try:
                registrar_auditoria("CORRIGIR_BH_PDFS", "Correção assíncrona do Banco de Horas pelo cache de PDFs", "INICIO")

                cache = carregar_cache_pdfs_banco_horas()
                if atualizar_cache_worker or not cache or not cache.get("catalogo"):
                    atualizacao = atualizar_cache_pdfs_banco_horas()
                    if not atualizacao.get("ok"):
                        resultado = {
                            "ok": False,
                            "mensagem": atualizacao.get("mensagem", "Falha ao atualizar histórico/cache de PDFs."),
                            "atualizacao_cache": atualizacao,
                            "modo": "assincrono_cache_sem_regravar_excel",
                        }
                    else:
                        cache = carregar_cache_pdfs_banco_horas()
                        atualizacao = atualizacao
                else:
                    atualizacao = {"ok": True, "mensagem": "Cache de PDFs já existente; atualização não necessária."}

                if resultado.get("ok") is False and resultado.get("atualizacao_cache"):
                    pass
                elif not cache or not cache.get("catalogo"):
                    resultado = {
                        "ok": False,
                        "mensagem": "Histórico/cache de PDFs vazio. Clique primeiro em Atualizar Histórico de PDFs.",
                        "modo": "assincrono_cache_sem_regravar_excel",
                    }
                else:
                    resumo_bh = gerar_resumo_banco_horas_pdf()
                    if not resumo_bh.get("ok"):
                        resultado = {
                            "ok": False,
                            "mensagem": resumo_bh.get("mensagem", "Falha ao gerar resumo do Banco de Horas."),
                            "resumo_banco_horas_pdf": resumo_bh,
                            "atualizacao_cache": atualizacao,
                            "modo": "assincrono_cache_sem_regravar_excel",
                        }
                    else:
                        linhas = resumo_bh.get("linhas", []) or []
                        resultado = {
                            "ok": True,
                            "executado": True,
                            "modo": "assincrono_cache_sem_regravar_excel",
                            "mensagem": "Banco de Horas corrigido pelo histórico/cache dos PDFs em segundo plano. " + resumo_bh.get("mensagem", ""),
                            "linhas_corrigidas": len(linhas),
                            "linhas_sem_match": 0,
                            "pdfs_lidos": len((cache or {}).get("catalogo", []) or []),
                            "cache_gerado_em": (cache or {}).get("gerado_em", ""),
                            "resumo_banco_horas_pdf": resumo_bh,
                            "atualizacao_cache": atualizacao,
                        }

                resultado["segundos"] = round(time.monotonic() - t0, 1)
                registrar_auditoria("CORRIGIR_BH_PDFS", resultado.get("mensagem", ""), "OK" if resultado.get("ok") else "ERRO")
            except Exception as e:
                resultado = {
                    "ok": False,
                    "erro": f"Falha inesperada ao corrigir Banco de Horas pelos PDFs: {e}",
                    "mensagem": f"Falha inesperada ao corrigir Banco de Horas pelos PDFs: {e}",
                    "modo": "assincrono_cache_sem_regravar_excel",
                    "segundos": round(time.monotonic() - t0, 1),
                }
                try:
                    registrar_auditoria("CORRIGIR_BH_PDFS", resultado.get("mensagem", ""), "ERRO")
                except Exception:
                    pass
            finally:
                with BH_CORRECAO_LOCK:
                    BH_CORRECAO_JOB.update({
                        "status": "concluido" if resultado.get("ok") else "erro",
                        "ok": bool(resultado.get("ok")),
                        "mensagem": resultado.get("mensagem", ""),
                        "finalizado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "segundos": resultado.get("segundos", round(time.monotonic() - t0, 1)),
                        "resultado": resultado,
                    })

        thread = threading.Thread(target=_worker_corrigir_bh_pdf, args=(atualizar_cache,), daemon=True)
        thread.start()

        return jsonify({
            "ok": True,
            "status": "rodando",
            "em_processamento": True,
            "mensagem": "Correção do Banco de Horas iniciada em segundo plano. Acompanhe o andamento na tela.",
            "job": dict(BH_CORRECAO_JOB),
        }), 202
    except Exception as e:
        return jsonify({
            "ok": False,
            "erro": f"Falha ao iniciar correção do Banco de Horas pelos PDFs: {e}",
            "mensagem": f"Falha ao iniciar correção do Banco de Horas pelos PDFs: {e}",
        }), 500


@app.route("/api/banco-horas/corrigir-pelos-pdfs/status", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_status_corrigir_banco_horas_pelos_pdfs():
    with BH_CORRECAO_LOCK:
        job = dict(BH_CORRECAO_JOB)
    return jsonify({
        "ok": bool(job.get("ok", True)),
        "status": job.get("status", "ocioso"),
        "em_processamento": job.get("status") == "rodando",
        "mensagem": job.get("mensagem", ""),
        "job": job,
        "resultado": job.get("resultado", {}),
    }), 200


@app.route("/api/pdfs-banco-horas/cache", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_status_cache_pdfs_banco_horas():
    cache = carregar_cache_pdfs_banco_horas()
    if not cache:
        return jsonify({"ok": True, "existe": False, "mensagem": "Histórico/cache de PDFs ainda não criado."})
    return jsonify({
        "ok": True,
        "existe": True,
        "cache": str(ARQUIVO_CACHE_PDFS_BANCO_HORAS),
        "gerado_em": cache.get("gerado_em"),
        "pasta_pdfs_robo": cache.get("pasta_pdfs_robo"),
        "pdfs_encontrados": cache.get("pdfs_encontrados"),
        "quadros_lidos": cache.get("quadros_lidos") or len(cache.get("catalogo", [])),
        "segundos": cache.get("segundos"),
    })



# ============================================================
# v8.19.0 - Central de Operacao
# ============================================================
def _agora_iso_v8190():
    return datetime.now().isoformat(timespec="seconds")


def _contar_pdfs_pasta_v8190(pasta):
    try:
        p = Path(pasta)
        if not p.exists():
            return 0
        return len([x for x in p.glob("*.pdf") if x.is_file()])
    except Exception:
        return 0


def _resumir_resultado_operacao_v8190(resultado):
    resultado = resultado or {}
    launcher = resultado.get("launcher") or {}
    imp = resultado.get("importacao_excel") or resultado.get("importacao") or {}
    partes = []
    if launcher:
        if launcher.get("returncode") is not None:
            partes.append(f"Launcher código {launcher.get('returncode')}")
        if launcher.get("ok") is not None:
            partes.append("launcher ok" if launcher.get("ok") else "launcher com erro")
    if imp:
        msg = imp.get("mensagem") or imp.get("erro") or "importação registrada"
        partes.append(str(msg)[:180])
    return " · ".join(partes)[:500]


def registrar_historico_operacao_v8190(tipo, status, mensagem="", resultado=None, segundos=None, origem="Tempo Fechado"):
    """Registra uma linha de histórico para a Central de Operação.
    Formato JSONL para ser simples, robusto e tolerante a falhas.
    """
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        base = Path.home() / "ponto_pdfs"
        entrada = base / "entrada"
        saida = base / "saida"
        processados = base / "processados"
        payload = {
            "quando": _agora_iso_v8190(),
            "tipo": str(tipo or "Operação"),
            "status": str(status or "INFO"),
            "mensagem": str(mensagem or "")[:1000],
            "resumo": _resumir_resultado_operacao_v8190(resultado),
            "segundos": segundos,
            "origem": str(origem or "Tempo Fechado"),
            "pdfs_entrada": _contar_pdfs_pasta_v8190(entrada),
            "pdfs_saida": _contar_pdfs_pasta_v8190(saida),
            "pdfs_processados": _contar_pdfs_pasta_v8190(processados),
        }
        if resultado is not None:
            payload["resultado"] = resultado
        with ARQUIVO_HISTORICO_OPERACAO_V8190.open("a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False, default=str) + "\n")
    except Exception:
        pass


def _ler_historico_operacao_v8190(limite=30):
    try:
        if not ARQUIVO_HISTORICO_OPERACAO_V8190.exists():
            return []
        linhas = ARQUIVO_HISTORICO_OPERACAO_V8190.read_text(encoding="utf-8", errors="ignore").splitlines()
        eventos = []
        for linha in linhas[-max(int(limite or 30), 1):]:
            try:
                eventos.append(json.loads(linha))
            except Exception:
                continue
        return list(reversed(eventos))
    except Exception:
        return []


def _idade_arquivo_v8191(caminho):
    try:
        p = Path(caminho)
        if not p.exists():
            return None
        segundos = max(0, int(time.time() - p.stat().st_mtime))
        if segundos < 60:
            return "agora"
        if segundos < 3600:
            return f"{segundos // 60} min"
        if segundos < 86400:
            return f"{segundos // 3600} h"
        return f"{segundos // 86400} dia(s)"
    except Exception:
        return None


def _info_caminho_v8191(caminho, tipo="arquivo"):
    try:
        p = Path(caminho)
        existe = p.exists()
        info = {
            "caminho": str(p),
            "existe": existe,
            "tipo": tipo,
            "idade": _idade_arquivo_v8191(p),
        }
        if existe and p.is_file():
            info["tamanho_kb"] = round(p.stat().st_size / 1024, 1)
        return info
    except Exception as e:
        return {"caminho": str(caminho), "existe": False, "tipo": tipo, "erro": str(e)}


def _status_saude_operacional_v8190(eventos):
    """v8.19.1 - Diagnóstico ampliado da Saúde do Robô.
    Mantém o contrato anterior do endpoint e adiciona sinais objetivos para a Central.
    """
    eventos = eventos or []
    cfg = normalizar_config_robo_runtime_v82123(carregar_config_robo()) if "carregar_config_robo" in globals() and "normalizar_config_robo_runtime_v82123" in globals() else (carregar_config_robo() if "carregar_config_robo" in globals() else {})

    base = Path(cfg.get("pasta_pdfs_robo") or Path.home() / "ponto_pdfs")
    entrada = base / "entrada"
    saida = base / "saida"
    processados = base / "processados"
    erro = base / "erro"
    excel = _normalizar_path_windows(cfg.get("arquivo_excel_robo"), _excel_robo_padrao_v81737()) if "_normalizar_path_windows" in globals() else Path(cfg.get("arquivo_excel_robo") or Path.home() / "ponto_pdfs" / "saida" / "resultado_lote_consultas_validado.xlsx")
    launcher = _normalizar_path_windows(cfg.get("script_launcher_robo"), BASE_DIR / "Script_Launcher_v5_filtrado.py") if "_normalizar_path_windows" in globals() else Path(cfg.get("script_launcher_robo") or BASE_DIR / "Script_Launcher_v5_filtrado.py")

    qtd_entrada = _contar_pdfs_pasta_v8190(entrada)
    qtd_saida = _contar_pdfs_pasta_v8190(saida)
    qtd_processados = _contar_pdfs_pasta_v8190(processados)
    qtd_erro = _contar_pdfs_pasta_v8190(erro)
    ultimo = eventos[0] if eventos else None
    recentes = eventos[:10]
    erros_recentes = [e for e in recentes if str(e.get("status", "")).upper() in {"ERRO", "FALHA"}]

    alertas = []
    diagnosticos = []
    nivel_rank = {"ok": 0, "atencao": 1, "erro": 2, "sem_historico": 0}
    nivel = "ok"

    def elevar(novo):
        nonlocal nivel
        if nivel_rank.get(novo, 0) > nivel_rank.get(nivel, 0):
            nivel = novo

    def add_diag(nome, estado, detalhe, nivel_item="ok"):
        diagnosticos.append({
            "nome": nome,
            "estado": estado,
            "detalhe": detalhe,
            "nivel": nivel_item,
        })

    def add_alerta(nivel_item, titulo, descricao):
        alertas.append({"nivel": nivel_item, "titulo": titulo, "descricao": descricao})
        if nivel_item == "alto":
            elevar("erro")
        elif nivel_item == "medio":
            elevar("atencao")

    # Pastas operacionais
    for nome, pasta in (("Entrada", entrada), ("Saída", saida), ("Processados", processados), ("Erro", erro)):
        if pasta.exists():
            add_diag(f"Pasta {nome}", "OK", str(pasta), "ok")
        else:
            add_diag(f"Pasta {nome}", "Atenção", f"Pasta não encontrada: {pasta}", "warning")
            add_alerta("medio", f"Pasta {nome} ausente", f"A pasta operacional não foi localizada: {pasta}")

    if qtd_entrada > 0:
        add_alerta("medio", "PDFs aguardando entrada", f"Há {qtd_entrada} PDF(s) aguardando processamento na pasta de entrada.")
    else:
        add_diag("Fila de entrada", "Livre", "Nenhum PDF aguardando na entrada.", "ok")

    if qtd_erro > 0:
        add_alerta("alto", "PDFs na pasta de erro", f"Há {qtd_erro} PDF(s) na pasta de erro. Vale revisar os arquivos antes da próxima rodada.")
    else:
        add_diag("Pasta de erro", "Limpa", "Nenhum PDF parado na pasta de erro.", "ok")

    # Arquivos-chave
    if excel.exists():
        add_diag("Excel do Robô", "OK", f"Localizado · atualizado há {_idade_arquivo_v8191(excel) or '-'}", "ok")
    else:
        add_diag("Excel do Robô", "Atenção", f"Arquivo não localizado: {excel}", "warning")
        add_alerta("alto", "Excel do robô ausente", "O arquivo de saída padrão ainda não foi encontrado.")

    if launcher.exists():
        add_diag("Script Launcher", "OK", f"Localizado · atualizado há {_idade_arquivo_v8191(launcher) or '-'}", "ok")
    else:
        add_diag("Script Launcher", "Erro", f"Script não encontrado: {launcher}", "danger")
        add_alerta("alto", "Script Launcher ausente", "O script configurado para o robô não foi encontrado.")

    politica = "inclui e-mails lidos recentes" if bool(cfg.get("capturar_emails_lidos", False)) else "somente e-mails não lidos"
    add_diag("Política Outlook", "Configurada", politica, "info")
    add_diag("Histórico operacional", "Ativo" if eventos else "Inicial", f"{len(eventos)} evento(s) carregado(s) na consulta atual.", "info")

    # Histórico e última execução
    if ultimo and str(ultimo.get("status", "")).upper() in {"ERRO", "FALHA"}:
        add_alerta("alto", "Última operação com erro", str(ultimo.get("mensagem") or "A última operação registrada terminou com erro.")[:220])
    elif ultimo:
        add_diag("Última operação", str(ultimo.get("status", "INFO")).upper(), f"{ultimo.get('tipo', 'Operação')} em {ultimo.get('quando', '-')}", "ok")

    if len(erros_recentes) >= 3:
        add_alerta("alto", "Erros recorrentes", f"Foram encontrados {len(erros_recentes)} erros nas últimas {len(recentes)} operações registradas.")
    elif erros_recentes:
        add_alerta("medio", "Erro recente", f"Há {len(erros_recentes)} erro(s) nas últimas {len(recentes)} operações.")

    if not eventos:
        nivel = "sem_historico"
        add_alerta("baixo", "Histórico inicial", "A Central de Operação começa a registrar eventos a partir da v8.19.0.")

    if nivel == "erro":
        titulo = "Ação recomendada"
        descricao = "A Saúde do Robô encontrou item crítico que merece revisão antes da próxima execução."
    elif nivel == "atencao":
        titulo = "Atenção operacional"
        descricao = "Há sinais para acompanhar, mas o robô não está necessariamente parado."
    elif nivel == "sem_historico":
        titulo = "Sem histórico registrado"
        descricao = "Execute uma importação ou processamento para formar o primeiro diário de bordo."
    else:
        titulo = "Operando normalmente"
        descricao = "Pastas, configuração e últimos eventos não indicam bloqueio operacional."

    score = 100
    score -= 25 if nivel == "erro" else 0
    score -= 10 if nivel == "atencao" else 0
    score -= min(len(erros_recentes) * 8, 24)
    score -= 10 if qtd_entrada > 0 else 0
    score -= 15 if qtd_erro > 0 else 0
    score = max(0, min(100, score))

    return {
        "nivel": nivel,
        "titulo": titulo,
        "descricao": descricao,
        "score": score,
        "alertas": alertas[:10],
        "diagnosticos": diagnosticos[:16],
        "pdfs_entrada": qtd_entrada,
        "pdfs_saida": qtd_saida,
        "pdfs_processados": qtd_processados,
        "pdfs_erro": qtd_erro,
        "excel_existe": excel.exists(),
        "excel": str(excel),
        "launcher_existe": launcher.exists(),
        "launcher": str(launcher),
        "pasta_base": str(base),
        "politica_outlook": politica,
        "ultima_operacao": ultimo,
        "erros_recentes": len(erros_recentes),
        "arquivos": {
            "excel": _info_caminho_v8191(excel, "arquivo"),
            "launcher": _info_caminho_v8191(launcher, "arquivo"),
            "historico": _info_caminho_v8191(ARQUIVO_HISTORICO_OPERACAO_V8190, "arquivo"),
        },
    }



# ============================================================
# v8.19.2 - Alertas Operacionais Inteligentes
# ============================================================
def _parse_iso_v8192(valor):
    try:
        if not valor:
            return None
        return datetime.fromisoformat(str(valor).replace('Z', '+00:00').replace('+00:00', ''))
    except Exception:
        return None


def _gerar_alertas_inteligentes_v8192(saude, eventos):
    """Gera alertas priorizados com acao sugerida.
    Nao altera a regra operacional: apenas interpreta sinais ja disponiveis na Central.
    """
    saude = saude or {}
    eventos = eventos or []
    alertas = []
    vistos = set()

    prioridade_rank = {"alto": 3, "medio": 2, "baixo": 1, "info": 0}

    def add(prioridade, categoria, titulo, descricao, acao, origem="Central de Operação"):
        chave = (str(titulo).lower(), str(categoria).lower())
        if chave in vistos:
            return
        vistos.add(chave)
        alertas.append({
            "prioridade": prioridade,
            "categoria": categoria,
            "titulo": titulo,
            "descricao": descricao,
            "acao": acao,
            "origem": origem,
        })

    # Reaproveita alertas da Saúde do Robô, mas traduz para linguagem de ação.
    for a in saude.get("alertas", []) or []:
        nivel = str(a.get("nivel", "baixo")).lower()
        prioridade = "alto" if nivel == "alto" else "medio" if nivel == "medio" else "baixo"
        titulo = str(a.get("titulo") or "Alerta operacional")
        desc = str(a.get("descricao") or "Há um sinal operacional para acompanhamento.")
        acao = "Revisar o item indicado antes da próxima execução."
        tl = titulo.lower()
        if "entrada" in tl or "aguardando" in tl:
            acao = "Abrir a pasta de entrada e executar o processamento ou remover arquivos indevidos."
        elif "erro" in tl:
            acao = "Abrir a pasta de erro, identificar o PDF problemático e reprocessar após correção."
        elif "excel" in tl:
            acao = "Conferir o caminho do Excel do Robô na Administração do Robô e gerar/importar novamente se necessário."
        elif "launcher" in tl or "script" in tl:
            acao = "Conferir o caminho do Script Launcher na Administração do Robô."
        elif "histórico" in tl or "historico" in tl:
            acao = "Executar uma importação ou processamento para iniciar o diário de bordo."
        add(prioridade, "Saúde do Robô", titulo, desc, acao, "Diagnóstico")

    qtd_entrada = int(saude.get("pdfs_entrada") or 0)
    qtd_erro = int(saude.get("pdfs_erro") or 0)
    erros_recentes = int(saude.get("erros_recentes") or 0)

    if qtd_entrada >= 20:
        add("alto", "Fila", "Fila de entrada elevada", f"Há {qtd_entrada} PDF(s) aguardando processamento.", "Priorizar o Processamento Manual ou Integrado e verificar se a captura Outlook está acumulando arquivos.")
    elif qtd_entrada >= 5:
        add("medio", "Fila", "Fila de entrada em crescimento", f"Há {qtd_entrada} PDF(s) aguardando processamento.", "Executar processamento ainda nesta rodada para evitar acúmulo.")

    if qtd_erro >= 5:
        add("alto", "Erros", "Muitos PDFs na pasta de erro", f"Há {qtd_erro} PDF(s) na pasta de erro.", "Revisar os arquivos da pasta de erro antes de novas execuções em lote.")
    elif qtd_erro > 0:
        add("medio", "Erros", "PDFs pendentes de revisão", f"Há {qtd_erro} PDF(s) na pasta de erro.", "Abrir a pasta de erro e decidir se os arquivos devem ser corrigidos, arquivados ou reprocessados.")

    if erros_recentes >= 3:
        add("alto", "Histórico", "Erros recorrentes detectados", f"Foram encontrados {erros_recentes} erros no histórico recente.", "Verificar as últimas operações no diário de bordo e corrigir a causa antes de novas execuções.")

    ultimo = saude.get("ultima_operacao") or (eventos[0] if eventos else None)
    if ultimo:
        dt = _parse_iso_v8192(ultimo.get("quando"))
        if dt:
            horas = (datetime.now() - dt).total_seconds() / 3600
            if horas >= 48:
                add("medio", "Rotina", "Robô sem execução recente", f"A última operação registrada ocorreu há cerca de {int(horas)} hora(s).", "Executar uma verificação manual ou confirmar se a rotina automática está ativa.")
            elif horas >= 24:
                add("baixo", "Rotina", "Última execução há mais de 24 horas", f"A última operação registrada ocorreu há cerca de {int(horas)} hora(s).", "Acompanhar se haverá nova captura/processamento no próximo ciclo operacional.")

    # Leitura semântica simples das mensagens recentes para capturar sinais úteis.
    textos_recentes = " \n".join([str(e.get("mensagem", "")) + " " + str(e.get("resumo", "")) for e in eventos[:8]]).lower()
    if "nenhum" in textos_recentes and ("email" in textos_recentes or "e-mail" in textos_recentes) and "não lido" in textos_recentes:
        add("baixo", "Outlook", "Sem e-mails não lidos recentes", "O histórico recente indica busca sem novos e-mails não lidos.", "Confirmar se o remetente/assunto configurado está correto ou aguardar nova remessa.")
    if "limite" in textos_recentes and "pdf" in textos_recentes:
        add("medio", "Limite", "Limite de PDFs possivelmente atingido", "O histórico recente menciona limite de PDFs.", "Conferir o limite configurado na Administração do Robô antes de nova captura em lote.")

    if not alertas:
        add("info", "Operação", "Sem alertas inteligentes ativos", "Nenhum sinal operacional relevante foi encontrado neste momento.", "Manter acompanhamento normal pela Central de Operação.")

    alertas.sort(key=lambda a: prioridade_rank.get(a.get("prioridade", "info"), 0), reverse=True)
    resumo = {
        "total": len(alertas),
        "altos": sum(1 for a in alertas if a.get("prioridade") == "alto"),
        "medios": sum(1 for a in alertas if a.get("prioridade") == "medio"),
        "baixos": sum(1 for a in alertas if a.get("prioridade") == "baixo"),
        "infos": sum(1 for a in alertas if a.get("prioridade") == "info"),
    }
    return alertas[:12], resumo



# ============================================================
# v8.21.38 - Painel de Implantacao / Log Amigavel
# ============================================================
def _status_classe_log_amigavel_v8193(status):
    st = str(status or "INFO").upper()
    if st == "OK":
        return "ok"
    if st in {"ERRO", "FALHA"}:
        return "danger"
    return "info"


def _quando_amigavel_v8193(valor):
    dt = _parse_iso_v8192(valor) if "_parse_iso_v8192" in globals() else None
    if not dt:
        return str(valor or "-")
    try:
        delta = datetime.now() - dt
        segundos = max(0, int(delta.total_seconds()))
        if segundos < 60:
            return "agora há pouco"
        if segundos < 3600:
            return f"há {segundos // 60} min"
        if segundos < 86400:
            return f"há {segundos // 3600} h"
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return dt.strftime("%d/%m/%Y %H:%M")


def _frase_evento_amigavel_v8193(e):
    tipo = str(e.get("tipo") or "Operação")
    status = str(e.get("status") or "INFO").upper()
    mensagem = str(e.get("mensagem") or "").strip()
    resumo = str(e.get("resumo") or "").strip()
    origem = str(e.get("origem") or "Tempo Fechado")
    segundos = e.get("segundos")
    pdfs_entrada = e.get("pdfs_entrada", 0)
    tipo_l = tipo.lower()

    if status == "OK":
        if "manual" in tipo_l:
            titulo = "Processamento manual concluído"
            frase = "O robô manual terminou a execução e registrou o resultado no diário de bordo."
        elif "integrado" in tipo_l or "launcher" in tipo_l:
            titulo = "Processamento integrado concluído"
            frase = "O processamento integrado terminou sem erro registrado."
        elif "excel" in tipo_l or "import" in tipo_l:
            titulo = "Importação concluída"
            frase = "A importação foi concluída e os dados ficaram disponíveis para consulta."
        else:
            titulo = f"{tipo} concluído"
            frase = "A operação foi registrada como concluída."
        acao = "Nenhuma ação imediata necessária."
    elif status in {"ERRO", "FALHA"}:
        titulo = f"Atenção em {tipo}"
        frase = mensagem or "A operação terminou com erro."
        acao = "Abrir o detalhe técnico ou repetir a execução após revisar a configuração indicada."
    else:
        titulo = f"{tipo} registrado"
        frase = mensagem or "O evento foi registrado para acompanhamento."
        acao = "Acompanhar o próximo ciclo operacional."

    detalhe_partes = []
    if resumo:
        detalhe_partes.append(resumo)
    elif mensagem and mensagem != frase:
        detalhe_partes.append(mensagem)
    if segundos not in (None, "", 0):
        detalhe_partes.append(f"duração aproximada: {segundos}s")
    detalhe_partes.append(f"entrada no momento: {pdfs_entrada} PDF(s)")
    detalhe = " · ".join(str(x) for x in detalhe_partes if x)[:500]

    return {
        "quando": e.get("quando"),
        "quando_amigavel": _quando_amigavel_v8193(e.get("quando")),
        "tipo": tipo,
        "status": status,
        "classe": _status_classe_log_amigavel_v8193(status),
        "titulo": titulo[:140],
        "frase": frase[:360],
        "detalhe": detalhe,
        "acao": acao,
        "origem": origem,
    }


def _gerar_log_amigavel_v8193(eventos, saude, alertas_inteligentes):
    """Monta um resumo operacional em linguagem simples para a Central de Operação.
    Não substitui o log técnico: apenas traduz os últimos eventos em texto de uso diário.
    """
    eventos = eventos or []
    saude = saude or {}
    alertas_inteligentes = alertas_inteligentes or []

    if not eventos:
        return {
            "titulo": "Diário de bordo ainda vazio",
            "resumo": "A Central de Operação está pronta, mas ainda não há execuções registradas nesta instalação.",
            "tom": "info",
            "linhas": [],
            "proximos_passos": ["Executar uma importação ou processamento para iniciar o Log Amigável."],
            "metricas": {"eventos": 0, "sucessos": 0, "erros": 0},
        }

    total = len(eventos)
    sucessos = sum(1 for e in eventos if str(e.get("status", "")).upper() == "OK")
    erros = sum(1 for e in eventos if str(e.get("status", "")).upper() in {"ERRO", "FALHA"})
    infos = max(0, total - sucessos - erros)
    ultimo = eventos[0]
    ultimo_status = str(ultimo.get("status") or "INFO").upper()
    linhas = [_frase_evento_amigavel_v8193(e) for e in eventos[:8]]

    if erros:
        tom = "danger" if ultimo_status in {"ERRO", "FALHA"} else "warning"
        titulo = "Operação com pontos de atenção"
        resumo = f"Nas últimas {total} operação(ões), houve {sucessos} sucesso(s), {erros} erro(s) e {infos} registro(s) informativo(s)."
    else:
        tom = "ok"
        titulo = "Operação sem erros recentes"
        resumo = f"Nas últimas {total} operação(ões), não há erro registrado. O robô está com diário de bordo saudável."

    proximos = []
    for a in alertas_inteligentes[:3]:
        acao = str(a.get("acao") or "").strip()
        if acao and acao not in proximos:
            proximos.append(acao)
    if not proximos:
        if int(saude.get("pdfs_entrada") or 0) > 0:
            proximos.append("Há PDFs na entrada: executar Processamento Integrado ou Manual quando apropriado.")
        else:
            proximos.append("Manter acompanhamento normal pela Central de Operação.")

    return {
        "titulo": titulo,
        "resumo": resumo,
        "tom": tom,
        "linhas": linhas,
        "proximos_passos": proximos[:3],
        "metricas": {"eventos": total, "sucessos": sucessos, "erros": erros, "informativos": infos},
    }



# ============================================================
# v8.21.38 - Painel de Implantacao
# ============================================================
def _agora_implantacao_v8210():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _id_implantacao_v8210():
    return uuid.uuid4().hex[:12]


def _normalizar_status_implantacao_v8210(valor):
    valor = str(valor or "pendente").strip().lower()
    permitidos = {"pendente", "em_configuracao", "homologado", "atencao", "erro", "desativado"}
    return valor if valor in permitidos else "pendente"


def _ler_painel_implantacao_v8210():
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        if not ARQUIVO_PAINEL_IMPLANTACAO_V8210.exists():
            payload = {"notebooks": [], "criado_em": _agora_implantacao_v8210(), "atualizado_em": _agora_implantacao_v8210()}
            _salvar_painel_implantacao_v8210(payload)
            return payload
        bruto = json.loads(ARQUIVO_PAINEL_IMPLANTACAO_V8210.read_text(encoding="utf-8") or "{}")
        if not isinstance(bruto, dict):
            bruto = {"notebooks": []}
        bruto.setdefault("notebooks", [])
        return bruto
    except Exception:
        return {"notebooks": [], "erro_leitura": True}


def _salvar_painel_implantacao_v8210(payload):
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    payload = payload if isinstance(payload, dict) else {"notebooks": []}
    payload["atualizado_em"] = _agora_implantacao_v8210()
    tmp = ARQUIVO_PAINEL_IMPLANTACAO_V8210.with_suffix(".tmp")
    tmp.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    tmp.replace(ARQUIVO_PAINEL_IMPLANTACAO_V8210)
    return payload


def _notebook_local_implantacao_v8210():
    try:
        host = socket.gethostname()
    except Exception:
        host = "Notebook local"
    return {
        "hostname": host,
        "sistema": platform.platform(),
        "usuario_windows": os.environ.get("USERNAME") or os.environ.get("USER") or "",
        "perfil": str(Path.home()),
        "versao_detectada": APP_VERSION,
    }


def _normalizar_notebook_implantacao_v8210(dados, existente=None):
    dados = dados or {}
    existente = existente or {}
    checklist = existente.get("checklist") if isinstance(existente.get("checklist"), dict) else {}
    checklist.update(dados.get("checklist") if isinstance(dados.get("checklist"), dict) else {})
    base_check = {
        "instalador": bool(checklist.get("instalador", False)),
        "login": bool(checklist.get("login", False)),
        "outlook": bool(checklist.get("outlook", False)),
        "pastas": bool(checklist.get("pastas", False)),
        "banco_horas": bool(checklist.get("banco_horas", False)),
        "central_operacao": bool(checklist.get("central_operacao", False)),
        "processamento": bool(checklist.get("processamento", False)),
        "backup_seguro": bool(checklist.get("backup_seguro", False)),
    }
    agora = _agora_implantacao_v8210()
    return {
        "id": str(dados.get("id") or existente.get("id") or _id_implantacao_v8210()),
        "nome": str(dados.get("nome") or existente.get("nome") or "Notebook").strip()[:120],
        "responsavel": str(dados.get("responsavel") or existente.get("responsavel") or "").strip()[:120],
        "setor": str(dados.get("setor") or existente.get("setor") or "").strip()[:120],
        "status": _normalizar_status_implantacao_v8210(dados.get("status") or existente.get("status")),
        "versao": str(dados.get("versao") or existente.get("versao") or APP_VERSION).strip()[:80],
        "ultima_homologacao": str(dados.get("ultima_homologacao") or existente.get("ultima_homologacao") or "").strip()[:40],
        "observacoes": str(dados.get("observacoes") or existente.get("observacoes") or "").strip()[:1000],
        "checklist": base_check,
        "criado_em": existente.get("criado_em") or agora,
        "atualizado_em": agora,
    }


def _resumo_painel_implantacao_v8210(notebooks):
    notebooks = notebooks or []
    total = len(notebooks)
    homologados = sum(1 for n in notebooks if n.get("status") == "homologado")
    atencao = sum(1 for n in notebooks if n.get("status") in {"atencao", "erro"})
    pendentes = sum(1 for n in notebooks if n.get("status") in {"pendente", "em_configuracao"})
    checklist_total = 0
    checklist_ok = 0
    for n in notebooks:
        checks = (n.get("checklist") or {}).values()
        for v in checks:
            checklist_total += 1
            checklist_ok += 1 if bool(v) else 0
    progresso = round((checklist_ok / checklist_total) * 100) if checklist_total else 0
    return {
        "total": total,
        "homologados": homologados,
        "pendentes": pendentes,
        "atencao": atencao,
        "progresso_checklist": progresso,
    }


@app.route("/api/painel-implantacao", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_painel_implantacao_v8210():
    payload = _ler_painel_implantacao_v8210()
    notebooks = payload.get("notebooks", []) if isinstance(payload.get("notebooks", []), list) else []
    return jsonify({
        "ok": True,
        "versao": APP_FULL_NAME,
        "gerado_em": _agora_txt() if "_agora_txt" in globals() else _agora_implantacao_v8210(),
        "arquivo": str(ARQUIVO_PAINEL_IMPLANTACAO_V8210),
        "notebook_local": _notebook_local_implantacao_v8210(),
        "resumo": _resumo_painel_implantacao_v8210(notebooks),
        "notebooks": notebooks,
    })


@app.route("/api/painel-implantacao/notebooks", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_salvar_notebook_implantacao_v8210():
    dados = request.get_json(silent=True) or {}
    payload = _ler_painel_implantacao_v8210()
    notebooks = payload.get("notebooks", []) if isinstance(payload.get("notebooks", []), list) else []
    alvo_id = str(dados.get("id") or "").strip()
    novo = []
    atualizado = False
    for n in notebooks:
        if alvo_id and str(n.get("id")) == alvo_id:
            novo.append(_normalizar_notebook_implantacao_v8210(dados, n))
            atualizado = True
        else:
            novo.append(n)
    if not atualizado:
        novo.append(_normalizar_notebook_implantacao_v8210(dados))
    payload["notebooks"] = novo
    _salvar_painel_implantacao_v8210(payload)
    registrar_auditoria("PAINEL_IMPLANTACAO", "Notebook salvo no painel de implantação", "OK")
    return jsonify({"ok": True, "mensagem": "Notebook salvo no Painel de Implantação.", "resumo": _resumo_painel_implantacao_v8210(novo), "notebooks": novo})


@app.route("/api/painel-implantacao/notebooks/<notebook_id>", methods=["DELETE"])
@login_obrigatorio
@admin_obrigatorio
def api_excluir_notebook_implantacao_v8210(notebook_id):
    payload = _ler_painel_implantacao_v8210()
    notebooks = payload.get("notebooks", []) if isinstance(payload.get("notebooks", []), list) else []
    novo = [n for n in notebooks if str(n.get("id")) != str(notebook_id)]
    payload["notebooks"] = novo
    _salvar_painel_implantacao_v8210(payload)
    registrar_auditoria("PAINEL_IMPLANTACAO", f"Notebook removido do painel: {notebook_id}", "OK")
    return jsonify({"ok": True, "mensagem": "Notebook removido do Painel de Implantação.", "resumo": _resumo_painel_implantacao_v8210(novo), "notebooks": novo})


@app.route("/api/central-operacao", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_central_operacao_v8190():
    eventos = _ler_historico_operacao_v8190(limite=int(request.args.get("limite", 30) or 30))
    saude = _status_saude_operacional_v8190(eventos)
    alertas_inteligentes, resumo_alertas = _gerar_alertas_inteligentes_v8192(saude, eventos)
    log_amigavel = _gerar_log_amigavel_v8193(eventos, saude, alertas_inteligentes)
    job = _job_processamento_integrado_atual() if "_job_processamento_integrado_atual" in globals() else {}
    kpis = {
        "eventos": len(eventos),
        "sucessos": sum(1 for e in eventos if str(e.get("status", "")).upper() == "OK"),
        "erros": sum(1 for e in eventos if str(e.get("status", "")).upper() in {"ERRO", "FALHA"}),
        "pdfs_entrada": saude.get("pdfs_entrada", 0),
        "alertas_altos": resumo_alertas.get("altos", 0),
    }
    return jsonify({
        "ok": True,
        "versao": APP_FULL_NAME,
        "gerado_em": _agora_txt(),
        "kpis": kpis,
        "saude": saude,
        "alertas_inteligentes": alertas_inteligentes,
        "resumo_alertas": resumo_alertas,
        "log_amigavel": log_amigavel,
        "eventos": eventos,
        "processamento_integrado": job,
    })

def _agora_txt():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _tail_texto(texto, limite=6000):
    texto = str(texto or "")
    if len(texto) <= limite:
        return texto
    return texto[-limite:]


def _salvar_log_processamento_integrado(titulo, conteudo):
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with ARQUIVO_LOG_PROCESSAMENTO_INTEGRADO.open("a", encoding="utf-8") as f:
            f.write("\n" + "=" * 80 + "\n")
            f.write(f"{_agora_txt()} | {titulo}\n")
            f.write("=" * 80 + "\n")
            f.write(str(conteudo or ""))
            f.write("\n")
    except Exception:
        pass


def _ler_tail_log_processamento_integrado(limite=6000):
    try:
        if not ARQUIVO_LOG_PROCESSAMENTO_INTEGRADO.exists():
            return ""
        return _tail_texto(ARQUIVO_LOG_PROCESSAMENTO_INTEGRADO.read_text(encoding="utf-8", errors="ignore"), limite)
    except Exception:
        return ""


def _atualizar_job_processamento_integrado(**kwargs):
    with PROCESSAMENTO_INTEGRADO_LOCK:
        PROCESSAMENTO_INTEGRADO_JOB.update(kwargs)
        PROCESSAMENTO_INTEGRADO_JOB["log_tail"] = _ler_tail_log_processamento_integrado()
        return dict(PROCESSAMENTO_INTEGRADO_JOB)


def _job_processamento_integrado_atual():
    with PROCESSAMENTO_INTEGRADO_LOCK:
        payload = dict(PROCESSAMENTO_INTEGRADO_JOB)
    payload["log_tail"] = _ler_tail_log_processamento_integrado()
    return payload


def _localizar_script_launcher(cfg):
    informado = str((cfg or {}).get("script_launcher_robo") or "").strip().strip('"').strip("'")
    candidatos = []
    if informado and not _script_launcher_persistido_invalido_v82123(informado):
        candidatos.append(Path(informado).expanduser())
    for pasta in _diretorios_runtime_app_v82122() if "_diretorios_runtime_app_v82122" in globals() else [BASE_DIR, Path.cwd()]:
        candidatos.extend([
            pasta / "Script_Launcher_v5_filtrado.py",
            pasta / "Script_Launcher_v5.py",
            pasta / "Baixar_Anexos_Disparar_Robo.py",
        ])
    vistos = set()
    for c in candidatos:
        try:
            chave = str(c.resolve()).lower()
        except Exception:
            chave = str(c).lower()
        if chave in vistos:
            continue
        vistos.add(chave)
        if c.exists() and c.is_file():
            return c
    return candidatos[0] if candidatos else (BASE_DIR / "Script_Launcher_v5_filtrado.py")


def _python_para_subprocesso_sem_console():
    """Retorna um Python adequado para scripts sem abrir console no Windows."""
    try:
        exe = Path(sys.executable)
        if getattr(sys, "frozen", False):
            return str(exe)
        if os.name == "nt" and exe.name.lower() == "python.exe":
            pythonw = exe.with_name("pythonw.exe")
            if pythonw.exists():
                return str(pythonw)
        if not getattr(sys, "frozen", False):
            return str(exe)
    except Exception:
        pass
    return "py"


def _kwargs_subprocess_sem_console():
    kwargs = {}
    if os.name == "nt":
        try:
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = 0
            kwargs["startupinfo"] = startupinfo
            kwargs["creationflags"] = getattr(subprocess, "CREATE_NO_WINDOW", 0)
        except Exception:
            pass
    return kwargs



def _montar_comando_launcher(launcher):
    """Monta comando de execução do launcher de forma mais tolerante."""
    launcher = Path(launcher)
    suf = launcher.suffix.lower()
    if suf == ".py":
        if getattr(sys, "frozen", False):
            return [str(sys.executable)]
        return [_python_para_subprocesso_sem_console(), "-u", str(launcher)]
    if suf in {".bat", ".cmd"}:
        return ["cmd", "/c", str(launcher)]
    return [str(launcher)]


def _preparar_env_script_embarcado_v82125(env, script):
    script = Path(script)
    if getattr(sys, "frozen", False) and script.suffix.lower() == ".py":
        env["TEMPO_FECHADO_RUN_SCRIPT"] = str(script)
        env["TEMPO_FECHADO_RUN_SCRIPT_CWD"] = str(script.parent)
        env["TEMPO_FECHADO_RUN_SCRIPT_MODE"] = "launcher_integrado"
    return env


def _executar_processamento_integrado_worker(opcoes):
    t0 = time.monotonic()
    resultado = {
        "launcher": {},
        "importacao_excel": {},
        "historico_pdfs": {},
        "banco_horas": {},
    }
    try:
        _atualizar_job_processamento_integrado(
            status="rodando", ok=True, fase="Preparando",
            mensagem="Preparando execução integrada do Tempo Fechado...",
            iniciado_em=_agora_txt(), finalizado_em="", segundos=0, resultado=resultado,
        )
        cfg = carregar_config_robo()
        aplicar_config_robo(cfg)
        launcher = _localizar_script_launcher(cfg)
        resultado["launcher"]["caminho"] = str(launcher)

        if not launcher.exists():
            raise FileNotFoundError(
                f"Script Launcher não encontrado: {launcher}. Coloque o Script_Launcher_v5_filtrado.py na pasta do Tempo Fechado ou configure o caminho em Configuração do Robô."
            )

        timeout = int(opcoes.get("timeout_segundos") or 3600)
        _atualizar_job_processamento_integrado(
            fase="Executando Launcher",
            mensagem=f"Executando {launcher.name}. Essa etapa baixa/processa os PDFs e pode levar alguns minutos.",
            resultado=resultado,
        )

        cmd = _montar_comando_launcher(launcher)
        env = os.environ.copy()
        env.setdefault("PYTHONUTF8", "1")
        env.setdefault("PYTHONIOENCODING", "utf-8")
        env = _preparar_env_script_embarcado_v82125(env, launcher)
        resultado["launcher"].update({
            "comando": " ".join(str(x) for x in cmd),
            "cwd": str(launcher.parent),
            "python_tempo_fechado": sys.executable,
            "executor_embarcado": bool(env.get("TEMPO_FECHADO_RUN_SCRIPT")),
        })
        proc = subprocess.run(
            cmd,
            cwd=str(launcher.parent),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
            env=env,
            **_kwargs_subprocess_sem_console(),
        )
        saida = (
            "[COMANDO]\n" + " ".join(str(x) for x in cmd) + "\n\n"
            "[PASTA DE EXECUÇÃO]\n" + str(launcher.parent) + "\n\n"
            "[PYTHON DO TEMPO FECHADO]\n" + str(sys.executable) + "\n\n"
            "[STDOUT]\n" + str(proc.stdout or "") + "\n\n[STDERR]\n" + str(proc.stderr or "")
        )
        _salvar_log_processamento_integrado("Execução do Script Launcher - v8.21.38 Painel de Implantacao", saida)
        stdout_tail = _tail_texto(proc.stdout, 3000)
        stderr_tail = _tail_texto(proc.stderr, 3000)
        resultado["launcher"].update({
            "ok": proc.returncode == 0,
            "returncode": proc.returncode,
            "stdout_tail": stdout_tail,
            "stderr_tail": stderr_tail,
        })
        if proc.returncode != 0:
            detalhe = stderr_tail or stdout_tail or "Sem saída de erro capturada pelo launcher."
            raise RuntimeError(
                f"O Script Launcher terminou com erro. Código de retorno: {proc.returncode}. "
                f"Resumo do erro: {detalhe[:1200]}"
            )

        if opcoes.get("importar_excel", True):
            _atualizar_job_processamento_integrado(
                fase="Importando Excel",
                mensagem="Importando automaticamente o Excel gerado pelo robô para o Tempo Fechado...",
                resultado=resultado,
            )
            imp = sincronizar_excel_robo(forcar=True, corrigir_saldo_pdf=False)
            resultado["importacao_excel"] = imp
            if imp.get("ok"):
                _ativar_base_operacional("resultado_lote_consultas_validado.xlsx")
                try:
                    _preaquecer_base_excel_em_segundo_plano()
                except Exception:
                    pass
                resultado["banco_horas"] = {"ok": True, "executado": False, "mensagem": "Resumo será montado sob demanda a partir do Excel oficial."}
            else:
                raise RuntimeError(imp.get("mensagem") or "Falha ao importar Excel gerado pelo robô.")

        # v8.11.21: Processamento Integrado trabalha em modo Excel oficial.
        # Não dispara histórico/correção por PDF para não contaminar Banco/Extrato.
        opcoes["atualizar_historico_pdfs"] = False
        opcoes["corrigir_banco_horas"] = False

        if opcoes.get("atualizar_historico_pdfs", False):
            _atualizar_job_processamento_integrado(
                fase="Atualizando PDFs",
                mensagem="Atualizando histórico/cache dos PDFs do Banco de Horas...",
                resultado=resultado,
            )
            hist = atualizar_cache_pdfs_banco_horas()
            resultado["historico_pdfs"] = hist

        if opcoes.get("corrigir_banco_horas", False):
            _atualizar_job_processamento_integrado(
                fase="Gerando Banco de Horas",
                mensagem="Gerando resumo do Banco de Horas pelos PDFs já catalogados...",
                resultado=resultado,
            )
            bh = gerar_resumo_banco_horas_pdf()
            resultado["banco_horas"] = bh

        segundos = round(time.monotonic() - t0, 1)
        _atualizar_job_processamento_integrado(
            status="concluido", ok=True, fase="Concluído",
            mensagem=f"Processamento integrado concluído em {segundos}s.",
            finalizado_em=_agora_txt(), segundos=segundos, resultado=resultado,
        )
        registrar_auditoria("PROCESSAMENTO_INTEGRADO", "Execução concluída", "OK")
        registrar_historico_operacao_v8190(
            "Processamento Integrado", "OK",
            f"Processamento integrado concluído em {segundos}s.",
            resultado, segundos=segundos, origem="Processamento"
        )
    except Exception as e:
        segundos = round(time.monotonic() - t0, 1)
        _salvar_log_processamento_integrado("Erro no Processamento Integrado", repr(e))
        _atualizar_job_processamento_integrado(
            status="erro", ok=False, fase="Erro",
            mensagem=str(e), finalizado_em=_agora_txt(), segundos=segundos, resultado=resultado,
        )
        registrar_auditoria("PROCESSAMENTO_INTEGRADO", str(e), "ERRO")
        registrar_historico_operacao_v8190(
            "Processamento Integrado", "ERRO",
            str(e), resultado, segundos=segundos, origem="Processamento"
        )



# v8.17.25 - prova de chamada do launcher pelo Tempo Fechado.
def _registrar_prova_chamada_launcher_v81725(script_path="", comando=""):
    try:
        controle = Path.home() / "ponto_pdfs" / "controle"
        controle.mkdir(parents=True, exist_ok=True)
        with open(controle / "tempo_fechado_chamou_launcher_em.txt", "a", encoding="utf-8") as f:
            f.write(f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | script={script_path} | comando={comando}\n")
    except Exception:
        pass


# ============================================================
# v8.17.37 - Processamento Manual do Robô | Opção B
# ============================================================
def _base_ponto_pdfs_v81737():
    base = Path.home() / "ponto_pdfs"
    for nome in ["entrada", "saida", "processados", "logs", "backup", "controle", "erro"]:
        (base / nome).mkdir(parents=True, exist_ok=True)
    return base

def _diretorios_runtime_app_v82122():
    candidatos = []
    for valor in [
        BASE_DIR,
        getattr(sys, "_MEIPASS", ""),
        Path(sys.executable).resolve().parent if getattr(sys, "executable", "") else "",
        Path.cwd(),
    ]:
        if not valor:
            continue
        try:
            p = Path(valor).resolve()
        except Exception:
            p = Path(valor)
        candidatos.extend([p, p / "_internal"])

    unicos = []
    vistos = set()
    for p in candidatos:
        try:
            chave = str(Path(p).resolve()).lower()
        except Exception:
            chave = str(p).lower()
        if chave in vistos:
            continue
        vistos.add(chave)
        unicos.append(Path(p))
    return unicos

def _localizar_script_robo_v81737():
    candidatos = [p / "Script_Robo_Ponto_v7_PRO.py" for p in _diretorios_runtime_app_v82122()]
    for c in candidatos:
        try:
            if Path(c).exists():
                return Path(c)
        except Exception:
            pass
    return candidatos[0]

def _candidatos_script_robo_v82122():
    return [str(p / "Script_Robo_Ponto_v7_PRO.py") for p in _diretorios_runtime_app_v82122()]

def _excel_robo_padrao_v81737():
    return _base_ponto_pdfs_v81737() / "saida" / "resultado_lote_consultas_validado.xlsx"


def _ler_resultado_robo_manual_v82127():
    arquivo = _base_ponto_pdfs_v81737() / "controle" / "robo_ponto_resultado.json"
    try:
        if arquivo.exists():
            return json.loads(arquivo.read_text(encoding="utf-8") or "{}")
    except Exception:
        pass
    return {}


def _excel_gerado_por_resultado_v82127(excel_padrao=None):
    excel_padrao = Path(excel_padrao or _excel_robo_padrao_v81737())
    resultado = _ler_resultado_robo_manual_v82127()
    candidatos = [excel_padrao]
    caminho_resultado = str(resultado.get("caminho_excel") or "").strip()
    if caminho_resultado:
        candidatos.insert(0, Path(caminho_resultado))
    for candidato in candidatos:
        try:
            if candidato.exists() and candidato.is_file():
                return candidato, resultado
        except Exception:
            pass
    return excel_padrao, resultado


def _diag_item_v82128(nome, ok, detalhe="", acao="", nivel=None, dados=None):
    if nivel is None:
        nivel = "ok" if ok else "erro"
    return {
        "nome": nome,
        "ok": bool(ok),
        "nivel": nivel,
        "estado": "OK" if ok else ("Atenção" if nivel == "atencao" else "Erro"),
        "detalhe": str(detalhe or ""),
        "acao": str(acao or ""),
        "dados": dados or {},
    }


def _testar_escrita_pasta_v82128(pasta):
    try:
        p = Path(pasta)
        p.mkdir(parents=True, exist_ok=True)
        teste = p / f".tempo_fechado_write_test_{uuid.uuid4().hex}.tmp"
        teste.write_text("ok", encoding="utf-8")
        teste.unlink(missing_ok=True)
        return True, "escrita liberada"
    except Exception as e:
        return False, str(e)


def _diagnosticar_outlook_v82128():
    if not sys.platform.startswith("win"):
        return _diag_item_v82128("Outlook", False, "Windows/Outlook COM indisponível nesta plataforma.", "Validar em notebook Windows com Outlook instalado.", "atencao")
    try:
        import win32com.client
    except Exception as e:
        return _diag_item_v82128("Outlook", False, f"pywin32/win32com indisponível: {e}", "Regerar o EXE com pywin32 ou revisar instalação.", "atencao")
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        nome = ""
        try:
            nome = str(outlook.Name)
        except Exception:
            nome = "Outlook"
        return _diag_item_v82128("Outlook", True, f"{nome} disponível para automação.", "Envio automático pode ser usado se o perfil estiver configurado.")
    except Exception as e:
        return _diag_item_v82128("Outlook", False, f"COM indisponível: {e}", "Manter Outlook opcional ou registrar/configurar Outlook neste notebook.", "atencao")


@app.route("/api/diagnostico-pos-instalacao", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_diagnostico_pos_instalacao_v82128():
    base = _base_ponto_pdfs_v81737()
    cfg = carregar_config_robo()
    launcher = _localizar_script_launcher(cfg)
    script_robo = _localizar_script_robo_v81737()
    excel_padrao, resultado_robo = _excel_gerado_por_resultado_v82127(_excel_robo_padrao_v81737())
    itens = []

    versao_ok = APP_VERSION == "v8.21.38"
    itens.append(_diag_item_v82128("Versão ativa", versao_ok, APP_FULL_NAME, "Fechar instâncias antigas se a versão exibida não for a esperada."))

    itens.append(_diag_item_v82128("Executável", bool(getattr(sys, "frozen", False)), str(sys.executable), "Em produção, deve apontar para TempoFechado.exe.", "ok" if getattr(sys, "frozen", False) else "atencao"))
    itens.append(_diag_item_v82128("Backend Flask", True, str(Path(__file__).resolve()), "Arquivo backend carregado."))
    itens.append(_diag_item_v82128("Script Launcher", launcher.exists(), str(launcher), "Regerar instalador se o script não estiver no pacote."))
    itens.append(_diag_item_v82128("Script Robô", script_robo.exists(), str(script_robo), "Regerar instalador se o script não estiver no pacote."))
    itens.append(_diag_item_v82128("Excel do Robô", excel_padrao.exists(), str(excel_padrao), "Executar coleta ou processamento manual para gerar o Excel.", "ok" if excel_padrao.exists() else "atencao", {"resultado_robo": resultado_robo}))
    itens.append(_diag_item_v82128("Base do app", ARQUIVO_PADRAO.exists(), str(ARQUIVO_PADRAO), "Importar o Excel do robô para ativar a base.", "ok" if ARQUIVO_PADRAO.exists() else "atencao"))

    for nome in ["entrada", "saida", "processados", "erro", "controle", "config", "backup"]:
        pasta = base / nome
        ok_escrita, detalhe_escrita = _testar_escrita_pasta_v82128(pasta)
        itens.append(_diag_item_v82128(f"Pasta {nome}", pasta.exists() and ok_escrita, f"{pasta} | {detalhe_escrita}", "Corrigir permissão de escrita na pasta ponto_pdfs."))

    itens.append(_diag_item_v82128("pdfplumber", pdfplumber is not None, "Disponível" if pdfplumber is not None else "Não instalado no pacote.", "Regerar EXE com dependências do executor."))
    itens.append(_diag_item_v82128("openpyxl", load_workbook is not None, "Disponível" if load_workbook is not None else "Não instalado no pacote.", "Regerar EXE com openpyxl."))
    itens.append(_diagnosticar_outlook_v82128())

    usuario = {
        "usuario": session.get("usuario"),
        "nome": session.get("nome"),
        "perfil": session.get("perfil"),
        "admin": session.get("perfil") == "admin",
    }
    itens.append(_diag_item_v82128("Usuário admin", usuario["admin"], f"{usuario.get('usuario') or '-'} | {usuario.get('perfil') or '-'}", "Entrar com perfil admin para validar instalação.", "ok" if usuario["admin"] else "erro", usuario))

    erros = sum(1 for item in itens if item.get("nivel") == "erro")
    atencoes = sum(1 for item in itens if item.get("nivel") == "atencao")
    total = len(itens)
    ok_count = sum(1 for item in itens if item.get("ok"))
    pronto = erros == 0
    nivel = "ok" if pronto and atencoes == 0 else ("atencao" if pronto else "erro")
    return jsonify({
        "ok": True,
        "erro": "",
        "pronto": pronto,
        "nivel": nivel,
        "gerado_em": _agora_txt() if "_agora_txt" in globals() else datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "versao": APP_VERSION,
        "titulo": APP_FULL_NAME,
        "base": str(base),
        "resumo": {"total": total, "ok": ok_count, "atencoes": atencoes, "erros": erros},
        "itens": itens,
        "proximos_passos": [
            item.get("acao") for item in itens
            if item.get("acao") and item.get("nivel") in {"erro", "atencao"}
        ][:8],
    })

@app.route("/api/processamento-manual-robo/status", methods=["GET"])
@login_obrigatorio
def api_processamento_manual_robo_status_v81737():
    base = _base_ponto_pdfs_v81737()
    script = _localizar_script_robo_v81737()
    entrada = base / "entrada"
    excel = _excel_robo_padrao_v81737()
    excel_detectado, resultado_robo = _excel_gerado_por_resultado_v82127(excel)
    pdfs = sorted(entrada.glob("*.pdf"))
    return jsonify({"ok": True, "erro": "", "base": str(base), "entrada": str(entrada), "saida": str(base / "saida"), "script_robo": str(script), "script_existe": script.exists(), "script_candidatos": _candidatos_script_robo_v82122(), "qtd_pdfs_entrada": len(pdfs), "pdfs_entrada": [p.name for p in pdfs[:50]], "excel_gerado": str(excel_detectado), "excel_existe": excel_detectado.exists(), "resultado_robo": resultado_robo})

@app.route("/api/processamento-manual-robo/abrir-entrada", methods=["POST"])
@login_obrigatorio
def api_processamento_manual_robo_abrir_entrada_v81740():
    entrada = _base_ponto_pdfs_v81737() / "entrada"
    try:
        entrada.mkdir(parents=True, exist_ok=True)
        caminho = str(entrada.resolve())
        if sys.platform.startswith("win"):
            try:
                os.startfile(caminho)
            except AttributeError:
                subprocess.Popen(["explorer", caminho], shell=False)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", caminho], shell=False)
        else:
            subprocess.Popen(["xdg-open", caminho], shell=False)
        return jsonify({"ok": True, "erro": "", "mensagem": "Pasta de entrada aberta.", "entrada": caminho})
    except Exception as e:
        return jsonify({"ok": False, "erro": f"Falha ao abrir pasta de entrada: {e}", "entrada": str(entrada)}), 500


@app.route("/api/processamento-manual-robo/upload-pdfs", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_processamento_manual_robo_upload_pdfs_v82138():
    entrada = _base_ponto_pdfs_v81737() / "entrada"
    entrada.mkdir(parents=True, exist_ok=True)
    arquivos = request.files.getlist("pdfs")
    if not arquivos:
        return jsonify({"ok": False, "erro": "Selecione ao menos um PDF para upload."}), 400

    salvos = []
    ignorados = []
    for arquivo in arquivos:
        nome_original = str(getattr(arquivo, "filename", "") or "").strip()
        if not nome_original:
            ignorados.append({"arquivo": "", "motivo": "nome vazio"})
            continue
        if not nome_original.lower().endswith(".pdf"):
            ignorados.append({"arquivo": nome_original, "motivo": "extensão diferente de .pdf"})
            continue

        nome_seguro = secure_filename(nome_original) or f"upload_{uuid.uuid4().hex}.pdf"
        if not nome_seguro.lower().endswith(".pdf"):
            nome_seguro += ".pdf"
        destino = entrada / nome_seguro
        if destino.exists():
            destino = entrada / f"{destino.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:6]}{destino.suffix}"
        try:
            arquivo.save(str(destino))
            salvos.append(destino.name)
        except Exception as e:
            ignorados.append({"arquivo": nome_original, "motivo": str(e)})

    if not salvos:
        return jsonify({"ok": False, "erro": "Nenhum PDF válido foi salvo.", "ignorados": ignorados, "entrada": str(entrada)}), 400

    registrar_historico_operacao_v8190(
        "Upload de PDFs",
        "OK",
        f"{len(salvos)} PDF(s) enviado(s) para a entrada do robô.",
        {"arquivos": salvos[:80], "ignorados": ignorados[:40], "entrada": str(entrada)},
        origem="Processamento Manual",
    )
    return jsonify({
        "ok": True,
        "erro": "",
        "mensagem": f"{len(salvos)} PDF(s) carregado(s) para processamento manual.",
        "entrada": str(entrada),
        "salvos": salvos,
        "ignorados": ignorados,
        "qtd_salvos": len(salvos),
        "qtd_ignorados": len(ignorados),
    })


def _importar_excel_robo_manual_v81737():
    excel = _excel_robo_padrao_v81737()
    global ARQUIVO_EXCEL_ROBO
    antigo_excel_robo = ARQUIVO_EXCEL_ROBO
    ARQUIVO_EXCEL_ROBO = excel
    try:
        imp = sincronizar_excel_robo(forcar=True, corrigir_saldo_pdf=False)
        if imp.get("ok"):
            _ativar_base_operacional("resultado_lote_consultas_validado.xlsx")
            try:
                _preaquecer_base_excel_em_segundo_plano()
            except Exception:
                pass
        return imp
    finally:
        ARQUIVO_EXCEL_ROBO = antigo_excel_robo

def _montar_execucao_script_robo_manual_v82124(script):
    env = os.environ.copy()
    env.setdefault("PYTHONUTF8", "1")
    env.setdefault("PYTHONIOENCODING", "utf-8")
    script = Path(script)
    if getattr(sys, "frozen", False):
        env["TEMPO_FECHADO_RUN_SCRIPT"] = str(script)
        env["TEMPO_FECHADO_RUN_SCRIPT_CWD"] = str(script.parent)
        return [str(sys.executable)], env
    return [str(sys.executable), "-u", str(script)], env

@app.route("/api/processamento-manual-robo/executar", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_processamento_manual_robo_executar_v81737():
    base = _base_ponto_pdfs_v81737()
    entrada = base / "entrada"
    controle = base / "controle"
    script = _localizar_script_robo_v81737()
    if not script.exists():
        return jsonify({"ok": False, "erro": f"Script_Robo_Ponto_v7_PRO.py nao encontrado: {script}", "script_candidatos": _candidatos_script_robo_v82122()}), 404
    pdfs = sorted(entrada.glob("*.pdf"))
    if not pdfs:
        return jsonify({"ok": False, "erro": f"Nenhum PDF encontrado na pasta de entrada: {entrada}"}), 400
    comando, env_exec = _montar_execucao_script_robo_manual_v82124(script)
    log_exec = controle / "processamento_manual_robo.log"
    try:
        with open(log_exec, "a", encoding="utf-8") as f:
            f.write("\n" + "=" * 80 + "\n")
            f.write(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + " | Execução manual do Script_Robo_Ponto_v7_PRO\n")
            f.write("[COMANDO]\n" + " ".join(comando) + "\n")
            f.write("[PASTA]\n" + str(script.parent) + "\n")
            f.write("[PDFS ENTRADA]\n" + "\n".join(p.name for p in pdfs) + "\n")
        r = subprocess.run(
            comando,
            cwd=str(script.parent),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=1800,
            env=env_exec,
            **_kwargs_subprocess_sem_console(),
        )
        with open(log_exec, "a", encoding="utf-8") as f:
            f.write("\n[RETURN CODE]\n" + str(r.returncode) + "\n")
            f.write("\n[STDOUT]\n" + (r.stdout or "") + "\n")
            f.write("\n[STDERR]\n" + (r.stderr or "") + "\n")
        excel = _excel_robo_padrao_v81737()
        if r.returncode != 0:
            excel_detectado, resultado_robo = _excel_gerado_por_resultado_v82127(excel)
            if excel_detectado.exists():
                imp = _importar_excel_robo_manual_v81737()
                aviso = resultado_robo.get("detalhe") or f"Robô retornou código {r.returncode}, mas o Excel foi gerado e importado."
                registrar_auditoria("PROCESSAMENTO_MANUAL_ROBO", str(script), "OK_COM_AVISO")
                registrar_historico_operacao_v8190(
                    "Processamento Manual", "OK",
                    "Robô gerou Excel e importação automática foi concluída com aviso.",
                    {"importacao": imp, "aviso": aviso, "resultado_robo": resultado_robo, "pdfs_enviados": [p.name for p in pdfs[:80]], "script": str(script)},
                    origem="Processamento Manual"
                )
                return jsonify({"ok": True, "erro": "", "mensagem": "Robô gerou Excel e importação automática foi concluída com aviso.", "aviso": aviso, "stdout": r.stdout[-4000:], "stderr": r.stderr[-4000:], "log": str(log_exec), "excel": str(excel_detectado), "excel_existe": excel_detectado.exists(), "importacao": imp, "resultado_robo": resultado_robo})
            return jsonify({"ok": False, "erro": f"Robô retornou erro. Código: {r.returncode}", "stdout": r.stdout[-4000:], "stderr": r.stderr[-4000:], "log": str(log_exec), "excel": str(excel), "excel_existe": excel.exists()}), 500
        imp = _importar_excel_robo_manual_v81737()
        registrar_auditoria("PROCESSAMENTO_MANUAL_ROBO", str(script), "OK")
        registrar_historico_operacao_v8190(
            "Processamento Manual", "OK",
            "Robô executado e importação automática concluída.",
            {"importacao": imp, "pdfs_enviados": [p.name for p in pdfs[:80]], "script": str(script)},
            origem="Processamento Manual"
        )
        return jsonify({"ok": True, "erro": "", "mensagem": "Robô executado e importação automática concluída.", "stdout": r.stdout[-4000:], "stderr": r.stderr[-4000:], "log": str(log_exec), "excel": str(excel), "excel_existe": excel.exists(), "importacao": imp})
    except subprocess.TimeoutExpired:
        registrar_historico_operacao_v8190("Processamento Manual", "ERRO", "Tempo limite excedido ao executar o robô.", {"script": str(script)}, origem="Processamento Manual")
        return jsonify({"ok": False, "erro": "Tempo limite excedido ao executar o robô.", "log": str(log_exec)}), 500
    except Exception as e:
        registrar_historico_operacao_v8190("Processamento Manual", "ERRO", f"Falha ao executar/importar robô manualmente: {e}", {"script": str(script)}, origem="Processamento Manual")
        return jsonify({"ok": False, "erro": f"Falha ao executar/importar robô manualmente: {e}", "log": str(log_exec)}), 500

@app.route("/api/processamento-manual-robo/importar-excel", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_processamento_manual_robo_importar_excel_v81737():
    imp = _importar_excel_robo_manual_v81737()
    excel = _excel_robo_padrao_v81737()
    registrar_historico_operacao_v8190(
        "Importação Manual",
        "OK" if imp.get("ok") else "ERRO",
        imp.get("mensagem", ""),
        {"importacao": imp, "excel": str(excel), "excel_existe": excel.exists()},
        origem="Processamento Manual"
    )
    return jsonify({"ok": bool(imp.get("ok")), "erro": "" if imp.get("ok") else imp.get("mensagem", ""), "importacao": imp, "excel": str(excel), "excel_existe": excel.exists(), "mensagem": imp.get("mensagem", "")})

@app.route("/api/processamento-manual-robo/log", methods=["GET"])
@login_obrigatorio
def api_processamento_manual_robo_log_v81737():
    base = _base_ponto_pdfs_v81737()
    logs = [base / "controle" / "processamento_manual_robo.log", base / "controle" / "robo_ponto.log", base / "controle" / "robo_ponto_resultado.json"]
    partes = []
    for log in logs:
        try:
            if log.exists():
                partes.append(f"===== {log} =====\n" + log.read_text(encoding="utf-8", errors="ignore")[-12000:])
        except Exception as e:
            partes.append(f"===== {log} =====\nFalha ao ler: {e}")
    return jsonify({"ok": True, "erro": "", "log": "\n\n".join(partes) or "Nenhum log encontrado."})


def _contar_arquivos_por_padroes_v82117(pasta, padroes):
    try:
        p = Path(pasta)
        if not p.exists() or not p.is_dir():
            return 0
        vistos = set()
        for padrao in padroes:
            for item in p.glob(padrao):
                if item.is_file():
                    vistos.add(str(item).lower())
        return len(vistos)
    except Exception:
        return 0


def _montar_painel_ciclo_operacional_v82117(limite_historico=8):
    cfg = carregar_config_robo()
    aplicar_config_robo(cfg)
    raiz = Path(PASTA_RAIZ_PDFS_ROBO)
    entrada = raiz / "entrada"
    saida = raiz / "saida"
    processados = raiz / "processados"
    erro = raiz / "erro"
    controle = raiz / "controle"
    launcher = _localizar_script_launcher(cfg)

    caches = {
        "pdfs_banco_horas": _info_caminho_v8191(ARQUIVO_CACHE_PDFS_BANCO_HORAS),
        "resumo_banco_horas": _info_caminho_v8191(ARQUIVO_RESUMO_BANCO_HORAS_PDF),
        "calculo_banco_horas": _info_caminho_v8191(ARQUIVO_CACHE_BH2_CALCULO),
    }
    base_app = {
        "base_ativa": _base_operacional_ativa(),
        "arquivo": str(ARQUIVO_PADRAO),
        "excel_existe": ARQUIVO_PADRAO.exists(),
        "excel_app": _info_caminho_v8191(ARQUIVO_PADRAO),
    }
    robo = {
        "raiz": _info_caminho_v8191(raiz, "pasta"),
        "entrada": {**_info_caminho_v8191(entrada, "pasta"), "pdfs": _contar_arquivos_por_padroes_v82117(entrada, ["*.pdf", "*.PDF"])},
        "saida": {**_info_caminho_v8191(saida, "pasta"), "planilhas": _contar_arquivos_por_padroes_v82117(saida, ["*.xlsx", "*.xls", "*.csv"])},
        "processados": {**_info_caminho_v8191(processados, "pasta"), "pdfs": _contar_arquivos_por_padroes_v82117(processados, ["*.pdf", "*.PDF"])},
        "erro": {**_info_caminho_v8191(erro, "pasta"), "pdfs": _contar_arquivos_por_padroes_v82117(erro, ["*.pdf", "*.PDF"])},
        "controle": _info_caminho_v8191(controle, "pasta"),
        "excel": _info_caminho_v8191(ARQUIVO_EXCEL_ROBO),
        "launcher": _info_caminho_v8191(launcher),
    }
    return {
        "ok": True,
        "erro": "",
        "versao": APP_VERSION,
        "gerado_em": _agora_txt(),
        "base_app": base_app,
        "robo": robo,
        "caches": caches,
        "job": _job_processamento_integrado_atual(),
        "historico": _ler_historico_operacao_v8190(limite_historico),
    }


def _montar_prevalidacao_processamento_v82117():
    cfg = carregar_config_robo()
    aplicar_config_robo(cfg)
    raiz = Path(PASTA_RAIZ_PDFS_ROBO)
    launcher = _localizar_script_launcher(cfg)
    script_robo = _localizar_script_robo_v81737()
    entrada = raiz / "entrada"
    saida = raiz / "saida"
    processados = raiz / "processados"
    erro = raiz / "erro"
    checks = []

    def add(nome, ok, detalhe, nivel=None):
        checks.append({
            "nome": nome,
            "ok": bool(ok),
            "nivel": nivel or ("ok" if ok else "erro"),
            "detalhe": detalhe,
        })

    add("Launcher", launcher.exists(), str(launcher))
    add("Script do Robo", script_robo.exists(), str(script_robo))
    for nome, pasta in [("Entrada", entrada), ("Saida", saida), ("Processados", processados), ("Erro", erro)]:
        add(f"Pasta {nome}", pasta.exists(), str(pasta))
    qtd_entrada = _contar_arquivos_por_padroes_v82117(entrada, ["*.pdf", "*.PDF"])
    qtd_erro = _contar_arquivos_por_padroes_v82117(erro, ["*.pdf", "*.PDF"])
    add("PDFs para coleta", qtd_entrada > 0, f"{qtd_entrada} PDF(s) na entrada", "ok" if qtd_entrada > 0 else "warning")
    add("Pasta de erro", qtd_erro == 0, f"{qtd_erro} PDF(s) em erro", "ok" if qtd_erro == 0 else "warning")
    add("Excel do Robo", ARQUIVO_EXCEL_ROBO.exists(), str(ARQUIVO_EXCEL_ROBO), "ok" if ARQUIVO_EXCEL_ROBO.exists() else "warning")
    add("Base do Tempo Fechado", ARQUIVO_PADRAO.exists(), str(ARQUIVO_PADRAO), "ok" if ARQUIVO_PADRAO.exists() else "warning")

    erros = sum(1 for c in checks if c.get("nivel") == "erro")
    avisos = sum(1 for c in checks if c.get("nivel") == "warning")
    pronto = erros == 0
    return {
        "ok": True,
        "erro": "",
        "pronto": pronto,
        "nivel": "erro" if erros else ("warning" if avisos else "ok"),
        "mensagem": "Pre-validacao concluida. Ambiente pronto." if pronto else "Pre-validacao encontrou bloqueios.",
        "checks": checks,
        "resumo": {"erros": erros, "avisos": avisos, "total": len(checks)},
    }


@app.route("/api/processamento-integrado/status", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_processamento_integrado_status():
    return jsonify(_job_processamento_integrado_atual())


@app.route("/api/ciclo-operacional", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_ciclo_operacional_v82117():
    return jsonify(_montar_painel_ciclo_operacional_v82117())


@app.route("/api/processamento/prevalidacao", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_processamento_prevalidacao_v82117():
    return jsonify(_montar_prevalidacao_processamento_v82117())


@app.route("/api/processamento/historico", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_processamento_historico_v82117():
    try:
        limite = int(request.args.get("limite", 20) or 20)
    except Exception:
        limite = 20
    return jsonify({"ok": True, "erro": "", "historico": _ler_historico_operacao_v8190(limite)})


@app.route("/api/processamento-integrado/iniciar", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_processamento_integrado_iniciar():
    payload = request.get_json(silent=True) or {}
    with PROCESSAMENTO_INTEGRADO_LOCK:
        if PROCESSAMENTO_INTEGRADO_JOB.get("status") == "rodando":
            return jsonify({"ok": True, "status": "rodando", "mensagem": "Já existe um processamento integrado em andamento.", "job": dict(PROCESSAMENTO_INTEGRADO_JOB)}), 202
        PROCESSAMENTO_INTEGRADO_JOB.update({
            "status": "rodando", "ok": True, "fase": "Fila",
            "mensagem": "Processamento integrado iniciado.",
            "iniciado_em": _agora_txt(), "finalizado_em": "", "segundos": 0,
            "resultado": {}, "log_tail": _ler_tail_log_processamento_integrado(),
        })
    opcoes = {
        "importar_excel": bool(payload.get("importar_excel", True)),
        "atualizar_historico_pdfs": bool(payload.get("atualizar_historico_pdfs", False)),
        "corrigir_banco_horas": bool(payload.get("corrigir_banco_horas", False)),
        "timeout_segundos": int(payload.get("timeout_segundos", 3600) or 3600),
    }
    th = threading.Thread(target=_executar_processamento_integrado_worker, args=(opcoes,), daemon=True)
    th.start()
    return jsonify({"ok": True, "status": "rodando", "mensagem": "Processamento integrado iniciado em segundo plano.", "job": _job_processamento_integrado_atual()}), 202


@app.route("/api/processamento/nova-rodada", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_processamento_nova_rodada_v82117():
    payload = request.get_json(silent=True) or {}
    with PROCESSAMENTO_INTEGRADO_LOCK:
        if PROCESSAMENTO_INTEGRADO_JOB.get("status") == "rodando":
            return jsonify({"ok": True, "status": "rodando", "mensagem": "Ja existe um processamento integrado em andamento.", "job": dict(PROCESSAMENTO_INTEGRADO_JOB)}), 202

    cleanup_resultado = None
    if bool(payload.get("executar_cleanup", True)):
        cleanup_payload = payload.get("cleanup") or {}
        cleanup_resultado = _executar_cleanup_operacional(cleanup_payload, validar_job=True)
        status_code = int(cleanup_resultado.pop("status_code", 200) or 200)
        if status_code >= 400 or not cleanup_resultado.get("ok"):
            return jsonify(cleanup_resultado), status_code

    with PROCESSAMENTO_INTEGRADO_LOCK:
        PROCESSAMENTO_INTEGRADO_JOB.update({
            "status": "rodando", "ok": True, "fase": "Fila",
            "mensagem": "Nova rodada iniciada.",
            "iniciado_em": _agora_txt(), "finalizado_em": "", "segundos": 0,
            "resultado": {"cleanup": cleanup_resultado or {}}, "log_tail": _ler_tail_log_processamento_integrado(),
        })
    opcoes = {
        "importar_excel": True,
        "atualizar_historico_pdfs": False,
        "corrigir_banco_horas": False,
        "timeout_segundos": int(payload.get("timeout_segundos", 3600) or 3600),
    }
    th = threading.Thread(target=_executar_processamento_integrado_worker, args=(opcoes,), daemon=True)
    th.start()
    try:
        registrar_historico_operacao_v8190(
            "Nova Rodada",
            "OK",
            "Nova rodada iniciada com cleanup previo." if cleanup_resultado else "Nova rodada iniciada sem cleanup previo.",
            {"cleanup": cleanup_resultado or {}},
            origem="Processamento"
        )
    except Exception:
        pass
    return jsonify({
        "ok": True,
        "status": "rodando",
        "mensagem": "Nova rodada iniciada em segundo plano.",
        "cleanup": cleanup_resultado,
        "job": _job_processamento_integrado_atual(),
    }), 202




@app.route("/api/configuracoes-robo", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_configuracoes_robo_get():
    cfg = normalizar_config_robo_runtime_v82123(carregar_config_robo())
    aplicar_config_robo(cfg)
    return jsonify({
        "ok": True,
        "config": cfg,
        "arquivo_excel_robo_existe": ARQUIVO_EXCEL_ROBO.exists(),
        "pasta_pdfs_robo_existe": PASTA_RAIZ_PDFS_ROBO.exists(),
        "pastas_verificadas": [str(p) for p in PASTAS_PDFS_ROBO],
        "cache_pdfs_banco_horas_existe": ARQUIVO_CACHE_PDFS_BANCO_HORAS.exists(),
    })


@app.route("/api/configuracoes-robo", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_configuracoes_robo_post():
    payload = request.get_json(silent=True) or {}
    arquivo_excel = str(payload.get("arquivo_excel_robo", "")).strip()
    pasta_pdfs = str(payload.get("pasta_pdfs_robo", "")).strip()
    script_launcher = str(payload.get("script_launcher_robo", "")).strip()

    if not arquivo_excel:
        return jsonify({"ok": False, "erro": "Informe o caminho do Excel gerado pelo Robô."}), 400
    if not pasta_pdfs:
        return jsonify({"ok": False, "erro": "Informe a pasta raiz dos PDFs."}), 400

    try:
        max_pdfs = int(payload.get("max_pdfs_correcao_saldo", 1500) or 1500)
        max_segundos = int(payload.get("max_segundos_correcao_saldo", 180) or 180)
        limite_emails_outlook = _normalizar_limite_emails_outlook_v8216(payload.get("limite_emails_outlook", 500))
    except Exception:
        return jsonify({"ok": False, "erro": "Limites de leitura devem ser números inteiros."}), 400

    cfg = salvar_config_robo({
        "arquivo_excel_robo": arquivo_excel,
        "pasta_pdfs_robo": pasta_pdfs,
        "script_launcher_robo": script_launcher,
        "auto_importar_excel_robo": bool(payload.get("auto_importar_excel_robo", True)),
        "corrigir_saldo_atual_por_pdf_ao_importar": bool(payload.get("corrigir_saldo_atual_por_pdf_ao_importar", True)),
        "max_pdfs_correcao_saldo": max_pdfs,
        "max_segundos_correcao_saldo": max_segundos,
        "capturar_emails_lidos": bool(payload.get("capturar_emails_lidos", False)),
        "limite_emails_outlook": limite_emails_outlook,
    })

    registrar_auditoria(
        "CONFIGURAR_PASTA_PDFS",
        f"PDFs={cfg.get('pasta_pdfs_robo')} | Excel={cfg.get('arquivo_excel_robo')}",
        "OK",
    )
    return jsonify({
        "ok": True,
        "mensagem": "Configurações do Robô salvas com sucesso.",
        "config": cfg,
        "arquivo_excel_robo_existe": ARQUIVO_EXCEL_ROBO.exists(),
        "pasta_pdfs_robo_existe": PASTA_RAIZ_PDFS_ROBO.exists(),
        "pastas_verificadas": [str(p) for p in PASTAS_PDFS_ROBO],
    })


@app.route("/api/configuracoes-robo/testar", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_configuracoes_robo_testar():
    payload = request.get_json(silent=True) or {}
    arquivo_excel = _normalizar_path_windows(payload.get("arquivo_excel_robo"), ARQUIVO_EXCEL_ROBO)
    pasta_pdfs = _normalizar_path_windows(payload.get("pasta_pdfs_robo"), PASTA_RAIZ_PDFS_ROBO)
    pastas = _montar_pastas_pdfs_robo(pasta_pdfs)

    total_pdfs = 0
    exemplos = []
    vistos = set()
    for pasta in pastas:
        try:
            if not pasta.exists():
                continue
            for pdf in pasta.rglob("*.pdf"):
                chave = str(pdf.resolve()).lower()
                if chave in vistos:
                    continue
                vistos.add(chave)
                total_pdfs += 1
                if len(exemplos) < 8:
                    exemplos.append(str(pdf))
        except Exception:
            continue

    return jsonify({
        "ok": True,
        "arquivo_excel": str(arquivo_excel),
        "arquivo_excel_existe": arquivo_excel.exists(),
        "pasta_pdfs": str(pasta_pdfs),
        "pasta_pdfs_existe": pasta_pdfs.exists(),
        "pastas_verificadas": [str(p) for p in pastas],
        "pdfs_encontrados": total_pdfs,
        "exemplos": exemplos,
    })



# v8.17.27 - Configuração de destinatários do e-mail do robô
def _lista_emails_config_v81727(valor):
    if isinstance(valor, list):
        base = valor
    else:
        base = str(valor or "").replace(";", ",").split(",")
    return [str(x).strip() for x in base if str(x).strip()]

def _cfg_email_robo_v81727():
    cfg = {
        "enviar_email_automaticamente": True,
        "exibir_email_antes_de_enviar": True,
        "destinatarios_email": [],
        "cc_email": [],
        "cco_email": [],
        "assunto_email_base": "Arquivo consolidado",
    }
    try:
        ARQUIVO_CONFIG_EMAIL_ROBO_V81727.parent.mkdir(parents=True, exist_ok=True)
        if ARQUIVO_CONFIG_EMAIL_ROBO_V81727.exists():
            with open(ARQUIVO_CONFIG_EMAIL_ROBO_V81727, "r", encoding="utf-8") as f:
                dados = json.load(f) or {}
            for k in cfg:
                if k in dados:
                    cfg[k] = dados[k]
    except Exception:
        pass
    cfg["destinatarios_email"] = _lista_emails_config_v81727(cfg.get("destinatarios_email"))
    cfg["cc_email"] = _lista_emails_config_v81727(cfg.get("cc_email"))
    cfg["cco_email"] = _lista_emails_config_v81727(cfg.get("cco_email"))
    cfg["enviar_email_automaticamente"] = bool(cfg.get("enviar_email_automaticamente"))
    cfg["exibir_email_antes_de_enviar"] = bool(cfg.get("exibir_email_antes_de_enviar"))
    cfg["assunto_email_base"] = str(cfg.get("assunto_email_base") or "Arquivo consolidado").strip() or "Arquivo consolidado"
    return cfg

def _salvar_cfg_email_robo_v81727(dados):
    cfg = _cfg_email_robo_v81727()
    for k in cfg.keys():
        if k in (dados or {}):
            cfg[k] = dados[k]
    cfg["destinatarios_email"] = _lista_emails_config_v81727(cfg.get("destinatarios_email"))
    cfg["cc_email"] = _lista_emails_config_v81727(cfg.get("cc_email"))
    cfg["cco_email"] = _lista_emails_config_v81727(cfg.get("cco_email"))
    cfg["enviar_email_automaticamente"] = bool(cfg.get("enviar_email_automaticamente"))
    cfg["exibir_email_antes_de_enviar"] = bool(cfg.get("exibir_email_antes_de_enviar"))
    cfg["assunto_email_base"] = str(cfg.get("assunto_email_base") or "Arquivo consolidado").strip() or "Arquivo consolidado"
    ARQUIVO_CONFIG_EMAIL_ROBO_V81727.parent.mkdir(parents=True, exist_ok=True)
    with open(ARQUIVO_CONFIG_EMAIL_ROBO_V81727, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)
    return cfg



@app.route("/admin-config-robo-view-v81729")
@login_obrigatorio
@admin_obrigatorio
def admin_config_robo_view_v81729():
    return redirect(url_for("index"))

@app.route("/api/admin/config-robo-painel", methods=["GET", "POST"])
@login_obrigatorio
@admin_obrigatorio
def api_admin_config_robo_painel_v81728():
    if request.method == "GET":
        try:
            cfg_email = _cfg_email_robo_v81727() if "_cfg_email_robo_v81727" in globals() else {}
        except Exception:
            cfg_email = {}
        try:
            cfg_robo = carregar_config_robo()
        except Exception:
            cfg_robo = {}
        payload = {
            "launcher": {
                "script_launcher_robo": str((cfg_robo or {}).get("script_launcher_robo", "")),
                "arquivo_config_robo": str(ARQUIVO_CONFIG_ROBO),
            },
            "outlook": {
                "remetente_filtro": str((cfg_robo or {}).get("remetente_filtro") or os.environ.get("TEMPO_FECHADO_REMETENTE_FILTRO") or ""),
                "assunto_filtro": str((cfg_robo or {}).get("assunto_filtro") or "Análise de controle de marcações"),
                "capturar_emails_lidos": bool((cfg_robo or {}).get("capturar_emails_lidos", False)),
                "limite_emails_outlook": _normalizar_limite_emails_outlook_v8216((cfg_robo or {}).get("limite_emails_outlook", 500)),
                "limite_emails_outlook_min": 50,
                "limite_emails_outlook_max": 5000,
                "politica_captura": "Incluir e-mails lidos recentes" if bool((cfg_robo or {}).get("capturar_emails_lidos", False)) else "Somente e-mails não lidos",
                "observacao": "A política de captura e o limite de e-mails são aplicados pelo Script_Launcher_v5_filtrado.py na próxima execução do robô.",
            },
            "email_robo": cfg_email,
            "arquivos": {
                "config_email_robo": str(ARQUIVO_CONFIG_EMAIL_ROBO_V81727) if "ARQUIVO_CONFIG_EMAIL_ROBO_V81727" in globals() else "",
                "config_robo": str(ARQUIVO_CONFIG_ROBO),
            },
        }
        return jsonify({"ok": True, "erro": "", "painel": payload})

    try:
        dados = request.get_json(silent=True) or {}
        email_robo = dados.get("email_robo", {})
        robo_cfg_payload = dados.get("robo", {}) if isinstance(dados.get("robo", {}), dict) else {}
        cfg_email = _salvar_cfg_email_robo_v81727(email_robo)
        limite_emails_outlook = _normalizar_limite_emails_outlook_v8216(robo_cfg_payload.get("limite_emails_outlook", 500))
        cfg_robo = salvar_config_robo({
            "capturar_emails_lidos": bool(robo_cfg_payload.get("capturar_emails_lidos", False)),
            "limite_emails_outlook": limite_emails_outlook,
            "remetente_filtro": str(robo_cfg_payload.get("remetente_filtro") or os.environ.get("TEMPO_FECHADO_REMETENTE_FILTRO") or "").strip(),
            "assunto_filtro": str(robo_cfg_payload.get("assunto_filtro") or "").strip() or "Análise de controle de marcações",
        })
        politica = "inclui e-mails lidos recentes" if cfg_robo.get("capturar_emails_lidos") else "somente e-mails não lidos"
        registrar_auditoria("ADMIN_CONFIG_ROBO_PAINEL", f"Configurações administrativas do robô atualizadas | Outlook: {politica} | limite={limite_emails_outlook}", "OK")
        return jsonify({
            "ok": True,
            "erro": "",
            "mensagem": "Configurações do robô salvas com sucesso.",
            "email_robo": cfg_email,
            "config_robo": cfg_robo,
            "arquivo": str(ARQUIVO_CONFIG_EMAIL_ROBO_V81727),
            "arquivo_config_robo": str(ARQUIVO_CONFIG_ROBO_USUARIO_V81743),
        })
    except Exception as e:
        return jsonify({"ok": False, "erro": f"Falha ao salvar painel de configuração do robô: {e}"}), 500

@app.route("/api/config-email-robo", methods=["GET", "POST"])
@login_obrigatorio
@admin_obrigatorio
def api_config_email_robo_v81727():
    if request.method == "GET":
        cfg = _cfg_email_robo_v81727()
        return jsonify({"ok": True, "erro": "", "config": cfg, "arquivo": str(ARQUIVO_CONFIG_EMAIL_ROBO_V81727)})
    try:
        dados = request.get_json(silent=True) or {}
        cfg = _salvar_cfg_email_robo_v81727(dados)
        registrar_auditoria("CONFIG_EMAIL_ROBO", str(ARQUIVO_CONFIG_EMAIL_ROBO_V81727), "OK")
        return jsonify({"ok": True, "erro": "", "mensagem": "Configuração de e-mail do robô salva com sucesso.", "config": cfg, "arquivo": str(ARQUIVO_CONFIG_EMAIL_ROBO_V81727)})
    except Exception as e:
        return jsonify({"ok": False, "erro": f"Falha ao salvar configuração de e-mail do robô: {e}"}), 500

@app.route("/api/notificacoes-jornada/status")
@login_obrigatorio
def api_notificacoes_jornada_status():
    cfg = _cfg_notificacoes_jornada()
    controle = _carregar_alertas_jornada_enviados()
    return jsonify({
        "erro": "",
        "email_ativo": bool(cfg.get("EMAIL_ATIVO")),
        "whatsapp_ativo": bool(cfg.get("WHATSAPP_ATIVO")),
        "enviar_ao_importar_excel": bool(cfg.get("ENVIAR_AO_IMPORTAR_EXCEL", True)),
        "alertas_ja_registrados": len(controle.get("ids", [])),
        "ultimo_historico": (controle.get("historico", [])[-1] if controle.get("historico") else None),
    })



@app.route("/api/notificacoes-jornada/config", methods=["GET", "POST"])
@login_obrigatorio
@admin_obrigatorio
def api_notificacoes_jornada_config():
    if request.method == "GET":
        cfg = _cfg_notificacoes_jornada()
        seguro = dict(cfg)
        if seguro.get("SMTP_SENHA"):
            seguro["SMTP_SENHA"] = "********"
        if seguro.get("WHATSAPP_TOKEN"):
            seguro["WHATSAPP_TOKEN"] = "********"
        seguro["_arquivo_config"] = str(ARQUIVO_CONFIG_NOTIFICACOES_JORNADA)
        return jsonify({"ok": True, "erro": "", "config": seguro})

    try:
        dados = request.get_json(silent=True) or {}
        atual = _cfg_notificacoes_jornada()
        if dados.get("SMTP_SENHA") == "********":
            dados["SMTP_SENHA"] = atual.get("SMTP_SENHA", "")
        if dados.get("WHATSAPP_TOKEN") == "********":
            dados["WHATSAPP_TOKEN"] = atual.get("WHATSAPP_TOKEN", "")
        cfg = _salvar_cfg_notificacoes_jornada(dados)
        registrar_auditoria("CONFIG_NOTIFICACOES_JORNADA", str(ARQUIVO_CONFIG_NOTIFICACOES_JORNADA), "OK")
        seguro = dict(cfg)
        if seguro.get("SMTP_SENHA"):
            seguro["SMTP_SENHA"] = "********"
        if seguro.get("WHATSAPP_TOKEN"):
            seguro["WHATSAPP_TOKEN"] = "********"
        seguro["_arquivo_config"] = str(ARQUIVO_CONFIG_NOTIFICACOES_JORNADA)
        return jsonify({"ok": True, "erro": "", "config": seguro, "mensagem": "Configuração de notificações salva com sucesso."})
    except Exception as e:
        return jsonify({"ok": False, "erro": f"Falha ao salvar configuração de notificações: {e}"}), 500


@app.route("/api/notificacoes-jornada/enviar-agora", methods=["POST", "GET"])
@login_obrigatorio
@admin_obrigatorio
def api_notificacoes_jornada_enviar_agora():
    try:
        forcar = str(request.args.get("forcar", "")).strip().lower() in {"1", "true", "sim", "yes"}
        registrar_auditoria("ENVIAR_ALERTAS_JORNADA", f"forcar_reenvio={forcar}", "INICIO")
        resultado = verificar_e_enviar_alertas_jornada(
            motivo="Envio manual pela interface web",
            forcar_reenvio=forcar,
        )
        registrar_auditoria(
            "ENVIAR_ALERTAS_JORNADA",
            f"novas={resultado.get('novas')} total={resultado.get('total_detectadas')}",
            "OK" if resultado.get("ok", True) else "ERRO",
        )
        return jsonify(resultado), (200 if resultado.get("ok", True) else 500)
    except Exception as e:
        msg = f"Falha inesperada ao enviar alertas de jornada: {e}"
        _log_notificacao_jornada(msg)
        return jsonify({
            "ok": False,
            "erro": msg,
            "mensagem": msg,
            "novas": 0,
            "total_detectadas": 0,
            "envios": [],
        }), 500


@app.route("/download/excel")
@login_obrigatorio
def download_excel():
    registrar_auditoria("DOWNLOAD_EXCEL", "Download da base ativa", "OK")
    if not ARQUIVO_PADRAO.exists():
        return "Arquivo não encontrado.", 404
    return send_file(ARQUIVO_PADRAO, as_attachment=True)


@app.route("/api/usuarios", methods=["GET"])
@login_obrigatorio
@admin_obrigatorio
def api_usuarios_listar():
    try:
        return jsonify({"ok": True, "usuarios": listar_usuarios(ARQUIVO_USUARIOS_DB)})
    except Exception as e:
        return jsonify({"ok": False, "erro": f"Falha ao listar usuários: {e}", "usuarios": []}), 500


@app.route("/api/usuarios", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_usuarios_salvar():
    payload = request.get_json(silent=True) or {}
    usuario = payload.get("usuario", "")
    nome = payload.get("nome", "")
    perfil = payload.get("perfil", "consulta")
    senha = payload.get("senha", "")
    ativo = bool(payload.get("ativo", True))
    trocar_senha = bool(payload.get("trocar_senha", True))

    ok, msg = criar_ou_atualizar_usuario(ARQUIVO_USUARIOS_DB, usuario, nome, perfil, senha, ativo, trocar_senha)
    registrar_auditoria("SALVAR_USUARIO", f"usuario={usuario}; perfil={perfil}; ativo={ativo}", "OK" if ok else "ERRO")
    status = 200 if ok else 400
    return jsonify({"ok": ok, "mensagem": msg, "erro": "" if ok else msg}), status


@app.route("/api/usuarios/<usuario>/status", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_usuarios_status(usuario):
    payload = request.get_json(silent=True) or {}
    ativo = bool(payload.get("ativo", True))
    if usuario == session.get("usuario") and not ativo:
        return jsonify({"ok": False, "erro": "Você não pode desativar o próprio usuário logado."}), 400
    ok, msg = definir_status_usuario(ARQUIVO_USUARIOS_DB, usuario, ativo)
    registrar_auditoria("ALTERAR_STATUS_USUARIO", f"usuario={usuario}; ativo={ativo}", "OK" if ok else "ERRO")
    status = 200 if ok else 400
    return jsonify({"ok": ok, "mensagem": msg, "erro": "" if ok else msg}), status


@app.route("/api/usuarios/<usuario>", methods=["DELETE"])
@login_obrigatorio
@admin_obrigatorio
def api_usuarios_excluir(usuario):
    usuario_normalizado = (usuario or "").strip().lower()
    if usuario_normalizado == (session.get("usuario") or "").strip().lower():
        return jsonify({"ok": False, "erro": "Voce nao pode excluir o proprio usuario logado."}), 400

    ok, msg = excluir_usuario(ARQUIVO_USUARIOS_DB, usuario_normalizado)
    registrar_auditoria("EXCLUIR_USUARIO", f"usuario={usuario_normalizado}", "OK" if ok else "ERRO")
    status = 200 if ok else 400
    return jsonify({"ok": ok, "mensagem": msg, "erro": "" if ok else msg}), status


@app.route("/api/usuarios/<usuario>/senha", methods=["POST"])
@login_obrigatorio
@admin_obrigatorio
def api_usuarios_senha_admin(usuario):
    payload = request.get_json(silent=True) or {}
    nova_senha = payload.get("nova_senha", "")
    ok, msg = alterar_senha_usuario(ARQUIVO_USUARIOS_DB, usuario, nova_senha, exigir_senha_atual=False)
    registrar_auditoria("REDEFINIR_SENHA_USUARIO", f"usuario={usuario}", "OK" if ok else "ERRO")
    status = 200 if ok else 400
    return jsonify({"ok": ok, "mensagem": msg, "erro": "" if ok else msg}), status


@app.route("/api/minha-senha", methods=["POST"])
@login_obrigatorio
def api_minha_senha():
    payload = request.get_json(silent=True) or {}
    usuario = session.get("usuario", "")
    ok, msg = alterar_senha_usuario(
        ARQUIVO_USUARIOS_DB,
        usuario,
        payload.get("nova_senha", ""),
        senha_atual=payload.get("senha_atual", ""),
        exigir_senha_atual=True,
    )
    if ok:
        session["trocar_senha"] = False
    registrar_auditoria("ALTERAR_MINHA_SENHA", "Troca de senha pelo próprio usuário", "OK" if ok else "ERRO")
    status = 200 if ok else 400
    return jsonify({"ok": ok, "mensagem": msg, "erro": "" if ok else msg}), status


@app.route("/api/auditoria")
@login_obrigatorio
@admin_obrigatorio
def api_auditoria():
    linhas = []
    if ARQUIVO_AUDITORIA_MULTIUSUARIO.exists():
        try:
            raw = ARQUIVO_AUDITORIA_MULTIUSUARIO.read_text(encoding="utf-8", errors="ignore").splitlines()
            for linha in raw[-200:]:
                partes = linha.split("	")
                while len(partes) < 8:
                    partes.append("")
                linhas.append({
                    "quando": partes[0],
                    "usuario": partes[1],
                    "nome": partes[2],
                    "perfil": partes[3],
                    "ip": partes[4],
                    "acao": partes[5],
                    "status": partes[6],
                    "detalhe": partes[7],
                })
        except Exception as e:
            return jsonify({"ok": False, "erro": f"Falha ao ler auditoria: {e}", "dados": []}), 500
    return jsonify({"ok": True, "erro": "", "dados": list(reversed(linhas))})


# ============================================================
# DIAGNÓSTICO CIRÚRGICO - BANCO DE HORAS (v8.10.15)
# ============================================================
def _canon_coluna_diag(coluna):
    texto = str(coluna or "").strip().lower()
    mapa = str.maketrans("áàãâéêíóôõúç", "aaaaeeiooouc")
    texto = texto.translate(mapa)
    texto = texto.replace("_", " ").replace("-", " ")
    texto = " ".join(texto.split())
    return texto


def _valor_diag(valor):
    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass
    texto = str(valor).strip()
    return "" if texto.lower() in {"nan", "none", "nat"} else texto


def _linha_raw_para_dict(linha):
    if linha is None:
        return {}
    return {str(col): _valor_diag(linha.get(col, "")) for col in linha.index}


def _filtrar_colunas_saldo_diag(row_dict):
    chaves = []
    for col in row_dict:
        c = _canon_coluna_diag(col)
        if any(tok in c for tok in ["saldo", "debito", "credito", "origem"]):
            chaves.append(col)
    return {col: row_dict.get(col, "") for col in chaves}


def _diagnosticar_pdf_para_linha(nome, data_linha, matricula=""):
    diag = {
        "pdfplumber_instalado": pdfplumber is not None,
        "pastas_verificadas": [str(p) for p in PASTAS_PDFS_ROBO],
        "pdfs_encontrados": 0,
        "quadros_lidos": 0,
        "match_encontrado": False,
        "pdf_escolhido": "",
        "saldo_anterior_pdf": "",
        "debito_bh_pdf": "",
        "credito_bh_pdf": "",
        "saldo_atual_pdf": "",
        "motivo": "",
        "candidatos_mesmo_nome_periodo": [],
        "matricula_usada": _normalizar_matricula_bh(matricula),
    }
    if pdfplumber is None:
        diag["motivo"] = "pdfplumber não está instalado; não foi possível ler PDFs."
        return diag
    try:
        pdfs = _listar_pdfs_robo_limitado(max_pdfs=MAX_PDFS_CORRECAO_SALDO)
        diag["pdfs_encontrados"] = len(pdfs)
        catalogo = []
        inicio = time.monotonic()
        for pdf in pdfs:
            if time.monotonic() - inicio > MAX_SEGUNDOS_CORRECAO_SALDO:
                break
            catalogo.extend(_extrair_espelhos_banco_horas_pdf(pdf))
        diag["quadros_lidos"] = len(catalogo)

        nome_norm = _canon_nome_banco(nome)
        mat_norm = _normalizar_matricula_bh(matricula)
        data_dt = _parse_data_pdf(data_linha)
        candidatos = []
        for item in catalogo:
            nj = item.get("nome_norm", "")
            match_matricula = bool(mat_norm and item.get("matricula_norm") and mat_norm == item.get("matricula_norm"))
            match_nome = _nomes_compativeis_bh(nome_norm, nj)
            if not (match_matricula or match_nome):
                continue
            ini, fim = item.get("inicio"), item.get("fim")
            if pd.isna(ini) or pd.isna(fim) or pd.isna(data_dt):
                continue
            if ini <= data_dt <= fim:
                datas = item.get("datas_lancadas") or []
                contem_data_real = bool(datas and data_dt in datas)
                candidatos.append({
                    "arquivo_pdf": item.get("arquivo_pdf", ""),
                    "match": "matricula" if match_matricula else "nome",
                    "nome_pdf": item.get("nome", ""),
                    "matricula_pdf": item.get("matricula", ""),
                    "periodo": f"{pd.to_datetime(ini).strftime('%d/%m/%Y')} a {pd.to_datetime(fim).strftime('%d/%m/%Y')}",
                    "emissao": "" if pd.isna(item.get("emissao")) else pd.to_datetime(item.get("emissao")).strftime("%d/%m/%Y"),
                    "datas_reais_pdf": bool(item.get("datas_reais_pdf")),
                    "contem_data_real": contem_data_real,
                    "data_max_lancada": "" if pd.isna(item.get("data_max_lancada")) else pd.to_datetime(item.get("data_max_lancada")).strftime("%d/%m/%Y"),
                    "saldo_anterior_pdf": item.get("saldo_anterior_pdf", ""),
                    "saldo_atual_pdf": item.get("saldo_atual_pdf", ""),
                })
        diag["candidatos_mesmo_nome_periodo"] = candidatos[:15]

        escolhido = _escolher_saldo_pdf_para_linha(nome, data_linha, catalogo, mat_norm)
        if escolhido:
            diag.update({
                "match_encontrado": True,
                "pdf_escolhido": escolhido.get("arquivo_pdf", ""),
                "saldo_anterior_pdf": escolhido.get("saldo_anterior_pdf", ""),
                "debito_bh_pdf": escolhido.get("debito_bh_pdf", ""),
                "credito_bh_pdf": escolhido.get("credito_bh_pdf", ""),
                "saldo_atual_pdf": escolhido.get("saldo_atual_pdf", ""),
                "motivo": "PDF escolhido por matrícula ou nome tolerante, respeitando período/data do espelho.",
            })
        else:
            diag["motivo"] = "Nenhum PDF casou com matrícula/nome + data pela regra atual. Se a matrícula estiver vazia, verifique se o Excel possui coluna Matrícula/Chapa; se houver candidatos sem data real, a leitura do corpo do PDF pode estar incompleta."
        return diag
    except Exception as e:
        diag["motivo"] = f"Falha no diagnóstico dos PDFs: {e}"
        return diag


# ============================================================
# CATÁLOGO DOS PDFs - BANCO DE HORAS (v8.10.18)
# ============================================================
def _fmt_data_catalogo(valor):
    try:
        if valor is None or pd.isna(valor):
            return ""
        return pd.to_datetime(valor).strftime("%d/%m/%Y")
    except Exception:
        return ""


def _montar_catalogo_pdfs_banco_horas(nome_filtro="", matricula_filtro="", data_filtro=""):
    catalogo_resp = {
        "pdfplumber_instalado": pdfplumber is not None,
        "pastas_verificadas": [str(p) for p in PASTAS_PDFS_ROBO],
        "pdfs_encontrados": 0,
        "quadros_lidos": 0,
        "nome_filtro": nome_filtro or "",
        "matricula_filtro": _normalizar_matricula_bh(matricula_filtro),
        "data_filtro": _fmt_data_catalogo(_parse_data_pdf(data_filtro)),
        "linhas": [],
        "erro": "",
    }
    if pdfplumber is None:
        catalogo_resp["erro"] = "pdfplumber não está instalado; não foi possível catalogar PDFs."
        return catalogo_resp

    nome_norm = _canon_nome_banco(nome_filtro)
    mat_norm = _normalizar_matricula_bh(matricula_filtro)
    data_dt = _parse_data_pdf(data_filtro)

    try:
        pdfs = _listar_pdfs_robo_limitado(max_pdfs=MAX_PDFS_CORRECAO_SALDO)
        catalogo_resp["pdfs_encontrados"] = len(pdfs)
        entradas = []
        inicio = time.monotonic()
        for pdf in pdfs:
            if time.monotonic() - inicio > MAX_SEGUNDOS_CORRECAO_SALDO:
                break
            try:
                entradas.extend(_extrair_espelhos_banco_horas_pdf(pdf))
            except Exception:
                continue
        catalogo_resp["quadros_lidos"] = len(entradas)

        linhas = []
        for item in entradas:
            datas = item.get("datas_lancadas") or []
            datas_fmt = [_fmt_data_catalogo(d) for d in datas if _fmt_data_catalogo(d)]
            match_nome = _nomes_compativeis_bh(nome_norm, item.get("nome_norm", "")) if nome_norm else False
            match_matricula = bool(mat_norm and item.get("matricula_norm") and mat_norm == item.get("matricula_norm"))
            dentro_periodo = False
            contem_data_real = False
            ini, fim = item.get("inicio"), item.get("fim")
            if not pd.isna(data_dt) and not pd.isna(ini) and not pd.isna(fim):
                dentro_periodo = bool(ini <= data_dt <= fim)
                contem_data_real = bool(datas and data_dt in datas)

            linhas.append({
                "arquivo_pdf": item.get("arquivo_pdf", ""),
                "caminho_pdf": item.get("caminho_pdf", ""),
                "nome_pdf": item.get("nome", ""),
                "nome_norm_pdf": item.get("nome_norm", ""),
                "matricula_pdf": item.get("matricula", ""),
                "matricula_norm_pdf": item.get("matricula_norm", ""),
                "periodo": f"{_fmt_data_catalogo(ini)} a {_fmt_data_catalogo(fim)}",
                "inicio": _fmt_data_catalogo(ini),
                "fim": _fmt_data_catalogo(fim),
                "emissao": _fmt_data_catalogo(item.get("emissao")),
                "datas_reais_pdf": bool(item.get("datas_reais_pdf")),
                "qtd_datas_lancadas": len(datas_fmt),
                "primeira_data_lancada": datas_fmt[0] if datas_fmt else "",
                "ultima_data_lancada": datas_fmt[-1] if datas_fmt else "",
                "datas_lancadas_amostra": ", ".join(datas_fmt[:10]) + ("..." if len(datas_fmt) > 10 else ""),
                "saldo_anterior_pdf": item.get("saldo_anterior_pdf", ""),
                "debito_bh_pdf": item.get("debito_bh_pdf", ""),
                "credito_bh_pdf": item.get("credito_bh_pdf", ""),
                "saldo_atual_pdf": item.get("saldo_atual_pdf", ""),
                "match_nome": match_nome,
                "match_matricula": match_matricula,
                "dentro_periodo_data": dentro_periodo,
                "contem_data_real": contem_data_real,
                "score_debug": (100 if match_matricula else 0) + (50 if match_nome else 0) + (20 if dentro_periodo else 0) + (30 if contem_data_real else 0),
            })

        linhas.sort(key=lambda x: (
            -int(x.get("score_debug", 0)),
            x.get("nome_pdf", ""),
            x.get("inicio", ""),
            x.get("arquivo_pdf", ""),
        ))
        catalogo_resp["linhas"] = linhas[:300]
        return catalogo_resp
    except Exception as e:
        catalogo_resp["erro"] = f"Falha ao montar catálogo dos PDFs: {e}"
        return catalogo_resp


@app.route("/api/catalogo-pdfs-banco-horas")
@login_obrigatorio
def api_catalogo_pdfs_banco_horas():
    nome = request.args.get("nome", "").strip()
    matricula = request.args.get("matricula", "").strip()
    data = request.args.get("data", "").strip()
    dados = _montar_catalogo_pdfs_banco_horas(nome, matricula, data)
    status = 200 if not dados.get("erro") else 500
    return jsonify({"ok": not bool(dados.get("erro")), "erro": dados.get("erro", ""), "dados": dados}), status


@app.route("/api/diagnostico-banco-horas")
@login_obrigatorio
def api_diagnostico_banco_horas():
    nome_req = request.args.get("nome", "").strip().lower()
    if not nome_req:
        return jsonify({"ok": False, "erro": "Selecione um colaborador para diagnosticar.", "dados": {}}), 400

    if not ARQUIVO_PADRAO.exists():
        return jsonify({"ok": False, "erro": "Excel ativo não encontrado na pasta data.", "dados": {}}), 404

    try:
        aba_diagnostico = _nome_aba_historico_v6(ARQUIVO_PADRAO) or "Consolidado"
        df_raw = pd.read_excel(ARQUIVO_PADRAO, sheet_name=aba_diagnostico)
        if "Nome" not in df_raw.columns or "Data" not in df_raw.columns:
            return jsonify({"ok": False, "erro": f"Aba {aba_diagnostico} não possui colunas Nome/Data.", "dados": {}}), 400

        nomes_lower = df_raw["Nome"].astype(str).str.lower()
        df_raw_nome = df_raw[nomes_lower == nome_req].copy()
        if df_raw_nome.empty:
            df_raw_nome = df_raw[nomes_lower.str.contains(nome_req, na=False)].copy()
        if df_raw_nome.empty:
            return jsonify({"ok": False, "erro": "Colaborador não encontrado no Excel bruto.", "dados": {}}), 404

        nome_real = str(df_raw_nome["Nome"].iloc[0])
        df_raw_nome = df_raw[df_raw["Nome"].astype(str) == nome_real].copy()
        df_raw_nome["data_dt"] = pd.to_datetime(df_raw_nome["Data"].astype(str), format="%d/%m/%Y", errors="coerce")
        df_raw_nome = df_raw_nome[df_raw_nome["data_dt"].notna()].sort_values("data_dt")
        if df_raw_nome.empty:
            return jsonify({"ok": False, "erro": "Colaborador sem datas válidas no Excel bruto.", "dados": {}}), 400

        inicio = _data_inicio_medicao(df_raw_nome["data_dt"].max()) or df_raw_nome["data_dt"].min()
        fim = _data_fim_medicao(inicio) or df_raw_nome["data_dt"].max()
        ciclo_raw = df_raw_nome[(df_raw_nome["data_dt"] >= inicio) & (df_raw_nome["data_dt"] <= fim)].copy()
        if ciclo_raw.empty:
            ciclo_raw = df_raw_nome.copy()
        row_raw = ciclo_raw.iloc[0]
        row_raw_dict = _linha_raw_para_dict(row_raw.drop(labels=["data_dt"], errors="ignore"))

        df_norm = normalizar_colunas(df_raw)
        df_norm_nome = df_norm[df_norm["nome"].astype(str) == nome_real].copy() if "nome" in df_norm.columns else pd.DataFrame()
        if not df_norm_nome.empty:
            df_norm_nome["data_dt"] = pd.to_datetime(df_norm_nome["data"].astype(str), format="%d/%m/%Y", errors="coerce")
            df_norm_nome = df_norm_nome[df_norm_nome["data_dt"].notna()].sort_values("data_dt")
            ciclo_norm = df_norm_nome[(df_norm_nome["data_dt"] >= inicio) & (df_norm_nome["data_dt"] <= fim)].copy()
            if ciclo_norm.empty:
                ciclo_norm = df_norm_nome.copy()
            row_norm = ciclo_norm.iloc[0] if not ciclo_norm.empty else None
        else:
            row_norm = None

        data_base = pd.to_datetime(row_raw["data_dt"]).strftime("%d/%m/%Y")
        matricula_base = _matricula_da_linha_bh(row_raw)
        pdf_diag = _diagnosticar_pdf_para_linha(nome_real, data_base, matricula_base)

        # v8.10.32: o diagnóstico passa a separar claramente duas coisas:
        # 1) PDF base do Saldo Calculado = fechamento do ciclo anterior;
        # 2) PDF da fotografia atual = saldo informado do ciclo corrente.
        try:
            cache_diag_bh = carregar_cache_pdfs_banco_horas()
            catalogo_diag_bh = cache_diag_bh.get("catalogo", []) if cache_diag_bh else []
        except Exception:
            catalogo_diag_bh = []
        pdf_base_diag = _buscar_pdf_fechamento_ciclo_anterior(nome_real, matricula_base, inicio, catalogo_diag_bh, permitir_busca_direta=True)
        pdf_corrente_diag = _escolher_saldo_pdf_para_linha(nome_real, data_base, catalogo_diag_bh, matricula_base)

        dados = {
            "nome": nome_real,
            "matricula_base": matricula_base,
            "data_base": data_base,
            "inicio_ciclo": pd.to_datetime(inicio).strftime("%d/%m/%Y"),
            "fim_ciclo": pd.to_datetime(fim).strftime("%d/%m/%Y"),
            "arquivo_excel": str(ARQUIVO_PADRAO),
            "colunas_saldo_no_excel": [str(c) for c in df_raw.columns if any(tok in _canon_coluna_diag(c) for tok in ["saldo", "debito", "credito", "origem"])],
            "linha_bruta_saldos": _filtrar_colunas_saldo_diag(row_raw_dict),
            "linha_normalizada": {
                "saldo_atual": _valor_diag(row_norm.get("saldo_atual", "")) if row_norm is not None else "",
                "saldo": _valor_diag(row_norm.get("saldo", "")) if row_norm is not None else "",
                "saldo_anterior": _valor_diag(row_norm.get("saldo_anterior", "")) if row_norm is not None else "",
                "debito_bh": _valor_diag(row_norm.get("debito_bh", "")) if row_norm is not None else "",
                "credito_bh": _valor_diag(row_norm.get("credito_bh", "")) if row_norm is not None else "",
                "origem_saldo_importado": _valor_diag(row_norm.get("origem_saldo_importado", "")) if row_norm is not None else "",
                "matricula_normalizada_excel": _normalizar_matricula_bh(row_norm.get("matricula", "")) if row_norm is not None else "",
            },
            "pdf_base_ciclo_anterior": {
                "pdf": (pdf_base_diag or {}).get("arquivo_pdf", ""),
                "periodo": ((pd.to_datetime((pdf_base_diag or {}).get("inicio")).strftime("%d/%m/%Y") + " a " + pd.to_datetime((pdf_base_diag or {}).get("fim")).strftime("%d/%m/%Y")) if pdf_base_diag is not None and not pd.isna((pdf_base_diag or {}).get("inicio")) and not pd.isna((pdf_base_diag or {}).get("fim")) else ""),
                "saldo_atual_pdf": (pdf_base_diag or {}).get("saldo_atual_pdf", ""),
                "match": (pdf_base_diag or {}).get("_match_base_pdf_v25", ""),
            },
            "pdf_fotografia_ciclo_atual": {
                "pdf": (pdf_corrente_diag or {}).get("arquivo_pdf", ""),
                "periodo": ((pd.to_datetime((pdf_corrente_diag or {}).get("inicio")).strftime("%d/%m/%Y") + " a " + pd.to_datetime((pdf_corrente_diag or {}).get("fim")).strftime("%d/%m/%Y")) if pdf_corrente_diag is not None and not pd.isna((pdf_corrente_diag or {}).get("inicio")) and not pd.isna((pdf_corrente_diag or {}).get("fim")) else ""),
                "saldo_atual_pdf": (pdf_corrente_diag or {}).get("saldo_atual_pdf", ""),
                "match": (pdf_corrente_diag or {}).get("_match_v21", ""),
            },
            "diagnostico_pdf": pdf_diag,
            "leitura": "Compare 'linha_bruta_saldos' com 'diagnostico_pdf'. Se Saldo Atual PDF estiver correto, mas linha_normalizada continuar vindo de saldo_atual_coluna_exata, o problema está na prioridade do Extrato. Se Saldo Atual PDF vier errado/vazio, o problema está na associação/leitura do PDF.",
        }
        return jsonify({"ok": True, "erro": "", "dados": dados})
    except Exception as e:
        return jsonify({"ok": False, "erro": f"Falha no diagnóstico do Banco de Horas: {e}", "dados": {}}), 500



@app.route("/api/score-operacional")
@login_obrigatorio
def api_score_operacional():
    """v8.16.22 - Score Operacional calculado."""
    cc_filtro = request.args.get("cc", "").strip()
    nome_filtro = request.args.get("nome", "").strip().lower()
    turno_filtro = request.args.get("turno", "").strip().lower()
    data_filtro = request.args.get("data", "").strip()
    datas_filtro = request.args.get("datas", "").strip()
    cache_key = (
        str(_assinatura_arquivo(ARQUIVO_PADRAO)),
        _assinatura_config_bh2(),
        cc_filtro,
        nome_filtro,
        turno_filtro,
        data_filtro,
        datas_filtro,
    )
    try:
        if (
            cache_key[0] != "None"
            and _CACHE_SCORE_OPERACIONAL_MEM.get("key") == cache_key
            and _CACHE_SCORE_OPERACIONAL_MEM.get("payload") is not None
        ):
            return jsonify(_CACHE_SCORE_OPERACIONAL_MEM["payload"])
    except Exception:
        pass

    df, erro = ler_consolidado()
    if erro:
        return jsonify({"erro": erro, "kpis": {}, "melhores": [], "piores": [], "por_cc": [], "por_turno": [], "colaboradores": []})

    df = normalizar_colunas(df)
    if df.empty:
        return jsonify({"erro": "Nenhuma base carregada.", "kpis": {}, "melhores": [], "piores": [], "por_cc": [], "por_turno": [], "colaboradores": []})

    df = _aplicar_filtros_executivos(df)
    if df.empty:
        return jsonify({
            "erro": "",
            "kpis": {"score_medio": 0, "excelente": 0, "controlado": 0, "atencao": 0, "critico": 0, "intervencao": 0, "colaboradores": 0},
            "melhores": [], "piores": [], "por_cc": [], "por_turno": [], "colaboradores": [],
        })

    df = _garantir_colunas_calculadas(df)
    df["texto_score"] = (
        df.get("status", "").astype(str).fillna("") + " " + df.get("inconsistencias", "").astype(str).fillna("")
    ).str.strip()

    try:
        mapa_bh2_nome, mapa_bh2_mat = _mapa_saldos_bh2_para_dashboard(
            data=request.args.get("data", "").strip(),
            datas=request.args.get("datas", "").strip(),
        )
    except Exception:
        mapa_bh2_nome, mapa_bh2_mat = {}, {}

    def classificar_score(score):
        score = int(score or 0)
        if score >= 90:
            return "Excelente", "success"
        if score >= 75:
            return "Controlado", "info"
        if score >= 60:
            return "Atenção", "warning"
        if score >= 40:
            return "Crítico", "danger"
        return "Intervenção Imediata", "danger"

    colaboradores = []
    for nome_colab, g in df.groupby("nome", dropna=False):
        nome_ref = str(nome_colab or "").strip()
        try:
            matricula_ref = str(g["matricula"].iloc[-1]).strip() if "matricula" in g.columns else ""
        except Exception:
            matricula_ref = ""

        saldo_min = None
        try:
            item_bh2 = None
            if matricula_ref and matricula_ref in mapa_bh2_mat:
                item_bh2 = mapa_bh2_mat.get(matricula_ref)
            else:
                item_bh2 = mapa_bh2_nome.get(_norm_bh2(nome_ref))
            if item_bh2:
                saldo_min = int(item_bh2.get("saldo_min", 0) or 0)
        except Exception:
            saldo_min = None

        if saldo_min is None:
            try:
                saldo_info = _calcular_saldo_base_medicao_por_colaborador(g)
                saldo_min = int(saldo_info.get("saldo_calculado_min", 0) or 0)
            except Exception:
                saldo_min = 0

        he_min = int(g["he_min"].sum()) if "he_min" in g.columns else 0
        absent_min = int(g["absent_min"].sum()) if "absent_min" in g.columns else 0
        dias_he = int((g["he_min"] > 0).sum()) if "he_min" in g.columns else 0
        dias_abs = int((g["absent_min"] > 0).sum()) if "absent_min" in g.columns else 0
        texto = " ".join(g["texto_score"].astype(str).tolist()).lower()

        penalidades = []
        score = 100

        if saldo_min >= 2400:
            penalidades.append({"motivo": "Banco acima de 40h", "pontos": -25, "detalhe": _minutos_para_saldo_virgula(saldo_min)})
            score -= 25
        elif saldo_min >= 1200:
            penalidades.append({"motivo": "Banco acima de 20h", "pontos": -10, "detalhe": _minutos_para_saldo_virgula(saldo_min)})
            score -= 10
        elif saldo_min < 0:
            penalidades.append({"motivo": "Banco negativo", "pontos": -15, "detalhe": _minutos_para_saldo_virgula(saldo_min)})
            score -= 15

        if he_min >= 600 or dias_he >= 3:
            penalidades.append({"motivo": "Horas extras elevadas", "pontos": -10, "detalhe": f"{minutos_para_hhmm(he_min)} em {dias_he} dia(s)"})
            score -= 10

        if absent_min >= 480 or dias_abs >= 2:
            penalidades.append({"motivo": "Ausência recorrente", "pontos": -15, "detalhe": f"{minutos_para_hhmm(absent_min)} em {dias_abs} dia(s)"})
            score -= 15

        tem_inconsistencia = False
        try:
            status_series = g["status"].astype(str).str.upper() if "status" in g.columns else pd.Series([], dtype=str)
            inc_series = g["inconsistencias"].astype(str).str.strip() if "inconsistencias" in g.columns else pd.Series([], dtype=str)
            if len(status_series) and len(inc_series):
                tem_inconsistencia = bool(((status_series != "OK") | (inc_series != "")).sum())
            elif len(status_series):
                tem_inconsistencia = bool((status_series != "OK").sum())
            elif len(inc_series):
                tem_inconsistencia = bool((inc_series != "").sum())
        except Exception:
            tem_inconsistencia = bool(str(texto).strip())

        if tem_inconsistencia:
            penalidades.append({"motivo": "Inconsistências para revisão", "pontos": -10, "detalhe": "Há registros com alerta/status"})
            score -= 10

        if any(p in texto for p in ["sem marca", "sem marcação", "ponto aberto", "marcação incompleta", "batida impar", "batida ímpar"]):
            penalidades.append({"motivo": "Ponto aberto ou sem marcação", "pontos": -15, "detalhe": "Há indicação de marcação incompleta"})
            score -= 15

        if any(p in texto for p in ["inter jornada", "interjornada", "violação", "violacao", "jornada"]):
            penalidades.append({"motivo": "Violação de jornada", "pontos": -20, "detalhe": "Há indicação de jornada sensível"})
            score -= 20

        score = max(0, min(100, int(score)))
        faixa, classe = classificar_score(score)

        colaboradores.append({
            "nome": nome_ref,
            "matricula": matricula_ref,
            "cc": valor_cc_display(g["cc"].iloc[-1]) if "cc" in g.columns and len(g) else "",
            "funcao": str(g["funcao"].iloc[-1]) if "funcao" in g.columns and len(g) else "",
            "turno": str(g["turno"].iloc[-1]) if "turno" in g.columns and len(g) else "",
            "score": score,
            "faixa": faixa,
            "classe": classe,
            "saldo": _minutos_para_saldo_virgula(saldo_min),
            "he": minutos_para_hhmm(he_min),
            "ausencia": minutos_para_hhmm(absent_min),
            "penalidades": penalidades,
        })

    colaboradores = sorted(colaboradores, key=lambda x: x["score"], reverse=True)
    scores = [c["score"] for c in colaboradores]
    score_medio = round(sum(scores) / max(len(scores), 1), 1) if scores else 0

    def media_grupo(campo):
        grupos = {}
        for c in colaboradores:
            chave = str(c.get(campo) or "-")
            grupos.setdefault(chave, []).append(int(c.get("score", 0)))
        linhas = []
        for chave, vals in grupos.items():
            linhas.append({"grupo": chave, "score_medio": round(sum(vals) / max(len(vals), 1), 1), "colaboradores": len(vals)})
        return sorted(linhas, key=lambda x: x["score_medio"])

    kpis = {
        "score_medio": score_medio,
        "colaboradores": len(colaboradores),
        "excelente": sum(1 for c in colaboradores if c["score"] >= 90),
        "controlado": sum(1 for c in colaboradores if 75 <= c["score"] < 90),
        "atencao": sum(1 for c in colaboradores if 60 <= c["score"] < 75),
        "critico": sum(1 for c in colaboradores if 40 <= c["score"] < 60),
        "intervencao": sum(1 for c in colaboradores if c["score"] < 40),
    }

    payload = {
        "erro": "",
        "kpis": kpis,
        "melhores": colaboradores[:10],
        "piores": sorted(colaboradores, key=lambda x: x["score"])[:10],
        "por_cc": media_grupo("cc"),
        "por_turno": media_grupo("turno"),
        "colaboradores": colaboradores[:80],
    }
    try:
        if cache_key[0] != "None":
            _CACHE_SCORE_OPERACIONAL_MEM["key"] = cache_key
            _CACHE_SCORE_OPERACIONAL_MEM["payload"] = payload
            _CACHE_SCORE_OPERACIONAL_MEM["gerado_em_ts"] = time.monotonic()
    except Exception:
        pass
    return jsonify(payload)


if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5050, threaded=True)
